from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app import app, build_teaching_week_schedule, get_level_syllabus_lessons_per_week
from models import Level, SyllabusPlanEntry, db


WORKBOOK_PATH = ROOT_DIR / "Weekly Yearly Curriculum plan .xlsx"
SHEET_NAME = "Arabic"

PRIMARY_PLAN_CONFIG = {
    "Primary Beginner": {
        "level_id": 2,
        "column": 5,
        "book_name": "العربية بين يدي أولادنا - المسار الابتدائي المبتدئ",
    },
    "Primary Intermediate": {
        "level_id": 3,
        "column": 6,
        "book_name": "المدينة للقراءة 2 - المسار الابتدائي المتوسط",
    },
    "Primary Advance": {
        "level_id": 4,
        "column": 7,
        "book_name": "المدينة للقراءة 3 - المسار الابتدائي المتقدم",
    },
}

PRIMARY_WEEK_TITLE_OVERRIDES = {
    "Primary Beginner": {
        1: "الوحدة 1 - التحية والتعارف",
        2: "الوحدة 2 - الوداع وصيغة المؤنث",
        3: "الوحدة 3 - السؤال عن الحال",
        4: "مراجعة الوحدات 1-3",
        5: "الوحدة 4 - السؤال عن الحال للمؤنث",
        6: "الوحدة 4 - التحية والتعارف الشفهي",
        7: "الوحدة 5 - الأعداد 1-5 وفعل هات",
        8: "مراجعة الوحدة 5",
        9: "الوحدة 6 - الأعداد 6-10 وأدوات السؤال",
        10: "الوحدة 7 - كم حقيبة؟",
        11: "الوحدة 8 - أين وهنا وقليل وكثير",
        12: "الوحدة 8 - مراجعة الأعداد والمكان",
        13: "مراجعة عامة",
        14: "مراجعة عامة",
        15: "الكتاب 2 - الوحدة 1 - التعريف بالنفس والعمر",
        16: "الكتاب 2 - الوحدة 2 - الأعداد الترتيبية للمذكر",
        17: "الكتاب 2 - الوحدة 3 - الأعداد الترتيبية للمؤنث",
        18: "مراجعة الكتاب 2 - الوحدات 1-3",
        19: "الكتاب 2 - الوحدة 4 - الاسم والعمر وأسماء الإشارة",
        20: "الكتاب 2 - الوحدة 5 - أين ومتى؟",
        21: "الكتاب 2 - الوحدة 6 - المفردات والأسئلة القصيرة",
        22: "مراجعة",
        23: "الكتاب 2 - الوحدة 7 - الأسرة والتحية",
        24: "الكتاب 2 - الوحدة 8 - أيام الأسبوع",
        25: "مراجعة",
        26: "المستوى B2 - الوحدة 1 - الألوان للمذكر",
        27: "المستوى B2 - الوحدة 1 - الشدة والألوان",
        28: "المستوى B2 - الوحدة 2 - الألوان للمؤنث",
        29: "المستوى B2 - الوحدة 2 - التنوين وصياغة السؤال",
        30: "المستوى B2 - الوحدة 3 - الاتجاهات والصفات",
        31: "المستوى B2 - الوحدة 3 - الهمزة وأدوات النفي",
        32: "مراجعة",
        33: "المستوى B2 - الوحدة 4 - الأدوات المدرسية",
        34: "المستوى B2 - الوحدة 4 - أدوات الاستفهام والصفات",
        35: "المستوى B2 - الوحدة 5 - الطائر",
        36: "المستوى B2 - الوحدة 5 - حروف الجر والمدود",
        37: "المستوى B2 - الوحدة 6 - المفردات والأعداد الترتيبية",
        38: "المستوى B2 - الوحدة 6 - الياء والمد",
        39: "المستوى B2 - الوحدة 7 - القطة والظروف المكانية",
        40: "المستوى B2 - الوحدة 8 - المفردات والمراجعة التطبيقية",
        41: "مراجعة نهائية وتقييم ختامي",
    },
    "Primary Intermediate": {
        1: "التهيئة وأهمية اللغة العربية",
        2: "الألوان وحروف الجر",
        3: "المفردات والجموع",
        4: "المذكر والمؤنث وبناء الجملة",
        5: "الصفات وحروف الجر والبيت",
        6: "الضمائر المتصلة وتصريف الفعل",
        7: "الجموع والتدريب الكتابي",
        8: "بناء الجملة والمفردات",
        9: "أسماء الإشارة والضمائر",
        10: "المذكر والمؤنث والصفات",
        11: "الجموع وأسماء الإشارة",
        12: "الأسرة واللام الشمسية والقمرية",
        13: "تدريب على الجموع والتراكيب",
        14: "السفر والعمرة وأدوات الاستفهام",
        15: "تثبيت الجموع والضمائر",
        16: "الضمائر والإضافة والمراجعة",
        17: "الصف الدراسي وأسماء الإشارة",
        18: "الحيوان المفضل والفهم القرائي",
        19: "الجملة الاسمية والجملة الفعلية",
        20: "المبتدأ والخبر",
        21: "العيد والتدريب الكتابي",
        22: "العيد والتطبيقات",
        23: "المثنى وبدايات الإعراب",
        24: "الأعداد والقراءة",
        25: "المفردات والجموع",
        26: "الأعداد وبناء الجمل",
        27: "التراكيب الأساسية",
        28: "الأعداد 3-10 والمعدود",
        29: "أنشطة الأعداد واختبار إملائي",
        30: "المدرسة والمراجعة",
        31: "أسماء البلدان والمفردات",
        32: "الجمل والتراكيب",
        33: "الفهم القرائي والنشاط",
        34: "الممنوع من الصرف والإعراب",
        35: "الجموع والتطبيق",
        36: "الفهم القرائي والنشاط",
        37: "الجار والمجرور",
        38: "الإضافة",
        39: "الأعداد 3-10 والإعراب",
        40: "مراجعة وعروض ومشاريع",
        41: "مراجعة نهائية وتقييم ختامي",
    },
    "Primary Advance": {
        1: "التهيئة وأهمية اللغة العربية",
        2: "الشراب والتعبيرات الزمنية",
        3: "هاشم والمدرس",
        4: "المفردات الوصفية والفهم",
        5: "إن وأخواتها",
        6: "المبتدأ والخبر وإن",
        7: "لعل والأعداد الكبيرة",
        8: "ذو وذوات وأسلوب أ...أم",
        9: "ليس والمفرد والجمع",
        10: "تدريب على ليس وإن",
        11: "أحمد وعلي واسم التفضيل",
        12: "تدريب على اسم التفضيل",
        13: "لكن وكأن",
        14: "الأعداد 11-20",
        15: "الأعداد الترتيبية",
        16: "أليس كذلك؟ بلى",
        17: "أيهما والجموع",
        18: "إبراهيم مع خاله يوسف والفعل الماضي",
        19: "لأن وأدوات الجواب",
        20: "المدرس والتلاميذ والفاعل والمفعول",
        21: "تدريب على الفاعل والمفعول",
        22: "سعيد ومريم وأمهما وتأنيث الفاعل",
        23: "تاء الفاعل وتاء التأنيث",
        24: "أن وأظن والأعداد",
        25: "أوزان فعلان وفعلى وفعال",
        26: "الاستفهام بلم",
        27: "حالات الإعراب",
        28: "فعل هات مع الضمائر",
        29: "مراجعة",
        30: "الأب وأبناؤه",
        31: "تحويل الجمل إلى المؤنث",
        32: "كان والجملة الاسمية وظرف الزمان",
        33: "أنتم مع الفعل الماضي وذو",
        34: "الاستفهام بـ أ...أم",
        35: "كتابة الكسور بالحروف",
        36: "الفعل مع المذكر والمؤنث",
        37: "مراجعة",
        38: "الفعل: ماض ومضارع وأمر",
        39: "الأعداد 21-30 والتنوين",
        40: "مراجعة وعروض ومشاريع",
        41: "مراجعة نهائية وتقييم ختامي",
    },
}

TOPIC_KEYWORDS = [
    (("الشّراب", "drinks and time expressions"), "الشراب والتعبيرات الزمنية"),
    (("هاشم", "المدرّس"), "هاشم والمدرس"),
    (("ذكيّ", "مجتهد", "زميل"), "المفردات الوصفية"),
    (("greetings", "introductions", "self-introduction", "مع السلامة"), "التحية والتعارف"),
    (("كيف حالك", "wellbeing"), "السؤال عن الحال"),
    (("كم حقيبة", "حقيبة"), "الأعداد والحقائب"),
    (("numbers 1", "numbers 1–5", "هاتِ"), "الأعداد الأولى"),
    (("numbers 6", "numbers 6–10", "feminine numbers"), "الأعداد والاستفهام"),
    (("family members", "family", "my family"), "الأسرة"),
    (("days of the week", "متى"), "أيام الأسبوع"),
    (("colours (masculine)", "colors (masculine)", "colours", "colors"), "الألوان"),
    (("colours (feminine)", "colors (feminine)"), "الألوان المؤنثة"),
    (("directions", "adjectives"), "الاتجاهات والصفات"),
    (("school supplies", "my classroom"), "الأدوات المدرسية"),
    (("dialogue about a bird", "a bird"), "الطائر"),
    (("dialogue about a cat", " قطة", "قطة،"), "القطة"),
    (("where", "أين"), "أدوات المكان"),
    (("ordinal numbers",), "الأعداد الترتيبية"),
    (("demonstrative", "هذا", "هذه"), "أسماء الإشارة"),
    (("colors", "prepositions"), "الألوان وحروف الجر"),
    (("plurals", "broken plurals"), "الجموع"),
    (("sun/moon letters",), "اللام الشمسية والقمرية"),
    (("travel", "umrah"), "السفر والعمرة"),
    (("nominal sentence", "verbal sentence"), "الجملة الاسمية والفعلية"),
    (("subject+predicate", "subject+predicate"), "المبتدأ والخبر"),
    (("dual", "dual form"), "المثنى"),
    (("i‘raab", "i'rab", "i‘raab"), "الإعراب"),
    (("mamnu",), "الممنوع من الصرف"),
    (("jarr majroor",), "الجار والمجرور"),
    (("mudhaaf",), "الإضافة"),
    (("inna", "إنّ"), "إن وأخواتها"),
    (("لعل",), "لعل وأخوات إن"),
    (("ليس",), "ليس والجملة الاسمية"),
    (("ذو، ذوو، ذات، ذوات", "ذوو", "ذوات"), "ذو وذوات"),
    (("هشام وبلال",), "هشام وبلال"),
    (("أحمد وعلي",), "أحمد وعلي"),
    (("إبراهيم مع خاله يوسف",), "إبراهيم مع خاله يوسف"),
    (("المدرّس والتلاميذ",), "المدرس والتلاميذ"),
    (("سعيد ومريم",), "سعيد ومريم وأمهما"),
    (("comparative", "superlative", "صيغة التفضيل"), "اسم التفضيل"),
    (("lakinna", "لكنّ"), "لكن وأخوات إن"),
    (("ka’anna", "كأن"), "كأن وأخوات إن"),
    (("11–20", "11-20"), "الأعداد 11-20"),
    (("21–30", "21-30"), "الأعداد 21-30"),
    (("أليس كذلك",), "أسلوب أليس كذلك"),
    (("أيّهما",), "أسلوب أيهما"),
    (("الفاعل والمفعول",), "الفاعل والمفعول به"),
    (("تأنيث الفاعل",), "تأنيث الفاعل"),
    (("الكسور", "fractions"), "الكسور بالحروف"),
    (("past, present-future and imperative", "three forms of the verb"), "أزمنة الفعل وصيغه"),
    (("tanween",), "التنوين"),
]

COMMON_REPLACEMENTS = [
    ("Final Full Review. Final assessment for the academic year.", "مراجعة ختامية شاملة وتقييم نهائي للعام الدراسي."),
    ("Final Review. Final assessment for the academic year.", "مراجعة ختامية وتقييم نهائي للعام الدراسي."),
    ("Final Review", "مراجعة ختامية"),
    ("Revision+ Oral tests+ Presentations/Projects", "مراجعة عامة واختبارات شفهية وعروض أو مشاريع."),
    ("Review", "مراجعة"),
    ("Revision", "مراجعة"),
    ("Icebreaking", "تهيئة وكسر الجليد"),
    ("Who am I?", "من أنا؟"),
    ("Introduce yourselves", "تعارف وتمرين تقديم النفس"),
    ("ice breaker games", "ألعاب تعارف وتمهيد"),
    ("Importance of the Arabic Language", "أهمية اللغة العربية"),
    ("History of Arabic Grammar", "تاريخ النحو العربي"),
    ("The importance of the Arabic language", "أهمية اللغة العربية"),
    ("The purpose of learning the Arabic language", "أهداف تعلم اللغة العربية"),
    ("Practice on some Arabic constructions", "تدريب على بعض التراكيب العربية"),
    ("Read and translate", "قراءة وترجمة"),
    ("Reading and Translating", "قراءة وترجمة"),
    ("Read and identify", "قراءة وتمييز"),
    ("Read and practise a dialogue", "قراءة الحوار والتدرّب عليه"),
    ("Read and practise", "قراءة والتدرّب"),
    ("Read and practise a dialogue", "قراءة الحوار والتدرّب عليه"),
    ("Read the main dialogue and learn vocabulary", "قراءة الحوار الرئيس وتعلّم المفردات"),
    ("Read a conversation and practise", "قراءة الحوار والتدرّب"),
    ("Read a conversation", "قراءة الحوار"),
    ("Read dialogue", "قراءة الحوار"),
    ("Recognise short dialogues", "التعرّف على الحوارات القصيرة"),
    ("Recognise", "التعرّف على"),
    ("Practise", "التدرّب على"),
    ("Practice", "التدرّب على"),
    ("Write", "كتابة"),
    ("Speak using", "التحدث باستخدام"),
    ("Introduce", "التعريف بـ"),
    ("Comprehend", "فهم"),
    ("Consolidate", "ترسيخ"),
    ("Count objects", "عدّ الأشياء"),
    ("Use", "استخدام"),
    ("Identify", "تحديد"),
    ("Differentiate between", "التمييز بين"),
    ("Differentiate", "التمييز بين"),
    ("Listen to", "الاستماع إلى"),
    ("Sentence formation", "تكوين الجمل"),
    ("Word meanings", "معاني الكلمات"),
    ("plurals", "الجموع"),
    ("Plurals", "الجموع"),
    ("opposites", "الأضداد"),
    ("handwriting practice", "تدريب الخط"),
    ("Handwriting practice", "تدريب الخط"),
    ("spelling test", "اختبار إملائي"),
    ("Spelling test", "اختبار إملائي"),
    ("oral presentation", "عرض شفهي"),
    ("days of the week", "أيام الأسبوع"),
    ("colors", "الألوان"),
    ("directions", "الاتجاهات"),
    ("adjectives", "الصفات"),
    ("family members", "أفراد الأسرة"),
    ("numbers", "الأعداد"),
    ("question words", "أدوات الاستفهام"),
    ("prepositions", "حروف الجر"),
]


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def latin_ratio(text: str) -> float:
    raw = text or ""
    if not raw:
        return 0.0
    latin_count = sum(1 for char in raw if "a" <= char.lower() <= "z")
    return latin_count / max(len(raw), 1)


def translate_text(text: str) -> str:
    translated = (text or "").replace("\n", " / ")
    for source, target in COMMON_REPLACEMENTS:
        translated = translated.replace(source, target)
    translated = translated.replace("Unit", "الوحدة")
    translated = translated.replace("Lesson", "الدرس")
    translated = translated.replace("Text p.", "النص ص.")
    translated = translated.replace("B2 U", "المستوى B2 - الوحدة ")
    translated = translated.replace(" + ", " / ")
    return normalize_spaces(translated)


def extract_topic_label(text: str) -> str:
    lowered = (text or "").lower()
    for keywords, label in TOPIC_KEYWORDS:
        if any(keyword.lower() in lowered for keyword in keywords):
            return label
    arabic_words = re.findall(r"[\u0600-\u06FF]{3,}", text or "")
    if arabic_words:
        return " / ".join(arabic_words[:2])
    return ""


def extract_unit_name(text: str, week_number: int) -> str:
    match = re.search(r"(B2\s*U\s*\d+|Unit\s*\d+|Lesson\s*\d+)", text or "", flags=re.IGNORECASE)
    if match:
        return translate_text(match.group(1))
    if "Review" in (text or "") or "Revision" in (text or ""):
        return f"مراجعة الأسبوع {week_number}"
    return f"الأسبوع {week_number}"


def extract_lesson_title(text: str, week_number: int, level_name: str) -> str:
    override = PRIMARY_WEEK_TITLE_OVERRIDES.get(level_name, {}).get(week_number)
    if override:
        return override
    segments = [normalize_spaces(part) for part in re.split(r"[/\n]+", text or "") if normalize_spaces(part)]
    if not segments:
        return f"خطة الأسبوع {week_number}"
    title = translate_text(segments[0])
    unit_name = extract_unit_name(text, week_number)
    topic_label = extract_topic_label(text)
    if "مراجعة" in (text or "") or "Review" in (text or "") or "Revision" in (text or ""):
        return topic_label or unit_name
    if topic_label:
        if "الوحدة" in unit_name:
            return f"{unit_name} - {topic_label}"
        return topic_label
    if re.search(r"[A-Za-z]", title) or latin_ratio(title) > 0.15:
        if "مراجعة" in unit_name:
            return unit_name
        if "الوحدة" in unit_name:
            return f"{unit_name} - قراءة وفهم وتطبيقات لغوية"
        if "الدرس" in unit_name:
            return f"{unit_name} - قراءة وفهم وتطبيقات كتابية"
        return f"الأسبوع {week_number} - قراءة وفهم ومهارات لغوية"
    if len(title) > 120:
        title = title[:117].rstrip(" ،-/") + "..."
    return title


def derive_unit_name_from_title(lesson_title: str, fallback_unit_name: str) -> str:
    title = normalize_spaces(lesson_title)
    if " - " in title:
        return normalize_spaces(title.split(" - ", 1)[0])
    if title.startswith("مراجعة"):
        return title
    if title:
        return title
    return fallback_unit_name


def simplify_topic_text(title: str) -> str:
    normalized = normalize_spaces(title)
    simplified = re.sub(r"^(الوحدة|الكتاب|المستوى)\s*[^-–:]*[-–:]\s*", "", normalized).strip()
    simplified = re.sub(r"^\d+\s*[-–:]\s*", "", simplified).strip()
    simplified = simplified.replace(" / ", " و")
    simplified = simplified.replace(" /", " ")
    simplified = simplified.replace("/ ", " ")
    simplified = simplified.replace("/", " ")
    simplified = normalize_spaces(simplified)
    return simplified or normalized


def build_learning_objective(text: str) -> str:
    segments = [translate_text(part) for part in re.split(r"[/\n]+", text or "") if normalize_spaces(part)]
    merged: list[str] = []
    for segment in segments[:4]:
        if segment and segment not in merged:
            merged.append(segment)
    objective = " / ".join(merged)
    if latin_ratio(objective) > 0.20:
        topic_label = extract_topic_label(text)
        if topic_label:
            return f"تنمية مهارات القراءة والفهم والتعبير حول موضوع: {topic_label}، مع تثبيت المفردات والتراكيب المستهدفة خلال هذا الأسبوع."
        return "تنمية القراءة والفهم والتعبير الشفهي أو الكتابي وفق محتوى هذا الأسبوع، مع تثبيت المفردات والتراكيب المستهدفة."
    return objective


def build_homework(text: str) -> str:
    if "Review" in (text or "") or "Revision" in (text or "") or "مراجعة" in (text or ""):
        return "مراجعة محتوى الأسبوع السابق والاستعداد لتقويم قصير."
    if "Write" in (text or "") or "كتابة" in (text or ""):
        return "إتمام تدريب كتابي قصير وتثبيت مفردات الأسبوع."
    return "مراجعة مفردات الأسبوع والتدرّب الشفهي أو الكتابي على مهارة الدرس."


def build_homework_from_title(title: str) -> str:
    normalized = title or ""
    topic = simplify_topic_text(normalized)
    if any(keyword in topic for keyword in ["التهيئة", "أهمية اللغة العربية"]):
        return "مراجعة رسالة الدرس الأولى وكتابة أو قول سبب واحد يبين أهمية تعلم العربية."
    if "مراجعة" in normalized:
        return "إكمال مراجعة مركزة لمحتوى الأسابيع السابقة والاستعداد لتقويم قصير."
    if any(keyword in topic for keyword in ["الوداع", "المؤنث", "التحية", "الحال"]):
        return "تطبيق مفردات الدرس في حوار شفهي قصير مع التركيز على الصيغة الصحيحة."
    if any(keyword in topic for keyword in ["الأعداد", "الأعداد الترتيبية"]):
        return "حل تدريب قصير على الأعداد واستخدامها في جمل مفهومة."
    if any(keyword in topic for keyword in ["التحية", "الحال", "الأسرة", "الشراب", "أيام الأسبوع"]):
        return "مراجعة مفردات الدرس وتطبيقها في حوار شفهي قصير في البيت."
    if any(keyword in topic for keyword in ["الألوان", "الصفات", "أسماء الإشارة", "الاتجاهات", "المفردات", "الأسرة", "البيت", "المدرسة"]):
        return "مراجعة مفردات الدرس واستخدامها في ثلاث جمل قصيرة صحيحة."
    if any(keyword in topic for keyword in ["الجموع", "الضمائر", "الإعراب", "إن", "ليس", "المبتدأ", "الفاعل", "الاستفهام", "الإضافة", "الجملة الاسمية", "الجملة الفعلية", "المؤنث", "المذكر", "المثنى", "لكن", "كأن", "أن ", "أظن", "هات", "اسم التفضيل"]):
        return "حل تطبيقات قصيرة على القاعدة ومراجعة الأمثلة الأساسية للدرس."
    return "مراجعة مفردات الأسبوع والتدرّب الشفهي أو الكتابي على مهارة الدرس."


def build_objective_from_title(title: str, raw_text: str) -> str:
    normalized = title or ""
    topic = simplify_topic_text(normalized)
    if any(keyword in topic for keyword in ["التهيئة", "أهمية اللغة العربية"]):
        return "تهيئة الطالب للمساق، وتعزيز دافعيته لتعلم العربية، مع بناء تصور أولي عن أهداف المادة ومسارها خلال العام."
    if "مراجعة" in normalized:
        return "مراجعة شاملة لمهارات ومفردات وتراكيب هذا الجزء، مع قياس جاهزية الطالب قبل الانتقال لما بعده."
    if any(keyword in topic for keyword in ["الوداع", "المؤنث", "التحية", "الحال", "أيام الأسبوع", "الأسرة", "الشراب", "الطائر", "القطة", "البيت"]):
        return f"تنمية القراءة والفهم والتعبير الشفهي حول موضوع: {topic}، مع تثبيت المفردات الأساسية والتراكيب المستهدفة."
    if any(keyword in topic for keyword in ["التحية", "الحال", "الأسرة", "أيام الأسبوع", "الشراب", "الطائر", "القطة", "البيت", "المدرسة", "العمرة", "السفر", "الأب", "المدرس", "هاشم", "أحمد", "إبراهيم", "سعيد"]):
        return f"تنمية القراءة والفهم والتعبير الشفهي حول موضوع: {topic}، مع تثبيت المفردات الأساسية والتراكيب المستهدفة."
    if any(keyword in topic for keyword in ["الأعداد", "الأعداد الترتيبية"]):
        return f"تمكين الطالب من فهم واستعمال {topic} في مواقف لغوية بسيطة، مع ربطها بالكتابة والتمييز السمعي."
    if any(keyword in topic for keyword in ["الألوان", "الصفات", "أسماء الإشارة", "الاتجاهات", "المفردات"]):
        return f"تنمية المفردات والفهم اللغوي في موضوع: {topic}، مع تدريب الطالب على الاستخدام الشفهي والكتابي السليم."
    if any(keyword in topic for keyword in ["الجموع", "الضمائر", "الإعراب", "المبتدأ", "ليس", "إن", "الفاعل", "الاستفهام", "الإضافة", "اسم التفضيل", "الجملة الاسمية", "الجملة الفعلية", "المؤنث", "المذكر", "المثنى", "لكن", "كأن", "أن ", "أظن", "هات"]):
        return f"فهم القاعدة المرتبطة بموضوع: {topic}، ثم تطبيقها في أمثلة وجمل قصيرة قراءةً وكتابةً."
    return build_learning_objective(raw_text)


def get_primary_week_slots() -> list[dict]:
    schedule = build_teaching_week_schedule(total_weeks=40, lessons_per_week=1)
    return [
        {
            "teaching_week_number": meta["teaching_week_number"],
            "session_number": 1,
            "slot_order": meta["teaching_week_number"],
        }
        for meta in schedule.values()
        if meta.get("is_instructional_week") and meta.get("teaching_week_number")
    ]


def build_level_rows(ws, column: int, book_name: str, level_name: str) -> list[dict]:
    rows: list[dict] = []
    for row_idx in range(3, 80):
        week_value = ws.cell(row_idx, 1).value
        content = ws.cell(row_idx, column).value
        if week_value is None or content is None:
            continue
        week_number = int(float(week_value))
        raw_text = normalize_spaces(str(content))
        if not raw_text:
            continue
        lesson_title = extract_lesson_title(raw_text, week_number, level_name)
        unit_name = derive_unit_name_from_title(
            lesson_title,
            extract_unit_name(raw_text, week_number),
        )
        rows.append(
            {
                "week_number": week_number,
                "book_name": book_name,
                "unit_name": unit_name,
                "lesson_title": lesson_title,
                "source_reference": f"الخطة الأسبوعية - الأسبوع {week_number}",
                "learning_objective": build_objective_from_title(lesson_title, raw_text),
                "planned_homework": build_homework_from_title(lesson_title),
                "note_text": f"مستورد من ملف الخطة الأسبوعية والسنوية - أسبوع {week_number}",
            }
        )
    return rows


def fill_level_plan(level: Level, rows: list[dict]) -> tuple[int, int]:
    for existing in SyllabusPlanEntry.query.filter_by(level_id=level.id).all():
        db.session.delete(existing)

    slots = get_primary_week_slots() if get_level_syllabus_lessons_per_week(level) == 1 else []
    created = 0
    unscheduled = 0
    for index, row in enumerate(rows, start=1):
        slot = slots[index - 1] if index - 1 < len(slots) else None
        if slot is None:
            unscheduled += 1
        db.session.add(
            SyllabusPlanEntry(
                level_id=level.id,
                week_number=slot["teaching_week_number"] if slot else None,
                session_number=slot["session_number"] if slot else None,
                book_name=row["book_name"],
                unit_name=row["unit_name"],
                lesson_title=row["lesson_title"],
                source_reference=row["source_reference"],
                learning_objective=row["learning_objective"],
                planned_homework=row["planned_homework"],
                note_text=row["note_text"],
                status="planned",
                order_index=slot["slot_order"] if slot else index,
            )
        )
        created += 1
    return created, unscheduled


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    with app.app_context():
        for _, config in PRIMARY_PLAN_CONFIG.items():
            level = Level.query.get(config["level_id"])
            if not level:
                continue
            rows = build_level_rows(ws, config["column"], config["book_name"], level.name)
            created, unscheduled = fill_level_plan(level, rows)
            print(f"IMPORTED {level.name}: rows={created} unscheduled={unscheduled}")
        db.session.commit()


if __name__ == "__main__":
    main()
