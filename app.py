from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import json
import os
import re
import shutil
import smtplib
from email.message import EmailMessage
from functools import lru_cache, wraps
from uuid import uuid4

import pandas as pd
from flask import Flask, Response, abort, redirect, render_template, request, send_file, session, url_for, has_request_context
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import text

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

from config import Config
from models import (
    AcademicCalendarEvent,
    ATTENDANCE_STATUSES,
    Announcement,
    ActionLog,
    Assignment,
    AssignmentSubmission,
    Attendance,
    ClassRecording,
    CurriculumItem,
    CurriculumProgress,
    ExamImportIssue,
    ExamPublication,
    ExamResult,
    ExamTemplate,
    ExamTemplateBranch,
    HolidayPeriod,
    Level,
    Student,
    StudentExamVisibility,
    StudentNameAlias,
    StudentMonthlyNote,
    Subject,
    SyllabusPlanEntry,
    Teacher,
    UpcomingExam,
    db,
)

LEVEL_DISPLAY_TO_CODES = {
    "Qaeda": ["Qaeda", "Qaeda-B", "Qaeda-B "],
    "Primary Beginner": ["Primary Beginner", "Pri-Beg", "Pri-Beg-B"],
    "Primary Intermediate": ["Primary Intermediate", "Pri-Int", "Pri-Int-B", "Pri-Int-B."],
    "Primary Advance": ["Primary Advance", "Pri-Adv", "Pri-Adv-B"],
    "Secondary Beginner 1": ["Secondary Beginner 1", "Sec-Beg-B1"],
    "Secondary Beginner 2": ["Secondary Beginner 2", "Sec-Beg-B2"],
    "Secondary Intermediate": ["Secondary Intermediate", "Sec-Int", "Sec-Int-B"],
    "Secondary Advance": ["Secondary Advance", "Sec-Adv", "Sec-Adv-B"],
    "iGCSE": ["iGCSE"],
}

LEVEL_LEGACY_NAME_TO_DISPLAY = {
    "Primary Beginner (yr5-6)": "Primary Beginner",
    "Primary Intermediate (yr5-6)": "Primary Intermediate",
    "Primary Advance (Yr5-6)": "Primary Advance",
}

LEVEL_DISPLAY_ORDER = [
    "Qaeda",
    "Primary Beginner",
    "Primary Intermediate",
    "Primary Advance",
    "Secondary Beginner 1",
    "Secondary Beginner 2",
    "Secondary Intermediate",
    "Secondary Advance",
    "iGCSE",
]

ARABIC_ATTENDANCE_SHEET_TO_LEVEL = {
    "Qaeda-B ": "Qaeda",
    "Pri-Beg-B": "Primary Beginner",
    "Pri-Int-B.": "Primary Intermediate",
    "Pri-Adv-B": "Primary Advance",
    "Sec-Beg-B1": "Secondary Beginner 1",
    "Sec-Beg-B2": "Secondary Beginner 2",
    "Sec-Int-B": "Secondary Intermediate",
    "Sec-Adv-B": "Secondary Advance",
}

WEEKDAY_NAME_TO_INDEX = {
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "saturday": 5,
    "sunday": 6,
}

ENGLISH_WEEKDAY_LABELS = {
    0: "Monday",
    1: "Tuesday",
    2: "Wednesday",
    3: "Thursday",
    4: "Friday",
    5: "Saturday",
    6: "Sunday",
}

ARABIC_WEEKDAY_LABELS = {
    0: "الاثنين",
    1: "الثلاثاء",
    2: "الأربعاء",
    3: "الخميس",
    4: "الجمعة",
    5: "السبت",
    6: "الأحد",
}

ACADEMIC_NON_TEACHING_EVENT_TYPES = {
    "holiday",
    "exam",
    "parents_meeting",
    "teacher_training",
}

ACADEMIC_STATUS_ORDER = {
    "holiday": 1,
    "exam": 2,
    "parents_meeting": 3,
    "teacher_training": 4,
    "school_resume": 5,
    "term_start": 6,
    "event_day": 7,
}

ACADEMIC_EVENT_TYPE_OPTIONS = [
    "term_start",
    "school_resume",
    "holiday",
    "exam",
    "parents_meeting",
    "teacher_training",
    "event_day",
]

DEFAULT_ACADEMIC_CALENDAR_2025_26 = [
    {
        "title": "Academic Year Begins",
        "event_type": "term_start",
        "start_date": date(2025, 9, 8),
        "end_date": date(2025, 9, 8),
        "is_instructional": True,
        "sort_order": 10,
        "note_text": "First teaching week of the 2025-26 academic year.",
    },
    {
        "title": "Half-Term Break",
        "event_type": "holiday",
        "start_date": date(2025, 10, 26),
        "end_date": date(2025, 10, 30),
        "is_instructional": False,
        "sort_order": 20,
        "note_text": "Autumn half-term break.",
    },
    {
        "title": "School Resumes",
        "event_type": "school_resume",
        "start_date": date(2025, 11, 3),
        "end_date": date(2025, 11, 3),
        "is_instructional": True,
        "sort_order": 30,
        "note_text": "Classes resume after the autumn break.",
    },
    {
        "title": "Winter Break",
        "event_type": "holiday",
        "start_date": date(2025, 12, 25),
        "end_date": date(2026, 1, 4),
        "is_instructional": False,
        "sort_order": 40,
        "note_text": "Winter holiday.",
    },
    {
        "title": "School Resumes",
        "event_type": "school_resume",
        "start_date": date(2026, 1, 5),
        "end_date": date(2026, 1, 5),
        "is_instructional": True,
        "sort_order": 50,
        "note_text": "Classes resume after winter break.",
    },
    {
        "title": "Mid-term Exams",
        "event_type": "exam",
        "start_date": date(2026, 2, 4),
        "end_date": date(2026, 2, 5),
        "is_instructional": False,
        "sort_order": 60,
        "note_text": "Mid-term assessment window.",
    },
    {
        "title": "Parents' Meeting",
        "event_type": "parents_meeting",
        "start_date": date(2026, 2, 6),
        "end_date": date(2026, 2, 7),
        "is_instructional": False,
        "sort_order": 70,
        "note_text": "Parent meeting days.",
    },
    {
        "title": "Teacher Training",
        "event_type": "teacher_training",
        "start_date": date(2026, 2, 8),
        "end_date": date(2026, 2, 8),
        "is_instructional": False,
        "sort_order": 80,
        "note_text": "Teacher training day from the academic calendar.",
    },
    {
        "title": "Ramadan Holiday",
        "event_type": "holiday",
        "start_date": date(2026, 3, 9),
        "end_date": date(2026, 3, 26),
        "is_instructional": False,
        "sort_order": 90,
        "note_text": "Ramadan and Eid holiday period based on the academy calendar.",
    },
    {
        "title": "School Resumes",
        "event_type": "school_resume",
        "start_date": date(2026, 3, 29),
        "end_date": date(2026, 3, 29),
        "is_instructional": True,
        "sort_order": 100,
        "note_text": "Classes resume after Ramadan and Eid holiday.",
    },
    {
        "title": "Half-Term Break",
        "event_type": "holiday",
        "start_date": date(2026, 5, 24),
        "end_date": date(2026, 5, 28),
        "is_instructional": False,
        "sort_order": 110,
        "note_text": "Late spring half-term break.",
    },
    {
        "title": "School Resumes",
        "event_type": "school_resume",
        "start_date": date(2026, 6, 1),
        "end_date": date(2026, 6, 1),
        "is_instructional": True,
        "sort_order": 120,
        "note_text": "Classes resume after the spring half-term break.",
    },
    {
        "title": "End of Year Exam Week",
        "event_type": "exam",
        "start_date": date(2026, 7, 7),
        "end_date": date(2026, 7, 10),
        "is_instructional": False,
        "sort_order": 130,
        "note_text": "End of year exams.",
    },
    {
        "title": "Summer Holiday",
        "event_type": "holiday",
        "start_date": date(2026, 7, 13),
        "end_date": date(2026, 8, 31),
        "is_instructional": False,
        "sort_order": 140,
        "note_text": "Summer break after the end of year exams.",
    },
]


HEAD_DASHBOARD_COPY = {
    "en": {
        "logout": "Logout",
        "main_dashboard": "Main Dashboard",
        "data_root": "Data Root",
        "title": "Head Dashboard",
        "subtitle": "Overview for level operations and quick actions.",
        "total_students": "Total Students",
        "total_levels": "Total Levels",
        "total_teachers": "Total Teachers",
        "assessment_control": "Assessment Control",
        "exam_management": "Exam Management",
        "exam_management_desc": "Create exam templates, define branch marks, and control when results become visible.",
        "templates": "Templates",
        "imported_exams": "Imported Exams",
        "visibility_rules": "Visibility Rules",
        "workflow_label": "Workflow:",
        "workflow_text": "Create template -> teachers enter scores -> publish result -> hide individual students if needed.",
        "tip_label": "Tip:",
        "tip_text": "Use one branch per line and add the mark after | so teachers receive a ready-made score table.",
        "import_exam_results": "Import Exam Results From Excel",
        "import_exam_desc_1": "Upload one Excel sheet for all levels. The system matches rows by student_code first, or by full_name with level_name.",
        "import_exam_desc_2": "Recommended columns: student_code, full_name, level_name, exam_title, subject_name, score, max_score, exam_date, notes.",
        "upload_excel": "Upload Excel Sheet",
        "step_1": "Step 1",
        "step_2": "Step 2",
        "step_3": "Step 3",
        "create_exam_template": "Create Exam Template",
        "create_exam_template_desc": "Create the exam title and its branches here. Teachers will only enter scores for these branches.",
        "branches_format": "Branches format",
        "branches_example": "Use one branch per line, for example Reading (Fluency)|10",
        "exam_title": "Exam title",
        "add_exam_template": "Add Exam Template",
        "publish_or_restrict": "Publish Or Restrict Results",
        "publish_or_restrict_desc": "Publish an exam for all students, then hide it for individual students if needed.",
        "exam": "Exam",
        "published_for_all": "Published For All",
        "visible": "Visible",
        "save": "Save",
        "no_exams_found": "No exams found yet.",
        "individual_student_visibility": "Individual Student Visibility",
        "select_student": "Select student",
        "select_exam": "Select exam",
        "hide_for_student": "Hide For This Student",
        "show_for_student": "Remove Individual Hide",
        "save_visibility": "Save Visibility",
        "existing_exam_templates": "Existing Exam Templates",
        "existing_exam_templates_desc": "Review and update any template before teachers start entering scores.",
        "no_date_set": "No date set",
        "active": "Active",
        "update_template": "Update Template",
        "no_exam_templates_yet": "No exam templates yet.",
        "curriculum": "Curriculum",
        "subjects_and_books": "Books And Shared Links",
        "subjects_and_books_desc": "Add level links directly. Books appear for students and teachers. Shared links appear for teachers only.",
        "add_subject": "Add Subject",
        "add_subject_desc": "Use this once for each subject in the level, such as Arabic or Islamic Studies.",
        "select_level": "Select level",
        "subject_name": "Subject name",
        "short_subject_description": "Short subject description",
        "order": "Order",
        "add_book_link": "Add Book Link",
        "add_book_desc": "After adding the subject above, choose it here and paste the Google Drive link for the book.",
        "select_subject": "Select subject",
        "book_title": "Book title",
        "resource_title": "Link title",
        "google_drive_link": "Google Drive link",
        "short_book_description": "Short description",
        "add_book": "Add Book",
        "add_subject_first": "Add at least one subject first. The book list stays empty until a subject is created.",
        "no_subject_available": "No subject available yet",
        "no_subjects_or_books": "No subjects or books added yet.",
        "book": "Book",
        "teacher_resource": "Teacher Resource",
        "link_type": "Link type",
        "add_level_link": "Add Level Link",
        "save_link": "Save Link",
        "books": "Books",
        "teacher_resources": "Teacher Resources",
        "no_teacher_resources_yet": "No teacher resources added yet.",
        "add_teacher_links_here": "Add shared teacher links here.",
        "apply_to_all_levels": "Apply To All Levels",
        "delete_link": "Delete",
        "delete_link_confirm": "Delete this link?",
        "no_books_yet": "No books yet",
        "add_drive_link": "Add a Drive link for this subject.",
        "open_on_drive": "Open on Google Drive",
        "level_name": "Level Name",
        "teacher_name": "Teacher Name",
        "students_count": "Students Count",
        "join_zoom": "Join Zoom",
        "details": "Details",
        "attendance_report": "Attendance Report",
        "no_levels_found": "No levels found.",
        "language": "Language",
    },
    "ar": {
        "logout": "تسجيل الخروج",
        "main_dashboard": "الواجهة الرئيسية",
        "data_root": "جذر البيانات",
        "title": "لوحة الإشراف",
        "subtitle": "نظرة عامة على الفصول والعمليات السريعة.",
        "total_students": "إجمالي الطلاب",
        "total_levels": "إجمالي الفصول",
        "total_teachers": "إجمالي المعلمين",
        "assessment_control": "إدارة التقييم",
        "exam_management": "إدارة الامتحانات",
        "exam_management_desc": "إنشاء قوالب الامتحانات وتحديد الفروع والتحكم في إظهار النتائج.",
        "templates": "القوالب",
        "imported_exams": "الامتحانات المستوردة",
        "visibility_rules": "قواعد الإظهار",
        "workflow_label": "الخطوات:",
        "workflow_text": "إنشاء القالب ثم إدخال الدرجات من المعلمين ثم نشر النتيجة ثم إخفاؤها فردياً عند الحاجة.",
        "tip_label": "ملاحظة:",
        "tip_text": "اكتب كل فرع في سطر مستقل ثم ضع الدرجة بعد علامة | حتى يصل القالب جاهزاً للمعلمين.",
        "import_exam_results": "استيراد نتائج الامتحان من إكسل",
        "import_exam_desc_1": "ارفع ملف إكسل واحداً لجميع الفصول. يحاول النظام المطابقة أولاً عبر student_code ثم عبر الاسم والمستوى.",
        "import_exam_desc_2": "الأعمدة المقترحة: student_code, full_name, level_name, exam_title, subject_name, score, max_score, exam_date, notes.",
        "upload_excel": "رفع ملف الإكسل",
        "step_1": "الخطوة 1",
        "step_2": "الخطوة 2",
        "step_3": "الخطوة 3",
        "create_exam_template": "إنشاء قالب امتحان",
        "create_exam_template_desc": "أنشئ اسم الامتحان وفروعه هنا، وسيقوم المعلم فقط بإدخال الدرجات.",
        "branches_format": "صيغة الفروع",
        "branches_example": "اكتب كل فرع في سطر، مثل Reading (Fluency)|10",
        "exam_title": "اسم الامتحان",
        "add_exam_template": "إضافة قالب امتحان",
        "publish_or_restrict": "نشر النتائج أو تقييدها",
        "publish_or_restrict_desc": "انشر النتيجة لجميع الطلاب ثم أخفها عن طالب محدد إذا احتجت.",
        "exam": "الامتحان",
        "published_for_all": "ظاهر للجميع",
        "visible": "ظاهر",
        "save": "حفظ",
        "no_exams_found": "لا توجد امتحانات بعد.",
        "individual_student_visibility": "إظهار فردي للطالب",
        "select_student": "اختر الطالب",
        "select_exam": "اختر الامتحان",
        "hide_for_student": "إخفاء لهذا الطالب",
        "show_for_student": "إزالة الإخفاء الفردي",
        "save_visibility": "حفظ الإظهار",
        "existing_exam_templates": "قوالب الامتحانات الحالية",
        "existing_exam_templates_desc": "راجع القوالب وعدّلها قبل أن يبدأ المعلمون بإدخال الدرجات.",
        "no_date_set": "لا يوجد تاريخ",
        "active": "نشط",
        "update_template": "تحديث القالب",
        "no_exam_templates_yet": "لا توجد قوالب امتحان بعد.",
        "curriculum": "المنهج",
        "subjects_and_books": "الكتب والروابط العامة",
        "subjects_and_books_desc": "أضف الروابط مباشرة على مستوى الفصل. الكتب تظهر للطالب والمعلم، وروابط الاستفادة العامة تظهر للمعلمين فقط.",
        "add_subject": "إضافة مادة",
        "add_subject_desc": "استخدم هذا مرة واحدة لكل مادة في الفصل مثل العربية أو الدراسات الإسلامية.",
        "select_level": "اختر الفصل",
        "subject_name": "اسم المادة",
        "short_subject_description": "وصف مختصر للمادة",
        "order": "الترتيب",
        "add_book_link": "إضافة رابط كتاب",
        "add_book_desc": "بعد إضافة المادة أعلاه اخترها هنا ثم الصق رابط Google Drive للكتاب.",
        "select_subject": "اختر المادة",
        "book_title": "اسم الكتاب",
        "resource_title": "عنوان الرابط",
        "google_drive_link": "رابط Google Drive",
        "short_book_description": "وصف مختصر",
        "add_book": "إضافة كتاب",
        "add_subject_first": "أضف مادة واحدة على الأقل أولاً. ستبقى قائمة الكتب فارغة حتى يتم إنشاء مادة.",
        "no_subject_available": "لا توجد مادة بعد",
        "no_subjects_or_books": "لا توجد مواد أو كتب بعد.",
        "book": "كتاب",
        "teacher_resource": "رابط للمعلمين",
        "link_type": "نوع الرابط",
        "add_level_link": "إضافة رابط للفصل",
        "save_link": "حفظ الرابط",
        "books": "الكتب",
        "teacher_resources": "روابط المعلمين",
        "no_teacher_resources_yet": "لا توجد روابط للمعلمين بعد.",
        "add_teacher_links_here": "أضف هنا روابط الاستفادة العامة للمعلمين.",
        "apply_to_all_levels": "تطبيق على جميع الفصول",
        "delete_link": "حذف",
        "delete_link_confirm": "هل تريد حذف هذا الرابط؟",
        "no_books_yet": "لا توجد كتب بعد",
        "add_drive_link": "أضف رابط الدرايف لهذه المادة.",
        "open_on_drive": "فتح على Google Drive",
        "level_name": "اسم الفصل",
        "teacher_name": "اسم المعلم",
        "students_count": "عدد الطلاب",
        "join_zoom": "دخول زوم",
        "details": "التفاصيل",
        "attendance_report": "تقرير الحضور",
        "no_levels_found": "لا توجد فصول.",
        "language": "اللغة",
    },
}


def get_head_dashboard_copy(lang: str) -> dict:
    return HEAD_DASHBOARD_COPY.get(lang, HEAD_DASHBOARD_COPY["en"])


GLOBAL_UI_COPY = {
    "en": {
        "language": "Language",
        "role_selection": "Role Selection",
        "logout": "Logout",
        "head_dashboard": "Head Dashboard",
        "data_root": "Data Root",
        "teacher_dashboard": "Teacher Dashboard",
        "admin_dashboard": "Supervisor Dashboard",
        "attendance": "Attendance",
        "attendance_records": "Attendance Records",
        "daily_attendance": "Daily Attendance",
        "select_a_level": "Select a level",
        "selected_date": "Selected Date",
        "students": "Students",
        "preset_default": "Preset Default",
        "present": "Present",
        "absent": "Absent",
        "late": "Late",
        "excused": "Excused",
        "choose_class_date": "Choose a class, load students, then mark attendance for the selected date.",
        "mark_students_for_date": "Mark students for {date} and save updates for this class.",
        "start_by_selecting_class": "Start by selecting the class and date you want to review.",
        "load_attendance": "Load Attendance",
        "class_and_date": "Class And Date",
        "load_students": "Load Students",
        "quick_actions": "Quick Actions",
        "fast_attendance_shortcuts": "Fast Attendance Shortcuts",
        "mark_all_present": "Mark All Present",
        "mark_all_absent": "Mark All Absent",
        "mark_all_late": "Mark All Late",
        "mark_all_excused": "Mark All Excused",
        "student_register": "Student Register",
        "attendance_list_title": "Attendance List",
        "save_attendance": "Save Attendance",
        "saved_summary": "Saved Summary",
        "attendance_summary_after_saving": "Attendance Summary After Saving",
        "students_marked_present": "Students marked present",
        "students_marked_absent": "Students marked absent",
        "students_marked_late": "Students marked late",
        "students_marked_excused": "Students marked excused",
        "present_students_out_of_class": "Present students out of the full class",
        "present_rate": "Present Rate",
        "no_students_for_level": "No students found for this level.",
        "attendance_saved_successfully": "Attendance saved successfully.",
        "supervisor_dashboard": "Supervisor Dashboard",
        "executive_control": "Executive Control",
        "one_view_supervisor": "One view for attendance, assignment follow-up, exam readiness, and student risk.",
        "levels": "Levels",
        "teachers": "Teachers",
        "week_start": "Week Start",
        "today": "Today",
        "student_interface": "Student Interface",
        "enter_student_code": "Enter student code to view report links and class info.",
        "find_student": "Find Student",
        "student_code_placeholder": "Student code (e.g. STD-000001)",
        "student_not_found": "Student not found for code: {code}",
        "teacher_signed_in_as": "Signed in as: {name}",
        "quick_actions": "Quick Actions",
        "today_tasks": "Today's Tasks",
        "focus_on_these_first": "Focus on these first",
        "before_class": "Before Class",
        "during_class": "During Class",
        "after_class": "After Class",
        "daily_workflow": "Daily Workflow",
        "daily_workflow_help": "Start here so you know what to do before, during, and after the lesson.",
        "prepare_zoom_and_plan": "Prepare Zoom, links, and lesson plan.",
        "take_attendance_and_teach": "Take attendance and teach the planned lesson.",
        "upload_recording_and_homework": "Upload the lesson record and assign or review homework.",
        "open_attendance": "Open Attendance",
        "open_first_workspace": "Open First Workspace",
        "open_review_queue": "Open Review Queue",
        "teaching_day": "Teaching Day",
        "teacher_dashboard_subtitle": "Your classes, attendance follow-up, review queue, and student priorities in one place.",
        "plan_main_fields_hint": "Book, unit, and lesson first. Open more details only when needed.",
        "more_details": "More Details",
        "welcome_back": "Welcome Back",
        "monthly_report": "Monthly Report",
        "weekly_report_label": "Weekly Report",
        "weekly_pdf_label": "Weekly PDF",
        "open_matching": "Open Matching",
        "data_cleanup": "Data Cleanup",
        "critical_records_snapshot": "Critical Records Snapshot",
        "open_cleanup": "Open Cleanup",
        "critical_label": "Critical",
        "students_only": "Students Only",
        "teachers_only": "Teachers Only",
        "levels_only": "Levels Only",
        "all_cleanup": "All Cleanup",
        "critical_only": "Critical Only",
        "students_cleanup": "Students Cleanup",
        "teachers_cleanup": "Teachers Cleanup",
        "levels_cleanup": "Levels Cleanup",
        "fix_record": "Fix Record",
        "no_critical_cleanup_blockers": "No critical cleanup blockers",
        "remaining_cleanup_review_only": "The remaining data cleanup items are review-level and not currently blocking daily operation.",
        "operating_calendar": "Operating Calendar",
        "duty_time_and_weekly_followup": "Duty Time And Weekly Follow-up",
        "calendar_settings_help": "Set the duty period and the weekly reminder day. The system will skip generation during active holidays.",
        "save_calendar_settings": "Save Calendar Settings",
        "weekly_followup_status": "Weekly follow-up status",
        "last_successful_scheduled_run": "Last successful scheduled run: {value}",
        "not_yet_run": "Not yet run",
        "generate_weekly_followup_now": "Generate Weekly Follow-up Now",
        "holiday_calendar": "Holiday Calendar",
        "academic_holidays": "Academic Holidays",
        "holiday_title": "Holiday title",
        "add_holiday": "Add Holiday",
        "active_label": "Active",
        "inactive_label": "Inactive",
        "no_holidays_added": "No holidays added",
        "weekly_generator_runs_normally": "The weekly generator will run normally until holiday periods are added here.",
        "broadcast_center": "Broadcast Center",
        "post_administrative_announcement": "Post Administrative Announcement",
        "announcement_posting_help": "Use this for administrative circulars, reminders, or follow-up alerts such as missing Zoom links, missing homework, or student preparation tasks.",
        "announcement_title": "Announcement title",
        "all_audience": "All",
        "teachers_only_audience": "Teachers Only",
        "students_only_audience": "Students Only",
        "general_category": "General",
        "reminder_category": "Reminder",
        "alert_category": "Alert",
        "follow_up_category": "Follow Up",
        "all_levels": "All Levels",
        "announcement_body_placeholder": "Write the announcement, expectation, or operational follow-up here.",
        "pin_this_announcement": "Pin this announcement",
        "publish_announcement": "Publish Announcement",
        "live_feed": "Live Feed",
        "recent_announcements": "Recent Announcements",
        "archived_state": "Archived",
        "no_announcements_yet": "No announcements yet",
        "published_messages_here": "Your published administrative messages will appear here.",
        "student_risk_radar": "Student Risk Radar",
        "students_needing_followup": "Students Needing Follow-up",
        "student_label": "Student",
        "risk_label": "Risk",
        "teacher_label": "Teacher",
        "weekly_standing": "Weekly Standing",
        "monday_to_thursday_status": "Monday To Thursday Completion Status",
        "needs_completion": "Needs Completion",
        "on_track_for_appreciation": "On Track For Appreciation",
        "review_window": "Review window: {start} to {end}.",
        "levels_followup_appreciation": "Levels on follow-up path: {follow_up}. Levels on appreciation path: {praise}.",
        "level": "Level",
        "status": "Status",
        "weekly_review": "Weekly Review",
        "follow_up": "Follow-up",
        "appreciation": "Appreciation",
        "all_weekly_tasks_complete": "All required weekly tasks are complete.",
        "no_weekly_review_data": "No weekly review data yet.",
        "administrative_board": "Administrative Board",
        "supervisor_announcements": "Supervisor Announcements",
        "pinned": "Pinned",
        "no_administrative_announcements": "No administrative announcements",
        "supervisor_messages_here": "Supervisor messages and follow-up instructions will appear here.",
        "priority_students": "Priority Students",
        "needs_followup": "Needs Follow-up",
        "attendance_14d": "14-Day Attendance",
        "arabic_attendance_snapshot": "Arabic Attendance Snapshot",
        "attendance_source_label": "Attendance Source",
        "attendance_source_recent": "Recent 14-Day Log",
        "attendance_source_imported": "Imported Arabic Workbook",
        "attendance_source_none": "No Attendance Data Yet",
        "current_student_status": "Current Student Status",
        "attendance_snapshot_caption": "Attendance currently used in your status calculation",
        "arabic_attendance_import": "Arabic Attendance Import",
        "arabic_attendance_import_subtitle": "Refresh the shared Arabic attendance workbook and sync strong name matches.",
        "refresh_arabic_attendance": "Refresh Arabic Attendance",
        "upload_arabic_attendance_file": "Upload Arabic Attendance File",
        "current_attendance_file": "Current Attendance File",
        "download_current_attendance_file": "Download Current File",
        "attendance_import_steps": "1. Upload a new workbook if needed. 2. Refresh Arabic attendance. 3. Review reports and follow-up tables.",
        "no_arabic_attendance_history": "No Arabic attendance uploads or refreshes have been recorded yet.",
        "download_latest_backup": "Download Latest Backup",
        "arabic_attendance_name_review": "Arabic Attendance Name Review",
        "arabic_attendance_name_review_subtitle": "If any workbook names fail to match, review them here before trusting the imported attendance fully.",
        "best_system_match": "Best System Match",
        "match_score": "Match Score",
        "imported_rows": "Imported Rows",
        "matched_aliases": "Matched Aliases",
        "matched_rows": "Matched Rows",
        "unmatched_rows": "Unmatched Rows",
        "last_arabic_attendance_refresh": "Last Arabic Attendance Refresh",
        "open_assignments": "Open Assignments",
        "latest_exam": "Latest Exam",
        "why": "Why",
        "reason_low_attendance": "Low attendance",
        "reason_pending_assignments": "Pending assignments",
        "reason_one_open_assignment": "One open assignment",
        "reason_low_exam_result": "Low exam result",
        "reason_attendance": "Attendance",
        "reason_assignments": "Assignments",
        "reason_exam": "Exam",
        "reason_result": "Result",
        "reason_monthly_note": "Monthly note",
        "no_students_flagged": "No students currently flagged for follow-up.",
        "class_operations": "Class Operations",
        "my_levels": "My Levels",
        "my_levels_caption": "Assigned teaching groups on your dashboard",
        "attendance_logged_today": "Attendance Logged Today",
        "attendance_logged_today_caption": "{marked} of {total} student attendance records captured today",
        "review_queue_label": "Review Queue",
        "review_queue_caption": "Submitted assignments waiting for your review",
        "upcoming_exams_caption": "Scheduled exams still ahead for your levels",
        "plan_progress_label": "Plan Progress",
        "plan_progress_caption": "Your yearly syllabus completion across all assigned levels",
        "assigned_level": "Assigned Level",
        "on_track": "On Track",
        "attendance_today": "Attendance Today",
        "not_started": "Not started",
        "review_queue": "Review Queue",
        "assignments": "Assignments",
        "present_label": "Present",
        "all_weekly_teaching_complete": "All weekly teaching requirements are currently complete.",
        "next_exam": "Next Exam",
        "open_workspace": "Open Workspace",
        "open_tasks_for_this_class": "Open Tasks For This Class",
        "class_ready": "This class is on track today.",
        "class_needs_action": "This class still needs action today.",
        "task_attendance_missing_title": "Attendance still missing",
        "task_attendance_missing_body": "{count} class(es) still need attendance today.",
        "task_review_queue_title": "Assignments need review",
        "task_review_queue_body": "{count} submission(s) are still waiting for your review.",
        "task_syllabus_delay_title": "Plan progress needs updating",
        "task_syllabus_delay_body": "{count} lesson(s) are still delayed against the current teaching week.",
        "task_clear_title": "You are clear for now",
        "task_clear_body": "Your main daily tasks are currently in a good state.",
        "take_attendance": "Take Attendance",
        "join_zoom": "Join Zoom",
        "no_assigned_levels": "No assigned levels found.",
        "code_label": "Code",
        "what_needs_attention": "What Needs Attention",
        "student_updates_now": "Your most important updates right now.",
        "all_clear": "All clear",
        "no_student_alerts": "No urgent student-side alerts right now.",
        "study_snapshot": "Study Snapshot",
        "latest_result": "Latest Result",
        "no_result_yet": "No result yet",
        "next_published_result_here": "Your next published result will appear here.",
        "no_exam_announced": "No exam announced",
        "check_back_exam_announcements": "Check back later for new exam announcements.",
        "latest_recording": "Latest Recording",
        "no_recording_yet": "No recording yet",
        "recordings_appear_after_lessons": "New class recordings will appear here after lessons.",
        "administrative_reminders_and_notices": "Administrative reminders and class notices from the supervisor.",
        "no_announcements_right_now": "No announcements right now",
        "announcements_appear_here": "Administrative reminders and notices will appear here.",
        "homework_board": "Homework Board",
        "open_homework": "Open Homework",
        "no_homework_board_yet": "No homework board yet.",
        "class_board": "Class Board",
        "open_class_board": "Open Class Board",
        "no_class_board_yet": "No class board yet.",
        "books": "Books",
        "study_books_for_class": "Study books for your class",
        "book": "Book",
        "open_book": "Open book",
        "no_books_yet": "No books yet",
        "supervisor_add_books_soon": "Your supervisor will add them soon.",
        "no_books_available_yet": "No books available yet.",
        "extra_learning": "Extra Learning",
        "extra_learning_note": "Helpful links and extra practice. Your main books stay above in the Books section.",
        "helpful_links": "Helpful Links",
        "extra_practice_links": "Extra practice and support links",
        "extra": "Extra",
        "open_resource": "Open resource",
        "no_extra_resources_yet": "No extra resources yet",
        "no_extra_learning_links_yet": "No extra learning links yet.",
        "class_recordings": "Class Recordings",
        "lesson_date": "Lesson Date",
        "title_label": "Title",
        "summary": "Summary",
        "weekly_homework": "Weekly Homework",
        "recording": "Recording",
        "watch": "Watch",
        "no_recordings_available": "No recordings available yet.",
        "assignments_section": "Assignments",
        "student_assignments_note": "Read the task, send your answer, and check your score after the teacher reviews it.",
        "due": "Due",
        "pending": "Pending",
        "score": "Score",
        "file": "File",
        "uploaded": "Uploaded",
        "no_file": "No File",
        "what_to_do": "What To Do",
        "open_resource_link": "Open Resource Link",
        "download_resource": "Download Resource",
        "my_uploaded_file": "My Uploaded File",
        "teacher_notes": "Teacher Notes",
        "write_your_answer": "Write your answer",
        "write_your_answer_here": "Write your answer here",
        "or_add_a_link": "Or add a link",
        "paste_answer_link": "Paste your answer link",
        "or_upload_a_file": "Or upload a file",
        "send_my_homework": "Send My Homework",
        "my_results": "My Results",
        "choose_exam": "Choose exam",
        "choose_result": "Choose Result",
        "open_result_pdf": "Open Result PDF",
        "choose_exam_then_pdf": "Choose an exam, then open the PDF.",
        "no_exam_results_yet": "No exam results available yet.",
        "upcoming_exams": "Upcoming Exams",
        "date": "Date",
        "subject": "Subject",
        "time": "Time",
        "notes": "Notes",
        "academic_week_label": "Academic Week",
        "teaching_week_label": "Teaching Week",
        "non_teaching_week": "Non-Teaching Week",
        "no_lessons_scheduled_this_week": "No lessons are scheduled in this week.",
        "no_upcoming_exams": "No upcoming exams announced yet.",
        "no_assignments_available_yet": "No assignments available yet.",
        "teacher_weekly_message_complete": "All weekly teaching requirements are currently complete.",
        "levels_on_track": "On Track",
        "teacher_level_workspace": "Teacher Level Workspace",
        "workspace_subtitle": "Level students, Zoom details, and preparation entry point.",
        "teacher_control_center": "Teacher Control Center",
        "upcoming_exams_label": "Upcoming Exams",
        "recordings_label": "Recordings",
        "zoom_link_label": "Zoom Link",
        "zoom_email_label": "Zoom Email",
        "zoom_meeting_id_label": "Zoom Meeting ID",
        "zoom_passcode_label": "Zoom Passcode",
        "resources": "Resources",
        "class_tools": "Class Tools",
        "padlet_zoom_setup": "Padlet and Zoom setup",
        "templates_reports_scores": "Templates, reports, and scores",
        "plan_dates_announcements": "Plan dates and announcements",
        "manage_lesson_recordings": "Manage lesson recordings",
        "create_tasks_review_submissions": "Create tasks and review submissions",
        "see_student_list_reports": "See student list and reports",
        "padlet_links": "Padlet Links",
        "homework_padlet_url": "Homework Padlet URL",
        "announcements_padlet_url": "Announcements Padlet URL",
        "save_padlet_links": "Save Padlet Links",
        "books_and_curriculum": "Books And Curriculum",
        "level_curriculum": "Level curriculum",
        "supervisor_not_added_links": "Supervisor has not added links yet.",
        "no_books_added_level": "No books added for this level yet.",
        "shared_links_for_teachers": "Shared links for teachers",
        "supervisor_not_added_shared_links": "Supervisor has not added shared links yet.",
        "no_shared_teacher_resources": "No shared teacher resources yet.",
        "add_lesson_link": "Add Lesson Link",
        "lesson_title": "Lesson title",
        "lesson_zoom_link": "Lesson / Zoom link",
        "short_lesson_summary": "Short lesson summary",
        "add_lesson": "Add Lesson",
        "lesson_links_and_homework": "Lesson Links And Homework",
        "url": "URL",
        "delete": "Delete",
        "no_lesson_links_yet": "No lesson links yet.",
        "update_recordings": "Update Recordings",
        "create_assignment": "Create Assignment",
        "assignment_title": "Assignment title",
        "optional_resource_link": "Optional resource link",
        "instructions_questions_tasks": "Instructions, questions, or task details",
        "add_assignment": "Add Assignment",
        "assignment_review": "Assignment Review",
        "select_assignment": "Select assignment",
        "open_assignment": "Open Assignment",
        "resource": "Resource",
        "link": "Link",
        "instructions": "Instructions",
        "submission": "Submission",
        "feedback": "Feedback",
        "submission_link": "Submission Link",
        "active_state": "Active",
        "closed_state": "Closed",
        "update_assignment_reviews": "Update Assignment Reviews",
        "no_student_submissions_yet": "No student submissions yet.",
        "select_assignment_review_marks": "Select an assignment to review submissions and enter marks.",
        "no_assignments_created_yet": "No assignments created yet.",
        "students_section": "Students",
        "report": "Report",
        "class_exam_overview": "Class Exam Overview",
        "template_exam_entry": "Template Exam Entry",
        "exam_results_label": "Exam Results",
        "show_class_report": "Show Class Report",
        "open_exam_table": "Open Exam Table",
        "open_results_table": "Open Results Table",
        "save_class_results": "Save Class Results",
        "update_exam_results": "Update Exam Results",
        "add_upcoming_exam": "Add Upcoming Exam",
        "update_upcoming_exams": "Update Upcoming Exams",
        "monthly_reports_review": "Monthly Reports Review",
        "students_page": "Students",
        "month_label": "Month",
        "monthly_review_subtitle": "Review teacher notes, exclude any student, then send the approved monthly reports.",
        "total_students": "Total Students",
        "ready_to_send": "Ready To Send",
        "missing_notes": "Missing Notes",
        "missing_email": "Missing Email",
        "missing_whatsapp": "Missing WhatsApp",
        "all_students": "All Students",
        "ready_only": "Ready Only",
        "not_ready_only": "Not Ready Only",
        "apply": "Apply",
        "reset": "Reset",
        "select_all_ready": "Select All Ready",
        "unselect_all": "Unselect All",
        "send": "Send",
        "parent_email": "Parent Email",
        "parent_whatsapp": "Parent WhatsApp",
        "teacher_note": "Teacher Note",
        "review": "Review",
        "ready": "Ready",
        "open_report": "Open Report",
        "no_students_found": "No students found.",
        "send_selected_monthly_reports": "Send Selected Monthly Reports",
        "student_report": "Student Report",
        "period": "Period",
        "this_week": "This Week",
        "this_month": "This Month",
        "export_pdf": "Export PDF",
        "export_excel": "Export Excel",
        "student_name": "Student Name",
        "date_range": "Date Range",
        "present_count": "Present Count",
        "absent_count": "Absent Count",
        "late_count": "Late Count",
        "excused_count": "Excused Count",
        "attendance_percentage": "Attendance Percentage",
        "imported_arabic_attendance": "Imported Arabic Attendance",
        "imported_arabic_attendance_subtitle": "Snapshot extracted from the shared Arabic attendance workbook.",
        "attendance_snapshot_matched_name": "Matched Workbook Name",
        "source_sheet": "Source Sheet",
        "total_sessions": "Total Sessions",
        "homework_given": "Homework Given",
        "homework_submitted": "Homework Submitted",
        "homework_reviewed": "Homework Reviewed",
        "homework_waiting_review": "Homework Waiting Review",
        "homework_missing": "Homework Missing",
        "homework_completion": "Homework Completion",
        "monthly_teacher_note": "Monthly Teacher Note",
        "monthly_teacher_note_for_student": "Monthly Teacher Note (for this student)",
        "save_monthly_note": "Save Monthly Note",
        "send_report_to_parent_email": "Send Report to Parent Email",
        "send_report": "Send Report",
        "view_only": "View Only",
        "student_view_only_report": "Student account can only view report details.",
        "back_to_students": "Back to students",
        "back_to_student_interface": "Back to student interface",
        "students_management": "Students",
        "students_subtitle": "Reports, parent emails, parent WhatsApp numbers, and quick student updates.",
        "add_student": "Add Student",
        "student_full_name": "Student full name",
        "status_optional": "Status (optional)",
        "student_year": "Student year",
        "parent_email_optional": "Parent email (optional)",
        "parent_whatsapp_optional": "Parent WhatsApp (optional)",
        "open_monthly_review": "Open Monthly Review",
        "full_name": "Full Name",
        "level_name": "Level Name",
        "parent_contact": "Parent Contact",
        "update": "Update",
        "save": "Save",
        "attendance_records_title": "Attendance Records",
        "student_name_label": "Student Name",
        "no_attendance_records_found": "No attendance records found.",
        "back_to_attendance_entry": "Back to attendance entry",
        "back_to_dashboard": "Back to dashboard",
        "level_attendance_report": "Level Attendance Report",
        "summary_label": "Summary",
        "total_present": "Total Present",
        "total_absent": "Total Absent",
        "total_late": "Total Late",
        "total_excused": "Total Excused",
        "back_to_level_details": "Back to level details",
        "back_to_levels": "Back to levels",
        "missing_attendance": "Missing Attendance",
        "partial_attendance": "Partial Attendance",
        "students_needing_attention": "Students Needing Attention",
        "active_assignments": "Active Assignments",
        "teachers_on_track": "Teachers On Track",
        "teachers_need_follow_up": "Teachers Need Follow-up",
        "cleanup_critical": "Cleanup Critical",
        "cleanup_total": "Cleanup Total",
        "deployment_checklist": "Deployment Checklist",
        "todays_gaps": "Today's Gaps",
        "immediate_supervisor_actions": "Immediate Supervisor Actions",
        "open_level": "Open Level",
        "edit_level": "Edit Level",
        "no_urgent_daily_gaps": "No urgent daily gaps",
        "urgent_daily_gaps_clear": "Attendance, Zoom links, active homework, and review queues are currently in good shape.",
        "supervisor_tools": "Supervisor Tools",
        "core_access": "Core Access",
        "core_links_access_management": "Core links and access management without leaving the main dashboard.",
        "new_access_code": "New access code",
        "update_code": "Update Code",
        "confirm_change_admin_code": "Confirm changing admin access code?",
        "supervisor_review_before_sending": "Supervisor Review Before Sending",
        "monthly_report_ready_for_supervisor": "Monthly report is ready for supervisor sending.",
        "weekly_archive": "Weekly Archive",
        "saved_weekly_reports": "Saved Weekly Reports",
        "updated": "Updated",
        "size": "Size",
        "download": "Download",
        "no_archived_weekly_reports": "No archived weekly reports yet.",
        "roster": "Roster",
        "student_code": "Student Code",
        "student_year_label": "Student Year",
        "total": "Total",
        "percentage": "Percentage",
        "grade": "Grade",
        "pdf": "PDF",
        "no_branches_defined": "No branches defined for this template yet.",
        "no_exam_results_level": "No exam results available for this level yet.",
        "exam_title_placeholder": "Exam title",
        "submitted_state": "Submitted",
        "reviewed_state": "Reviewed",
        "needs_update_state": "Needs Update",
        "delete_upcoming_exam_confirm": "Delete this upcoming exam?",
        "delete_recording_confirm": "Delete this recording?",
        "recent_system_actions": "Recent System Actions",
        "latest_logged_updates": "Latest logged updates from teachers and supervisors.",
        "no_recent_actions": "No recent actions have been logged yet.",
        "actor": "Actor",
        "action": "Action",
        "action_details": "Details",
        "logged_time": "Time",
        "system_action_log": "System Action Log",
        "action_log_subtitle": "Filter recent teacher and supervisor operations by actor, class, or action type.",
        "all_roles": "All Roles",
        "all_actions": "All Actions",
        "all_levels": "All Levels",
        "search_label": "Search",
        "search_actions_placeholder": "Search actor, class, or details",
        "open_action_log": "Open Action Log",
        "supervisor_role": "Supervisor",
        "teacher_role": "Teacher",
        "system_role": "System",
        "entity": "Entity",
        "open_link": "Open",
        "teacher_thanks_sent": "Teacher Thanks Sent",
        "admin_access_code_updated": "Admin Access Code Updated",
        "announcement_published": "Announcement Published",
        "announcement_updated": "Announcement Updated",
        "calendar_settings_updated": "Calendar Settings Updated",
        "holiday_added": "Holiday Added",
        "holiday_updated": "Holiday Updated",
        "academic_calendar_event_added": "Academic Calendar Event Added",
        "academic_calendar_event_updated": "Academic Calendar Event Updated",
        "weekly_followup_generated": "Weekly Follow-up Generated",
        "student_added": "Student Added",
        "student_updated": "Student Updated",
        "parent_contact_saved": "Parent Contact Saved",
        "bulk_reports_sent": "Bulk Reports Sent",
        "monthly_reports_sent": "Monthly Reports Sent",
        "student_report_sent": "Student Report Sent",
        "curriculum_progress_updated": "Curriculum Progress Updated",
        "attendance_saved": "Attendance Saved",
        "curriculum_plan_followup": "Curriculum Plan Follow-up",
        "curriculum_plan_progress": "Curriculum Plan Progress",
        "completed_items_label": "Completed Items",
        "pending_items_label": "Pending Items",
        "in_progress_items_label": "In Progress Items",
        "last_completed_item": "Last Completed Item",
        "plan_note": "Plan Note",
        "save_progress": "Save Progress",
        "pending_state": "Not Started",
        "in_progress_state": "In Progress",
        "completed_state": "Completed",
        "progress_percentage": "Progress",
        "no_curriculum_plan_items": "No curriculum plan items yet.",
        "teacher_plan_checkpoint": "Teacher plan checkpoint",
        "start_here": "Start Here",
        "syllabus_plan": "Syllabus Plan",
        "plan_sheet_subtitle": "An Excel-like yearly plan sheet for this class.",
        "week_number": "Week",
        "lesson_reference": "Reference",
        "learning_objective": "Learning Objective",
        "planned_homework_label": "Planned Homework",
        "planned_state": "Planned",
        "postponed_state": "Postponed",
        "add_plan_row": "Add Plan Row",
        "save_plan_sheet": "Save Plan Sheet",
        "open_syllabus_plan": "Open Syllabus Plan",
        "download_plan_pdf": "Download Plan PDF",
        "download_plan_template": "Download Excel Template",
        "import_plan_template": "Import Excel Plan",
        "excel_label": "Excel",
        "plan_template_hint": "Download the official template, fill it, then upload it here. Importing replaces the current plan for this class.",
        "plan_template_file_label": "Excel plan file",
        "import_preview_ready": "Import Preview Ready",
        "import_preview_hint": "Review these counts and sample rows before applying the import. Applying the preview will replace the current plan for this class.",
        "preview_scheduled_rows": "Preview Scheduled Rows",
        "preview_reserve_rows": "Preview Reserve Rows",
        "apply_import": "Apply Import",
        "discard_preview": "Discard Preview",
        "changed_plan_rows": "Changed Plan Rows",
        "current_value": "Current",
        "new_value": "New",
        "plan_comparison_preview": "Plan Comparison Preview",
        "plan_import_matches_current": "This import matches the current plan. No lesson rows will change.",
        "editing_locked": "Editing Locked",
        "editing_open": "Editing Open",
        "allow_teacher_plan_editing": "Allow Teacher Editing",
        "lock_teacher_plan_editing": "Lock Teacher Editing",
        "read_only_plan": "Read-Only Plan",
        "teacher_can_mark_completed_only": "Teachers can only mark lessons completed while editing is locked.",
        "no_plan_rows_yet": "No plan rows added yet.",
        "plan_rows_count": "Plan Rows",
        "plan_entry_added": "Plan Entry Added",
        "plan_sheet_updated": "Plan Sheet Updated",
        "teacher_plan_overview": "Teacher Plan Overview",
        "select_level_to_review_plan": "Select a class to review its syllabus plan.",
        "weeks_count_label": "40 Weeks",
        "lessons_per_week_label": "4 Lessons",
        "levels_on_track": "Classes On Track",
        "levels_delayed": "Delayed Classes",
        "view_full_plan_followup": "View Full Plan Follow-up",
        "class_followup_register": "Class Follow-up Register",
        "student_followup_subtitle": "One row per student for attendance, homework, latest exam, and monthly note readiness.",
        "monthly_note_status": "Monthly Note",
        "note_ready": "Ready",
        "note_missing": "Missing",
        "open_followup": "Open Follow-up",
        "supervisor_level_followup_subtitle": "Supervisor view of student attendance, homework, latest exam, and monthly note readiness for this class.",
        "followup_attention_levels": "Follow-up Attention Levels",
        "classes_with_student_flags": "Classes with student flags from attendance, homework, exam, and monthly note readiness.",
        "students_flagged": "Students Flagged",
        "student_overall_status": "Overall Status",
        "excellent_status": "Excellent",
        "stable_status": "Stable",
        "needs_attention_status": "Needs Attention",
        "sort_by": "Sort By",
        "lowest_attendance_first": "Lowest Attendance First",
        "most_open_assignments_first": "Most Open Assignments First",
        "missing_monthly_note_first": "Missing Monthly Note First",
        "excellent_students": "Excellent Students",
        "stable_students": "Stable Students",
        "students_need_attention": "Students Need Attention",
        "review_lesson": "Review Lesson",
        "empty_plan_cell": "Empty planning cell",
        "fill_plan_cell_hint": "Add the book, unit, and lesson when you are ready to build this part of the yearly plan.",
        "complete_previous_prompt": "Previous lessons are still incomplete. Press OK to mark this lesson and all previous lessons as completed.",
        "complete_only_current_prompt": "Press OK to mark only this lesson as completed. Press Cancel to go back without any change.",
        "lesson_slot_1": "Lesson 1",
        "lesson_slot_2": "Lesson 2",
        "lesson_slot_3": "Lesson 3",
        "lesson_slot_4": "Lesson 4",
        "book_name": "Book",
        "unit_name": "Unit",
        "lesson_name": "Lesson",
        "execution_status": "Execution Status",
        "completed_lessons": "Completed Lessons",
        "planned_lessons": "Planned Lessons",
        "current_teaching_week": "Current Teaching Week",
        "expected_by_now": "Expected By Now",
        "delayed_lessons": "Delayed Lessons",
        "unscheduled_plan_rows": "Unscheduled Plan Rows",
        "unscheduled_plan_caption": "These rows were imported from the workbook but did not fit into the remaining instructional weeks in the academic calendar.",
        "reserve_plan_rows": "Reserve / Unscheduled Rows",
        "mark_completed": "Mark Completed",
        "completed_on": "Completed On",
        "on_track_status": "On Track",
        "delayed_status": "Delayed",
        "plan_entry_completed": "Plan Entry Completed",
        "save_current_week": "Save Current Week",
        "recording_added": "Recording Added",
        "recordings_updated": "Recordings Updated",
        "assignment_added": "Assignment Added",
        "assignment_reviews_updated": "Assignment Reviews Updated",
        "template_results_saved": "Template Results Saved",
        "exam_results_updated": "Exam Results Updated",
        "upcoming_exam_added": "Upcoming Exam Added",
        "upcoming_exams_updated": "Upcoming Exams Updated",
        "monthly_note_saved": "Monthly Note Saved",
        "close_menu": "Close menu",
        "attendance_recorded_missing": "Attendance recorded for {recorded}/{expected} teaching day(s) - missing: {days}",
        "zoom_recordings_uploaded_missing": "Zoom recordings uploaded for {uploaded}/{expected} teaching day(s) - missing: {days}",
        "no_homework_assigned_week": "No homework was assigned during the teaching week",
        "all_assigned_levels_on_track": "All assigned levels are currently on track.",
        "teachers_weekly_followup": "Teacher Weekly Follow-up",
        "teacher_weekly_report": "Teacher Report This Week",
        "teacher_label_short": "Teacher",
        "completed_status": "Completed",
        "levels_count_label": "Levels",
        "students_count_label": "Students",
        "levels_need_followup": "Levels Needing Follow-up",
        "week_status_label": "Weekly Status",
        "weekly_summary_label": "Weekly Summary",
        "missing_attendance_label": "Missing Attendance",
        "attendance_complete_label": "Attendance Complete",
        "missing_recordings_label": "Missing Recordings",
        "recordings_complete_label": "Recordings Complete",
        "no_teacher_weekly_data": "No teacher weekly follow-up data yet.",
        "teacher_monthly_report": "Teacher Monthly Report",
        "ideal_teacher_and_thanks": "Ideal Teacher And Appreciation Messages",
        "month_label": "Month",
        "teachers_count_label": "Teachers",
        "ready_for_thanks": "Ready For Appreciation",
        "ideal_teacher": "Ideal Teacher",
        "score_label": "Score",
        "monthly_report_label": "Monthly Report",
        "attendance_percentage_label": "Attendance",
        "recordings_percentage_label": "Recordings",
        "weekly_homework_label": "Weekly Homework",
        "pending_reviews_label": "Pending Reviews",
        "email_ready_label": "Email",
        "ready_label": "Ready",
        "not_ready_label": "Not Ready",
        "level_breakdown_label": "Level Breakdown",
        "attendance_word": "Attendance",
        "recordings_word": "Recordings",
        "homework_word": "Homework",
        "send_appreciation_message": "Send Appreciation Message",
        "teachers_master_excel": "Teachers Master Excel",
        "teachers_master_pdf": "Teachers Master PDF",
        "level_full_excel": "Level Full Excel",
        "level_full_pdf": "Level Full PDF",
        "student_full_excel": "Student Full Excel",
        "student_full_pdf": "Student Full PDF",
        "reports_center": "Reports Center",
        "open_reports_center": "Open Reports Center",
        "executive_reports_exports": "Executive Reports And Exports",
        "teachers_master_report": "Teachers Master Report",
        "download_unified_teacher_sheet": "Download the unified monthly teacher sheet.",
        "level_full_report": "Level Full Report",
        "select_level_to_export": "Select a level, then export the full follow-up sheet.",
        "student_full_report": "Student Full Report",
        "select_student_to_export": "Select a student, then export the full student report workbook.",
        "export_selected_level_excel": "Export Selected Level Excel",
        "export_selected_student_excel": "Export Selected Student Excel",
        "ideal_status": "Ideal",
        "excellent_status": "Excellent",
        "good_status": "Good",
        "no_teacher_monthly_data": "No monthly teacher data yet.",
    },
    "ar": {
        "language": "اللغة",
        "role_selection": "اختيار الدور",
        "logout": "تسجيل الخروج",
        "head_dashboard": "لوحة الإشراف",
        "data_root": "جذر البيانات",
        "teacher_dashboard": "لوحة المعلم",
        "admin_dashboard": "لوحة المشرف",
        "attendance": "الحضور",
        "attendance_records": "سجلات الحضور",
        "daily_attendance": "الحضور اليومي",
        "select_a_level": "اختر فصلاً",
        "selected_date": "التاريخ المحدد",
        "students": "الطلاب",
        "preset_default": "الحالة الافتراضية",
        "present": "حاضر",
        "absent": "غائب",
        "late": "متأخر",
        "excused": "بعذر",
        "choose_class_date": "اختر الفصل ثم حمّل الطلاب وبعدها سجّل الحضور للتاريخ المحدد.",
        "mark_students_for_date": "سجّل حضور الطلاب ليوم {date} ثم احفظ التحديثات لهذا الفصل.",
        "start_by_selecting_class": "ابدأ باختيار الفصل والتاريخ الذي تريد مراجعته.",
        "load_attendance": "تحميل الحضور",
        "class_and_date": "الفصل والتاريخ",
        "load_students": "تحميل الطلاب",
        "quick_actions": "إجراءات سريعة",
        "fast_attendance_shortcuts": "اختصارات الحضور السريعة",
        "mark_all_present": "تحديد الجميع حاضر",
        "mark_all_absent": "تحديد الجميع غائب",
        "mark_all_late": "تحديد الجميع متأخر",
        "mark_all_excused": "تحديد الجميع بعذر",
        "student_register": "سجل الطلاب",
        "attendance_list_title": "قائمة الحضور",
        "save_attendance": "حفظ الحضور",
        "saved_summary": "ملخص الحفظ",
        "attendance_summary_after_saving": "ملخص الحضور بعد الحفظ",
        "students_marked_present": "طلاب تم تسجيلهم حضوراً",
        "students_marked_absent": "طلاب تم تسجيلهم غياباً",
        "students_marked_late": "طلاب تم تسجيلهم متأخرين",
        "students_marked_excused": "طلاب تم تسجيلهم بعذر",
        "present_students_out_of_class": "نسبة الحضور من إجمالي طلاب الفصل",
        "present_rate": "نسبة الحضور",
        "no_students_for_level": "لا يوجد طلاب في هذا الفصل.",
        "attendance_saved_successfully": "تم حفظ الحضور بنجاح.",
        "supervisor_dashboard": "لوحة المشرف",
        "executive_control": "التحكم التنفيذي",
        "one_view_supervisor": "واجهة واحدة لمتابعة الحضور والواجبات والامتحانات والطلاب الذين يحتاجون متابعة.",
        "levels": "الفصول",
        "teachers": "المعلمون",
        "week_start": "بداية الأسبوع",
        "today": "اليوم",
        "student_interface": "واجهة الطالب",
        "enter_student_code": "أدخل كود الطالب لعرض الروابط والتقارير ومعلومات الفصل.",
        "find_student": "بحث عن الطالب",
        "student_code_placeholder": "كود الطالب مثل STD-000001",
        "student_not_found": "لم يتم العثور على طالب بهذا الكود: {code}",
        "teacher_signed_in_as": "تم تسجيل الدخول باسم: {name}",
        "quick_actions": "إجراءات سريعة",
        "today_tasks": "مهام اليوم",
        "focus_on_these_first": "ابدأ بهذه أولاً",
        "before_class": "قبل الحصة",
        "during_class": "أثناء الحصة",
        "after_class": "بعد الحصة",
        "daily_workflow": "سير العمل اليومي",
        "daily_workflow_help": "ابدأ من هنا لتعرف ما المطلوب قبل الدرس وأثناءه وبعده.",
        "prepare_zoom_and_plan": "جهز الزوم والروابط وخطة الدرس.",
        "take_attendance_and_teach": "سجل الحضور وابدأ تنفيذ الدرس المخطط.",
        "upload_recording_and_homework": "ارفع سجل الحصة وأضف الواجب أو راجع تسليماته.",
        "open_attendance": "فتح الحضور",
        "open_first_workspace": "فتح أول مساحة عمل",
        "open_review_queue": "فتح طابور المراجعة",
        "teaching_day": "اليوم الدراسي",
        "teacher_dashboard_subtitle": "فصولك، ومتابعة الحضور، وطابور المراجعة، والطلاب الذين يحتاجون متابعة في مكان واحد.",
        "plan_main_fields_hint": "ابدأ بالكتاب والوحدة والدرس. افتح التفاصيل الإضافية فقط عند الحاجة.",
        "more_details": "تفاصيل إضافية",
        "welcome_back": "مرحباً بعودتك",
        "monthly_report": "التقرير الشهري",
        "weekly_report_label": "التقرير الأسبوعي",
        "weekly_pdf_label": "PDF أسبوعي",
        "open_matching": "فتح المطابقة",
        "data_cleanup": "تنظيف البيانات",
        "critical_records_snapshot": "ملخص السجلات الحرجة",
        "open_cleanup": "فتح التنظيف",
        "critical_label": "حرج",
        "students_only": "الطلاب فقط",
        "teachers_only": "المعلمون فقط",
        "levels_only": "الفصول فقط",
        "all_cleanup": "كل التنظيف",
        "critical_only": "الحرج فقط",
        "students_cleanup": "تنظيف الطلاب",
        "teachers_cleanup": "تنظيف المعلمين",
        "levels_cleanup": "تنظيف الفصول",
        "fix_record": "إصلاح السجل",
        "no_critical_cleanup_blockers": "لا توجد عوائق حرجة في التنظيف",
        "remaining_cleanup_review_only": "العناصر المتبقية في التنظيف هي للمراجعة فقط ولا تعيق التشغيل اليومي حالياً.",
        "operating_calendar": "تقويم التشغيل",
        "duty_time_and_weekly_followup": "أوقات الدوام والمتابعة الأسبوعية",
        "calendar_settings_help": "حدد فترة الدوام ويوم التذكير الأسبوعي. سيتجاوز النظام التوليد أثناء الإجازات النشطة.",
        "save_calendar_settings": "حفظ إعدادات التقويم",
        "weekly_followup_status": "حالة المتابعة الأسبوعية",
        "last_successful_scheduled_run": "آخر تشغيل ناجح مجدول: {value}",
        "not_yet_run": "لم يُشغّل بعد",
        "generate_weekly_followup_now": "تشغيل المتابعة الأسبوعية الآن",
        "holiday_calendar": "تقويم الإجازات",
        "academic_holidays": "الإجازات الأكاديمية",
        "holiday_title": "عنوان الإجازة",
        "add_holiday": "إضافة إجازة",
        "active_label": "نشط",
        "inactive_label": "غير نشط",
        "no_holidays_added": "لم تتم إضافة إجازات بعد",
        "weekly_generator_runs_normally": "سيعمل مولد المتابعة الأسبوعية بشكل طبيعي حتى تتم إضافة فترات إجازة هنا.",
        "broadcast_center": "مركز التعميم",
        "post_administrative_announcement": "نشر إعلان إداري",
        "announcement_posting_help": "استخدم هذا للتعاميم الإدارية أو التذكيرات أو تنبيهات المتابعة مثل نقص رابط زوم أو الواجب أو تجهيز الطلاب.",
        "announcement_title": "عنوان الإعلان",
        "all_audience": "الكل",
        "teachers_only_audience": "للمعلمين فقط",
        "students_only_audience": "للطلاب فقط",
        "general_category": "عام",
        "reminder_category": "تذكير",
        "alert_category": "تنبيه",
        "follow_up_category": "متابعة",
        "all_levels": "كل الفصول",
        "announcement_body_placeholder": "اكتب الإعلان أو التوقع أو المتابعة التشغيلية هنا.",
        "pin_this_announcement": "تثبيت هذا الإعلان",
        "publish_announcement": "نشر الإعلان",
        "live_feed": "البث المباشر",
        "recent_announcements": "أحدث الإعلانات",
        "archived_state": "مؤرشف",
        "no_announcements_yet": "لا توجد إعلانات بعد",
        "published_messages_here": "ستظهر هنا رسائلك الإدارية المنشورة.",
        "student_risk_radar": "رادار مخاطر الطلاب",
        "students_needing_followup": "الطلاب الذين يحتاجون متابعة",
        "student_label": "الطالب",
        "risk_label": "المخاطر",
        "teacher_label": "المعلم",
        "weekly_standing": "الحالة الأسبوعية",
        "monday_to_thursday_status": "حالة الإنجاز من الاثنين إلى الخميس",
        "needs_completion": "بحاجة إكمال",
        "on_track_for_appreciation": "على مسار الشكر",
        "review_window": "نافذة المراجعة: {start} إلى {end}.",
        "levels_followup_appreciation": "الفصول التي تحتاج متابعة: {follow_up}. الفصول على مسار الشكر: {praise}.",
        "level": "الفصل",
        "status": "الحالة",
        "weekly_review": "المراجعة الأسبوعية",
        "follow_up": "متابعة",
        "appreciation": "شكر",
        "all_weekly_tasks_complete": "كل المهام الأسبوعية المطلوبة مكتملة.",
        "no_weekly_review_data": "لا توجد بيانات مراجعة أسبوعية بعد.",
        "administrative_board": "اللوحة الإدارية",
        "supervisor_announcements": "إعلانات المشرف",
        "pinned": "مثبت",
        "no_administrative_announcements": "لا توجد إعلانات إدارية",
        "supervisor_messages_here": "ستظهر هنا رسائل المشرف وتعليمات المتابعة.",
        "priority_students": "الطلاب ذوو الأولوية",
        "needs_followup": "يحتاجون متابعة",
        "attendance_14d": "حضور 14 يوماً",
        "arabic_attendance_snapshot": "ملخص حضور العربية",
        "attendance_source_label": "مصدر الحضور",
        "attendance_source_recent": "سجل آخر 14 يوماً",
        "attendance_source_imported": "ملف حضور العربية المستورد",
        "attendance_source_none": "لا توجد بيانات حضور بعد",
        "current_student_status": "الحالة الحالية للطالب",
        "attendance_snapshot_caption": "الحضور المستخدم حالياً في حساب حالتك العامة",
        "arabic_attendance_import": "استيراد حضور العربية",
        "arabic_attendance_import_subtitle": "حدّث ملف حضور العربية المشترك وثبّت مطابقة الأسماء القوية.",
        "refresh_arabic_attendance": "تحديث حضور العربية",
        "upload_arabic_attendance_file": "رفع ملف حضور العربية",
        "current_attendance_file": "ملف الحضور الحالي",
        "download_current_attendance_file": "تنزيل الملف الحالي",
        "attendance_import_steps": "1. ارفع ملفاً جديداً عند الحاجة. 2. اضغط تحديث حضور العربية. 3. راجع التقارير ودفاتر المتابعة.",
        "no_arabic_attendance_history": "لا توجد بعد أي عمليات رفع أو تحديث مسجلة لحضور العربية.",
        "download_latest_backup": "تنزيل آخر نسخة احتياطية",
        "arabic_attendance_name_review": "مراجعة أسماء حضور العربية",
        "arabic_attendance_name_review_subtitle": "إذا فشل أي اسم في المطابقة فراجعه هنا قبل الاعتماد الكامل على الحضور المستورد.",
        "best_system_match": "أقرب اسم في النظام",
        "match_score": "درجة المطابقة",
        "imported_rows": "الصفوف المستوردة",
        "matched_aliases": "الأسماء المثبتة",
        "matched_rows": "الصفوف المطابقة",
        "unmatched_rows": "الصفوف غير المطابقة",
        "last_arabic_attendance_refresh": "آخر تحديث لحضور العربية",
        "open_assignments": "الواجبات المفتوحة",
        "latest_exam": "آخر اختبار",
        "why": "السبب",
        "reason_low_attendance": "ضعف في الحضور",
        "reason_pending_assignments": "واجبات بانتظار الإنجاز",
        "reason_one_open_assignment": "واجب مفتوح واحد",
        "reason_low_exam_result": "انخفاض في نتيجة الاختبار",
        "reason_attendance": "الحضور",
        "reason_assignments": "الواجبات",
        "reason_exam": "الاختبار",
        "reason_result": "النتيجة",
        "reason_monthly_note": "الملاحظة الشهرية",
        "no_students_flagged": "لا يوجد طلاب محددون للمتابعة حالياً.",
        "class_operations": "تشغيل الفصول",
        "my_levels": "فصولي",
        "my_levels_caption": "الفصول المسندة لك في هذه اللوحة",
        "attendance_logged_today": "الحضور المسجل اليوم",
        "attendance_logged_today_caption": "تم تسجيل {marked} من أصل {total} سجل حضور للطلاب اليوم",
        "review_queue_label": "طابور المراجعة",
        "review_queue_caption": "واجبات مسلمة ما زالت تنتظر مراجعتك",
        "upcoming_exams_caption": "اختبارات مجدولة قادمة لفصولك",
        "plan_progress_label": "تقدم الخطة",
        "plan_progress_caption": "نسبة إنجاز خطة المنهج السنوية عبر فصولك",
        "assigned_level": "الفصل المكلّف به",
        "on_track": "منجز",
        "attendance_today": "حضور اليوم",
        "not_started": "لم يبدأ",
        "review_queue": "طابور المراجعة",
        "assignments": "الواجبات",
        "present_label": "الحضور",
        "all_weekly_teaching_complete": "كل متطلبات التدريس الأسبوعية مكتملة حالياً.",
        "next_exam": "الاختبار القادم",
        "open_workspace": "فتح مساحة العمل",
        "open_tasks_for_this_class": "فتح مهام هذا الفصل",
        "class_ready": "هذا الفصل في وضع جيد اليوم.",
        "class_needs_action": "هذا الفصل ما زال يحتاج إجراء اليوم.",
        "task_attendance_missing_title": "الحضور ما زال ناقصاً",
        "task_attendance_missing_body": "يوجد {count} فصل/فصول لم يُسجل لها الحضور اليوم بعد.",
        "task_review_queue_title": "هناك واجبات تحتاج مراجعة",
        "task_review_queue_body": "يوجد {count} تسليم/تسليمات ما زالت بانتظار مراجعتك.",
        "task_syllabus_delay_title": "تقدم الخطة يحتاج تحديثاً",
        "task_syllabus_delay_body": "يوجد {count} حصة متأخرة عن الأسبوع التعليمي الحالي.",
        "task_clear_title": "لا توجد مهام عاجلة الآن",
        "task_clear_body": "مهامك الأساسية اليوم في وضع جيد حالياً.",
        "take_attendance": "تسجيل الحضور",
        "join_zoom": "دخول زوم",
        "no_assigned_levels": "لا توجد فصول مسندة حالياً.",
        "code_label": "الكود",
        "what_needs_attention": "ما الذي يحتاج انتباهاً",
        "student_updates_now": "أهم التحديثات التي تحتاجها الآن.",
        "all_clear": "كل شيء جيد",
        "no_student_alerts": "لا توجد تنبيهات عاجلة للطالب حالياً.",
        "study_snapshot": "الملخص الدراسي",
        "latest_result": "آخر نتيجة",
        "no_result_yet": "لا توجد نتيجة بعد",
        "next_published_result_here": "ستظهر نتيجتك المنشورة القادمة هنا.",
        "no_exam_announced": "لا يوجد اختبار معلن",
        "check_back_exam_announcements": "عد لاحقاً لعرض إعلانات الاختبارات الجديدة.",
        "latest_recording": "آخر تسجيل",
        "no_recording_yet": "لا يوجد تسجيل بعد",
        "recordings_appear_after_lessons": "ستظهر تسجيلات الحصص الجديدة هنا بعد الدروس.",
        "administrative_reminders_and_notices": "التذكيرات الإدارية والتنبيهات الصفية من المشرف.",
        "no_announcements_right_now": "لا توجد إعلانات حالياً",
        "announcements_appear_here": "ستظهر هنا التذكيرات والتنبيهات الإدارية.",
        "homework_board": "لوحة الواجب",
        "open_homework": "فتح الواجب",
        "no_homework_board_yet": "لا توجد لوحة واجب بعد.",
        "class_board": "لوحة الفصل",
        "open_class_board": "فتح لوحة الفصل",
        "no_class_board_yet": "لا توجد لوحة فصل بعد.",
        "books": "الكتب",
        "study_books_for_class": "كتب الدراسة الخاصة بفصلك",
        "book": "كتاب",
        "open_book": "فتح الكتاب",
        "no_books_yet": "لا توجد كتب بعد",
        "supervisor_add_books_soon": "سيضيف المشرف الكتب قريباً.",
        "no_books_available_yet": "لا توجد كتب متاحة بعد.",
        "extra_learning": "تعلم إضافي",
        "extra_learning_note": "روابط مفيدة وتمارين إضافية. الكتب الأساسية تبقى في قسم الكتب بالأعلى.",
        "helpful_links": "روابط مفيدة",
        "extra_practice_links": "روابط مساعدة وتدريب إضافي",
        "extra": "إضافي",
        "open_resource": "فتح الرابط",
        "no_extra_resources_yet": "لا توجد موارد إضافية بعد",
        "no_extra_learning_links_yet": "لا توجد روابط تعلم إضافي بعد.",
        "class_recordings": "تسجيلات الحصص",
        "lesson_date": "تاريخ الدرس",
        "title_label": "العنوان",
        "summary": "الملخص",
        "weekly_homework": "واجب الأسبوع",
        "recording": "التسجيل",
        "watch": "مشاهدة",
        "no_recordings_available": "لا توجد تسجيلات متاحة بعد.",
        "assignments_section": "الواجبات",
        "student_assignments_note": "اقرأ المهمة ثم أرسل إجابتك وتابع درجتك بعد مراجعة المعلم.",
        "due": "التسليم",
        "pending": "قيد الانتظار",
        "score": "الدرجة",
        "file": "الملف",
        "uploaded": "مرفوع",
        "no_file": "لا يوجد ملف",
        "what_to_do": "المطلوب",
        "open_resource_link": "فتح رابط المورد",
        "download_resource": "تنزيل المورد",
        "my_uploaded_file": "ملفي المرفوع",
        "teacher_notes": "ملاحظات المعلم",
        "write_your_answer": "اكتب إجابتك",
        "write_your_answer_here": "اكتب إجابتك هنا",
        "or_add_a_link": "أو أضف رابطاً",
        "paste_answer_link": "الصق رابط الإجابة",
        "or_upload_a_file": "أو ارفع ملفاً",
        "send_my_homework": "إرسال الواجب",
        "my_results": "نتائجي",
        "choose_exam": "اختر الاختبار",
        "choose_result": "اختيار النتيجة",
        "open_result_pdf": "فتح ملف النتيجة PDF",
        "choose_exam_then_pdf": "اختر اختباراً ثم افتح ملف النتيجة.",
        "no_exam_results_yet": "لا توجد نتائج اختبارات متاحة بعد.",
        "upcoming_exams": "الاختبارات القادمة",
        "date": "التاريخ",
        "subject": "المادة",
        "time": "الوقت",
        "notes": "الملاحظات",
        "academic_week_label": "الأسبوع الأكاديمي",
        "teaching_week_label": "الأسبوع التدريسي",
        "non_teaching_week": "أسبوع غير تدريسي",
        "no_lessons_scheduled_this_week": "لا توجد دروس مجدولة في هذا الأسبوع.",
        "no_upcoming_exams": "لا توجد اختبارات قادمة معلنة بعد.",
        "no_assignments_available_yet": "لا توجد واجبات متاحة بعد.",
        "teacher_weekly_message_complete": "كل متطلبات التدريس الأسبوعية مكتملة حالياً.",
        "levels_on_track": "منجز",
        "teacher_level_workspace": "مساحة عمل المعلم",
        "workspace_subtitle": "طلاب الفصل وتفاصيل زوم ونقطة الدخول للتحضير وإدارة الحصة.",
        "teacher_control_center": "مركز تحكم المعلم",
        "upcoming_exams_label": "الاختبارات القادمة",
        "recordings_label": "التسجيلات",
        "zoom_link_label": "رابط زوم",
        "zoom_email_label": "بريد زوم",
        "zoom_meeting_id_label": "معرّف اجتماع زوم",
        "zoom_passcode_label": "رمز مرور زوم",
        "resources": "الموارد",
        "class_tools": "أدوات الفصل",
        "padlet_zoom_setup": "إعدادات Padlet وZoom",
        "templates_reports_scores": "القوالب والتقارير والدرجات",
        "plan_dates_announcements": "تخطيط التواريخ والإعلانات",
        "manage_lesson_recordings": "إدارة تسجيلات الحصص",
        "create_tasks_review_submissions": "إنشاء المهام ومراجعة التسليمات",
        "see_student_list_reports": "عرض الطلاب وتقاريرهم",
        "padlet_links": "روابط Padlet",
        "homework_padlet_url": "رابط Padlet للواجب",
        "announcements_padlet_url": "رابط Padlet للإعلانات",
        "save_padlet_links": "حفظ روابط Padlet",
        "books_and_curriculum": "الكتب والمنهج",
        "level_curriculum": "منهج الفصل",
        "supervisor_not_added_links": "لم يضف المشرف الروابط بعد.",
        "no_books_added_level": "لا توجد كتب مضافة لهذا الفصل بعد.",
        "shared_links_for_teachers": "روابط مشتركة للمعلمين",
        "supervisor_not_added_shared_links": "لم يضف المشرف الروابط المشتركة بعد.",
        "no_shared_teacher_resources": "لا توجد موارد مشتركة للمعلمين بعد.",
        "add_lesson_link": "إضافة رابط الحصة",
        "lesson_title": "عنوان الدرس",
        "lesson_zoom_link": "رابط الدرس / زوم",
        "short_lesson_summary": "ملخص قصير للدرس",
        "add_lesson": "إضافة الدرس",
        "lesson_links_and_homework": "روابط الدروس والواجب",
        "url": "الرابط",
        "delete": "حذف",
        "no_lesson_links_yet": "لا توجد روابط دروس بعد.",
        "update_recordings": "تحديث التسجيلات",
        "create_assignment": "إنشاء واجب",
        "assignment_title": "عنوان الواجب",
        "optional_resource_link": "رابط مورد اختياري",
        "instructions_questions_tasks": "التعليمات أو الأسئلة أو تفاصيل المهمة",
        "add_assignment": "إضافة الواجب",
        "assignment_review": "مراجعة الواجب",
        "select_assignment": "اختر الواجب",
        "open_assignment": "فتح الواجب",
        "resource": "المورد",
        "link": "رابط",
        "instructions": "التعليمات",
        "submission": "التسليم",
        "feedback": "التغذية الراجعة",
        "submission_link": "رابط التسليم",
        "active_state": "نشط",
        "closed_state": "مغلق",
        "update_assignment_reviews": "تحديث مراجعات الواجب",
        "no_student_submissions_yet": "لا توجد تسليمات طلابية بعد.",
        "select_assignment_review_marks": "اختر واجباً لمراجعة التسليمات وإدخال الدرجات.",
        "no_assignments_created_yet": "لا توجد واجبات منشأة بعد.",
        "students_section": "الطلاب",
        "report": "التقرير",
        "class_exam_overview": "نظرة عامة على اختبار الفصل",
        "template_exam_entry": "إدخال درجات القالب",
        "exam_results_label": "نتائج الاختبارات",
        "show_class_report": "عرض تقرير الفصل",
        "open_exam_table": "فتح جدول الاختبار",
        "open_results_table": "فتح جدول النتائج",
        "save_class_results": "حفظ نتائج الفصل",
        "update_exam_results": "تحديث نتائج الاختبارات",
        "add_upcoming_exam": "إضافة اختبار قادم",
        "update_upcoming_exams": "تحديث الاختبارات القادمة",
        "monthly_reports_review": "مراجعة التقارير الشهرية",
        "students_page": "الطلاب",
        "month_label": "الشهر",
        "monthly_review_subtitle": "راجع ملاحظات المعلمين، واستبعد من تريد، ثم أرسل التقارير الشهرية المعتمدة.",
        "total_students": "إجمالي الطلاب",
        "ready_to_send": "جاهز للإرسال",
        "missing_notes": "نواقص الملاحظات",
        "missing_email": "نقص البريد",
        "missing_whatsapp": "نقص واتساب",
        "all_students": "كل الطلاب",
        "ready_only": "الجاهزون فقط",
        "not_ready_only": "غير الجاهزين فقط",
        "apply": "تطبيق",
        "reset": "إعادة ضبط",
        "select_all_ready": "تحديد كل الجاهزين",
        "unselect_all": "إلغاء تحديد الكل",
        "send": "إرسال",
        "parent_email": "بريد ولي الأمر",
        "parent_whatsapp": "واتساب ولي الأمر",
        "teacher_note": "ملاحظة المعلم",
        "review": "مراجعة",
        "ready": "جاهز",
        "open_report": "فتح التقرير",
        "no_students_found": "لم يتم العثور على طلاب.",
        "send_selected_monthly_reports": "إرسال التقارير الشهرية المحددة",
        "student_report": "تقرير الطالب",
        "period": "الفترة",
        "this_week": "هذا الأسبوع",
        "this_month": "هذا الشهر",
        "export_pdf": "تصدير PDF",
        "export_excel": "تصدير Excel",
        "student_name": "اسم الطالب",
        "date_range": "الفترة الزمنية",
        "present_count": "عدد الحضور",
        "absent_count": "عدد الغياب",
        "late_count": "عدد التأخر",
        "excused_count": "عدد الأعذار",
        "attendance_percentage": "نسبة الحضور",
        "imported_arabic_attendance": "الحضور المرحلي المستورد",
        "imported_arabic_attendance_subtitle": "ملخص مستخرج من ملف حضور قسم اللغة العربية المشارك.",
        "attendance_snapshot_matched_name": "الاسم المطابق في ملف الحضور",
        "source_sheet": "ورقة المصدر",
        "total_sessions": "إجمالي الجلسات",
        "homework_given": "الواجبات المطلوبة",
        "homework_submitted": "الواجبات المسلمة",
        "homework_reviewed": "الواجبات المراجعة",
        "homework_waiting_review": "واجبات بانتظار المراجعة",
        "homework_missing": "الواجبات الناقصة",
        "homework_completion": "نسبة إنجاز الواجب",
        "monthly_teacher_note": "ملاحظة المعلم الشهرية",
        "monthly_teacher_note_for_student": "ملاحظة المعلم الشهرية لهذا الطالب",
        "save_monthly_note": "حفظ الملاحظة الشهرية",
        "send_report_to_parent_email": "إرسال التقرير إلى بريد ولي الأمر",
        "send_report": "إرسال التقرير",
        "view_only": "عرض فقط",
        "student_view_only_report": "حساب الطالب يعرض تفاصيل التقرير فقط.",
        "back_to_students": "العودة إلى الطلاب",
        "back_to_student_interface": "العودة إلى واجهة الطالب",
        "students_management": "الطلاب",
        "students_subtitle": "التقارير وبريد أولياء الأمور وأرقام واتسابهم وتحديثات الطلاب السريعة.",
        "add_student": "إضافة طالب",
        "student_full_name": "الاسم الكامل للطالب",
        "status_optional": "الحالة (اختياري)",
        "student_year": "السنة الدراسية",
        "parent_email_optional": "بريد ولي الأمر (اختياري)",
        "parent_whatsapp_optional": "واتساب ولي الأمر (اختياري)",
        "open_monthly_review": "فتح المراجعة الشهرية",
        "full_name": "الاسم الكامل",
        "level_name": "اسم الفصل",
        "parent_contact": "تواصل ولي الأمر",
        "update": "تحديث",
        "save": "حفظ",
        "attendance_records_title": "سجلات الحضور",
        "student_name_label": "اسم الطالب",
        "no_attendance_records_found": "لا توجد سجلات حضور.",
        "back_to_attendance_entry": "العودة إلى إدخال الحضور",
        "back_to_dashboard": "العودة إلى الواجهة الرئيسية",
        "level_attendance_report": "تقرير حضور الفصل",
        "summary_label": "الملخص",
        "total_present": "إجمالي الحضور",
        "total_absent": "إجمالي الغياب",
        "total_late": "إجمالي التأخر",
        "total_excused": "إجمالي الأعذار",
        "back_to_level_details": "العودة إلى تفاصيل الفصل",
        "back_to_levels": "العودة إلى الفصول",
        "missing_attendance": "نقص الحضور",
        "partial_attendance": "حضور جزئي",
        "students_needing_attention": "طلاب يحتاجون انتباهاً",
        "active_assignments": "واجبات نشطة",
        "teachers_on_track": "معلمون منجزون",
        "teachers_need_follow_up": "معلمون يحتاجون متابعة",
        "cleanup_critical": "تنظيف حرج",
        "cleanup_total": "إجمالي التنظيف",
        "deployment_checklist": "قائمة جاهزية النشر",
        "todays_gaps": "نواقص اليوم",
        "immediate_supervisor_actions": "إجراءات المشرف العاجلة",
        "open_level": "فتح الفصل",
        "edit_level": "تعديل الفصل",
        "no_urgent_daily_gaps": "لا توجد نواقص يومية عاجلة",
        "urgent_daily_gaps_clear": "الحضور وروابط زوم والواجبات النشطة وطوابير المراجعة في وضع جيد حالياً.",
        "supervisor_tools": "أدوات المشرف",
        "core_access": "الوصول الأساسي",
        "core_links_access_management": "روابط أساسية وإدارة وصول بدون مغادرة لوحة المشرف.",
        "new_access_code": "كود وصول جديد",
        "update_code": "تحديث الكود",
        "confirm_change_admin_code": "هل تريد تأكيد تغيير كود وصول المشرف؟",
        "supervisor_review_before_sending": "مراجعة المشرف قبل الإرسال",
        "monthly_report_ready_for_supervisor": "التقرير الشهري جاهز لإرسال المشرف.",
        "weekly_archive": "أرشيف الأسبوع",
        "saved_weekly_reports": "التقارير الأسبوعية المحفوظة",
        "updated": "آخر تحديث",
        "size": "الحجم",
        "download": "تنزيل",
        "no_archived_weekly_reports": "لا توجد تقارير أسبوعية مؤرشفة بعد.",
        "roster": "كشف الطلاب",
        "student_code": "كود الطالب",
        "student_year_label": "السنة الدراسية",
        "total": "الإجمالي",
        "percentage": "النسبة",
        "grade": "التقدير",
        "pdf": "PDF",
        "no_branches_defined": "لا توجد فروع معرفة لهذا القالب بعد.",
        "no_exam_results_level": "لا توجد نتائج اختبارات متاحة لهذا الفصل بعد.",
        "exam_title_placeholder": "عنوان الاختبار",
        "submitted_state": "تم التسليم",
        "reviewed_state": "تمت المراجعة",
        "needs_update_state": "يحتاج تحديثاً",
        "delete_upcoming_exam_confirm": "هل تريد حذف هذا الاختبار القادم؟",
        "delete_recording_confirm": "هل تريد حذف هذا التسجيل؟",
        "recent_system_actions": "آخر إجراءات النظام",
        "latest_logged_updates": "آخر التحديثات المسجلة من المعلمين والمشرفين.",
        "no_recent_actions": "لا توجد إجراءات مسجلة بعد.",
        "actor": "المنفذ",
        "action": "الإجراء",
        "action_details": "التفاصيل",
        "logged_time": "الوقت",
        "system_action_log": "سجل إجراءات النظام",
        "action_log_subtitle": "تصفية آخر عمليات المعلمين والمشرفين حسب المنفذ أو الفصل أو نوع الإجراء.",
        "all_roles": "كل الأدوار",
        "all_actions": "كل الإجراءات",
        "all_levels": "كل الفصول",
        "search_label": "بحث",
        "search_actions_placeholder": "ابحث باسم المنفذ أو الفصل أو التفاصيل",
        "open_action_log": "فتح سجل الإجراءات",
        "supervisor_role": "المشرف",
        "teacher_role": "المعلم",
        "system_role": "النظام",
        "entity": "العنصر",
        "open_link": "فتح",
        "teacher_thanks_sent": "تم إرسال شكر للمعلم",
        "admin_access_code_updated": "تم تحديث كود دخول المشرف",
        "announcement_published": "تم نشر إعلان",
        "announcement_updated": "تم تحديث إعلان",
        "calendar_settings_updated": "تم تحديث إعدادات التقويم",
        "holiday_added": "تمت إضافة إجازة",
        "holiday_updated": "تم تحديث الإجازة",
        "academic_calendar_event_added": "تمت إضافة حدث أكاديمي",
        "academic_calendar_event_updated": "تم تحديث حدث أكاديمي",
        "weekly_followup_generated": "تم توليد المتابعة الأسبوعية",
        "student_added": "تمت إضافة طالب",
        "student_updated": "تم تحديث الطالب",
        "parent_contact_saved": "تم حفظ تواصل ولي الأمر",
        "bulk_reports_sent": "تم إرسال تقارير جماعية",
        "monthly_reports_sent": "تم إرسال التقارير الشهرية",
        "student_report_sent": "تم إرسال تقرير طالب",
        "curriculum_progress_updated": "تم تحديث تقدم المنهج",
        "attendance_saved": "تم حفظ الحضور",
        "curriculum_plan_followup": "متابعة خطة المنهج",
        "curriculum_plan_progress": "تقدم خطة المنهج",
        "completed_items_label": "العناصر المكتملة",
        "pending_items_label": "العناصر المتبقية",
        "in_progress_items_label": "قيد التنفيذ",
        "last_completed_item": "آخر عنصر مكتمل",
        "plan_note": "ملاحظة الخطة",
        "save_progress": "حفظ التقدم",
        "pending_state": "لم يبدأ",
        "in_progress_state": "قيد التنفيذ",
        "completed_state": "مكتمل",
        "progress_percentage": "نسبة الإنجاز",
        "no_curriculum_plan_items": "لا توجد عناصر خطة منهج بعد.",
        "teacher_plan_checkpoint": "محطة متابعة المعلم",
        "start_here": "ابدأ من هنا",
        "syllabus_plan": "خطة المنهج",
        "plan_sheet_subtitle": "كشف سنوي شبيه بجدول إكسل لخطة هذا الفصل.",
        "week_number": "الأسبوع",
        "lesson_reference": "المرجع",
        "learning_objective": "هدف التعلم",
        "planned_homework_label": "الواجب المخطط",
        "planned_state": "مخطط",
        "postponed_state": "مؤجل",
        "add_plan_row": "إضافة سطر للخطة",
        "save_plan_sheet": "حفظ كشف الخطة",
        "open_syllabus_plan": "فتح خطة المنهج",
        "download_plan_pdf": "تنزيل الخطة PDF",
        "download_plan_template": "تنزيل قالب Excel",
        "import_plan_template": "استيراد خطة Excel",
        "excel_label": "إكسل",
        "plan_template_hint": "نزّل القالب الرسمي، ثم عبّئه وارفعه هنا. الاستيراد يستبدل الخطة الحالية لهذا الفصل.",
        "plan_template_file_label": "ملف خطة Excel",
        "import_preview_ready": "معاينة الاستيراد جاهزة",
        "import_preview_hint": "راجع الأعداد وعينة الصفوف أولاً، ثم اعتمد الاستيراد. اعتماد المعاينة سيستبدل الخطة الحالية لهذا الفصل.",
        "preview_scheduled_rows": "معاينة الصفوف المجدولة",
        "preview_reserve_rows": "معاينة صفوف الاحتياط",
        "apply_import": "اعتماد الاستيراد",
        "discard_preview": "إلغاء المعاينة",
        "changed_plan_rows": "الصفوف التي ستتغير",
        "current_value": "الحالي",
        "new_value": "الجديد",
        "plan_comparison_preview": "معاينة مقارنة الخطة",
        "plan_import_matches_current": "هذا الملف مطابق للخطة الحالية. لن تتغير أي حصص عند الاعتماد.",
        "editing_locked": "التعديل مقفل",
        "editing_open": "التعديل مفتوح",
        "allow_teacher_plan_editing": "فتح التعديل للمعلم",
        "lock_teacher_plan_editing": "قفل التعديل على المعلم",
        "read_only_plan": "خطة للقراءة فقط",
        "teacher_can_mark_completed_only": "يمكن للمعلم فقط تحديد الدروس كمكتملة أثناء قفل التعديل.",
        "no_plan_rows_yet": "لا توجد أسطر خطة بعد.",
        "plan_rows_count": "أسطر الخطة",
        "plan_entry_added": "تمت إضافة سطر للخطة",
        "plan_sheet_updated": "تم تحديث كشف الخطة",
        "teacher_plan_overview": "نظرة على خطة المعلم",
        "select_level_to_review_plan": "اختر فصلاً لمراجعة خطة المنهج الخاصة به.",
        "weeks_count_label": "40 أسبوعاً",
        "lessons_per_week_label": "4 حصص",
        "levels_on_track": "الفصول على المسار",
        "levels_delayed": "الفصول المتأخرة",
        "view_full_plan_followup": "فتح متابعة الخطة كاملة",
        "class_followup_register": "دفتر متابعة الفصل",
        "student_followup_subtitle": "سطر واحد لكل طالب لمتابعة الحضور والواجبات وآخر اختبار وجهوزية الملاحظة الشهرية.",
        "monthly_note_status": "الملاحظة الشهرية",
        "note_ready": "جاهزة",
        "note_missing": "ناقصة",
        "open_followup": "فتح المتابعة",
        "supervisor_level_followup_subtitle": "عرض إشرافي لحضور الطلاب والواجبات وآخر اختبار وجهوزية الملاحظة الشهرية لهذا الفصل.",
        "followup_attention_levels": "الفصول الأكثر احتياجاً للمتابعة",
        "classes_with_student_flags": "فصول فيها مؤشرات متابعة من الحضور والواجبات والاختبار وجهوزية الملاحظة الشهرية.",
        "students_flagged": "طلاب يحتاجون متابعة",
        "student_overall_status": "الحالة العامة",
        "excellent_status": "ممتاز",
        "stable_status": "مستقر",
        "needs_attention_status": "يحتاج متابعة",
        "sort_by": "الفرز حسب",
        "lowest_attendance_first": "الأضعف حضوراً أولاً",
        "most_open_assignments_first": "الأكثر واجبات مفتوحة أولاً",
        "missing_monthly_note_first": "الناقصين ملاحظة شهرية أولاً",
        "excellent_students": "الطلاب الممتازون",
        "stable_students": "الطلاب المستقرون",
        "students_need_attention": "الطلاب الذين يحتاجون متابعة",
        "review_lesson": "حصة مراجعة",
        "empty_plan_cell": "خانة تخطيط فارغة",
        "fill_plan_cell_hint": "أضف الكتاب والوحدة والدرس عندما تكون جاهزاً لبناء هذا الجزء من الخطة السنوية.",
        "complete_previous_prompt": "هناك دروس سابقة غير مكتملة. اضغط موافق لاعتماد هذا الدرس وكل ما قبله كمكتمل.",
        "complete_only_current_prompt": "اضغط موافق لاعتماد هذا الدرس فقط. اضغط إلغاء للتراجع دون أي تغيير.",
        "lesson_slot_1": "الحصة 1",
        "lesson_slot_2": "الحصة 2",
        "lesson_slot_3": "الحصة 3",
        "lesson_slot_4": "الحصة 4",
        "book_name": "الكتاب",
        "unit_name": "الوحدة",
        "lesson_name": "الدرس",
        "execution_status": "حالة التنفيذ",
        "completed_lessons": "الحصص المكتملة",
        "planned_lessons": "الحصص المخططة",
        "current_teaching_week": "الأسبوع التعليمي الحالي",
        "expected_by_now": "المفترض حتى الآن",
        "delayed_lessons": "الحصص المتأخرة",
        "unscheduled_plan_rows": "الحصص غير المجدولة",
        "unscheduled_plan_caption": "تم استيراد هذه الصفوف من ملف الخطة، لكنها لم تجد مكاناً داخل أسابيع التدريس المتبقية في التقويم الأكاديمي.",
        "reserve_plan_rows": "حصص احتياط / غير موزعة بعد",
        "mark_completed": "تحديد كمكتمل",
        "completed_on": "تاريخ الإكمال",
        "on_track_status": "على المسار",
        "delayed_status": "متأخر",
        "plan_entry_completed": "تم إكمال حصة من الخطة",
        "save_current_week": "حفظ الأسبوع الحالي",
        "recording_added": "تمت إضافة تسجيل",
        "recordings_updated": "تم تحديث التسجيلات",
        "assignment_added": "تمت إضافة واجب",
        "assignment_reviews_updated": "تم تحديث مراجعات الواجب",
        "template_results_saved": "تم حفظ نتائج القالب",
        "exam_results_updated": "تم تحديث نتائج الاختبار",
        "upcoming_exam_added": "تمت إضافة اختبار قادم",
        "upcoming_exams_updated": "تم تحديث الاختبارات القادمة",
        "monthly_note_saved": "تم حفظ الملاحظة الشهرية",
        "close_menu": "إغلاق القائمة",
        "attendance_recorded_missing": "تم تسجيل الحضور في {recorded}/{expected} من أيام التدريس - الأيام الناقصة: {days}",
        "zoom_recordings_uploaded_missing": "تم رفع تسجيلات الزوم في {uploaded}/{expected} من أيام التدريس - الأيام الناقصة: {days}",
        "no_homework_assigned_week": "لم يتم تعيين واجب خلال أسبوع التدريس",
        "all_assigned_levels_on_track": "جميع المستويات المكلف بها تسير على المسار.",
        "teachers_weekly_followup": "المتابعة الأسبوعية للمعلمين",
        "teacher_weekly_report": "تقرير المعلمين هذا الأسبوع",
        "teacher_label_short": "المعلم",
        "completed_status": "منجز",
        "levels_count_label": "الفصول",
        "students_count_label": "الطلاب",
        "levels_need_followup": "فصول تحتاج متابعة",
        "week_status_label": "حالة الأسبوع",
        "weekly_summary_label": "الملخص الأسبوعي",
        "missing_attendance_label": "الحضور الناقص",
        "attendance_complete_label": "الحضور مكتمل",
        "missing_recordings_label": "التسجيلات الناقصة",
        "recordings_complete_label": "التسجيلات مكتملة",
        "no_teacher_weekly_data": "لا توجد بيانات متابعة أسبوعية للمعلمين حتى الآن.",
        "teacher_monthly_report": "التقرير الشهري للمعلمين",
        "ideal_teacher_and_thanks": "المعلم المثالي ورسائل الشكر",
        "month_label": "الشهر",
        "teachers_count_label": "المعلمين",
        "ready_for_thanks": "جاهزون للشكر",
        "ideal_teacher": "المعلم المثالي",
        "score_label": "الدرجة",
        "monthly_report_label": "تقرير شهري",
        "attendance_percentage_label": "الحضور",
        "recordings_percentage_label": "التسجيلات",
        "weekly_homework_label": "الواجب الأسبوعي",
        "pending_reviews_label": "مراجعات معلقة",
        "email_ready_label": "البريد",
        "ready_label": "جاهز",
        "not_ready_label": "غير جاهز",
        "level_breakdown_label": "تفصيل الفصول",
        "attendance_word": "حضور",
        "recordings_word": "تسجيلات",
        "homework_word": "واجب",
        "send_appreciation_message": "إرسال كلمة شكر",
        "teachers_master_excel": "كشف المعلمين Excel",
        "teachers_master_pdf": "كشف المعلمين PDF",
        "level_full_excel": "كشف الفصل Excel",
        "level_full_pdf": "كشف الفصل PDF",
        "student_full_excel": "كشف الطالب Excel",
        "student_full_pdf": "كشف الطالب PDF",
        "reports_center": "مركز التقارير",
        "open_reports_center": "فتح مركز التقارير",
        "executive_reports_exports": "تقارير الإدارة العليا والتصدير",
        "teachers_master_report": "كشف المعلمين الموحد",
        "download_unified_teacher_sheet": "تنزيل الكشف الشهري الموحد للمعلمين.",
        "level_full_report": "كشف الفصل الكامل",
        "select_level_to_export": "اختر فصلاً ثم صدّر كشف المتابعة الكامل.",
        "student_full_report": "كشف الطالب الكامل",
        "select_student_to_export": "اختر طالباً ثم صدّر ملف الطالب الكامل.",
        "export_selected_level_excel": "تصدير Excel للفصل المحدد",
        "export_selected_student_excel": "تصدير Excel للطالب المحدد",
        "ideal_status": "مثالي",
        "excellent_status": "ممتاز",
        "good_status": "جيد",
        "no_teacher_monthly_data": "لا توجد بيانات شهرية للمعلمين حتى الآن.",
    },
}

GLOBAL_UI_COPY["en"].update(
    {
        "yes": "Yes",
        "no": "No",
        "academic_calendar": "Academic Calendar",
        "open_academic_calendar": "Open Academic Calendar",
        "academic_calendar_subtitle": "Review the academic year timeline, current week, and upcoming non-teaching events.",
        "manage_calendar_events": "Manage Calendar Events",
        "add_calendar_event": "Add Calendar Event",
        "event_title_label": "Event Title",
        "event_note_label": "Event Note",
        "save_calendar_event": "Save Calendar Event",
        "update_calendar_event": "Update Calendar Event",
        "event_saved_success": "Academic calendar event saved successfully.",
        "event_updated_success": "Academic calendar event updated successfully.",
        "event_dates_required": "Event title, type, start date, and end date are required.",
        "event_dates_invalid": "Event end date cannot be before the start date.",
        "event_type_label": "Event Type",
        "start_date_label": "Start Date",
        "end_date_label": "End Date",
        "instructional_label": "Instructional",
        "academic_calendar_status": "Academic Calendar Status",
        "calendar_driven_week": "Calendar-Driven Week",
        "calendar_driven_week_hint": "The current teaching week is now derived automatically from the academic calendar.",
        "current_academic_event": "Current Academic Event",
        "current_calendar_status": "Current Status",
        "teaching_in_session": "Teaching In Session",
        "non_teaching_period": "Non-Teaching Period",
        "next_academic_event": "Next Academic Event",
        "upcoming_academic_events": "Upcoming Academic Events",
        "no_upcoming_academic_events": "No upcoming academic events scheduled yet.",
        "teaching_week_auto": "Teaching week is calculated automatically from the academic calendar.",
        "term_start_type": "Term Start",
        "school_resume_type": "School Resumes",
        "holiday_type": "Holiday",
        "exam_type": "Exam Period",
        "parents_meeting_type": "Parents' Meeting",
        "teacher_training_type": "Teacher Training",
        "event_day_type": "Event Day",
        "day_count_label": "{count} day(s)",
        "calendar_upcoming_alert_title": "Upcoming Calendar Event",
        "calendar_current_alert_title": "Current Academic Status",
        "calendar_upcoming_alert_body": "{title} starts on {date}.",
        "calendar_current_alert_body": "{title} is currently active.",
        "starts_today": "Starts today",
        "starts_tomorrow": "Starts tomorrow",
        "starts_in_days": "Starts in {count} days",
        "exam_week_status": "Exam Week",
        "holiday_status": "Holiday",
        "week_dates": "Week Dates",
        "current_week_badge": "Current Week",
    }
)

GLOBAL_UI_COPY["ar"].update(
    {
        "yes": "نعم",
        "no": "لا",
        "academic_calendar": "التقويم الأكاديمي",
        "open_academic_calendar": "فتح التقويم الأكاديمي",
        "academic_calendar_subtitle": "راجع تسلسل العام الدراسي والأسبوع الحالي والأحداث القادمة غير التدريسية.",
        "manage_calendar_events": "إدارة أحداث التقويم",
        "add_calendar_event": "إضافة حدث أكاديمي",
        "event_title_label": "عنوان الحدث",
        "event_note_label": "ملاحظة الحدث",
        "save_calendar_event": "حفظ الحدث",
        "update_calendar_event": "تحديث الحدث",
        "event_saved_success": "تم حفظ الحدث الأكاديمي بنجاح.",
        "event_updated_success": "تم تحديث الحدث الأكاديمي بنجاح.",
        "event_dates_required": "عنوان الحدث ونوعه وتاريخ البداية والنهاية مطلوبة.",
        "event_dates_invalid": "لا يمكن أن يكون تاريخ نهاية الحدث قبل بدايته.",
        "event_type_label": "نوع الحدث",
        "start_date_label": "تاريخ البداية",
        "end_date_label": "تاريخ النهاية",
        "instructional_label": "تدريسي",
        "academic_calendar_status": "حالة التقويم الأكاديمي",
        "calendar_driven_week": "الأسبوع المحسوب من التقويم",
        "calendar_driven_week_hint": "يُحسب الأسبوع التعليمي الحالي الآن تلقائياً من التقويم الأكاديمي.",
        "current_academic_event": "الحدث الأكاديمي الحالي",
        "current_calendar_status": "الحالة الحالية",
        "teaching_in_session": "الدراسة جارية",
        "non_teaching_period": "فترة غير تدريسية",
        "next_academic_event": "الحدث الأكاديمي القادم",
        "upcoming_academic_events": "الأحداث الأكاديمية القادمة",
        "no_upcoming_academic_events": "لا توجد أحداث أكاديمية قادمة حالياً.",
        "teaching_week_auto": "يُحسب الأسبوع التعليمي تلقائياً من التقويم الأكاديمي.",
        "term_start_type": "بداية الفصل",
        "school_resume_type": "استئناف الدراسة",
        "holiday_type": "إجازة",
        "exam_type": "فترة اختبارات",
        "parents_meeting_type": "اجتماع أولياء الأمور",
        "teacher_training_type": "تدريب المعلمين",
        "event_day_type": "فعالية",
        "day_count_label": "{count} يوم",
        "calendar_upcoming_alert_title": "حدث أكاديمي قادم",
        "calendar_current_alert_title": "الحالة الأكاديمية الحالية",
        "calendar_upcoming_alert_body": "{title} يبدأ في {date}.",
        "calendar_current_alert_body": "{title} جارٍ حالياً.",
        "starts_today": "يبدأ اليوم",
        "starts_tomorrow": "يبدأ غداً",
        "starts_in_days": "يبدأ بعد {count} أيام",
        "exam_week_status": "أسبوع اختبارات",
        "holiday_status": "إجازة",
        "week_dates": "تواريخ الأسبوع",
        "current_week_badge": "الأسبوع الحالي",
    }
)


def get_current_ui_language() -> str:
    if not has_request_context():
        return "en"
    return session.get("ui_lang", "en") if session.get("ui_lang", "en") in {"en", "ar"} else "en"


def get_global_ui_copy(lang: str) -> dict:
    return GLOBAL_UI_COPY.get(lang, GLOBAL_UI_COPY["en"])


def is_teacher_resource_subject(subject_name: str) -> bool:
    normalized = normalize_optional_text(subject_name).strip().lower()
    return normalized in {
        "teacher resources",
        "teachers resources",
        "shared teacher resources",
        "teacher links",
        "teacher materials",
        "مصادر المعلمين",
        "روابط المعلمين",
        "المصادر المشتركة للمعلمين",
    }


app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)


@app.before_request
def apply_requested_ui_language():
    requested_lang = (request.args.get("lang") or "").strip().lower()
    if requested_lang in {"en", "ar"}:
        session["ui_lang"] = requested_lang


@app.context_processor
def inject_global_ui_context():
    current_lang = get_current_ui_language()
    ui = get_global_ui_copy(current_lang)

    def tr(key: str, **kwargs):
        value = ui.get(key, key.replace("_", " ").title())
        if kwargs:
            try:
                return value.format(**kwargs)
            except Exception:
                return value
        return value

    def with_lang(lang_code: str) -> str:
        args = request.args.to_dict(flat=True)
        args["lang"] = lang_code
        return url_for(request.endpoint, **(request.view_args or {}), **args) if request.endpoint else request.path

    def static_asset(path: str) -> str:
        asset_path = os.path.join(app.static_folder or "static", path)
        version = "1"
        try:
            version = str(int(os.path.getmtime(asset_path)))
        except OSError:
            pass
        return url_for("static", filename=path, v=version)

    return {
        "current_lang": current_lang,
        "current_dir": "rtl" if current_lang == "ar" else "ltr",
        "tr": tr,
        "with_lang": with_lang,
        "static_asset": static_asset,
    }


def get_current_actor_context() -> dict:
    if session.get("is_admin"):
        return {
            "actor_role": "admin",
            "actor_id": None,
            "actor_name": "Supervisor",
        }

    teacher = get_current_teacher()
    if teacher:
        return {
            "actor_role": "teacher",
            "actor_id": teacher.id,
            "actor_name": teacher.full_name,
        }

    return {
        "actor_role": "system",
        "actor_id": None,
        "actor_name": "System",
    }


def ui_text(key: str, **kwargs) -> str:
    ui = get_global_ui_copy(get_current_ui_language())
    value = ui.get(key, key.replace("_", " ").title())
    if kwargs:
        try:
            return value.format(**kwargs)
        except Exception:
            return value
    return value


def log_action(
    action_type: str,
    entity_type: str,
    entity_id: int | None = None,
    entity_label: str | None = None,
    level: Level | None = None,
    details: str | None = None,
) -> None:
    actor = get_current_actor_context()
    db.session.add(
        ActionLog(
            actor_role=actor["actor_role"],
            actor_id=actor["actor_id"],
            actor_name=actor["actor_name"],
            action_type=action_type,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_label=entity_label,
            level_id=level.id if level else None,
            details=details or None,
        )
    )


def get_action_log_target(action_log: ActionLog) -> tuple[str | None, str]:
    if action_log.action_type == "attendance_saved" and action_log.level_id:
        return url_for("attendance_page", level_id=action_log.level_id), "attendance"
    if action_log.action_type in {
        "recording_added",
        "recordings_updated",
        "assignment_added",
        "assignment_reviews_updated",
        "template_results_saved",
        "exam_results_updated",
        "upcoming_exam_added",
        "upcoming_exams_updated",
        "curriculum_progress_updated",
    } and action_log.level_id:
        return url_for("teacher_level_workspace", level_id=action_log.level_id), "open_workspace"
    if action_log.action_type in {"plan_entry_added", "plan_sheet_updated"} and action_log.level_id:
        return url_for("teacher_syllabus_plan", level_id=action_log.level_id), "open_syllabus_plan"
    if action_log.action_type in {
        "monthly_note_saved",
        "student_report_sent",
    } and action_log.entity_id:
        return url_for("student_report", student_id=action_log.entity_id, period="month"), "open_report"
    if action_log.action_type in {
        "student_added",
        "student_updated",
        "parent_contact_saved",
        "bulk_reports_sent",
    }:
        return url_for("students_list"), "students"
    if action_log.action_type == "monthly_reports_sent":
        return url_for("monthly_reports_review"), "open_monthly_review"
    if action_log.action_type in {
        "teacher_thanks_sent",
        "admin_access_code_updated",
        "announcement_published",
        "announcement_updated",
        "calendar_settings_updated",
        "holiday_added",
        "holiday_updated",
        "weekly_followup_generated",
    }:
        return url_for("admin_dashboard"), "supervisor_dashboard"
    return None, "open_link"


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped


def teacher_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("teacher_id"):
            return redirect(url_for("teacher_login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped


def get_current_teacher() -> Teacher | None:
    teacher_id = session.get("teacher_id")
    if not teacher_id:
        return None
    return Teacher.query.get(teacher_id)


def get_teacher_level_or_403(level_id: int) -> Level:
    teacher = get_current_teacher()
    if not teacher:
        abort(403)
    level = Level.query.get_or_404(level_id)
    if level.teacher_id != teacher.id:
        abort(403)
    return level


def can_manage_student_report(student: Student) -> bool:
    if session.get("is_admin"):
        return True

    teacher = get_current_teacher()
    if not teacher:
        return False

    if student.level_id:
        return bool(
            Level.query.filter_by(id=student.level_id, teacher_id=teacher.id).first()
        )

    if student.level_name:
        level = Level.query.filter_by(
            name=normalize_level_display_name(student.level_name)
        ).first()
        return bool(level and level.teacher_id == teacher.id)

    return False


def is_safe_next(target: str) -> bool:
    return bool(target and target.startswith("/") and not target.startswith("//"))


def normalize_data_root_section(section: str) -> str:
    valid_sections = {"students", "teachers", "levels", "recordings", "cleanup"}
    section = (section or "").strip().lower()
    return section if section in valid_sections else "students"


def redirect_head_data_root(op_status: str, op_message: str, section: str = "students"):
    return redirect(
        url_for(
            "head_data_root",
            op_status=op_status,
            op_message=op_message,
            section=normalize_data_root_section(section),
        )
    )


def build_cleanup_priority(flags: list[str], critical_flags: set[str]) -> dict:
    is_critical = any(flag in critical_flags for flag in flags)
    return {
        "label": "Critical" if is_critical else "Needs Review",
        "tone": "critical" if is_critical else "review",
        "rank": 0 if is_critical else 1,
    }


def is_level_open_for_release(level: Level, students: list[Student]) -> bool:
    return any(student.level_id == level.id for student in students)


def build_cleanup_center_data(levels: list[Level], teachers: list[Teacher], students: list[Student]) -> dict:
    cleanup_students = []
    for student in students:
        flags = []
        if not student.parent_email:
            flags.append("missing_parent_email")
        if not student.parent_whatsapp:
            flags.append("missing_parent_whatsapp")
        if not student.level_id:
            flags.append("missing_level")
        if not student.status:
            flags.append("missing_status")
        if not student.student_year:
            flags.append("missing_year")
        if flags:
            cleanup_students.append(
                {
                    "student": student,
                    "flags": flags,
                    "priority": build_cleanup_priority(flags, {"missing_level"}),
                }
            )

    cleanup_teachers = []
    for teacher in teachers:
        flags = []
        assigned_levels = [level for level in levels if level.teacher_id == teacher.id]
        if not teacher.phone:
            flags.append("missing_phone")
        if not teacher.email:
            flags.append("missing_email")
        if not teacher.status:
            flags.append("missing_status")
        if not assigned_levels:
            flags.append("missing_level_assignment")
        if flags:
            cleanup_teachers.append(
                {
                    "teacher": teacher,
                    "flags": flags,
                    "assigned_levels": assigned_levels,
                    "priority": build_cleanup_priority(flags, {"missing_level_assignment", "missing_status"}),
                }
            )

    cleanup_levels = []
    for level in levels:
        if not is_level_open_for_release(level, students):
            continue
        flags = []
        if not level.teacher_id:
            flags.append("missing_teacher")
        if not level.zoom_link:
            flags.append("missing_zoom")
        if not level.homework_padlet_url:
            flags.append("missing_homework_board")
        if not level.announcements_padlet_url:
            flags.append("missing_class_board")
        if flags:
            cleanup_levels.append(
                {
                    "level": level,
                    "health_flags": flags,
                    "priority": build_cleanup_priority(flags, {"missing_teacher", "missing_zoom"}),
                }
            )

    cleanup_students.sort(key=lambda row: (row["priority"]["rank"], row["student"].full_name))
    cleanup_teachers.sort(key=lambda row: (row["priority"]["rank"], row["teacher"].full_name))
    cleanup_levels.sort(key=lambda row: (row["priority"]["rank"], row["level"].name))

    cleanup_summary = {
        "student_records": len(cleanup_students),
        "teacher_records": len(cleanup_teachers),
        "level_records": len(cleanup_levels),
        "critical_items": sum(
            1
            for row in cleanup_students + cleanup_teachers + cleanup_levels
            if row["priority"]["tone"] == "critical"
        ),
        "total_items": len(cleanup_students) + len(cleanup_teachers) + len(cleanup_levels),
    }

    return {
        "cleanup_students": cleanup_students,
        "cleanup_teachers": cleanup_teachers,
        "cleanup_levels": cleanup_levels,
        "cleanup_summary": cleanup_summary,
    }


def get_admin_access_code() -> str:
    row = db.session.execute(
        text("SELECT value FROM system_settings WHERE key = 'admin_access_code' LIMIT 1")
    ).fetchone()
    if row and row[0]:
        return str(row[0])
    return app.config["ADMIN_ACCESS_CODE"]


def get_system_setting(key: str, default: str = "") -> str:
    row = db.session.execute(
        text("SELECT value FROM system_settings WHERE key = :key LIMIT 1"),
        {"key": key},
    ).fetchone()
    if row and row[0] is not None:
        return str(row[0])
    return default


def set_system_setting(key: str, value: str) -> None:
    db.session.execute(
        text(
            """
            INSERT INTO system_settings(key, value)
            VALUES (:key, :value)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """
        ),
        {"key": key, "value": value},
    )


def get_academic_calendar_events() -> list[AcademicCalendarEvent]:
    return (
        AcademicCalendarEvent.query.order_by(
            AcademicCalendarEvent.start_date.asc(),
            AcademicCalendarEvent.sort_order.asc(),
            AcademicCalendarEvent.id.asc(),
        ).all()
    )


def get_active_academic_calendar_events(target_date: date) -> list[AcademicCalendarEvent]:
    return (
        AcademicCalendarEvent.query.filter(
            AcademicCalendarEvent.start_date <= target_date,
            AcademicCalendarEvent.end_date >= target_date,
        )
        .order_by(
            AcademicCalendarEvent.start_date.asc(),
            AcademicCalendarEvent.sort_order.asc(),
            AcademicCalendarEvent.id.asc(),
        )
        .all()
    )


def get_primary_academic_event(target_date: date) -> AcademicCalendarEvent | None:
    active_events = get_active_academic_calendar_events(target_date)
    if not active_events:
        return None
    return sorted(
        active_events,
        key=lambda item: (
            ACADEMIC_STATUS_ORDER.get(item.event_type, 999),
            item.sort_order,
            item.start_date,
            item.id,
        ),
    )[0]


def get_academic_year_start_date() -> date | None:
    row = (
        AcademicCalendarEvent.query.filter_by(event_type="term_start")
        .order_by(AcademicCalendarEvent.start_date.asc(), AcademicCalendarEvent.id.asc())
        .first()
    )
    return row.start_date if row else None


def is_non_teaching_date(target_date: date) -> bool:
    if is_holiday_date(target_date):
        return True
    return bool(
        AcademicCalendarEvent.query.filter(
            AcademicCalendarEvent.start_date <= target_date,
            AcademicCalendarEvent.end_date >= target_date,
            AcademicCalendarEvent.is_instructional.is_(False),
        ).first()
    )


def get_current_teaching_week(target_date: date | None = None) -> int:
    resolved_date = target_date or date.today()
    start_date = get_academic_year_start_date()
    if not start_date:
        week_value = get_system_setting("current_teaching_week", "1").strip()
        if week_value.isdigit():
            return max(1, min(40, int(week_value)))
        return 1

    if resolved_date < start_date:
        return 1

    teaching_days = build_teaching_days(start_date, resolved_date)
    if not teaching_days:
        return 1

    computed_week = ((len(teaching_days) - 1) // 4) + 1
    return max(1, min(40, computed_week))


def get_upcoming_academic_events(target_date: date, limit: int = 4) -> list[AcademicCalendarEvent]:
    return (
        AcademicCalendarEvent.query.filter(AcademicCalendarEvent.start_date > target_date)
        .order_by(
            AcademicCalendarEvent.start_date.asc(),
            AcademicCalendarEvent.sort_order.asc(),
            AcademicCalendarEvent.id.asc(),
        )
        .limit(limit)
        .all()
    )


def get_academic_event_type_label(event_type: str) -> str:
    return ui_text(f"{event_type}_type")


def get_next_event_countdown_label(next_event: AcademicCalendarEvent | None, target_date: date) -> str:
    if not next_event:
        return ""
    days_until = (next_event.start_date - target_date).days
    if days_until <= 0:
        return ui_text("starts_today")
    if days_until == 1:
        return ui_text("starts_tomorrow")
    return ui_text("starts_in_days", count=days_until)


def build_academic_calendar_status(target_date: date | None = None) -> dict:
    resolved_date = target_date or date.today()
    current_event = get_primary_academic_event(resolved_date)
    upcoming_events = get_upcoming_academic_events(resolved_date, limit=4)
    current_week = get_current_teaching_week(resolved_date)
    start_date = get_academic_year_start_date()
    is_teaching = (
        resolved_date.weekday() <= 3
        and not is_non_teaching_date(resolved_date)
        and bool(start_date and resolved_date >= start_date)
    )

    return {
        "today": resolved_date,
        "current_week": current_week,
        "is_teaching_day": is_teaching,
        "status_key": "teaching_in_session" if is_teaching else "non_teaching_period",
        "current_event": current_event,
        "current_event_type_label": get_academic_event_type_label(current_event.event_type) if current_event else "",
        "upcoming_events": upcoming_events,
        "start_date": start_date,
        "next_event": upcoming_events[0] if upcoming_events else None,
        "next_event_countdown_label": get_next_event_countdown_label(upcoming_events[0] if upcoming_events else None, resolved_date),
        "is_exam_period": bool(current_event and current_event.event_type == "exam"),
        "is_holiday_period": bool(current_event and current_event.event_type == "holiday"),
    }


def get_teacher_access_code() -> str:
    return str(app.config.get("TEACHER_ACCESS_CODE", "") or "").strip()


def teacher_access_code_enabled() -> bool:
    return bool(get_teacher_access_code())


def get_runtime_readiness_issues() -> list[str]:
    issues = []
    if app.config["SECRET_KEY"] == "hikmah-dev-secret-key":
        issues.append("FLASK_SECRET_KEY is still using the default development value.")
    if get_admin_access_code() == "1234":
        issues.append("ADMIN_ACCESS_CODE is still using the default value 1234.")
    if not teacher_access_code_enabled():
        issues.append("TEACHER_ACCESS_CODE is empty, so any listed teacher can open the teacher area.")

    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    teachers = Teacher.query.order_by(Teacher.full_name.asc()).all()
    students = Student.query.order_by(Student.full_name.asc()).all()
    cleanup_data = build_cleanup_center_data(levels, teachers, students)
    critical_cleanup_count = cleanup_data["cleanup_summary"]["critical_items"]
    if critical_cleanup_count:
        issues.append(f"There are still {critical_cleanup_count} critical data-cleanup item(s) to resolve before release.")

    levels_missing_teacher = sum(
        1 for level in levels
        if is_level_open_for_release(level, students) and not level.teacher_id
    )
    if levels_missing_teacher:
        issues.append(f"{levels_missing_teacher} level(s) still have no assigned teacher.")

    levels_missing_zoom = sum(
        1 for level in levels
        if is_level_open_for_release(level, students) and not (level.zoom_link or "").strip()
    )
    if levels_missing_zoom:
        issues.append(f"{levels_missing_zoom} level(s) still have no Zoom link.")

    calendar_settings = get_calendar_settings()
    if not calendar_settings.get("weekly_followup_weekday") or not calendar_settings.get("weekly_followup_time"):
        issues.append("Weekly follow-up scheduling is incomplete.")

    if not list_archived_weekly_reports():
        issues.append("No archived weekly report has been generated yet.")
    return issues


def set_admin_access_code(new_code: str) -> None:
    db.session.execute(
        text(
            """
            INSERT INTO system_settings(key, value)
            VALUES ('admin_access_code', :code)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """
        ),
        {"code": new_code},
    )
    db.session.commit()

def normalize_level_display_name(level_name: str) -> str:
    if not level_name:
        return level_name

    name = level_name.strip()
    if name in LEVEL_LEGACY_NAME_TO_DISPLAY:
        return LEVEL_LEGACY_NAME_TO_DISPLAY[name]

    for display_name, codes in LEVEL_DISPLAY_TO_CODES.items():
        if name in codes:
            return display_name

    return name


def get_level_codes_for_display_name(display_name: str) -> list[str]:
    normalized = normalize_level_display_name(display_name)
    return LEVEL_DISPLAY_TO_CODES.get(normalized, [normalized])


def validate_email(email: str) -> bool:
    return bool(email and "@" in email and "." in email.split("@")[-1])


def validate_whatsapp(value: str) -> bool:
    normalized = re.sub(r"[\s\-\(\)]", "", value or "")
    if normalized.startswith("+"):
        normalized = normalized[1:]
    return normalized.isdigit() and 7 <= len(normalized) <= 15


def normalize_import_header(value: str) -> str:
    normalized = (value or "").strip().lower()
    normalized = normalized.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    normalized = normalized.replace("ة", "ه").replace("ى", "ي")
    return re.sub(r"[^a-z0-9\u0600-\u06ff]+", "_", normalized).strip("_")


def normalize_optional_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalize_person_name(value: str) -> str:
    normalized = normalize_optional_text(value).lower()
    return re.sub(r"[^a-z0-9\u0600-\u06ff]+", "", normalized)


def split_name_tokens(value: str) -> list[str]:
    normalized = normalize_optional_text(value).lower()
    tokens = re.split(r"[^a-z0-9\u0600-\u06ff]+", normalized)
    return [token for token in tokens if token]


def token_subset_similarity(left_name: str, right_name: str) -> float:
    left_tokens = split_name_tokens(left_name)
    right_tokens = split_name_tokens(right_name)
    if not left_tokens or not right_tokens:
        return 0.0

    shorter, longer = (left_tokens, right_tokens) if len(left_tokens) <= len(right_tokens) else (right_tokens, left_tokens)
    matched = 0
    for token in shorter:
        if any(
            token == other
            or (len(token) >= 4 and len(other) >= 4 and SequenceMatcher(None, token, other).ratio() >= 0.84)
            for other in longer
        ):
            matched += 1

    return matched / max(len(shorter), 1)


def name_similarity_score(left_name: str, right_name: str) -> float:
    compact_ratio = SequenceMatcher(
        None,
        normalize_person_name(left_name),
        normalize_person_name(right_name),
    ).ratio()
    token_ratio = token_subset_similarity(left_name, right_name)
    return max(compact_ratio, token_ratio)


def parse_whole_number(value) -> int:
    text_value = normalize_optional_text(value)
    if not text_value:
        return 0
    try:
        return int(float(text_value))
    except (TypeError, ValueError):
        return 0


def is_arabic_attendance_subject(value) -> bool:
    normalized = normalize_optional_text(value).lower()
    return "arabic" in normalized


def get_arabic_attendance_workbook_path() -> Path:
    return Path(app.root_path) / "data" / "students.xlsx"


def ensure_arabic_attendance_upload_dir() -> Path:
    upload_dir = Path(app.root_path) / "data" / "attendance_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    return upload_dir


def save_uploaded_arabic_attendance_file(uploaded_file) -> dict:
    workbook_path = get_arabic_attendance_workbook_path()
    upload_dir = ensure_arabic_attendance_upload_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = workbook_path.suffix or ".xlsx"

    backup_path = None
    if workbook_path.exists():
        backup_path = upload_dir / f"students_backup_{timestamp}{suffix}"
        shutil.copy2(workbook_path, backup_path)

    uploaded_file.seek(0)
    uploaded_file.save(workbook_path)
    uploaded_file.seek(0)

    archived_upload_path = upload_dir / f"students_uploaded_{timestamp}{suffix}"
    shutil.copy2(workbook_path, archived_upload_path)

    return {
        "workbook_path": str(workbook_path),
        "backup_path": str(backup_path) if backup_path else "",
        "archived_upload_path": str(archived_upload_path),
    }


@lru_cache(maxsize=4)
def _load_imported_arabic_attendance_snapshot_cached(
    workbook_path: str,
    workbook_mtime_ns: int,
) -> dict[str, list[dict]]:
    del workbook_mtime_ns

    parsed_rows: dict[str, list[dict]] = {}
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)

    for sheet_name, level_name in ARABIC_ATTENDANCE_SHEET_TO_LEVEL.items():
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        sheet_rows: list[dict] = []

        for row in worksheet.iter_rows(min_row=6, values_only=True):
            student_name = normalize_optional_text(row[4] if len(row) > 4 else "")
            if not student_name:
                continue

            subject_name = normalize_optional_text(row[10] if len(row) > 10 else "")
            if not is_arabic_attendance_subject(subject_name):
                continue

            present_count = parse_whole_number(row[7] if len(row) > 7 else 0)
            late_count = parse_whole_number(row[8] if len(row) > 8 else 0)
            absent_count = parse_whole_number(row[9] if len(row) > 9 else 0)
            total_count = present_count + late_count + absent_count
            attendance_percentage = (
                round(((present_count + late_count) / total_count) * 100, 1)
                if total_count
                else 0.0
            )

            sheet_rows.append(
                {
                    "sheet_name": sheet_name,
                    "level_name": level_name,
                    "student_name": student_name,
                    "normalized_name": normalize_person_name(student_name),
                    "subject_name": subject_name,
                    "total_count": total_count,
                    "present_count": present_count,
                    "late_count": late_count,
                    "absent_count": absent_count,
                    "attendance_percentage": attendance_percentage,
                }
            )

        parsed_rows[level_name] = sheet_rows

    workbook.close()
    return parsed_rows


def load_imported_arabic_attendance_snapshot() -> dict[str, list[dict]]:
    workbook_path = get_arabic_attendance_workbook_path()
    if not workbook_path.exists():
        return {}
    return _load_imported_arabic_attendance_snapshot_cached(
        str(workbook_path),
        workbook_path.stat().st_mtime_ns,
    )


def build_imported_arabic_attendance_snapshot(student: Student) -> dict | None:
    level_name = normalize_level_display_name(student.level.name if student.level else student.level_name)
    level_rows = load_imported_arabic_attendance_snapshot().get(level_name, [])
    if not level_rows:
        return None

    candidate_names = {normalize_optional_text(student.full_name)}
    for alias in getattr(student, "name_aliases", []) or []:
        alias_name = normalize_optional_text(alias.alias_name)
        if alias_name:
            candidate_names.add(alias_name)

    normalized_candidates = {
        normalize_person_name(name): name
        for name in candidate_names
        if normalize_person_name(name)
    }

    for row in level_rows:
        if row["normalized_name"] in normalized_candidates:
            return row

    scored_rows = []
    for row in level_rows:
        best_score = max(
            name_similarity_score(candidate_name, row["student_name"])
            for candidate_name in candidate_names
            if candidate_name
        )
        scored_rows.append((best_score, row))

    scored_rows.sort(key=lambda item: item[0], reverse=True)
    if not scored_rows:
        return None

    best_score, best_row = scored_rows[0]
    second_score = scored_rows[1][0] if len(scored_rows) > 1 else 0.0
    if best_score >= 0.88 and (best_score - second_score >= 0.04 or second_score < 0.82):
        return best_row

    return None


def resolve_student_attendance_display(
    student: Student,
    recent_attendance_rate: float | None = None,
) -> dict:
    imported_attendance = build_imported_arabic_attendance_snapshot(student)
    attendance_rate_display = (
        recent_attendance_rate
        if recent_attendance_rate is not None
        else (imported_attendance["attendance_percentage"] if imported_attendance else None)
    )
    attendance_source_key = (
        "attendance_source_recent"
        if recent_attendance_rate is not None
        else ("attendance_source_imported" if imported_attendance else "")
    )
    return {
        "attendance_rate_display": attendance_rate_display,
        "attendance_source_key": attendance_source_key,
        "imported_attendance": imported_attendance,
    }


def ensure_student_alias(alias_name: str, level_name: str | None, student_id: int) -> str:
    normalized_alias = normalize_optional_text(alias_name)
    normalized_level_name = normalize_level_display_name(level_name) if level_name else None
    if not normalized_alias:
        return False

    existing_alias = StudentNameAlias.query.filter_by(
        alias_name=normalized_alias,
        level_name=normalized_level_name,
    ).first()
    if existing_alias:
        if existing_alias.student_id != student_id:
            existing_alias.student_id = student_id
            return "updated"
        return "unchanged"

    db.session.add(
        StudentNameAlias(
            alias_name=normalized_alias,
            level_name=normalized_level_name,
            student_id=student_id,
        )
    )
    return "created"


def sync_imported_arabic_attendance_aliases() -> dict:
    imported_rows_by_level = load_imported_arabic_attendance_snapshot()
    created_count = 0
    updated_count = 0

    for level_name, rows in imported_rows_by_level.items():
        students = Student.query.filter(
            db.or_(
                Student.level_name == level_name,
                Student.level.has(Level.name == level_name),
            )
        ).all()
        if not students:
            continue

        for row in rows:
            alias_name = row["student_name"]
            if any(normalize_optional_text(alias.alias_name) == alias_name for student in students for alias in getattr(student, "name_aliases", []) or []):
                continue

            best_matches = []
            for student in students:
                score = name_similarity_score(student.full_name, alias_name)
                best_matches.append((score, student))
            best_matches.sort(key=lambda item: item[0], reverse=True)
            if not best_matches:
                continue

            best_score, best_student = best_matches[0]
            second_score = best_matches[1][0] if len(best_matches) > 1 else 0.0
            if best_score < 0.88 or (second_score >= 0.82 and (best_score - second_score) < 0.04):
                continue

            alias_status = ensure_student_alias(alias_name, level_name, best_student.id)
            if alias_status == "created":
                created_count += 1
            elif alias_status == "updated":
                updated_count += 1

    if created_count or updated_count:
        db.session.commit()

    return {
        "created_count": created_count,
        "updated_count": updated_count,
    }


def build_imported_arabic_attendance_summary() -> dict:
    imported_rows_by_level = load_imported_arabic_attendance_snapshot()
    imported_rows_count = sum(len(rows) for rows in imported_rows_by_level.values())
    level_counts = {
        level_name: len(rows)
        for level_name, rows in imported_rows_by_level.items()
        if rows
    }
    matched_rows_count = 0
    unmatched_rows_count = 0
    unmatched_samples = []
    for level_name, rows in imported_rows_by_level.items():
        students = Student.query.filter(
            db.or_(
                Student.level_name == level_name,
                Student.level.has(Level.name == level_name),
            )
        ).all()
        for row in rows:
            best_score = 0.0
            best_name = ""
            for student in students:
                score = name_similarity_score(student.full_name, row["student_name"])
                if score > best_score:
                    best_score = score
                    best_name = student.full_name
            if best_score >= 0.88:
                matched_rows_count += 1
            else:
                unmatched_rows_count += 1
                if len(unmatched_samples) < 5:
                    unmatched_samples.append(
                        {
                            "level_name": level_name,
                            "workbook_name": row["student_name"],
                            "best_match_name": best_name,
                            "score": round(best_score, 3),
                        }
                    )
    alias_count = StudentNameAlias.query.filter(
        StudentNameAlias.level_name.in_(list(level_counts.keys()) if level_counts else [""])
    ).count() if level_counts else 0
    last_refresh = get_system_setting("last_arabic_attendance_refresh", "")
    workbook_path = get_arabic_attendance_workbook_path()
    return {
        "imported_rows_count": imported_rows_count,
        "matched_rows_count": matched_rows_count,
        "unmatched_rows_count": unmatched_rows_count,
        "unmatched_samples": unmatched_samples,
        "level_counts": level_counts,
        "matched_aliases_count": alias_count,
        "last_refresh": last_refresh,
        "current_file_name": workbook_path.name,
        "current_file_modified": datetime.fromtimestamp(workbook_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M") if workbook_path.exists() else "",
    }


def build_arabic_attendance_upload_history(limit: int = 5) -> tuple[list[dict], list[dict]]:
    upload_dir = ensure_arabic_attendance_upload_dir()
    archive_rows = []
    for path in sorted(upload_dir.glob("students_uploaded_*"), key=lambda item: item.stat().st_mtime, reverse=True)[:limit]:
        archive_rows.append(
            {
                "file_name": path.name,
                "modified_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "size_kb": max(1, round(path.stat().st_size / 1024)),
            }
        )

    action_rows = (
        ActionLog.query.filter(
            ActionLog.action_type.in_(["arabic_attendance_uploaded", "arabic_attendance_refreshed"])
        )
        .order_by(ActionLog.created_at.desc(), ActionLog.id.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "action_type": row.action_type,
            "entity_label": row.entity_label or "Arabic attendance workbook",
            "details": row.details or "",
            "created_at": row.created_at,
        }
        for row in action_rows
    ], archive_rows


def get_latest_arabic_attendance_backup() -> Path | None:
    upload_dir = ensure_arabic_attendance_upload_dir()
    backups = sorted(
        upload_dir.glob("students_backup_*"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    return backups[0] if backups else None


def parse_score_number(value) -> float | None:
    text_value = normalize_optional_text(value)
    if not text_value:
        return None
    try:
        return float(text_value)
    except ValueError:
        return None


def is_exam_total_subject(subject_name: str) -> bool:
    normalized = normalize_optional_text(subject_name).lower()
    return normalized.startswith("total")


def is_exam_percentage_subject(subject_name: str) -> bool:
    normalized = normalize_optional_text(subject_name).lower()
    return normalized.startswith("percentage")


def should_display_exam_component(subject_name: str) -> bool:
    return not is_exam_total_subject(subject_name) and not is_exam_percentage_subject(subject_name)


def build_exam_summary_groups(exam_results: list[ExamResult]) -> list[dict]:
    grouped = {}
    for result in exam_results:
        group = grouped.setdefault(
            result.exam_title,
            {
                "exam_title": result.exam_title,
                "exam_date": result.exam_date,
                "components": [],
                "stored_total": None,
                "stored_percentage": None,
            },
        )
        if result.exam_date and (group["exam_date"] is None or result.exam_date > group["exam_date"]):
            group["exam_date"] = result.exam_date

        if is_exam_total_subject(result.subject_name):
            group["stored_total"] = result
        elif is_exam_percentage_subject(result.subject_name):
            group["stored_percentage"] = result
        else:
            group["components"].append(result)

    summaries = []
    for group in grouped.values():
        components = sorted(group["components"], key=lambda item: item.id)
        total_score = 0.0
        total_max = 0.0
        has_numeric_score = False
        shared_note = next((result.notes for result in components if result.notes), None)
        for result in components:
            numeric_score = parse_score_number(result.score_value)
            numeric_max = parse_score_number(result.max_score)
            if numeric_score is not None:
                total_score += numeric_score
                has_numeric_score = True
            if numeric_max is not None:
                total_max += numeric_max

        if not components and group["stored_total"] is not None:
            total_score = parse_score_number(group["stored_total"].score_value) or 0.0
            total_max = parse_score_number(group["stored_total"].max_score) or total_max
            has_numeric_score = parse_score_number(group["stored_total"].score_value) is not None
            if not shared_note:
                shared_note = group["stored_total"].notes

        percentage = round((total_score / total_max) * 100, 1) if has_numeric_score and total_max else None
        if percentage is None and group["stored_percentage"] is not None:
            stored_percentage_value = parse_score_number(group["stored_percentage"].score_value)
            if stored_percentage_value is not None:
                percentage = round(stored_percentage_value * 100, 1) if stored_percentage_value <= 1 else round(stored_percentage_value, 1)
            if not shared_note:
                shared_note = group["stored_percentage"].notes
        summaries.append(
            {
                **group,
                "components": components,
                "computed_total": round(total_score, 2) if has_numeric_score else None,
                "computed_max_total": round(total_max, 2) if total_max else None,
                "computed_percentage": percentage,
                "computed_grade": get_exam_grade_label(percentage),
                "shared_note": shared_note,
            }
        )

    summaries.sort(key=lambda item: (item["exam_date"] or date.min, item["exam_title"]), reverse=True)
    return summaries


def get_exam_summary_for_title(exam_results: list[ExamResult], exam_title: str) -> dict | None:
    for summary in build_exam_summary_groups(exam_results):
        if summary["exam_title"] == exam_title:
            return summary
    return None


def get_exam_grade_label(percentage: float | None) -> str:
    if percentage is None:
        return "-"
    if percentage >= 90:
        return "ممتاز"
    if percentage >= 80:
        return "جيد جدا"
    if percentage >= 70:
        return "جيد"
    if percentage >= 60:
        return "مقبول"
    return "يحتاج تحسين"


def format_decimal_for_display(value: float | None) -> str:
    if value is None:
        return "-"
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def get_exam_publication(exam_title: str) -> ExamPublication | None:
    return ExamPublication.query.filter_by(exam_title=exam_title).first()


def is_exam_visible_to_student(student: Student, exam_title: str) -> bool:
    publication = get_exam_publication(exam_title)
    if publication and not publication.is_published:
        return False

    override = StudentExamVisibility.query.filter_by(
        student_id=student.id,
        exam_title=exam_title,
    ).first()
    return not bool(override and override.is_hidden)


def filter_exam_results_for_student(student: Student, exam_results: list[ExamResult]) -> list[ExamResult]:
    visible_titles = {
        exam_title
        for exam_title in {result.exam_title for result in exam_results}
        if is_exam_visible_to_student(student, exam_title)
    }
    return [result for result in exam_results if result.exam_title in visible_titles]


def get_all_exam_titles() -> list[str]:
    titles = {
        exam_title
        for (exam_title,) in db.session.query(ExamResult.exam_title).distinct().all()
        if exam_title
    }
    titles.update(
        title
        for (title,) in db.session.query(ExamTemplate.title).distinct().all()
        if title
    )
    return sorted(titles)


def parse_template_branches_input(raw_value: str) -> list[dict]:
    branches = []
    for index, line in enumerate((raw_value or "").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        branch_name, _, max_score = line.partition("|")
        branches.append(
            {
                "branch_name": branch_name.strip(),
                "max_score": max_score.strip() or None,
                "order_index": index,
            }
        )
    return branches


def get_template_branch_text(template: ExamTemplate) -> str:
    ordered_branches = sorted(template.branches, key=lambda branch: ((branch.order_index or 0), branch.id))
    return "\n".join(
        f"{branch.branch_name}|{branch.max_score or ''}".rstrip("|")
        for branch in ordered_branches
    )


def sync_exam_template_branches(template: ExamTemplate, branch_rows: list[dict]) -> None:
    ExamTemplateBranch.query.filter_by(exam_template_id=template.id).delete()
    for branch_row in branch_rows:
        db.session.add(
            ExamTemplateBranch(
                exam_template_id=template.id,
                branch_name=branch_row["branch_name"],
                max_score=branch_row["max_score"],
                order_index=branch_row["order_index"],
            )
        )


def ensure_exam_template_from_parsed_rows(parsed_rows: list[dict]) -> None:
    grouped = {}
    for row in parsed_rows:
        if not should_display_exam_component(row["subject_name"]):
            continue
        group = grouped.setdefault(
            row["exam_title"],
            {
                "exam_date": row.get("exam_date"),
                "branches": {},
            },
        )
        if row.get("exam_date") and not group["exam_date"]:
            group["exam_date"] = row["exam_date"]
        group["branches"][row["subject_name"]] = row.get("max_score") or None

    for exam_title, exam_data in grouped.items():
        template = ExamTemplate.query.filter_by(title=exam_title).first()
        if not template:
            template = ExamTemplate(title=exam_title, exam_date=exam_data["exam_date"], is_active=True)
            db.session.add(template)
            db.session.flush()
        elif exam_data["exam_date"] and not template.exam_date:
            template.exam_date = exam_data["exam_date"]

        if not template.branches:
            sync_exam_template_branches(
                template,
                [
                    {
                        "branch_name": branch_name,
                        "max_score": max_score,
                        "order_index": index,
                    }
                    for index, (branch_name, max_score) in enumerate(exam_data["branches"].items(), start=1)
                ],
            )


def ensure_exam_publication_row(exam_title: str, default_published: bool = True) -> None:
    publication = ExamPublication.query.filter_by(exam_title=exam_title).first()
    if not publication:
        db.session.add(ExamPublication(exam_title=exam_title, is_published=default_published))


def set_student_exam_visibility_override(student_id: int, exam_title: str, is_hidden: bool) -> None:
    override = StudentExamVisibility.query.filter_by(student_id=student_id, exam_title=exam_title).first()
    if is_hidden:
        if not override:
            db.session.add(StudentExamVisibility(student_id=student_id, exam_title=exam_title, is_hidden=True))
        else:
            override.is_hidden = True
    elif override:
        db.session.delete(override)


def pdf_escape_text(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def get_exam_grade_label_en(percentage: float | None) -> str:
    if percentage is None:
        return "-"
    if percentage >= 90:
        return "Excellent"
    if percentage >= 80:
        return "Very Good"
    if percentage >= 70:
        return "Good"
    if percentage >= 60:
        return "Pass"
    return "Needs Improvement"


def wrap_pdf_text(text: str, max_chars: int) -> list[str]:
    cleaned = normalize_optional_text(text)
    if not cleaned:
        return [""]
    words = cleaned.split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= max_chars:
            current = candidate
            continue
        if current:
            lines.append(current)
            current = word
        else:
            lines.append(word[:max_chars])
            current = word[max_chars:]
    if current:
        lines.append(current)
    return lines or [""]


def build_pdf_document(page_streams: list[str]) -> bytes:
    objects = []

    def add_object(payload: str | bytes) -> int:
        objects.append(payload)
        return len(objects)

    font_regular_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_bold_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
    content_ids = []
    page_ids = []
    for stream in page_streams:
        stream_data = stream.encode("latin-1", errors="replace")
        content_id = add_object(
            b"<< /Length " + str(len(stream_data)).encode("ascii") + b" >>\nstream\n" + stream_data + b"\nendstream"
        )
        content_ids.append(content_id)
        page_id = add_object(
            "<< /Type /Page /Parent PAGES_REF /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> "
            f"/Contents {content_id} 0 R >>"
        )
        page_ids.append(page_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    pages_id = add_object(f"<< /Type /Pages /Count {len(page_ids)} /Kids [{kids}] >>")
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

    rendered_objects = []
    for index, obj in enumerate(objects, start=1):
        payload = obj.replace("PAGES_REF", f"{pages_id} 0 R") if isinstance(obj, str) else obj
        if isinstance(payload, str):
            payload = payload.encode("latin-1", errors="replace")
        rendered_objects.append((index, payload))

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj_id, payload in rendered_objects:
        offsets.append(len(pdf))
        pdf.extend(f"{obj_id} 0 obj\n".encode("ascii"))
        pdf.extend(payload)
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(rendered_objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(rendered_objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode(
            "ascii"
        )
    )
    return bytes(pdf)


def build_simple_pdf(lines: list[str]) -> bytes:
    pages = []
    page_lines = []
    for line in lines:
        page_lines.append(line)
        if len(page_lines) >= 42:
            pages.append(page_lines)
            page_lines = []
    if page_lines or not pages:
        pages.append(page_lines)

    objects = []

    def add_object(payload: str | bytes) -> int:
        objects.append(payload)
        return len(objects)

    font_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    content_ids = []
    page_ids = []
    for page_lines in pages:
        stream_lines = ["BT", "/F1 12 Tf", "50 790 Td", "14 TL"]
        for line in page_lines:
            stream_lines.append(f"({pdf_escape_text(line)}) Tj")
            stream_lines.append("T*")
        stream_lines.append("ET")
        stream_data = "\n".join(stream_lines).encode("latin-1", errors="replace")
        content_id = add_object(b"<< /Length " + str(len(stream_data)).encode("ascii") + b" >>\nstream\n" + stream_data + b"\nendstream")
        content_ids.append(content_id)
        page_id = add_object(
            f"<< /Type /Page /Parent PAGES_REF /MediaBox [0 0 595 842] /Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>"
        )
        page_ids.append(page_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    pages_id = add_object(f"<< /Type /Pages /Count {len(page_ids)} /Kids [{kids}] >>")
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

    rendered_objects = []
    for index, obj in enumerate(objects, start=1):
        payload = obj.replace("PAGES_REF", f"{pages_id} 0 R") if isinstance(obj, str) else obj
        if isinstance(payload, str):
            payload = payload.encode("latin-1", errors="replace")
        rendered_objects.append((index, payload))

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj_id, payload in rendered_objects:
        offsets.append(len(pdf))
        pdf.extend(f"{obj_id} 0 obj\n".encode("ascii"))
        pdf.extend(payload)
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(rendered_objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(rendered_objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode(
            "ascii"
        )
    )
    return bytes(pdf)


def build_exam_report_pdf(report_title: str, student: Student, level_name: str | None, exam_summary: dict) -> bytes:
    page_width = 595
    page_height = 842
    left = 42
    content_width = page_width - (left * 2)
    page_streams = []
    commands = []

    palette = {
        "navy": (0.07, 0.19, 0.39),
        "sky": (0.18, 0.53, 0.86),
        "teal": (0.04, 0.58, 0.56),
        "gold": (0.95, 0.70, 0.16),
        "ink": (0.14, 0.17, 0.23),
        "muted": (0.39, 0.45, 0.55),
        "line": (0.84, 0.88, 0.93),
        "soft": (0.96, 0.98, 1.00),
        "white": (1.0, 1.0, 1.0),
        "row": (0.985, 0.99, 1.0),
    }

    def add_text(x: float, y: float, text: str, size: int = 12, font: str = "F1", color=(0, 0, 0)) -> None:
        commands.append(
            "BT "
            f"/{font} {size} Tf "
            f"{color[0]:.3f} {color[1]:.3f} {color[2]:.3f} rg "
            f"1 0 0 1 {x:.2f} {y:.2f} Tm "
            f"({pdf_escape_text(text)}) Tj ET"
        )

    def add_rect(x: float, y: float, width: float, height: float, fill=None, stroke=None, line_width: float = 1.0) -> None:
        commands.append("q")
        if fill is not None:
            commands.append(f"{fill[0]:.3f} {fill[1]:.3f} {fill[2]:.3f} rg")
        if stroke is not None:
            commands.append(f"{stroke[0]:.3f} {stroke[1]:.3f} {stroke[2]:.3f} RG")
            commands.append(f"{line_width:.2f} w")
        commands.append(f"{x:.2f} {y:.2f} {width:.2f} {height:.2f} re")
        if fill is not None and stroke is not None:
            commands.append("B")
        elif fill is not None:
            commands.append("f")
        else:
            commands.append("S")
        commands.append("Q")

    def start_page(page_number: int) -> float:
        commands.clear()
        add_rect(0, 0, page_width, page_height, fill=palette["white"])
        add_rect(left, 744, content_width, 66, fill=palette["navy"])
        add_text(left + 18, 785, report_title, size=22, font="F2", color=palette["white"])
        add_text(left + 18, 764, exam_summary["exam_title"], size=14, font="F1", color=palette["white"])
        add_text(page_width - 105, 764, f"Page {page_number}", size=10, font="F1", color=palette["white"])

        add_rect(left, 675, content_width, 56, fill=palette["soft"], stroke=palette["line"])
        add_text(left + 14, 712, f"Student: {student.full_name}", size=12, font="F2", color=palette["ink"])
        add_text(left + 14, 693, f"Student Code: {student.student_code}", size=11, color=palette["ink"])
        add_text(left + 250, 712, f"Level: {level_name or '-'}", size=11, color=palette["ink"])
        add_text(left + 250, 693, f"Date: {exam_summary['exam_date'] or '-'}", size=11, color=palette["ink"])

        card_width = (content_width - 20) / 3
        card_y = 610
        summary_cards = [
            (
                "Total",
                format_decimal_for_display(exam_summary["computed_total"])
                + (
                    f" / {format_decimal_for_display(exam_summary['computed_max_total'])}"
                    if exam_summary["computed_max_total"] is not None
                    else ""
                ),
                palette["sky"],
            ),
            (
                "Percentage",
                f"{format_decimal_for_display(exam_summary['computed_percentage'])}%",
                palette["teal"],
            ),
            (
                "Grade",
                get_exam_grade_label_en(exam_summary["computed_percentage"]),
                palette["gold"],
            ),
        ]
        for index, (label, value, color) in enumerate(summary_cards):
            x = left + (index * (card_width + 10))
            add_rect(x, card_y, card_width, 58, fill=color)
            add_text(x + 12, card_y + 38, label, size=10, font="F1", color=palette["white"])
            add_text(x + 12, card_y + 16, value, size=16, font="F2", color=palette["white"])

        table_y = 560
        column_widths = [250, 85, 85, content_width - 420]
        headers = ["Branch", "Score", "Out Of", "Notes"]
        x = left
        for width, header in zip(column_widths, headers):
            add_rect(x, table_y, width, 26, fill=palette["navy"])
            add_text(x + 8, table_y + 8, header, size=10, font="F2", color=palette["white"])
            x += width
        return table_y - 30

    def flush_page() -> None:
        page_streams.append("\n".join(commands))

    current_page = 1
    current_y = start_page(current_page)
    row_height = 28
    for index, component in enumerate(exam_summary["components"]):
        if current_y < 110:
            flush_page()
            current_page += 1
            current_y = start_page(current_page)

        x = left
        fill = palette["row"] if index % 2 == 0 else palette["white"]
        note_lines = wrap_pdf_text(component.notes or exam_summary.get("shared_note") or "-", 28)
        row_dynamic_height = max(row_height, 18 + (len(note_lines) * 12))
        widths = [250, 85, 85, content_width - 420]
        values = [
            component.subject_name,
            component.score_value or "-",
            component.max_score or "-",
            note_lines,
        ]
        for column_index, width in enumerate(widths):
            add_rect(x, current_y, width, row_dynamic_height, fill=fill, stroke=palette["line"], line_width=0.7)
            if column_index < 3:
                add_text(x + 8, current_y + row_dynamic_height - 18, str(values[column_index]), size=10, color=palette["ink"])
            else:
                for line_index, line in enumerate(values[column_index]):
                    add_text(x + 8, current_y + row_dynamic_height - 18 - (line_index * 12), line, size=9, color=palette["muted"])
            x += width
        current_y -= row_dynamic_height

    shared_note = exam_summary.get("shared_note")
    if shared_note:
        note_lines = wrap_pdf_text(shared_note, 86)
        note_height = 22 + (len(note_lines) * 12)
        if current_y - note_height < 70:
            flush_page()
            current_page += 1
            current_y = start_page(current_page)
        add_rect(left, current_y - note_height, content_width, note_height, fill=palette["soft"], stroke=palette["line"])
        add_text(left + 10, current_y - 18, "General Notes", size=10, font="F2", color=palette["ink"])
        for line_index, line in enumerate(note_lines):
            add_text(left + 10, current_y - 34 - (line_index * 12), line, size=10, color=palette["muted"])
        current_y -= note_height + 10

    add_text(left, 28, "Hikmah Academy", size=9, font="F2", color=palette["muted"])
    add_text(page_width - 190, 28, "Exam report generated automatically", size=9, color=palette["muted"])
    flush_page()
    return build_pdf_document(page_streams)


def parse_optional_exam_date(value):
    text_value = normalize_optional_text(value)
    if not text_value:
        return None
    parsed = pd.to_datetime(text_value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def extract_max_score_from_header(header_text: str) -> str:
    match = re.search(r"/\s*([0-9]+(?:\.[0-9]+)?)", header_text or "")
    return match.group(1) if match else ""


def infer_level_name_from_sheet(sheet_name: str) -> str:
    normalized_sheet = normalize_optional_text(sheet_name)
    prefix = normalized_sheet.split("(")[0].strip()

    for display_name, aliases in LEVEL_DISPLAY_TO_CODES.items():
        candidates = [display_name, *aliases]
        if any(prefix.lower() == alias.strip().lower() for alias in candidates):
            return display_name
    return normalize_level_display_name(prefix)


def ensure_exam_upload_dir() -> None:
    os.makedirs(app.config["EXAM_UPLOAD_DIR"], exist_ok=True)


def save_uploaded_exam_file(uploaded_file) -> str:
    ensure_exam_upload_dir()
    safe_name = os.path.basename(uploaded_file.filename or "exam_upload.xlsx").strip() or "exam_upload.xlsx"
    destination = os.path.join(app.config["EXAM_UPLOAD_DIR"], safe_name)
    uploaded_file.seek(0)
    uploaded_file.save(destination)
    uploaded_file.seek(0)
    return destination


def ensure_assignment_upload_dir() -> None:
    os.makedirs(app.config["ASSIGNMENT_UPLOAD_DIR"], exist_ok=True)
    os.makedirs(os.path.join(app.config["ASSIGNMENT_UPLOAD_DIR"], "resources"), exist_ok=True)
    os.makedirs(os.path.join(app.config["ASSIGNMENT_UPLOAD_DIR"], "submissions"), exist_ok=True)


def sanitize_uploaded_filename(filename: str) -> str:
    base_name = os.path.basename(filename or "").strip() or "upload"
    if "." in base_name:
        stem, extension = base_name.rsplit(".", 1)
        stem = re.sub(r"[^A-Za-z0-9_-]+", "_", stem).strip("_") or "file"
        extension = re.sub(r"[^A-Za-z0-9]+", "", extension).lower() or "bin"
        return f"{stem}.{extension}"
    return re.sub(r"[^A-Za-z0-9_-]+", "_", base_name).strip("_") or "file"


def save_assignment_uploaded_file(uploaded_file, bucket: str) -> tuple[str, str]:
    ensure_assignment_upload_dir()
    safe_name = sanitize_uploaded_filename(uploaded_file.filename or "upload")
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    stored_name = f"{timestamp}_{safe_name}"
    relative_path = os.path.join(bucket, stored_name)
    absolute_path = os.path.join(app.config["ASSIGNMENT_UPLOAD_DIR"], relative_path)
    uploaded_file.save(absolute_path)
    return relative_path, safe_name


def ensure_weekly_report_archive_dir() -> None:
    os.makedirs(app.config["WEEKLY_REPORT_ARCHIVE_DIR"], exist_ok=True)


def weekly_report_archive_path(file_name: str) -> str:
    return os.path.join(app.config["WEEKLY_REPORT_ARCHIVE_DIR"], os.path.basename(file_name))


def assignment_file_absolute_path(relative_path: str | None) -> str | None:
    if not relative_path:
        return None
    return os.path.join(app.config["ASSIGNMENT_UPLOAD_DIR"], relative_path)


def register_exam_import_issue(source_file_name: str, parsed_row: dict) -> None:
    alias_name = parsed_row["full_name"]
    level_name = normalize_level_display_name(parsed_row["level_name"]) if parsed_row["level_name"] else None
    if not alias_name:
        return

    existing_issue = ExamImportIssue.query.filter_by(
        source_file_name=source_file_name,
        alias_name=alias_name,
        level_name=level_name,
    ).first()
    if existing_issue:
        return

    db.session.add(
        ExamImportIssue(
            source_file_name=source_file_name,
            alias_name=alias_name,
            level_name=level_name,
            exam_title=parsed_row.get("exam_title") or None,
            notes="Student name could not be matched automatically.",
        )
    )


def clear_exam_import_issue(source_file_name: str, alias_name: str, level_name: str | None) -> None:
    ExamImportIssue.query.filter_by(
        source_file_name=source_file_name,
        alias_name=alias_name,
        level_name=level_name,
    ).delete()


def import_exam_rows(parsed_rows: list[dict], source_file_name: str) -> tuple[int, int, list[str]]:
    ExamImportIssue.query.filter_by(source_file_name=source_file_name).delete()
    ensure_exam_template_from_parsed_rows(parsed_rows)
    for exam_title in {row["exam_title"] for row in parsed_rows if row.get("exam_title")}:
        ensure_exam_publication_row(exam_title, default_published=True)

    imported_count = 0
    skipped_count = 0
    row_errors = []

    for parsed_row in parsed_rows:
        student = find_student_for_exam_row(parsed_row)
        if not student:
            register_exam_import_issue(source_file_name, parsed_row)
            skipped_count += 1
            continue

        level = student.level
        if not level and student.level_name:
            level = Level.query.filter_by(name=normalize_level_display_name(student.level_name)).first()

        existing_result = ExamResult.query.filter_by(
            student_id=student.id,
            exam_title=parsed_row["exam_title"],
            subject_name=parsed_row["subject_name"],
        ).first()

        if not existing_result:
            existing_result = ExamResult(
                student_id=student.id,
                exam_title=parsed_row["exam_title"],
                subject_name=parsed_row["subject_name"],
            )
            db.session.add(existing_result)

        existing_result.level_id = level.id if level else None
        existing_result.score_value = parsed_row["score_value"]
        existing_result.max_score = parsed_row["max_score"] or None
        existing_result.exam_date = parsed_row["exam_date"]
        existing_result.notes = parsed_row["notes"] or None
        existing_result.source_file_name = source_file_name
        imported_count += 1

    return imported_count, skipped_count, row_errors


def read_excel_exam_rows(uploaded_file) -> tuple[list[dict], list[str]]:
    uploaded_file.seek(0)
    workbook = pd.ExcelFile(uploaded_file)
    if len(workbook.sheet_names) > 1:
        return read_multi_sheet_exam_rows(workbook)

    dataframe = pd.read_excel(workbook, sheet_name=workbook.sheet_names[0])
    dataframe = dataframe.rename(columns=lambda value: normalize_import_header(str(value)))
    dataframe = dataframe.fillna("")

    column_aliases = {
        "student_code": [
            "student_code",
            "student_id",
            "code",
            "studentid",
            "id",
            "كود_الطالب",
            "رمز_الطالب",
            "رقم_الطالب",
            "رقم_القيد",
        ],
        "full_name": [
            "full_name",
            "student_name",
            "name",
            "student",
            "fullstudentname",
            "اسم_الطالب",
            "الاسم",
            "اسم",
        ],
        "level_name": [
            "level_name",
            "level",
            "class",
            "group",
            "grade_level",
            "المستوي",
            "المستوى",
            "الصف",
            "الفصل",
            "المجموعه",
            "المجموعة",
        ],
        "exam_title": [
            "exam_title",
            "exam",
            "test_name",
            "assessment",
            "assessment_name",
            "اسم_الاختبار",
            "الاختبار",
            "نوع_الاختبار",
        ],
        "subject_name": [
            "subject_name",
            "subject",
            "course",
            "material",
            "الماده",
            "المادة",
            "المقرر",
        ],
        "score_value": [
            "score_value",
            "score",
            "marks",
            "mark",
            "grade",
            "result",
            "الدرجه",
            "الدرجة",
            "درجه",
            "درجة",
            "النتيجه",
            "النتيجة",
            "المجموع",
        ],
        "max_score": [
            "max_score",
            "total",
            "out_of",
            "full_mark",
            "max_marks",
            "الدرجه_النهائيه",
            "الدرجة_النهائية",
            "النهايه_الكبري",
            "النهاية_الكبرى",
            "من",
        ],
        "exam_date": [
            "exam_date",
            "date",
            "test_date",
            "assessment_date",
            "تاريخ_الاختبار",
            "التاريخ",
            "تاريخ",
        ],
        "notes": [
            "notes",
            "note",
            "remarks",
            "remark",
            "comment",
            "ملاحظات",
            "ملاحظه",
            "ملاحظة",
        ],
    }

    resolved_columns = {
        target_name: next(
            (alias for alias in aliases if alias in dataframe.columns),
            None,
        )
        for target_name, aliases in column_aliases.items()
    }

    missing_required = [
        column_name
        for column_name in ("exam_title", "subject_name", "score_value")
        if not resolved_columns[column_name]
    ]
    if not resolved_columns["student_code"] and not resolved_columns["full_name"]:
        missing_required.append("student_code/full_name")

    if missing_required:
        return [], [
            "Missing required column(s): "
            + ", ".join(missing_required)
            + ". Expected columns like student_code or full_name, exam_title, subject_name, score."
        ]

    parsed_rows = []
    row_errors = []
    for row_index, row in dataframe.iterrows():
        parsed_row = {}
        for target_name, source_name in resolved_columns.items():
            parsed_row[target_name] = normalize_optional_text(row[source_name]) if source_name else ""

        parsed_row["exam_date"] = parse_optional_exam_date(
            row[resolved_columns["exam_date"]] if resolved_columns["exam_date"] else ""
        )

        if not any(value for key, value in parsed_row.items() if key != "exam_date") and not parsed_row["exam_date"]:
            continue

        if (
            not parsed_row["exam_title"]
            or not parsed_row["subject_name"]
            or not parsed_row["score_value"]
        ):
            row_errors.append(
                f"Row {row_index + 2}: exam_title, subject_name, and score are required."
            )
            continue

        parsed_rows.append(parsed_row)

    return parsed_rows, row_errors


def read_multi_sheet_exam_rows(workbook: pd.ExcelFile) -> tuple[list[dict], list[str]]:
    parsed_rows = []
    row_errors = []

    for sheet_name in workbook.sheet_names:
        sheet = pd.read_excel(workbook, sheet_name=sheet_name, header=None).fillna("")
        exam_title = ""
        header_row_index = None
        student_name_col = None

        for row_index in range(min(len(sheet), 8)):
            row_values = [normalize_optional_text(value) for value in sheet.iloc[row_index].tolist()]
            if not exam_title:
                exam_title = next((value for value in row_values if "exam" in value.lower()), exam_title)
            normalized_values = [normalize_import_header(value) for value in row_values]
            if "student_name" in normalized_values:
                header_row_index = row_index
                student_name_col = normalized_values.index("student_name")
                break

        if header_row_index is None or student_name_col is None:
            row_errors.append(f"Sheet '{sheet_name}': could not find the student header row.")
            continue

        if not exam_title:
            exam_title = "Imported Exam"

        header_values = [normalize_optional_text(value) for value in sheet.iloc[header_row_index].tolist()]
        level_name = infer_level_name_from_sheet(sheet_name)
        year_col = student_name_col + 1

        for row_index in range(header_row_index + 1, len(sheet)):
            row_values = [normalize_optional_text(value) for value in sheet.iloc[row_index].tolist()]
            student_name = row_values[student_name_col] if len(row_values) > student_name_col else ""
            year_value = row_values[year_col] if len(row_values) > year_col else ""

            if not student_name:
                continue

            has_score_cells = any(value for value in row_values[year_col + 1 :])
            if not has_score_cells:
                continue

            for column_index in range(year_col + 1, len(header_values)):
                subject_header = normalize_optional_text(header_values[column_index])
                score_value = row_values[column_index] if column_index < len(row_values) else ""

                if not subject_header or not score_value:
                    continue

                normalized_subject = subject_header.lower()
                if normalized_subject in {"percentage %", "remarks"}:
                    continue

                parsed_rows.append(
                    {
                        "student_code": "",
                        "full_name": student_name,
                        "level_name": level_name,
                        "exam_title": exam_title,
                        "subject_name": subject_header,
                        "score_value": score_value,
                        "max_score": extract_max_score_from_header(subject_header),
                        "exam_date": None,
                        "notes": year_value if year_value else "",
                    }
                )

    return parsed_rows, row_errors


def find_student_for_exam_row(parsed_row: dict) -> Student | None:
    student_code = parsed_row["student_code"]
    full_name = parsed_row["full_name"]
    level_name = normalize_level_display_name(parsed_row["level_name"]) if parsed_row["level_name"] else ""

    if student_code:
        return Student.query.filter_by(student_code=student_code).first()

    if not full_name:
        return None

    alias_match = StudentNameAlias.query.filter_by(
        alias_name=full_name,
        level_name=level_name or None,
    ).first()
    if not alias_match and level_name:
        alias_match = StudentNameAlias.query.filter_by(
            alias_name=full_name,
            level_name=None,
        ).first()
    if alias_match:
        return Student.query.get(alias_match.student_id)

    normalized_target = normalize_person_name(full_name)
    query = Student.query
    if level_name:
        query = query.filter(
            db.or_(
                Student.level_name == level_name,
                Student.level.has(Level.name == level_name),
            )
        )

    matches = query.all()
    normalized_matches = [
        student for student in matches if normalize_person_name(student.full_name) == normalized_target
    ]
    if len(normalized_matches) == 1:
        return normalized_matches[0]

    exact_casefold_matches = [
        student for student in matches if normalize_optional_text(student.full_name).casefold() == full_name.casefold()
    ]
    if len(exact_casefold_matches) == 1:
        return exact_casefold_matches[0]

    scored_matches = []
    target_tokens = split_name_tokens(full_name)
    for student in matches:
        candidate_name = student.full_name or ""
        candidate_tokens = split_name_tokens(candidate_name)
        score = name_similarity_score(full_name, candidate_name)
        shared_tokens = set(target_tokens) & set(candidate_tokens)
        scored_matches.append((score, len(shared_tokens), abs(len(target_tokens) - len(candidate_tokens)), student))

    scored_matches.sort(key=lambda item: (item[0], item[1], -item[2]), reverse=True)
    if scored_matches:
        best_score, best_shared, best_len_diff, best_student = scored_matches[0]
        next_score = scored_matches[1][0] if len(scored_matches) > 1 else 0.0

        if (
            best_score >= 0.93
            or (
                best_score >= 0.88
                and best_shared >= max(2, min(len(target_tokens), len(split_name_tokens(best_student.full_name or ""))) - 1)
                and best_len_diff <= 2
            )
        ) and best_score - next_score >= 0.03:
            return best_student

    return None


def generate_next_student_code() -> str:
    prefix = "STD-"
    max_num = 0
    for (code,) in db.session.query(Student.student_code).all():
        if not code or not code.startswith(prefix):
            continue
        suffix = code[len(prefix):]
        if suffix.isdigit():
            max_num = max(max_num, int(suffix))
    return f"{prefix}{max_num + 1:06d}"


def send_email_via_smtp(to_email: str, subject: str, body: str) -> None:
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM", smtp_username or "")
    smtp_use_tls = os.getenv("SMTP_USE_TLS", "true").strip().lower() in {"1", "true", "yes"}

    if not smtp_host or not smtp_username or not smtp_password or not smtp_from:
        raise ValueError("SMTP settings are missing. Set SMTP_HOST, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM.")

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_from
    message["To"] = to_email
    message.set_content(body)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        if smtp_use_tls:
            server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(message)


def count_study_days(start_date: date, end_date: date) -> int:
    # Study days are Monday-Thursday only.
    count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() in {0, 1, 2, 3}:
            count += 1
        current += timedelta(days=1)
    return count


def build_student_report_data(student: Student, period: str) -> dict:
    end_date = date.today()
    if period == "week":
        start_date = end_date - timedelta(days=end_date.weekday())
    else:
        start_date = end_date.replace(day=1)

    level = student.level
    if not level and student.level_name:
        normalized_level_name = normalize_level_display_name(student.level_name)
        level = Level.query.filter_by(name=normalized_level_name).first()

    teacher = level.teacher if level else None
    level_display_name = level.name if level else normalize_level_display_name(student.level_name)
    imported_arabic_attendance = build_imported_arabic_attendance_snapshot(student)

    level_assignments = []
    submissions_by_assignment_id = {}
    if level:
        all_level_assignments = (
            Assignment.query.filter_by(level_id=level.id)
            .order_by(Assignment.created_at.asc(), Assignment.id.asc())
            .all()
        )
        for assignment in all_level_assignments:
            assignment_anchor_date = assignment.due_date or (
                assignment.created_at.date() if assignment.created_at else None
            )
            if assignment_anchor_date and start_date <= assignment_anchor_date <= end_date:
                level_assignments.append(assignment)

        student_submissions = AssignmentSubmission.query.filter_by(student_id=student.id).all()
        submissions_by_assignment_id = {
            submission.assignment_id: submission
            for submission in student_submissions
        }

    counts = (
        db.session.query(
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.sum(db.case((Attendance.status == "Absent", 1), else_=0)).label("absent_count"),
            db.func.sum(db.case((Attendance.status == "Late", 1), else_=0)).label("late_count"),
            db.func.sum(db.case((Attendance.status == "Excused", 1), else_=0)).label("excused_count"),
            db.func.count(Attendance.id).label("total_count"),
        )
        .filter(
            Attendance.student_id == student.id,
            Attendance.attendance_date >= start_date,
            Attendance.attendance_date <= end_date,
        )
        .first()
    )

    present_count = int(counts.present_count or 0)
    absent_count = int(counts.absent_count or 0)
    late_count = int(counts.late_count or 0)
    excused_count = int(counts.excused_count or 0)
    total_count = int(counts.total_count or 0)
    if period == "month":
        expected_days = count_study_days(start_date, end_date)
        attendance_percentage = round((present_count / expected_days) * 100, 1) if expected_days else 0.0
    else:
        attendance_percentage = round((present_count / total_count) * 100, 1) if total_count else 0.0
    month_key = end_date.strftime("%Y-%m")
    monthly_note_text = ""
    if period == "month":
        note = StudentMonthlyNote.query.filter_by(student_id=student.id, month_key=month_key).first()
        monthly_note_text = (note.note_text or "").strip() if note else ""

    homework_total = len(level_assignments)
    homework_submitted = 0
    homework_reviewed = 0
    homework_pending_review = 0
    homework_missing = 0
    for assignment in level_assignments:
        submission = submissions_by_assignment_id.get(assignment.id)
        if submission and submission.submitted_at:
            homework_submitted += 1
            if submission.status == "Reviewed":
                homework_reviewed += 1
            else:
                homework_pending_review += 1
        else:
            homework_missing += 1

    homework_completion_rate = (
        round((homework_submitted / homework_total) * 100, 1)
        if homework_total
        else 0.0
    )

    return {
        "level_name": level_display_name,
        "teacher_name": teacher.full_name if teacher else "-",
        "start_date": start_date,
        "end_date": end_date,
        "present_count": present_count,
        "absent_count": absent_count,
        "late_count": late_count,
        "excused_count": excused_count,
        "attendance_percentage": attendance_percentage,
        "imported_arabic_attendance": imported_arabic_attendance,
        "month_key": month_key,
        "monthly_note_text": monthly_note_text,
        "homework_total": homework_total,
        "homework_submitted": homework_submitted,
        "homework_reviewed": homework_reviewed,
        "homework_pending_review": homework_pending_review,
        "homework_missing": homework_missing,
        "homework_completion_rate": homework_completion_rate,
    }


def build_monthly_report_review_rows() -> dict:
    students = Student.query.order_by(Student.full_name.asc()).all()
    month_key = date.today().strftime("%Y-%m")
    rows = []

    for student in students:
        report_data = build_student_report_data(student, "month")
        parent_email = (student.parent_email or "").strip()
        parent_whatsapp = (student.parent_whatsapp or "").strip()
        has_monthly_note = bool((report_data["monthly_note_text"] or "").strip())
        has_valid_email = validate_email(parent_email)

        issues = []
        if not has_monthly_note:
            issues.append("missing_monthly_note")
        if not has_valid_email:
            issues.append("missing_or_invalid_parent_email")
        if not parent_whatsapp:
            issues.append("missing_parent_whatsapp")

        rows.append(
            {
                "student": student,
                "report_data": report_data,
                "parent_email": parent_email,
                "parent_whatsapp": parent_whatsapp,
                "has_monthly_note": has_monthly_note,
                "has_valid_email": has_valid_email,
                "send_ready": has_monthly_note and has_valid_email,
                "issues": issues,
            }
        )

    summary = {
        "month_key": month_key,
        "total_students": len(rows),
        "send_ready": sum(1 for row in rows if row["send_ready"]),
        "missing_notes": sum(1 for row in rows if not row["has_monthly_note"]),
        "missing_email": sum(1 for row in rows if not row["has_valid_email"]),
        "missing_whatsapp": sum(1 for row in rows if not row["parent_whatsapp"]),
    }

    return {
        "rows": rows,
        "summary": summary,
    }


def format_percent_label(value: float | None) -> str:
    if value is None:
        return "-"
    if float(value).is_integer():
        return f"{int(value)}%"
    return f"{value:.1f}%"


def parse_optional_date_input(value: str) -> date | None:
    text_value = (value or "").strip()
    if not text_value:
        return None
    try:
        return datetime.strptime(text_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def normalize_weekday_name(value: str) -> str:
    normalized = (value or "").strip().lower()
    return normalized if normalized in WEEKDAY_NAME_TO_INDEX else "thursday"


def parse_time_hhmm(value: str) -> tuple[int, int]:
    text_value = (value or "").strip()
    try:
        hour_str, minute_str = text_value.split(":", 1)
        hour = min(max(int(hour_str), 0), 23)
        minute = min(max(int(minute_str), 0), 59)
        return hour, minute
    except (ValueError, AttributeError):
        return 14, 0


def get_calendar_settings() -> dict:
    return {
        "duty_start_time": get_system_setting("duty_start_time", "08:00"),
        "duty_end_time": get_system_setting("duty_end_time", "15:00"),
        "weekly_followup_weekday": normalize_weekday_name(get_system_setting("weekly_followup_weekday", "thursday")),
        "weekly_followup_time": get_system_setting("weekly_followup_time", "14:00"),
        "last_weekly_followup_run": get_system_setting("last_weekly_followup_run", ""),
    }


def is_holiday_date(target_date: date) -> HolidayPeriod | None:
    return HolidayPeriod.query.filter(
        HolidayPeriod.is_active.is_(True),
        HolidayPeriod.start_date <= target_date,
        HolidayPeriod.end_date >= target_date,
    ).order_by(HolidayPeriod.start_date.asc(), HolidayPeriod.id.asc()).first()


def should_run_weekly_followup(target_date: date) -> tuple[bool, str]:
    academic_event = get_primary_academic_event(target_date)
    if academic_event and not academic_event.is_instructional:
        return False, f"Skipped because {target_date.isoformat()} is inside {academic_event.title}."

    holiday = is_holiday_date(target_date)
    if holiday:
        return False, f"Skipped because {target_date.isoformat()} is داخل إجازة: {holiday.title}."

    weekday_name = get_calendar_settings()["weekly_followup_weekday"]
    if target_date.weekday() != WEEKDAY_NAME_TO_INDEX[weekday_name]:
        return False, f"Skipped because configured follow-up day is {weekday_name.title()}."

    return True, "Follow-up generation allowed."


def get_teaching_week_window(target_date: date) -> tuple[date, date]:
    week_start = target_date - timedelta(days=target_date.weekday())
    week_end = week_start + timedelta(days=3)
    return week_start, min(target_date, week_end)


def build_teaching_days(start_date: date, end_date: date) -> list[date]:
    teaching_days = []
    current = start_date
    while current <= end_date:
        if current.weekday() <= 3 and not is_non_teaching_date(current):
            teaching_days.append(current)
        current += timedelta(days=1)
    return teaching_days


def format_weekday_names(days: list[date], lang: str | None = None) -> str:
    if not days:
        return "-"
    active_lang = lang or get_current_ui_language()
    labels = ARABIC_WEEKDAY_LABELS if active_lang == "ar" else ENGLISH_WEEKDAY_LABELS
    separator = "، " if active_lang == "ar" else ", "
    return separator.join(labels.get(day.weekday(), day.isoformat()) for day in days)


def build_month_teaching_week_starts(start_date: date, end_date: date) -> list[date]:
    teaching_days = build_teaching_days(start_date, end_date)
    week_starts = []
    seen = set()
    for teaching_day in teaching_days:
        week_start = teaching_day - timedelta(days=teaching_day.weekday())
        if week_start not in seen:
            seen.add(week_start)
            week_starts.append(week_start)
    return week_starts


def collect_weekly_teacher_reviews(target_date: date) -> list[dict]:
    week_start, week_end = get_teaching_week_window(target_date)
    expected_teaching_dates = build_teaching_days(week_start, week_end)
    expected_teaching_days = len(expected_teaching_dates)
    active_lang = get_current_ui_language()
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    review_rows = []
    for level in levels:
        students_count = Student.query.filter_by(level_id=level.id).count()
        if students_count == 0 or not level.teacher_id:
            continue

        attendance_dates_recorded = {
            row[0]
            for row in db.session.query(db.distinct(Attendance.attendance_date))
            .filter(
                Attendance.level_id == level.id,
                Attendance.attendance_date >= week_start,
                Attendance.attendance_date <= week_end,
            )
            .all()
            if row[0] is not None
        }
        recording_dates_uploaded = {
            row[0]
            for row in db.session.query(db.distinct(ClassRecording.lesson_date))
            .filter(
                ClassRecording.class_id == level.id,
                ClassRecording.lesson_date >= week_start,
                ClassRecording.lesson_date <= week_end,
            )
            .all()
            if row[0] is not None
        }
        attendance_missing_days = [day for day in expected_teaching_dates if day not in attendance_dates_recorded]
        recording_missing_days = [day for day in expected_teaching_dates if day not in recording_dates_uploaded]
        attendance_days_recorded = len(attendance_dates_recorded)
        recording_days_uploaded = len(recording_dates_uploaded)
        weekly_assignments_count = Assignment.query.filter(
            Assignment.level_id == level.id,
            db.func.date(Assignment.created_at) >= week_start.isoformat(),
            db.func.date(Assignment.created_at) <= week_end.isoformat(),
        ).count()

        issues = []
        if attendance_days_recorded < expected_teaching_days:
            issues.append(
                ui_text(
                    "attendance_recorded_missing",
                    recorded=attendance_days_recorded,
                    expected=expected_teaching_days,
                    days=format_weekday_names(attendance_missing_days, active_lang),
                )
            )
        if recording_days_uploaded < expected_teaching_days:
            issues.append(
                ui_text(
                    "zoom_recordings_uploaded_missing",
                    uploaded=recording_days_uploaded,
                    expected=expected_teaching_days,
                    days=format_weekday_names(recording_missing_days, active_lang),
                )
            )
        if weekly_assignments_count == 0:
            issues.append(ui_text("no_homework_assigned_week"))

        review_rows.append(
            {
                "level": level,
                "teacher": level.teacher,
                "students_count": students_count,
                "week_start": week_start,
                "week_end": week_end,
                "expected_teaching_days": expected_teaching_days,
                "attendance_days_recorded": int(attendance_days_recorded),
                "recording_days_uploaded": int(recording_days_uploaded),
                "attendance_missing_days": attendance_missing_days,
                "recording_missing_days": recording_missing_days,
                "attendance_missing_days_label": format_weekday_names(attendance_missing_days, active_lang),
                "recording_missing_days_label": format_weekday_names(recording_missing_days, active_lang),
                "weekly_assignments_count": int(weekly_assignments_count),
                "issues": issues,
                "status": "follow_up" if issues else "praise",
            }
        )
    return review_rows


def build_teacher_monthly_report_rows(target_date: date | None = None) -> dict:
    today = target_date or date.today()
    active_lang = get_current_ui_language()
    month_start = today.replace(day=1)
    teaching_days = build_teaching_days(month_start, today)
    teaching_week_starts = build_month_teaching_week_starts(month_start, today)
    teachers = Teacher.query.order_by(Teacher.full_name.asc()).all()
    rows = []

    for teacher in teachers:
        teacher_levels = Level.query.filter_by(teacher_id=teacher.id).order_by(Level.order_index.asc(), Level.name.asc()).all()
        if not teacher_levels:
            continue

        expected_days_per_level = len(teaching_days)
        expected_weeks_per_level = len(teaching_week_starts)
        attendance_days_recorded = 0
        recording_days_uploaded = 0
        homework_weeks_completed = 0
        monthly_assignments_count = 0
        monthly_pending_reviews = 0
        level_checks = []

        for level in teacher_levels:
            attendance_dates = {
                row[0]
                for row in db.session.query(db.distinct(Attendance.attendance_date))
                .filter(
                    Attendance.level_id == level.id,
                    Attendance.attendance_date >= month_start,
                    Attendance.attendance_date <= today,
                )
                .all()
                if row[0] is not None
            }
            recording_dates = {
                row[0]
                for row in db.session.query(db.distinct(ClassRecording.lesson_date))
                .filter(
                    ClassRecording.class_id == level.id,
                    ClassRecording.lesson_date >= month_start,
                    ClassRecording.lesson_date <= today,
                )
                .all()
                if row[0] is not None
            }
            assignment_rows = (
                Assignment.query.filter(
                    Assignment.level_id == level.id,
                    db.func.date(Assignment.created_at) >= month_start.isoformat(),
                    db.func.date(Assignment.created_at) <= today.isoformat(),
                )
                .order_by(Assignment.created_at.desc(), Assignment.id.desc())
                .all()
            )
            assignment_week_starts = {
                assignment.created_at.date() - timedelta(days=assignment.created_at.date().weekday())
                for assignment in assignment_rows
                if assignment.created_at
            }
            pending_reviews = (
                AssignmentSubmission.query.join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
                .filter(
                    Assignment.level_id == level.id,
                    AssignmentSubmission.submitted_at.isnot(None),
                    AssignmentSubmission.status != "Reviewed",
                )
                .count()
            )

            attendance_missing_days = [day for day in teaching_days if day not in attendance_dates]
            recording_missing_days = [day for day in teaching_days if day not in recording_dates]
            missing_homework_weeks = [week_start for week_start in teaching_week_starts if week_start not in assignment_week_starts]

            attendance_days_recorded += len(attendance_dates)
            recording_days_uploaded += len(recording_dates)
            homework_weeks_completed += len(assignment_week_starts)
            monthly_assignments_count += len(assignment_rows)
            monthly_pending_reviews += pending_reviews

            level_checks.append(
                {
                    "level_name": level.name,
                    "attendance_label": f"{len(attendance_dates)}/{expected_days_per_level}",
                    "recordings_label": f"{len(recording_dates)}/{expected_days_per_level}",
                    "homework_label": f"{len(assignment_week_starts)}/{expected_weeks_per_level}",
                    "attendance_missing_days_label": format_weekday_names(attendance_missing_days, active_lang),
                    "recording_missing_days_label": format_weekday_names(recording_missing_days, active_lang),
                    "missing_homework_weeks_label": (
                        "، ".join(week_start.isoformat() for week_start in missing_homework_weeks)
                        if missing_homework_weeks
                        else "-"
                    ),
                }
            )

        expected_attendance_days = expected_days_per_level * len(teacher_levels)
        expected_recording_days = expected_days_per_level * len(teacher_levels)
        expected_homework_weeks = expected_weeks_per_level * len(teacher_levels)

        attendance_pct = round((attendance_days_recorded / expected_attendance_days) * 100, 1) if expected_attendance_days else 0.0
        recordings_pct = round((recording_days_uploaded / expected_recording_days) * 100, 1) if expected_recording_days else 0.0
        homework_pct = round((homework_weeks_completed / expected_homework_weeks) * 100, 1) if expected_homework_weeks else 0.0
        overall_score = round(
            max(
                0.0,
                min(
                    100.0,
                    (attendance_pct * 0.35) + (recordings_pct * 0.35) + (homework_pct * 0.30) - min(monthly_pending_reviews * 5, 20),
                ),
            ),
            1,
        )

        if overall_score >= 90 and monthly_pending_reviews == 0:
            performance_label_key = "ideal_status"
        elif overall_score >= 75:
            performance_label_key = "excellent_status"
        elif overall_score >= 60:
            performance_label_key = "good_status"
        else:
            performance_label_key = "needs_followup"

        rows.append(
            {
                "teacher": teacher,
                "levels_count": len(teacher_levels),
                "month_start": month_start,
                "month_end": today,
                "attendance_pct": attendance_pct,
                "recordings_pct": recordings_pct,
                "homework_pct": homework_pct,
                "monthly_pending_reviews": monthly_pending_reviews,
                "monthly_assignments_count": monthly_assignments_count,
                "overall_score": overall_score,
                "performance_label": get_global_ui_copy(active_lang).get(
                    performance_label_key,
                    performance_label_key.replace("_", " ").title(),
                ),
                "performance_label_key": performance_label_key,
                "has_valid_email": validate_email((teacher.email or "").strip()),
                "level_checks": level_checks,
            }
        )

    rows.sort(
        key=lambda row: (
            -row["overall_score"],
            row["monthly_pending_reviews"],
            row["teacher"].full_name,
        )
    )

    ideal_teacher_row = rows[0] if rows else None
    summary = {
        "month_key": month_start.strftime("%Y-%m"),
        "teachers_count": len(rows),
        "ideal_teacher_name": ideal_teacher_row["teacher"].full_name if ideal_teacher_row else "-",
        "ideal_teacher_score": ideal_teacher_row["overall_score"] if ideal_teacher_row else 0,
        "ready_to_thank": sum(1 for row in rows if row["has_valid_email"]),
    }
    return {
        "rows": rows,
        "summary": summary,
        "ideal_teacher_row": ideal_teacher_row,
    }


def build_teachers_master_export_frames(target_date: date | None = None) -> dict:
    report_data = build_teacher_monthly_report_rows(target_date)
    academic_calendar_status = build_academic_calendar_status(target_date or date.today())
    summary_rows = []
    breakdown_rows = []

    for row in report_data["rows"]:
        teacher = row["teacher"]
        summary_rows.append(
            {
                "Teacher Name": teacher.full_name,
                "Email": (teacher.email or "").strip(),
                "Phone": (teacher.phone or "").strip(),
                "Levels Count": row["levels_count"],
                "Attendance %": row["attendance_pct"],
                "Recordings %": row["recordings_pct"],
                "Weekly Homework %": row["homework_pct"],
                "Monthly Assignments": row["monthly_assignments_count"],
                "Pending Reviews": row["monthly_pending_reviews"],
                "Overall Score": row["overall_score"],
                "Performance": row["performance_label"],
                "Email Ready": "Yes" if row["has_valid_email"] else "No",
                "Month": report_data["summary"]["month_key"],
                "Academic Week": academic_calendar_status["current_week"],
                "Current Academic Event": academic_calendar_status["current_event"].title if academic_calendar_status["current_event"] else "Teaching In Session",
                "Next Academic Event": academic_calendar_status["next_event"].title if academic_calendar_status["next_event"] else "",
            }
        )

        for check in row["level_checks"]:
            breakdown_rows.append(
                {
                    "Teacher Name": teacher.full_name,
                    "Level": check["level_name"],
                    "Attendance": check["attendance_label"],
                    "Recordings": check["recordings_label"],
                    "Homework": check["homework_label"],
                    "Missing Attendance Days": check["attendance_missing_days_label"],
                    "Missing Recording Days": check["recording_missing_days_label"],
                    "Missing Homework Weeks": check["missing_homework_weeks_label"],
                    "Month": report_data["summary"]["month_key"],
                    "Academic Week": academic_calendar_status["current_week"],
                }
            )

    return {
        "summary_df": pd.DataFrame(summary_rows),
        "breakdown_df": pd.DataFrame(breakdown_rows),
        "summary": report_data["summary"],
    }


def build_teachers_master_excel_file(target_date: date | None = None) -> tuple[BytesIO, str]:
    export_data = build_teachers_master_export_frames(target_date)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_data["summary_df"].to_excel(writer, index=False, sheet_name="Teachers Summary")
        export_data["breakdown_df"].to_excel(writer, index=False, sheet_name="Level Breakdown")

        for sheet_name, dataframe in {
            "Teachers Summary": export_data["summary_df"],
            "Level Breakdown": export_data["breakdown_df"],
        }.items():
            worksheet = writer.sheets[sheet_name]
            for column_index, column_name in enumerate(dataframe.columns, start=1):
                max_length = max(
                    [len(str(column_name))]
                    + [len(str(value)) for value in dataframe[column_name].fillna("").tolist()]
                )
                worksheet.column_dimensions[chr(64 + column_index)].width = min(max(max_length + 2, 14), 36)

    output.seek(0)
    file_name = f"hikmah_teachers_master_{export_data['summary']['month_key']}.xlsx"
    return output, file_name


def build_teachers_master_pdf(target_date: date | None = None) -> bytes:
    report_data = build_teacher_monthly_report_rows(target_date)
    academic_calendar_status = build_academic_calendar_status(target_date or date.today())
    summary = report_data["summary"]
    lines = [
        "Hikmah Academy Teachers Master Report",
        "",
        f"Month: {summary['month_key']}",
        f"Academic Week: {academic_calendar_status['current_week']}",
        f"Current Academic Event: {academic_calendar_status['current_event'].title if academic_calendar_status['current_event'] else 'Teaching In Session'}",
        f"Next Academic Event: {academic_calendar_status['next_event'].title if academic_calendar_status['next_event'] else '-'}",
        f"Teachers Total: {summary['teachers_count']}",
        f"Ready For Appreciation: {summary['ready_to_thank']}",
        f"Ideal Teacher: {summary['ideal_teacher_name']}",
        f"Ideal Teacher Score: {summary['ideal_teacher_score']}",
        "",
        "Teacher Summary",
    ]
    if report_data["rows"]:
        for row in report_data["rows"]:
            lines.append(
                f"- {row['teacher'].full_name}: score={row['overall_score']}, attendance={row['attendance_pct']}%, recordings={row['recordings_pct']}%, homework={row['homework_pct']}%, status={row['performance_label']}"
            )
            for check in row["level_checks"]:
                lines.append(
                    f"  {check['level_name']}: attendance {check['attendance_label']}, recordings {check['recordings_label']}, homework {check['homework_label']}"
                )
    else:
        lines.append("- No teacher monthly data available.")
    return build_simple_pdf(lines)


def build_level_full_export_frames(level: Level, target_date: date | None = None) -> dict:
    today = target_date or date.today()
    academic_calendar_status = build_academic_calendar_status(today)
    followup_rows = build_level_followup_register(level, today)
    summary = summarize_level_followup_rows(followup_rows)
    detail_rows = []

    for row in followup_rows:
        student = row["student"]
        report_data = build_student_report_data(student, "month")
        detail_rows.append(
            {
                "Student Code": student.student_code,
                "Student Name": student.full_name,
                "Overall Status": ui_text(row["status_key"]),
                "Attendance 14D %": row["attendance_rate_14d"] if row["attendance_rate_14d"] is not None else "",
                "Present Count": report_data["present_count"],
                "Absent Count": report_data["absent_count"],
                "Late Count": report_data["late_count"],
                "Excused Count": report_data["excused_count"],
                "Monthly Attendance %": report_data["attendance_percentage"],
                "Open Assignments": row["open_assignments"],
                "Homework Given": report_data["homework_total"],
                "Homework Submitted": report_data["homework_submitted"],
                "Homework Reviewed": report_data["homework_reviewed"],
                "Homework Missing": report_data["homework_missing"],
                "Homework Completion %": report_data["homework_completion_rate"],
                "Latest Exam": row["latest_exam_title"] or "",
                "Latest Exam %": row["latest_exam_percentage"] if row["latest_exam_percentage"] is not None else "",
                "Monthly Note Ready": "Yes" if row["has_monthly_note"] else "No",
                "Monthly Note": report_data["monthly_note_text"] or "",
                "Parent Email": (student.parent_email or "").strip(),
                "Parent WhatsApp": (student.parent_whatsapp or "").strip(),
                "Teacher": report_data["teacher_name"],
                "Level": report_data["level_name"],
                "Month": report_data["month_key"],
            }
        )

    summary_rows = [
        {"Metric": "Level", "Value": level.name},
        {"Metric": "Teacher", "Value": level.teacher.full_name if level.teacher else "-"},
        {"Metric": "Students", "Value": summary["total_count"]},
        {"Metric": "Excellent", "Value": summary["excellent_count"]},
        {"Metric": "Stable", "Value": summary["stable_count"]},
        {"Metric": "Needs Attention", "Value": summary["needs_attention_count"]},
        {"Metric": "Month", "Value": today.strftime("%Y-%m")},
        {"Metric": "Academic Week", "Value": academic_calendar_status["current_week"]},
        {"Metric": "Current Academic Event", "Value": academic_calendar_status["current_event"].title if academic_calendar_status["current_event"] else "Teaching In Session"},
    ]

    return {
        "summary_df": pd.DataFrame(summary_rows),
        "details_df": pd.DataFrame(detail_rows),
        "month_key": today.strftime("%Y-%m"),
    }


def build_level_full_excel_file(level: Level, target_date: date | None = None) -> tuple[BytesIO, str]:
    export_data = build_level_full_export_frames(level, target_date)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_data["summary_df"].to_excel(writer, index=False, sheet_name="Level Summary")
        export_data["details_df"].to_excel(writer, index=False, sheet_name="Students Follow-up")

        for sheet_name, dataframe in {
            "Level Summary": export_data["summary_df"],
            "Students Follow-up": export_data["details_df"],
        }.items():
            worksheet = writer.sheets[sheet_name]
            for column_index, column_name in enumerate(dataframe.columns, start=1):
                values = dataframe[column_name].fillna("").tolist() if not dataframe.empty else []
                max_length = max([len(str(column_name))] + [len(str(value)) for value in values])
                column_letter = ""
                temp = column_index
                while temp > 0:
                    temp, remainder = divmod(temp - 1, 26)
                    column_letter = chr(65 + remainder) + column_letter
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 40)

    output.seek(0)
    safe_level_name = re.sub(r"[^A-Za-z0-9_-]+", "_", level.name or "level").strip("_") or "level"
    file_name = f"hikmah_level_{safe_level_name}_{export_data['month_key']}.xlsx"
    return output, file_name


def build_level_full_pdf(level: Level, target_date: date | None = None) -> bytes:
    today = target_date or date.today()
    academic_calendar_status = build_academic_calendar_status(today)
    followup_rows = build_level_followup_register(level, today)
    summary = summarize_level_followup_rows(followup_rows)
    lines = [
        "Hikmah Academy Level Full Report",
        "",
        f"Level: {level.name}",
        f"Teacher: {level.teacher.full_name if level.teacher else '-'}",
        f"Month: {today.strftime('%Y-%m')}",
        f"Academic Week: {academic_calendar_status['current_week']}",
        f"Current Academic Event: {academic_calendar_status['current_event'].title if academic_calendar_status['current_event'] else 'Teaching In Session'}",
        f"Students Total: {summary['total_count']}",
        f"Excellent: {summary['excellent_count']}",
        f"Stable: {summary['stable_count']}",
        f"Needs Attention: {summary['needs_attention_count']}",
        "",
        "Student Follow-up",
    ]
    if followup_rows:
        for row in followup_rows:
            report_data = build_student_report_data(row["student"], "month")
            lines.append(
                f"- {row['student'].full_name} ({row['student'].student_code}): status={ui_text(row['status_key'])}, attendance14d={row['attendance_rate_14d'] if row['attendance_rate_14d'] is not None else '-'}%, open_assignments={row['open_assignments']}, latest_exam={row['latest_exam_title'] or '-'}, latest_exam_pct={row['latest_exam_percentage'] if row['latest_exam_percentage'] is not None else '-'}"
            )
            lines.append(
                f"  Monthly attendance={report_data['attendance_percentage']}%, homework completion={report_data['homework_completion_rate']}%, note={'Ready' if row['has_monthly_note'] else 'Missing'}"
            )
    else:
        lines.append("- No students found for this level.")
    return build_simple_pdf(lines)


def build_student_full_export_frames(student: Student, period: str = "month") -> dict:
    report_data = build_student_report_data(student, period)
    academic_calendar_status = build_academic_calendar_status()
    imported_attendance = report_data.get("imported_arabic_attendance")
    level = student.level

    attendance_rows = (
        Attendance.query.filter_by(student_id=student.id)
        .order_by(Attendance.attendance_date.desc(), Attendance.id.desc())
        .all()
    )
    homework_rows = []
    if level:
        assignments = (
            Assignment.query.filter_by(level_id=level.id)
            .order_by(Assignment.created_at.desc(), Assignment.id.desc())
            .all()
        )
        submissions = {
            submission.assignment_id: submission
            for submission in AssignmentSubmission.query.filter_by(student_id=student.id).all()
        }
        for assignment in assignments:
            submission = submissions.get(assignment.id)
            homework_rows.append(
                {
                    "Title": assignment.title,
                    "Created At": assignment.created_at.date().isoformat() if assignment.created_at else "",
                    "Due Date": assignment.due_date.isoformat() if assignment.due_date else "",
                    "Active": "Yes" if assignment.is_active else "No",
                    "Submission Status": submission.status if submission else "Not Submitted",
                    "Submitted At": submission.submitted_at.isoformat(sep=" ") if submission and submission.submitted_at else "",
                    "Teacher Feedback": submission.teacher_feedback if submission else "",
                    "Grade": submission.grade if submission and submission.grade is not None else "",
                }
            )

    exam_rows = (
        ExamResult.query.filter_by(student_id=student.id)
        .order_by(ExamResult.exam_date.desc(), ExamResult.id.desc())
        .all()
    )

    summary_rows = [
        {"Field": "Student Code", "Value": student.student_code},
        {"Field": "Student Name", "Value": student.full_name},
        {"Field": "Level", "Value": report_data["level_name"]},
        {"Field": "Teacher", "Value": report_data["teacher_name"]},
        {"Field": "Period", "Value": period},
        {"Field": "Date Range", "Value": f"{report_data['start_date']} to {report_data['end_date']}"},
        {"Field": "Academic Week", "Value": academic_calendar_status["current_week"]},
        {"Field": "Current Academic Event", "Value": academic_calendar_status["current_event"].title if academic_calendar_status["current_event"] else "Teaching In Session"},
        {"Field": "Present Count", "Value": report_data["present_count"]},
        {"Field": "Absent Count", "Value": report_data["absent_count"]},
        {"Field": "Late Count", "Value": report_data["late_count"]},
        {"Field": "Excused Count", "Value": report_data["excused_count"]},
        {"Field": "Attendance Percentage", "Value": report_data["attendance_percentage"]},
        {"Field": "Imported Arabic Attendance Available", "Value": "Yes" if imported_attendance else "No"},
        {"Field": "Homework Given", "Value": report_data["homework_total"]},
        {"Field": "Homework Submitted", "Value": report_data["homework_submitted"]},
        {"Field": "Homework Reviewed", "Value": report_data["homework_reviewed"]},
        {"Field": "Homework Missing", "Value": report_data["homework_missing"]},
        {"Field": "Homework Completion %", "Value": report_data["homework_completion_rate"]},
        {"Field": "Monthly Note", "Value": report_data["monthly_note_text"] or ""},
        {"Field": "Parent Email", "Value": (student.parent_email or "").strip()},
        {"Field": "Parent WhatsApp", "Value": (student.parent_whatsapp or "").strip()},
    ]
    if imported_attendance:
        summary_rows.extend(
            [
                {"Field": "Imported Present Count", "Value": imported_attendance["present_count"]},
                {"Field": "Imported Absent Count", "Value": imported_attendance["absent_count"]},
                {"Field": "Imported Late Count", "Value": imported_attendance["late_count"]},
                {"Field": "Imported Total Sessions", "Value": imported_attendance["total_count"]},
                {"Field": "Imported Attendance Percentage", "Value": imported_attendance["attendance_percentage"]},
                {"Field": "Imported Source Sheet", "Value": imported_attendance["sheet_name"]},
                {"Field": "Imported Matched Workbook Name", "Value": imported_attendance["student_name"]},
            ]
        )

    attendance_df = pd.DataFrame(
        [
            {
                "Date": record.attendance_date.isoformat() if record.attendance_date else "",
                "Status": record.status,
                "Level": report_data["level_name"],
            }
            for record in attendance_rows
        ]
    )
    homework_df = pd.DataFrame(homework_rows)
    exams_df = pd.DataFrame(
        [
            {
                "Exam Title": result.exam_title,
                "Subject": result.subject_name,
                "Score": result.score_value,
                "Max Score": result.max_score if result.max_score is not None else "",
                "Exam Date": result.exam_date.isoformat() if result.exam_date else "",
                "Notes": result.notes or "",
            }
            for result in exam_rows
        ]
    )
    imported_attendance_df = pd.DataFrame(
        [
            {
                "Source Sheet": imported_attendance["sheet_name"],
                "Matched Workbook Name": imported_attendance["student_name"],
                "Present": imported_attendance["present_count"],
                "Absent": imported_attendance["absent_count"],
                "Late": imported_attendance["late_count"],
                "Total Sessions": imported_attendance["total_count"],
                "Attendance %": imported_attendance["attendance_percentage"],
            }
        ]
        if imported_attendance
        else []
    )

    return {
        "summary_df": pd.DataFrame(summary_rows),
        "attendance_df": attendance_df,
        "imported_attendance_df": imported_attendance_df,
        "homework_df": homework_df,
        "exams_df": exams_df,
        "period": period,
    }


def build_student_full_excel_file(student: Student, period: str = "month") -> tuple[BytesIO, str]:
    export_data = build_student_full_export_frames(student, period)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets = {
            "Student Summary": export_data["summary_df"],
            "Attendance": export_data["attendance_df"],
            "Arabic Snapshot": export_data["imported_attendance_df"],
            "Homework": export_data["homework_df"],
            "Exams": export_data["exams_df"],
        }
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            for column_index, column_name in enumerate(dataframe.columns, start=1):
                values = dataframe[column_name].fillna("").tolist() if not dataframe.empty else []
                max_length = max([len(str(column_name))] + [len(str(value)) for value in values])
                column_letter = ""
                temp = column_index
                while temp > 0:
                    temp, remainder = divmod(temp - 1, 26)
                    column_letter = chr(65 + remainder) + column_letter
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 40)

    output.seek(0)
    safe_student = re.sub(r"[^A-Za-z0-9_-]+", "_", student.full_name or "student").strip("_") or "student"
    file_name = f"hikmah_student_{safe_student}_{period}.xlsx"
    return output, file_name


def build_student_full_pdf(student: Student, period: str = "month") -> bytes:
    report_data = build_student_report_data(student, period)
    academic_calendar_status = build_academic_calendar_status()
    imported_attendance = report_data.get("imported_arabic_attendance")
    lines = [
        "Hikmah Academy Student Full Report",
        "",
        f"Student Code: {student.student_code}",
        f"Student Name: {student.full_name}",
        f"Level: {report_data['level_name']}",
        f"Teacher: {report_data['teacher_name']}",
        f"Period: {period}",
        f"Date Range: {report_data['start_date']} to {report_data['end_date']}",
        f"Academic Week: {academic_calendar_status['current_week']}",
        f"Current Academic Event: {academic_calendar_status['current_event'].title if academic_calendar_status['current_event'] else 'Teaching In Session'}",
        "",
        "Attendance Summary",
        f"- Present: {report_data['present_count']}",
        f"- Absent: {report_data['absent_count']}",
        f"- Late: {report_data['late_count']}",
        f"- Excused: {report_data['excused_count']}",
        f"- Attendance Percentage: {report_data['attendance_percentage']}%",
    ]
    if imported_attendance:
        lines.extend(
            [
                "",
                "Imported Arabic Attendance Snapshot",
                f"- Source Sheet: {imported_attendance['sheet_name']}",
                f"- Matched Workbook Name: {imported_attendance['student_name']}",
                f"- Present: {imported_attendance['present_count']}",
                f"- Absent: {imported_attendance['absent_count']}",
                f"- Late: {imported_attendance['late_count']}",
                f"- Total Sessions: {imported_attendance['total_count']}",
                f"- Attendance Percentage: {imported_attendance['attendance_percentage']}%",
            ]
        )
    lines.extend(
        [
        "",
        "Homework Summary",
        f"- Homework Given: {report_data['homework_total']}",
        f"- Homework Submitted: {report_data['homework_submitted']}",
        f"- Homework Reviewed: {report_data['homework_reviewed']}",
        f"- Homework Waiting Review: {report_data['homework_pending_review']}",
        f"- Homework Missing: {report_data['homework_missing']}",
        f"- Homework Completion: {report_data['homework_completion_rate']}%",
        "",
        "Monthly Teacher Note",
        report_data["monthly_note_text"] or "-",
        "",
        "Latest Exams",
        ]
    )
    exam_rows = (
        ExamResult.query.filter_by(student_id=student.id)
        .order_by(ExamResult.exam_date.desc(), ExamResult.id.desc())
        .limit(10)
        .all()
    )
    if exam_rows:
        for result in exam_rows:
            lines.append(
                f"- {result.exam_title} / {result.subject_name}: {result.score_value if result.score_value is not None else '-'} / {result.max_score if result.max_score is not None else '-'} on {result.exam_date.isoformat() if result.exam_date else '-'}"
            )
    else:
        lines.append("- No exam results available.")
    return build_simple_pdf(lines)


def generate_weekly_followup_announcements(target_date: date) -> tuple[int, int, list[str]]:
    review_rows = collect_weekly_teacher_reviews(target_date)
    follow_up_count = 0
    praise_count = 0
    messages = []

    for row in review_rows:
        level = row["level"]
        teacher = row["teacher"]
        is_follow_up = row["status"] == "follow_up"
        title_prefix = "Weekly Follow-up" if is_follow_up else "Weekly Appreciation"
        title = f"{title_prefix} - {level.name} - {target_date.isoformat()}"
        existing = Announcement.query.filter_by(
            title=title,
            audience="teachers",
            level_id=level.id,
        ).first()

        if is_follow_up:
            body = (
                f"Administrative follow-up for {teacher.full_name} in {level.name}.\n"
                f"Teaching week reviewed: {row['week_start']} to {row['week_end']} (Monday to Thursday).\n"
                "Please complete the following before the next cycle:\n"
                + "\n".join(f"- {issue}" for issue in row["issues"])
            )
        else:
            body = (
                f"Thank you {teacher.full_name}.\n"
                f"Your level {level.name} completed the required weekly tasks for the teaching week {row['week_start']} to {row['week_end']}.\n"
                "- Zoom link was available\n"
                "- Attendance was recorded\n"
                "- Homework was assigned\n"
                "May Allah reward your effort."
            )

        if existing:
            existing.body = body
            existing.category = "follow_up" if is_follow_up else "general"
            existing.is_active = True
            existing.is_pinned = is_follow_up
            existing.starts_on = target_date
            existing.expires_on = target_date + timedelta(days=3)
            messages.append(f"Updated {'follow-up' if is_follow_up else 'appreciation'} for {level.name}.")
        else:
            db.session.add(
                Announcement(
                    title=title,
                    body=body,
                    audience="teachers",
                    category="follow_up" if is_follow_up else "general",
                    level_id=level.id,
                    is_pinned=is_follow_up,
                    is_active=True,
                    starts_on=target_date,
                    expires_on=target_date + timedelta(days=3),
                )
            )
            messages.append(f"Created {'follow-up' if is_follow_up else 'appreciation'} for {level.name}.")

        if is_follow_up:
            follow_up_count += 1
        else:
            praise_count += 1

    db.session.commit()
    return follow_up_count, praise_count, messages


def run_scheduled_weekly_followup(now: datetime | None = None, force: bool = False) -> dict:
    current_dt = now or datetime.now()
    target_date = current_dt.date()
    calendar_settings = get_calendar_settings()

    if not force:
        can_run, status_message = should_run_weekly_followup(target_date)
        if not can_run:
            return {"status": "skipped", "message": status_message, "created_count": 0}

        scheduled_hour, scheduled_minute = parse_time_hhmm(calendar_settings["weekly_followup_time"])
        scheduled_time_value = scheduled_hour * 60 + scheduled_minute
        current_time_value = current_dt.hour * 60 + current_dt.minute
        if current_time_value < scheduled_time_value:
            return {
                "status": "skipped",
                "message": f"Skipped because scheduled follow-up time is {calendar_settings['weekly_followup_time']}.",
                "created_count": 0,
            }

        if calendar_settings["last_weekly_followup_run"] == target_date.isoformat():
            return {
                "status": "skipped",
                "message": f"Skipped because weekly follow-up already ran on {target_date.isoformat()}.",
                "created_count": 0,
            }

    follow_up_count, praise_count, messages = generate_weekly_followup_announcements(target_date)
    report_data = build_supervisor_weekly_report_data(target_date)
    archived_report_file = archive_supervisor_weekly_report(report_data)
    set_system_setting("last_weekly_followup_run", target_date.isoformat())
    db.session.commit()

    message = (
        f"Weekly review generated. Follow-up: {follow_up_count}, appreciation: {praise_count}."
        if (follow_up_count or praise_count)
        else "No weekly review announcements were needed."
    )
    if messages:
        message += " " + " | ".join(messages[:4])
    message += f" Archived report: {archived_report_file}."
    return {
        "status": "success",
        "message": message,
        "created_count": follow_up_count + praise_count,
        "follow_up_count": follow_up_count,
        "praise_count": praise_count,
        "archived_report_file": archived_report_file,
    }


def get_active_announcements(audience: str, level_ids: int | list[int] | None = None) -> list[Announcement]:
    today = date.today()
    allowed_audiences = ["all", audience]
    query = Announcement.query.filter(
        Announcement.is_active.is_(True),
        Announcement.audience.in_(allowed_audiences),
    ).filter(
        db.or_(Announcement.starts_on.is_(None), Announcement.starts_on <= today)
    ).filter(
        db.or_(Announcement.expires_on.is_(None), Announcement.expires_on >= today)
    )

    if isinstance(level_ids, int):
        level_ids = [level_ids]

    if level_ids:
        query = query.filter(
            db.or_(Announcement.level_id.is_(None), Announcement.level_id.in_(level_ids))
        )
    else:
        query = query.filter(Announcement.level_id.is_(None))

    return query.order_by(
        Announcement.is_pinned.desc(),
        Announcement.created_at.desc(),
        Announcement.id.desc(),
    ).all()


def build_announcement_view_rows(items: list[Announcement]) -> list[dict]:
    category_labels = {
        "general": "General",
        "reminder": "Reminder",
        "alert": "Alert",
        "follow_up": "Follow Up",
    }
    audience_labels = {
        "all": "All",
        "teachers": "Teachers",
        "students": "Students",
    }
    return [
        {
            "announcement": item,
            "category_label": category_labels.get(item.category, item.category.replace("_", " ").title()),
            "audience_label": audience_labels.get(item.audience, item.audience.title()),
            "tone": "danger" if item.category == "alert" else "warning" if item.category == "follow_up" else "info",
        }
        for item in items
    ]


def build_supervisor_dashboard_data() -> dict:
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    fourteen_days_ago = today - timedelta(days=13)
    academic_calendar_status = build_academic_calendar_status(today)
    arabic_attendance_import_summary = build_imported_arabic_attendance_summary()

    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    teachers = Teacher.query.order_by(Teacher.full_name.asc()).all()
    students = Student.query.order_by(Student.full_name.asc()).all()
    cleanup_data = build_cleanup_center_data(levels, teachers, students)
    monthly_review_data = build_monthly_report_review_rows()
    teacher_monthly_report_data = build_teacher_monthly_report_rows(today)
    weekly_teacher_reviews = collect_weekly_teacher_reviews(today)
    curriculum_plan_rows = sorted(
        build_syllabus_plan_summary(levels),
        key=lambda row: (-row["delayed_count"], row["level"].order_index or 0, row["level"].name.lower()),
    )
    curriculum_plan_summary = {
        "current_week": academic_calendar_status["current_week"],
        "levels_on_track": sum(1 for row in curriculum_plan_rows if row["delayed_count"] == 0),
        "levels_delayed": sum(1 for row in curriculum_plan_rows if row["delayed_count"] > 0),
        "completed_lessons": sum(row["completed_count"] for row in curriculum_plan_rows),
        "expected_count": sum(row["expected_count"] for row in curriculum_plan_rows),
        "delayed_count": sum(row["delayed_count"] for row in curriculum_plan_rows),
    }
    recent_action_logs = (
        ActionLog.query.order_by(ActionLog.created_at.desc(), ActionLog.id.desc())
        .limit(10)
        .all()
    )

    recent_action_rows = [
        {
            "actor_name": item.actor_name,
            "actor_role": item.actor_role,
            "action_key": item.action_type,
            "entity_label": item.entity_label or item.entity_type.replace("_", " ").title(),
            "level_name": item.level.name if item.level else "",
            "details": item.details or "",
            "created_at": item.created_at,
            "target_url": get_action_log_target(item)[0],
            "target_label_key": get_action_log_target(item)[1],
        }
        for item in recent_action_logs
    ]

    student_count_by_level = {
        level_id: count
        for level_id, count in (
            db.session.query(Student.level_id, db.func.count(Student.id))
            .filter(Student.level_id.isnot(None))
            .group_by(Student.level_id)
            .all()
        )
    }

    attendance_rows_today = (
        db.session.query(
            Attendance.level_id,
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.sum(db.case((Attendance.status == "Absent", 1), else_=0)).label("absent_count"),
            db.func.sum(db.case((Attendance.status == "Late", 1), else_=0)).label("late_count"),
            db.func.sum(db.case((Attendance.status == "Excused", 1), else_=0)).label("excused_count"),
            db.func.count(Attendance.id).label("marked_count"),
        )
        .filter(Attendance.attendance_date == today, Attendance.level_id.isnot(None))
        .group_by(Attendance.level_id)
        .all()
    )
    attendance_today_by_level = {row.level_id: row for row in attendance_rows_today}

    attendance_rows_week = (
        db.session.query(
            Attendance.student_id,
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.count(Attendance.id).label("total_count"),
        )
        .filter(
            Attendance.attendance_date >= fourteen_days_ago,
            Attendance.attendance_date <= today,
        )
        .group_by(Attendance.student_id)
        .all()
    )
    attendance_week_by_student = {row.student_id: row for row in attendance_rows_week}

    active_assignments = Assignment.query.filter_by(is_active=True).all()
    active_assignment_count_by_level = {}
    for assignment in active_assignments:
        active_assignment_count_by_level[assignment.level_id] = active_assignment_count_by_level.get(assignment.level_id, 0) + 1

    submission_rows = (
        AssignmentSubmission.query.join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .add_columns(Assignment.level_id)
        .all()
    )
    pending_review_count_by_level = {}
    submissions_by_assignment_student = {}
    for submission, level_id in submission_rows:
        submissions_by_assignment_student[(submission.assignment_id, submission.student_id)] = submission
        if submission.submitted_at and submission.status != "Reviewed":
            pending_review_count_by_level[level_id] = pending_review_count_by_level.get(level_id, 0) + 1

    upcoming_exam_rows = (
        UpcomingExam.query.filter(UpcomingExam.exam_date >= today)
        .order_by(UpcomingExam.exam_date.asc(), UpcomingExam.id.asc())
        .all()
    )
    next_exam_by_level = {}
    upcoming_exams_this_week = 0
    for exam in upcoming_exam_rows:
        if exam.exam_date <= today + timedelta(days=7):
            upcoming_exams_this_week += 1
        if exam.level_id not in next_exam_by_level:
            next_exam_by_level[exam.level_id] = exam

    exam_publications = ExamPublication.query.all()
    unpublished_exam_count = sum(1 for publication in exam_publications if not publication.is_published)
    unresolved_name_matches = ExamImportIssue.query.count()

    exam_results = ExamResult.query.order_by(ExamResult.student_id.asc(), ExamResult.exam_date.desc(), ExamResult.id.desc()).all()
    exam_results_by_student = {}
    for result in exam_results:
        exam_results_by_student.setdefault(result.student_id, []).append(result)

    student_rows_by_level = {}
    for student in students:
        if student.level_id:
            student_rows_by_level.setdefault(student.level_id, []).append(student)

    level_rows = []
    missing_attendance_levels = []
    partial_attendance_levels = []
    total_students = len(students)
    total_marked_today = 0
    total_present_today = 0
    total_pending_reviews = 0
    total_active_assignments = len(active_assignments)

    for level in levels:
        students_count = int(student_count_by_level.get(level.id, 0))
        attendance_row = attendance_today_by_level.get(level.id)
        marked_count = int(attendance_row.marked_count or 0) if attendance_row else 0
        present_count = int(attendance_row.present_count or 0) if attendance_row else 0
        absent_count = int(attendance_row.absent_count or 0) if attendance_row else 0
        late_count = int(attendance_row.late_count or 0) if attendance_row else 0
        excused_count = int(attendance_row.excused_count or 0) if attendance_row else 0

        total_marked_today += marked_count
        total_present_today += present_count

        if students_count == 0:
            attendance_state = "empty"
        elif marked_count == 0:
            attendance_state = "missing"
            missing_attendance_levels.append(level)
        elif marked_count < students_count:
            attendance_state = "partial"
            partial_attendance_levels.append(level)
        else:
            attendance_state = "complete"

        attendance_rate = round((present_count / marked_count) * 100, 1) if marked_count else None
        pending_reviews = int(pending_review_count_by_level.get(level.id, 0))
        total_pending_reviews += pending_reviews
        next_exam = next_exam_by_level.get(level.id)

        level_rows.append(
            {
                "level": level,
                "students_count": students_count,
                "marked_count": marked_count,
                "present_count": present_count,
                "absent_count": absent_count,
                "late_count": late_count,
                "excused_count": excused_count,
                "attendance_rate": attendance_rate,
                "attendance_state": attendance_state,
                "pending_reviews": pending_reviews,
                "active_assignments": int(active_assignment_count_by_level.get(level.id, 0)),
                "next_exam": next_exam,
            }
        )

    teacher_rows = []
    for teacher in teachers:
        teacher_levels = [row for row in level_rows if row["level"].teacher_id == teacher.id]
        if not teacher_levels:
            continue
        teacher_weekly_reviews = [row for row in weekly_teacher_reviews if row["teacher"].id == teacher.id]
        teacher_follow_up_levels = [row for row in teacher_weekly_reviews if row["status"] == "follow_up"]
        levels_with_complete_recordings = sum(
            1 for row in teacher_weekly_reviews if row["recording_days_uploaded"] >= row["expected_teaching_days"]
        )
        levels_with_weekly_homework = sum(1 for row in teacher_weekly_reviews if row["weekly_assignments_count"] > 0)
        total_weekly_assignments = sum(row["weekly_assignments_count"] for row in teacher_weekly_reviews)
        attendance_days_recorded = sum(row["attendance_days_recorded"] for row in teacher_weekly_reviews)
        expected_attendance_days = sum(row["expected_teaching_days"] for row in teacher_weekly_reviews)
        recording_days_uploaded = sum(row["recording_days_uploaded"] for row in teacher_weekly_reviews)
        expected_recording_days = sum(row["expected_teaching_days"] for row in teacher_weekly_reviews)
        teacher_rows.append(
            {
                "teacher": teacher,
                "levels_count": len(teacher_levels),
                "students_count": sum(row["students_count"] for row in teacher_levels),
                "missing_attendance_count": sum(1 for row in teacher_levels if row["attendance_state"] == "missing"),
                "pending_reviews": sum(row["pending_reviews"] for row in teacher_levels),
                "upcoming_exams": sum(1 for row in teacher_levels if row["next_exam"] is not None),
                "weekly_status": "follow_up" if teacher_follow_up_levels else "praise",
                "weekly_follow_up_count": len(teacher_follow_up_levels),
                "levels_with_complete_recordings": levels_with_complete_recordings,
                "levels_missing_recordings": max(len(teacher_weekly_reviews) - levels_with_complete_recordings, 0),
                "levels_with_weekly_homework": levels_with_weekly_homework,
                "levels_without_weekly_homework": max(len(teacher_weekly_reviews) - levels_with_weekly_homework, 0),
                "weekly_assignments_count": total_weekly_assignments,
                "attendance_days_recorded": attendance_days_recorded,
                "expected_attendance_days": expected_attendance_days,
                "attendance_completion_label": f"{attendance_days_recorded}/{expected_attendance_days}" if expected_attendance_days else "0/0",
                "recordings_completion_label": f"{recording_days_uploaded}/{expected_recording_days}" if expected_recording_days else "0/0",
                "weekly_level_checks": [
                    {
                        "level_name": row["level"].name,
                        "attendance_label": f"{row['attendance_days_recorded']}/{row['expected_teaching_days']}",
                        "recordings_label": f"{row['recording_days_uploaded']}/{row['expected_teaching_days']}",
                        "attendance_missing_days_label": row["attendance_missing_days_label"],
                        "recording_missing_days_label": row["recording_missing_days_label"],
                        "weekly_assignments_count": row["weekly_assignments_count"],
                        "status": row["status"],
                    }
                    for row in teacher_weekly_reviews
                ],
                "weekly_review_notes": ", ".join(
                    f"{row['level'].name}: {'; '.join(row['issues'])}"
                    for row in teacher_follow_up_levels[:2]
                ) or "All assigned levels are currently on track.",
            }
        )

    student_attention_rows = []
    for student in students:
        if not student.level_id:
            continue

        attendance_window = attendance_week_by_student.get(student.id)
        attendance_rate_14d = (
            round((int(attendance_window.present_count or 0) / int(attendance_window.total_count or 0)) * 100, 1)
            if attendance_window and int(attendance_window.total_count or 0) > 0
            else None
        )
        attendance_meta = resolve_student_attendance_display(student, attendance_rate_14d)
        attendance_rate_display = attendance_meta["attendance_rate_display"]
        attendance_source_key = attendance_meta["attendance_source_key"]

        active_level_assignments = [assignment for assignment in active_assignments if assignment.level_id == student.level_id]
        open_assignments = 0
        for assignment in active_level_assignments:
            submission = submissions_by_assignment_student.get((assignment.id, student.id))
            if not submission or submission.status != "Reviewed":
                open_assignments += 1

        latest_exam_percentage = None
        latest_exam_title = ""
        student_exam_results = exam_results_by_student.get(student.id, [])
        if student_exam_results:
            exam_summaries = build_exam_summary_groups(student_exam_results)
            if exam_summaries:
                latest_exam_percentage = exam_summaries[0]["computed_percentage"]
                latest_exam_title = exam_summaries[0]["exam_title"]

        risk_score = 0
        reasons = []
        if attendance_rate_display is not None and attendance_rate_display < 75:
            risk_score += 2
            reasons.append(ui_text("reason_low_attendance"))
        if open_assignments >= 2:
            risk_score += 2
            reasons.append(ui_text("reason_pending_assignments"))
        elif open_assignments == 1:
            risk_score += 1
            reasons.append(ui_text("reason_one_open_assignment"))
        if latest_exam_percentage is not None and latest_exam_percentage < 65:
            risk_score += 2
            reasons.append(ui_text("reason_low_exam_result"))

        if risk_score == 0:
            continue

        if risk_score >= 4:
            risk_label = "High"
        elif risk_score >= 2:
            risk_label = "Medium"
        else:
            risk_label = "Low"

        student_attention_rows.append(
            {
                "student": student,
                "level_name": student.level.name if student.level else normalize_level_display_name(student.level_name),
                "attendance_rate_14d": attendance_rate_14d,
                "attendance_rate_display": attendance_rate_display,
                "attendance_source_key": attendance_source_key,
                "open_assignments": open_assignments,
                "latest_exam_percentage": latest_exam_percentage,
                "latest_exam_title": latest_exam_title,
                "risk_score": risk_score,
                "risk_label": risk_label,
                "reasons": ", ".join(reasons),
            }
        )

    student_attention_rows.sort(key=lambda row: (-row["risk_score"], row["student"].full_name))

    dashboard_alerts = []
    if missing_attendance_levels:
        dashboard_alerts.append(
            {
                "tone": "danger",
                "title": "Attendance Missing Today",
                "body": f"{len(missing_attendance_levels)} level(s) still have no attendance record for {today.isoformat()}.",
            }
        )
    if partial_attendance_levels:
        dashboard_alerts.append(
            {
                "tone": "warning",
                "title": "Attendance Partially Filled",
                "body": f"{len(partial_attendance_levels)} level(s) have incomplete attendance entries today.",
            }
        )
    if total_pending_reviews:
        dashboard_alerts.append(
            {
                "tone": "warning",
                "title": "Assignments Need Review",
                "body": f"{total_pending_reviews} submitted assignment(s) still need teacher review.",
            }
        )
    if unresolved_name_matches:
        dashboard_alerts.append(
            {
                "tone": "warning",
                "title": "Exam Imports Need Resolution",
                "body": f"{unresolved_name_matches} imported exam name match issue(s) are waiting for admin action.",
            }
        )
    if unpublished_exam_count:
        dashboard_alerts.append(
            {
                "tone": "info",
                "title": "Exam Results Not Yet Published",
                "body": f"{unpublished_exam_count} exam publication rule(s) are still hidden from students.",
            }
        )
    if arabic_attendance_import_summary["unmatched_rows_count"] > 0:
        dashboard_alerts.append(
            {
                "tone": "danger",
                "title": "Arabic Attendance Names Need Review",
                "body": (
                    f"{arabic_attendance_import_summary['unmatched_rows_count']} Arabic attendance row(s) "
                    "did not match student records."
                ),
            }
        )

    attendance_capture_rate = round((total_marked_today / total_students) * 100, 1) if total_students else 0.0
    today_presence_rate = round((total_present_today / total_marked_today) * 100, 1) if total_marked_today else 0.0

    kpis = [
        {
            "label": "Attendance Logged Today",
            "value": format_percent_label(attendance_capture_rate),
            "caption": f"{total_marked_today} of {total_students} student attendance records captured today",
        },
        {
            "label": "Present Rate Today",
            "value": format_percent_label(today_presence_rate),
            "caption": f"{total_present_today} students marked present out of {total_marked_today} marked records",
        },
        {
            "label": "Open Review Queue",
            "value": str(total_pending_reviews),
            "caption": "Submitted assignments still waiting for teacher feedback",
        },
        {
            "label": "Upcoming Exams (7 Days)",
            "value": str(upcoming_exams_this_week),
            "caption": "Scheduled exams happening in the next seven days",
        },
    ]

    spotlight = {
        "levels_missing_attendance": len(missing_attendance_levels),
        "levels_partial_attendance": len(partial_attendance_levels),
        "students_needing_attention": len(student_attention_rows),
        "active_assignments": total_active_assignments,
        "week_start": week_start,
        "today": today,
        "teachers_on_track": sum(1 for row in teacher_rows if row["weekly_status"] == "praise"),
        "teachers_need_followup": sum(1 for row in teacher_rows if row["weekly_status"] == "follow_up"),
        "cleanup_total_items": cleanup_data["cleanup_summary"]["total_items"],
        "cleanup_critical_items": cleanup_data["cleanup_summary"]["critical_items"],
        "monthly_ready_reports": monthly_review_data["summary"]["send_ready"],
        "monthly_missing_notes": monthly_review_data["summary"]["missing_notes"],
        "current_academic_week": academic_calendar_status["current_week"],
    }

    todays_gap_rows = []
    for row in level_rows:
        level = row["level"]
        gaps = []
        if row["attendance_state"] == "missing":
            gaps.append("Attendance missing today")
        elif row["attendance_state"] == "partial":
            gaps.append("Attendance partially filled today")
        if not (level.zoom_link or "").strip():
            gaps.append("Zoom link missing")
        if row["active_assignments"] == 0:
            gaps.append("No active homework")
        if row["pending_reviews"] > 0:
            gaps.append(f"{row['pending_reviews']} assignment review(s) pending")
        if gaps:
            todays_gap_rows.append(
                {
                    "level": level,
                    "teacher_name": level.teacher.full_name if level.teacher else "-",
                    "gaps": gaps,
                }
            )

    level_followup_snapshot_rows = []
    for level in levels:
        if not Student.query.filter_by(level_id=level.id).count():
            continue
        followup_rows = build_level_followup_register(level, today)
        flagged_count = 0
        missing_monthly_note_count = 0
        open_assignments_count = 0
        low_attendance_count = 0
        imported_attendance_count = 0
        low_exam_count = 0
        for followup_row in followup_rows:
            reasons = []
            if (followup_row["attendance_rate_display"] or 100) < 75:
                reasons.append("attendance")
                low_attendance_count += 1
            if followup_row["imported_attendance"]:
                imported_attendance_count += 1
            if followup_row["open_assignments"] > 0:
                reasons.append("assignments")
                open_assignments_count += 1
            if (
                followup_row["latest_exam_percentage"] is not None
                and followup_row["latest_exam_percentage"] < 65
            ):
                reasons.append("exam")
                low_exam_count += 1
            if not followup_row["has_monthly_note"]:
                reasons.append("monthly_note")
                missing_monthly_note_count += 1
            if followup_row["status_key"] == "needs_attention_status":
                flagged_count += 1

        if flagged_count:
            level_followup_snapshot_rows.append(
                {
                    "level": level,
                    "teacher_name": level.teacher.full_name if level.teacher else "-",
                    "flagged_count": flagged_count,
                    "missing_monthly_note_count": missing_monthly_note_count,
                    "open_assignments_count": open_assignments_count,
                    "low_attendance_count": low_attendance_count,
                    "imported_attendance_count": imported_attendance_count,
                    "low_exam_count": low_exam_count,
                }
            )

    level_followup_snapshot_rows.sort(
        key=lambda row: (
            -row["flagged_count"],
            -row["missing_monthly_note_count"],
            -row["open_assignments_count"],
            row["level"].name.lower(),
        )
    )

    return {
        "kpis": kpis,
        "dashboard_alerts": dashboard_alerts,
        "level_rows": level_rows,
        "teacher_rows": teacher_rows,
        "student_attention_rows": student_attention_rows[:10],
        "spotlight": spotlight,
        "teacher_weekly_status_rows": teacher_rows[:],
        "todays_gap_rows": todays_gap_rows,
        "level_followup_snapshot_rows": level_followup_snapshot_rows[:6],
        "recent_action_rows": recent_action_rows,
        "curriculum_plan_rows": curriculum_plan_rows,
        "curriculum_plan_summary": curriculum_plan_summary,
        "cleanup_summary": cleanup_data["cleanup_summary"],
        "cleanup_snapshot_students": cleanup_data["cleanup_students"][:3],
        "cleanup_snapshot_teachers": cleanup_data["cleanup_teachers"][:3],
        "cleanup_snapshot_levels": cleanup_data["cleanup_levels"][:3],
        "monthly_review_summary": monthly_review_data["summary"],
        "monthly_review_rows": monthly_review_data["rows"][:6],
        "teacher_monthly_report_summary": teacher_monthly_report_data["summary"],
        "teacher_monthly_report_rows": teacher_monthly_report_data["rows"],
        "ideal_teacher_row": teacher_monthly_report_data["ideal_teacher_row"],
        "academic_calendar_status": academic_calendar_status,
        "arabic_attendance_import_summary": arabic_attendance_import_summary,
    }


def build_supervisor_weekly_report_data(target_date: date | None = None) -> dict:
    report_date = target_date or date.today()
    dashboard_data = build_supervisor_dashboard_data()
    week_start, week_end = get_teaching_week_window(report_date)

    teacher_status_rows = dashboard_data["teacher_weekly_status_rows"]
    follow_up_teachers = [row for row in teacher_status_rows if row["weekly_status"] == "follow_up"]
    on_track_teachers = [row for row in teacher_status_rows if row["weekly_status"] == "praise"]
    level_attention_rows = [
        row for row in dashboard_data["level_rows"]
        if row["attendance_state"] in {"missing", "partial"} or row["pending_reviews"] > 0
    ]

    summary = {
        "report_date": report_date,
        "week_start": week_start,
        "week_end": week_end,
        "teachers_total": len(teacher_status_rows),
        "teachers_on_track": len(on_track_teachers),
        "teachers_need_followup": len(follow_up_teachers),
        "levels_need_attention": len(level_attention_rows),
        "students_need_attention": len(dashboard_data["student_attention_rows"]),
    }

    return {
        "summary": summary,
        "teacher_status_rows": teacher_status_rows,
        "follow_up_teachers": follow_up_teachers,
        "on_track_teachers": on_track_teachers,
        "level_attention_rows": level_attention_rows,
        "student_attention_rows": dashboard_data["student_attention_rows"],
    }


def build_supervisor_weekly_report_pdf(report_data: dict) -> bytes:
    summary = report_data["summary"]
    lines = [
        "Hikmah Academy Weekly Supervisor Report",
        "",
        f"Report Date: {summary['report_date']}",
        f"Teaching Week: {summary['week_start']} to {summary['week_end']} (Monday to Thursday)",
        "",
        "Executive Summary",
        f"- Teachers total: {summary['teachers_total']}",
        f"- Teachers on track: {summary['teachers_on_track']}",
        f"- Teachers needing follow-up: {summary['teachers_need_followup']}",
        f"- Levels needing attention: {summary['levels_need_attention']}",
        f"- Students needing attention: {summary['students_need_attention']}",
        "",
        "Teacher Weekly Status",
    ]

    if report_data["teacher_status_rows"]:
        for row in report_data["teacher_status_rows"]:
            lines.append(
                f"- {row['teacher'].full_name}: {'Needs Follow-up' if row['weekly_status'] == 'follow_up' else 'On Track'}"
            )
            lines.append(f"  {row['weekly_review_notes']}")
    else:
        lines.append("- No teacher weekly data available.")

    lines.extend(["", "Levels Needing Attention"])
    if report_data["level_attention_rows"]:
        for row in report_data["level_attention_rows"]:
            level_attention_notes = []
            if row["attendance_state"] in {"missing", "partial"}:
                level_attention_notes.append(f"attendance={row['attendance_state']}")
            if row["pending_reviews"] > 0:
                level_attention_notes.append(f"review_queue={row['pending_reviews']}")
            lines.append(f"- {row['level'].name}: {', '.join(level_attention_notes)}")
    else:
        lines.append("- No levels currently need extra attention.")

    lines.extend(["", "Students Needing Follow-up"])
    if report_data["student_attention_rows"]:
        for row in report_data["student_attention_rows"]:
            lines.append(
                f"- {row['student'].full_name} ({row['level_name']}): {row['reasons']}"
            )
    else:
        lines.append("- No students are currently flagged.")

    return build_simple_pdf(lines)


def archive_supervisor_weekly_report(report_data: dict) -> str:
    ensure_weekly_report_archive_dir()
    report_date = report_data["summary"]["report_date"].isoformat()
    file_name = f"supervisor_weekly_report_{report_date}.pdf"
    file_path = weekly_report_archive_path(file_name)
    pdf_bytes = build_supervisor_weekly_report_pdf(report_data)
    with open(file_path, "wb") as archive_file:
        archive_file.write(pdf_bytes)
    return file_name


def list_archived_weekly_reports(limit: int = 12) -> list[dict]:
    ensure_weekly_report_archive_dir()
    file_rows = []
    for entry in sorted(
        os.scandir(app.config["WEEKLY_REPORT_ARCHIVE_DIR"]),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    ):
        if not entry.is_file() or not entry.name.endswith(".pdf"):
            continue
        file_rows.append(
            {
                "file_name": entry.name,
                "size_kb": max(1, round(entry.stat().st_size / 1024)),
                "updated_at": datetime.fromtimestamp(entry.stat().st_mtime),
            }
        )
        if len(file_rows) >= limit:
            break
    return file_rows


def build_supervisor_announcement_summary(levels: list[Level]) -> dict:
    announcement_rows = build_announcement_view_rows(
        Announcement.query.order_by(
            Announcement.is_active.desc(),
            Announcement.is_pinned.desc(),
            Announcement.created_at.desc(),
            Announcement.id.desc(),
        ).limit(12).all()
    )
    return {
        "announcement_rows": announcement_rows,
        "levels": levels,
    }


def build_calendar_admin_summary() -> dict:
    holidays = HolidayPeriod.query.order_by(HolidayPeriod.start_date.desc(), HolidayPeriod.id.desc()).all()
    target_date = date.today()
    can_run, status_message = should_run_weekly_followup(target_date)
    calendar_settings = get_calendar_settings()
    week_start, week_end = get_teaching_week_window(target_date)
    return {
        "calendar_settings": calendar_settings,
        "holiday_rows": holidays,
        "followup_ready": can_run,
        "followup_status_message": status_message,
        "detected_followup_reviews": collect_weekly_teacher_reviews(target_date) if can_run else [],
        "last_weekly_followup_run": calendar_settings["last_weekly_followup_run"],
        "teaching_week_start": week_start,
        "teaching_week_end": week_end,
        "archived_weekly_reports": list_archived_weekly_reports(),
    }


def build_teacher_dashboard_data(teacher: Teacher) -> dict:
    today = date.today()
    academic_calendar_status = build_academic_calendar_status(today)
    teacher_levels = (
        Level.query.filter_by(teacher_id=teacher.id)
        .order_by(Level.order_index.asc(), Level.name.asc())
        .all()
    )
    teacher_level_ids = [level.id for level in teacher_levels]

    if not teacher_level_ids:
        return {
            "kpis": [],
            "level_rows": [],
            "student_focus_rows": [],
            "teacher_levels": teacher_levels,
            "weekly_review_rows": [],
            "weekly_review_summary": None,
            "academic_calendar_status": academic_calendar_status,
        }

    students = (
        Student.query.filter(Student.level_id.in_(teacher_level_ids))
        .order_by(Student.full_name.asc())
        .all()
    )
    student_ids = [student.id for student in students]
    students_by_level = {}
    for student in students:
        students_by_level.setdefault(student.level_id, []).append(student)

    attendance_today_rows = (
        db.session.query(
            Attendance.level_id,
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.count(Attendance.id).label("marked_count"),
        )
        .filter(Attendance.level_id.in_(teacher_level_ids), Attendance.attendance_date == today)
        .group_by(Attendance.level_id)
        .all()
    )
    attendance_today_by_level = {row.level_id: row for row in attendance_today_rows}

    attendance_14d_rows = (
        db.session.query(
            Attendance.student_id,
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.count(Attendance.id).label("total_count"),
        )
        .filter(
            Attendance.student_id.in_(student_ids if student_ids else [-1]),
            Attendance.attendance_date >= today - timedelta(days=13),
            Attendance.attendance_date <= today,
        )
        .group_by(Attendance.student_id)
        .all()
    )
    attendance_14d_by_student = {row.student_id: row for row in attendance_14d_rows}

    assignments = (
        Assignment.query.filter(Assignment.level_id.in_(teacher_level_ids))
        .order_by(Assignment.due_date.asc(), Assignment.id.desc())
        .all()
    )
    active_assignments = [assignment for assignment in assignments if assignment.is_active]
    submissions = (
        AssignmentSubmission.query.join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .filter(Assignment.level_id.in_(teacher_level_ids))
        .all()
    )
    submissions_by_assignment_student = {
        (submission.assignment_id, submission.student_id): submission
        for submission in submissions
    }

    pending_review_count = sum(
        1 for submission in submissions if submission.submitted_at and submission.status != "Reviewed"
    )
    overdue_assignments_count = sum(
        1 for assignment in active_assignments if assignment.due_date and assignment.due_date < today
    )

    upcoming_exams = (
        UpcomingExam.query.filter(
            UpcomingExam.level_id.in_(teacher_level_ids),
            UpcomingExam.exam_date >= today,
        )
        .order_by(UpcomingExam.exam_date.asc(), UpcomingExam.id.asc())
        .all()
    )

    exam_results = (
        ExamResult.query.filter(ExamResult.level_id.in_(teacher_level_ids))
        .order_by(ExamResult.student_id.asc(), ExamResult.exam_date.desc(), ExamResult.id.desc())
        .all()
    )
    exam_results_by_student = {}
    for result in exam_results:
        exam_results_by_student.setdefault(result.student_id, []).append(result)

    weekly_review_rows = [
        row for row in collect_weekly_teacher_reviews(today)
        if row["teacher"].id == teacher.id
    ]
    weekly_review_by_level_id = {row["level"].id: row for row in weekly_review_rows}
    syllabus_plan_rows = build_syllabus_plan_summary(teacher_levels)
    syllabus_plan_by_level_id = {row["level"].id: row for row in syllabus_plan_rows}

    level_rows = []
    missing_attendance_levels = 0
    total_students = len(students)
    total_marked_today = 0

    for level in teacher_levels:
        level_students = students_by_level.get(level.id, [])
        attendance_row = attendance_today_by_level.get(level.id)
        marked_count = int(attendance_row.marked_count or 0) if attendance_row else 0
        present_count = int(attendance_row.present_count or 0) if attendance_row else 0
        total_marked_today += marked_count

        if level_students and marked_count == 0:
            missing_attendance_levels += 1

        level_assignments = [assignment for assignment in active_assignments if assignment.level_id == level.id]
        level_submissions = [
            submission for submission in submissions
            if submission.assignment_id in {assignment.id for assignment in level_assignments}
        ]
        review_queue = sum(
            1 for submission in level_submissions if submission.submitted_at and submission.status != "Reviewed"
        )
        next_exam = next((exam for exam in upcoming_exams if exam.level_id == level.id), None)
        weekly_review = weekly_review_by_level_id.get(level.id)
        syllabus_plan = syllabus_plan_by_level_id.get(level.id)

        level_rows.append(
            {
                "level": level,
                "students_count": len(level_students),
                "marked_count": marked_count,
                "present_count": present_count,
                "attendance_rate": round((present_count / marked_count) * 100, 1) if marked_count else None,
                "review_queue": review_queue,
                "active_assignments": len(level_assignments),
                "next_exam": next_exam,
                "weekly_status": weekly_review["status"] if weekly_review else "praise",
                "weekly_issues": weekly_review["issues"] if weekly_review else [],
                "syllabus_plan": syllabus_plan,
            }
        )

    student_focus_rows = []
    for student in students:
        level_assignments = [assignment for assignment in active_assignments if assignment.level_id == student.level_id]
        open_assignments = 0
        for assignment in level_assignments:
            submission = submissions_by_assignment_student.get((assignment.id, student.id))
            if not submission or submission.status != "Reviewed":
                open_assignments += 1

        attendance_window = attendance_14d_by_student.get(student.id)
        attendance_rate_14d = (
            round((int(attendance_window.present_count or 0) / int(attendance_window.total_count or 0)) * 100, 1)
            if attendance_window and int(attendance_window.total_count or 0) > 0
            else None
        )
        attendance_meta = resolve_student_attendance_display(student, attendance_rate_14d)
        attendance_rate_display = attendance_meta["attendance_rate_display"]
        attendance_source_key = attendance_meta["attendance_source_key"]

        latest_exam_title = ""
        latest_exam_percentage = None
        student_exam_results = exam_results_by_student.get(student.id, [])
        if student_exam_results:
            exam_summaries = build_exam_summary_groups(student_exam_results)
            if exam_summaries:
                latest_exam_title = exam_summaries[0]["exam_title"]
                latest_exam_percentage = exam_summaries[0]["computed_percentage"]

        needs_attention = False
        reasons = []
        if attendance_rate_display is not None and attendance_rate_display < 75:
            needs_attention = True
            reasons.append(ui_text("reason_attendance"))
        if open_assignments >= 2:
            needs_attention = True
            reasons.append(ui_text("reason_assignments"))
        if latest_exam_percentage is not None and latest_exam_percentage < 65:
            needs_attention = True
            reasons.append(ui_text("reason_result"))

        if needs_attention:
            student_focus_rows.append(
                {
                    "student": student,
                    "level_name": student.level.name if student.level else normalize_level_display_name(student.level_name),
                    "attendance_rate_14d": attendance_rate_14d,
                    "attendance_rate_display": attendance_rate_display,
                    "attendance_source_key": attendance_source_key,
                    "open_assignments": open_assignments,
                    "latest_exam_title": latest_exam_title,
                    "latest_exam_percentage": latest_exam_percentage,
                    "reasons": ", ".join(reasons),
                }
            )

    student_focus_rows.sort(
        key=lambda row: (
            row["attendance_rate_display"] if row["attendance_rate_display"] is not None else 999,
            -row["open_assignments"],
            row["student"].full_name,
        )
    )

    kpis = [
        {
            "label_key": "my_levels",
            "value": str(len(teacher_levels)),
            "caption_key": "my_levels_caption",
        },
        {
            "label_key": "attendance_logged_today",
            "value": format_percent_label(round((total_marked_today / total_students) * 100, 1) if total_students else 0.0),
            "caption_key": "attendance_logged_today_caption",
            "caption_params": {"marked": total_marked_today, "total": total_students},
        },
        {
            "label_key": "review_queue_label",
            "value": str(pending_review_count),
            "caption_key": "review_queue_caption",
        },
        {
            "label_key": "upcoming_exams_label",
            "value": str(len(upcoming_exams)),
            "caption_key": "upcoming_exams_caption",
        },
        {
            "label_key": "plan_progress_label",
            "value": format_percent_label(
                round(
                    (
                        sum(row["completed_count"] for row in syllabus_plan_rows)
                        / sum(row["total_rows"] for row in syllabus_plan_rows)
                    ) * 100,
                    1,
                ) if sum(row["total_rows"] for row in syllabus_plan_rows) else 0.0
            ),
            "caption_key": "plan_progress_caption",
        },
    ]

    follow_up_levels = [row for row in weekly_review_rows if row["status"] == "follow_up"]
    praise_levels = [row for row in weekly_review_rows if row["status"] == "praise"]
    week_start, week_end = get_teaching_week_window(today)
    weekly_review_summary = {
        "week_start": week_start,
        "week_end": week_end,
        "follow_up_count": len(follow_up_levels),
        "praise_count": len(praise_levels),
        "status": "follow_up" if follow_up_levels else "praise",
        "message": (
            f"You still have {len(follow_up_levels)} level(s) needing completion before the Thursday review."
            if follow_up_levels
            else "You are currently on track for the weekly appreciation message."
        ),
    }

    syllabus_plan_summary = {
        "current_week": academic_calendar_status["current_week"],
        "levels_count": len(syllabus_plan_rows),
        "levels_on_track": sum(1 for row in syllabus_plan_rows if row["delayed_count"] == 0),
        "levels_delayed": sum(1 for row in syllabus_plan_rows if row["delayed_count"] > 0),
        "completed_count": sum(row["completed_count"] for row in syllabus_plan_rows),
        "expected_count": sum(row["expected_count"] for row in syllabus_plan_rows),
        "delayed_count": sum(row["delayed_count"] for row in syllabus_plan_rows),
        "progress_pct": round(
            (
                sum(row["completed_count"] for row in syllabus_plan_rows)
                / sum(row["total_rows"] for row in syllabus_plan_rows)
            ) * 100,
            1,
        ) if sum(row["total_rows"] for row in syllabus_plan_rows) else 0.0,
    }

    quick_tasks = []
    if academic_calendar_status["current_event"] and not academic_calendar_status["is_teaching_day"]:
        quick_tasks.append(
            {
                "tone": "info",
                "title_key": "calendar_current_alert_title",
                "body_key": "calendar_current_alert_body",
                "body_params": {"title": academic_calendar_status["current_event"].title},
                "href": "/teacher/dashboard",
                "cta_key": "teacher_dashboard",
            }
        )
    elif academic_calendar_status["next_event"]:
        quick_tasks.append(
            {
                "tone": "info",
                "title_key": "calendar_upcoming_alert_title",
                "body_key": "calendar_upcoming_alert_body",
                "body_params": {
                    "title": academic_calendar_status["next_event"].title,
                    "date": academic_calendar_status["next_event"].start_date.isoformat(),
                },
                "href": "/teacher/dashboard",
                "cta_key": "teacher_dashboard",
            }
        )
    if missing_attendance_levels:
        quick_tasks.append(
            {
                "tone": "warning",
                "title_key": "task_attendance_missing_title",
                "body_key": "task_attendance_missing_body",
                "body_params": {"count": missing_attendance_levels},
                "href": "/attendance",
                "cta_key": "open_attendance",
            }
        )
    if pending_review_count:
        quick_tasks.append(
            {
                "tone": "warning",
                "title_key": "task_review_queue_title",
                "body_key": "task_review_queue_body",
                "body_params": {"count": pending_review_count},
                "href": f"/teacher/levels/{level_rows[0]['level'].id}?workspace=assignments" if level_rows else "/teacher/dashboard",
                "cta_key": "open_review_queue",
            }
        )
    if syllabus_plan_summary["delayed_count"]:
        quick_tasks.append(
            {
                "tone": "info",
                "title_key": "task_syllabus_delay_title",
                "body_key": "task_syllabus_delay_body",
                "body_params": {"count": syllabus_plan_summary["delayed_count"]},
                "href": f"/teacher/levels/{level_rows[0]['level'].id}/syllabus-plan" if level_rows else "/teacher/dashboard",
                "cta_key": "open_syllabus_plan",
            }
        )
    if not quick_tasks:
        quick_tasks.append(
            {
                "tone": "success",
                "title_key": "task_clear_title",
                "body_key": "task_clear_body",
                "body_params": {},
                "href": f"/teacher/levels/{level_rows[0]['level'].id}" if level_rows else "/teacher/dashboard",
                "cta_key": "open_first_workspace",
            }
        )

    return {
        "kpis": kpis,
        "level_rows": level_rows,
        "student_focus_rows": student_focus_rows[:8],
        "teacher_levels": teacher_levels,
        "weekly_review_rows": weekly_review_rows,
        "weekly_review_summary": weekly_review_summary,
        "syllabus_plan_rows": syllabus_plan_rows,
        "syllabus_plan_summary": syllabus_plan_summary,
        "quick_tasks": quick_tasks,
        "academic_calendar_status": academic_calendar_status,
    }


def build_level_followup_register(level: Level, today: date | None = None) -> list[dict]:
    target_date = today or date.today()
    students = Student.query.filter_by(level_id=level.id).order_by(Student.full_name.asc()).all()
    student_ids = [student.id for student in students]

    attendance_14d_rows = (
        db.session.query(
            Attendance.student_id,
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.count(Attendance.id).label("total_count"),
        )
        .filter(
            Attendance.level_id == level.id,
            Attendance.attendance_date >= target_date - timedelta(days=13),
            Attendance.attendance_date <= target_date,
        )
        .group_by(Attendance.student_id)
        .all()
    ) if student_ids else []
    attendance_14d_by_student = {row.student_id: row for row in attendance_14d_rows}

    assignments = (
        Assignment.query.filter_by(level_id=level.id)
        .order_by(Assignment.due_date.asc(), Assignment.id.desc())
        .all()
    )
    active_assignments = [assignment for assignment in assignments if assignment.is_active]
    submissions = (
        AssignmentSubmission.query.join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .filter(Assignment.level_id == level.id)
        .all()
    ) if student_ids else []
    submissions_by_assignment_student = {
        (submission.assignment_id, submission.student_id): submission
        for submission in submissions
    }

    exam_results = (
        ExamResult.query.filter_by(level_id=level.id)
        .join(Student, ExamResult.student_id == Student.id)
        .order_by(Student.full_name.asc(), ExamResult.exam_date.desc(), ExamResult.id.desc())
        .all()
    ) if student_ids else []
    exam_results_by_student = {}
    for result in exam_results:
        exam_results_by_student.setdefault(result.student_id, []).append(result)

    month_key = target_date.strftime("%Y-%m")
    monthly_notes = StudentMonthlyNote.query.filter(
        StudentMonthlyNote.student_id.in_(student_ids if student_ids else [-1]),
        StudentMonthlyNote.month_key == month_key,
    ).all() if student_ids else []
    monthly_note_by_student_id = {note.student_id: note for note in monthly_notes}

    rows = []
    for student in students:
        attendance_window = attendance_14d_by_student.get(student.id)
        attendance_rate_14d = (
            round((int(attendance_window.present_count or 0) / int(attendance_window.total_count or 0)) * 100, 1)
            if attendance_window and int(attendance_window.total_count or 0) > 0
            else None
        )
        attendance_meta = resolve_student_attendance_display(student, attendance_rate_14d)
        attendance_rate_display = attendance_meta["attendance_rate_display"]
        attendance_source_key = attendance_meta["attendance_source_key"]
        imported_attendance = attendance_meta["imported_attendance"]
        open_assignments = 0
        for assignment in active_assignments:
            submission = submissions_by_assignment_student.get((assignment.id, student.id))
            if not submission or submission.status != "Reviewed":
                open_assignments += 1

        exam_groups = build_exam_summary_groups(exam_results_by_student.get(student.id, []))
        latest_exam = exam_groups[0] if exam_groups else None
        monthly_note = monthly_note_by_student_id.get(student.id)
        has_monthly_note = bool((monthly_note.note_text or "").strip()) if monthly_note else False
        status_key = "excellent_status"
        if (
            (attendance_rate_display is not None and attendance_rate_display < 75)
            or open_assignments > 0
            or (latest_exam and latest_exam["computed_percentage"] is not None and latest_exam["computed_percentage"] < 65)
            or not has_monthly_note
        ):
            status_key = "needs_attention_status"
        elif (
            (attendance_rate_display is not None and attendance_rate_display < 90)
            or (latest_exam and latest_exam["computed_percentage"] is not None and latest_exam["computed_percentage"] < 80)
        ):
            status_key = "stable_status"

        rows.append(
            {
                "student": student,
                "attendance_rate_14d": attendance_rate_14d,
                "attendance_rate_display": attendance_rate_display,
                "attendance_source_key": attendance_source_key,
                "imported_attendance": imported_attendance,
                "open_assignments": open_assignments,
                "latest_exam_title": latest_exam["exam_title"] if latest_exam else "",
                "latest_exam_percentage": latest_exam["computed_percentage"] if latest_exam else None,
                "has_monthly_note": has_monthly_note,
                "status_key": status_key,
            }
        )
    return rows


def filter_level_followup_rows(
    rows: list[dict],
    status_filter: str = "",
    search_query: str = "",
    sort_by: str = "",
) -> list[dict]:
    normalized_status = (status_filter or "").strip()
    normalized_search = (search_query or "").strip().lower()
    filtered_rows = rows

    if normalized_status in {"excellent_status", "stable_status", "needs_attention_status"}:
        filtered_rows = [row for row in filtered_rows if row["status_key"] == normalized_status]

    if normalized_search:
        filtered_rows = [
            row for row in filtered_rows
            if normalized_search in (row["student"].full_name or "").lower()
            or normalized_search in (row["student"].student_code or "").lower()
        ]

    normalized_sort = (sort_by or "").strip()
    if normalized_sort == "lowest_attendance":
        filtered_rows = sorted(
            filtered_rows,
            key=lambda row: (
                row["attendance_rate_display"] if row["attendance_rate_display"] is not None else 999,
                -row["open_assignments"],
                row["student"].full_name.lower(),
            ),
        )
    elif normalized_sort == "open_assignments":
        filtered_rows = sorted(
            filtered_rows,
            key=lambda row: (
                -row["open_assignments"],
                row["attendance_rate_display"] if row["attendance_rate_display"] is not None else 999,
                row["student"].full_name.lower(),
            ),
        )
    elif normalized_sort == "missing_monthly_note":
        filtered_rows = sorted(
            filtered_rows,
            key=lambda row: (
                0 if not row["has_monthly_note"] else 1,
                row["attendance_rate_display"] if row["attendance_rate_display"] is not None else 999,
                -row["open_assignments"],
                row["student"].full_name.lower(),
            ),
        )

    return filtered_rows


def summarize_level_followup_rows(rows: list[dict]) -> dict:
    return {
        "excellent_count": sum(1 for row in rows if row["status_key"] == "excellent_status"),
        "stable_count": sum(1 for row in rows if row["status_key"] == "stable_status"),
        "needs_attention_count": sum(1 for row in rows if row["status_key"] == "needs_attention_status"),
        "total_count": len(rows),
    }


def build_student_dashboard_summary(student: Student, level: Level | None, assignment_rows: list[dict], exam_summary_groups: list[dict], upcoming_exams: list[UpcomingExam], recordings: list[ClassRecording]) -> dict:
    today = date.today()
    academic_calendar_status = build_academic_calendar_status(today)
    pending_assignments = 0
    reviewed_assignments = 0
    overdue_assignments = 0
    submitted_waiting_review = 0

    for row in assignment_rows:
        assignment = row["assignment"]
        submission = row["submission"]
        status = submission.status if submission else "Pending"
        if status == "Reviewed":
            reviewed_assignments += 1
        else:
            pending_assignments += 1
        if submission and status == "Submitted":
            submitted_waiting_review += 1
        if assignment.due_date and assignment.due_date < today and status != "Reviewed":
            overdue_assignments += 1

    latest_exam = exam_summary_groups[0] if exam_summary_groups else None
    next_exam = next((exam for exam in upcoming_exams if exam.exam_date >= today), None)
    latest_recording = recordings[0] if recordings else None
    attendance_window = (
        db.session.query(
            db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
            db.func.count(Attendance.id).label("total_count"),
        )
        .filter(
            Attendance.student_id == student.id,
            Attendance.attendance_date >= today - timedelta(days=13),
            Attendance.attendance_date <= today,
        )
        .first()
    )
    attendance_rate_14d = (
        round((int(attendance_window.present_count or 0) / int(attendance_window.total_count or 0)) * 100, 1)
        if attendance_window and int(attendance_window.total_count or 0) > 0
        else None
    )
    attendance_meta = resolve_student_attendance_display(student, attendance_rate_14d)
    attendance_rate_display = attendance_meta["attendance_rate_display"]
    attendance_source_key = attendance_meta["attendance_source_key"] or "attendance_source_none"
    imported_attendance = attendance_meta["imported_attendance"]

    alerts = []
    if academic_calendar_status["current_event"] and not academic_calendar_status["is_teaching_day"]:
        alerts.append(
            {
                "tone": "info",
                "title": ui_text("calendar_current_alert_title"),
                "body": academic_calendar_status["current_event"].title,
            }
        )
    elif academic_calendar_status["next_event"]:
        alerts.append(
            {
                "tone": "info",
                "title": ui_text("calendar_upcoming_alert_title"),
                "body": f"{academic_calendar_status['next_event'].title} - {academic_calendar_status['next_event'].start_date.isoformat()}",
            }
        )
    if overdue_assignments:
        alerts.append(
            {
                "tone": "danger",
                "title": "Overdue Work",
                "body": f"You have {overdue_assignments} assignment(s) past their due date.",
            }
        )
    if submitted_waiting_review:
        alerts.append(
            {
                "tone": "info",
                "title": "Waiting For Review",
                "body": f"{submitted_waiting_review} submitted assignment(s) are waiting for teacher feedback.",
            }
        )
    if attendance_rate_display is not None and attendance_rate_display < 75:
        alerts.append(
            {
                "tone": "warning",
                "title": ui_text("current_student_status"),
                "body": f"{ui_text('attendance_percentage')}: {attendance_rate_display}% • {ui_text(attendance_source_key)}",
            }
        )
    if next_exam:
        days_left = (next_exam.exam_date - today).days
        alerts.append(
            {
                "tone": "warning" if days_left <= 3 else "info",
                "title": "Upcoming Exam",
                "body": f"{next_exam.title} is scheduled on {next_exam.exam_date}.",
            }
        )

    quick_actions = []
    if level and level.zoom_link:
        quick_actions.append({"label": "Join Class", "href": level.zoom_link, "kind": "primary", "external": True})
    if next_exam:
        quick_actions.append({"label": "Upcoming Exams", "href": "#student-upcoming-exams", "kind": "secondary", "external": False})
    if latest_exam:
        quick_actions.append(
            {
                "label": "Latest Result PDF",
                "href": f"/students/{student.id}/exam-report.pdf?student_code={student.student_code}&exam_title={latest_exam['exam_title']}",
                "kind": "secondary",
                "external": False,
            }
        )

    metrics = [
        {
            "label": ui_text("current_student_status"),
            "value": ui_text(
                "needs_attention_status"
                if (
                    (attendance_rate_display is not None and attendance_rate_display < 75)
                    or overdue_assignments > 0
                    or (latest_exam and latest_exam["computed_percentage"] is not None and latest_exam["computed_percentage"] < 65)
                )
                else (
                    "stable_status"
                    if (
                        (attendance_rate_display is not None and attendance_rate_display < 90)
                        or (latest_exam and latest_exam["computed_percentage"] is not None and latest_exam["computed_percentage"] < 80)
                    )
                    else "excellent_status"
                )
            ),
            "caption": ui_text("attendance_snapshot_caption"),
        },
        {
            "label": ui_text("attendance_percentage"),
            "value": f"{attendance_rate_display}%" if attendance_rate_display is not None else "-",
            "caption": ui_text(attendance_source_key),
        },
        {"label": "Open Assignments", "value": str(pending_assignments), "caption": "Tasks that still need completion or review"},
        {"label": "Reviewed Work", "value": str(reviewed_assignments), "caption": "Assignments already checked by your teacher"},
        {"label": "Upcoming Exams", "value": str(len(upcoming_exams)), "caption": "Announced exams for your class"},
        {"label": "Class Recordings", "value": str(len(recordings)), "caption": "Recent class recordings available for replay"},
        {"label": ui_text("current_teaching_week"), "value": str(academic_calendar_status["current_week"]), "caption": ui_text(academic_calendar_status["status_key"])},
    ]

    return {
        "alerts": alerts,
        "quick_actions": quick_actions[:4],
        "metrics": metrics,
        "latest_exam": latest_exam,
        "next_exam": next_exam,
        "latest_recording": latest_recording,
        "pending_assignments": pending_assignments,
        "attendance_rate_display": attendance_rate_display,
        "attendance_source_key": attendance_source_key,
        "imported_attendance": imported_attendance,
        "academic_calendar_status": academic_calendar_status,
    }


def ensure_teacher_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(teachers)"))
    for row in result:
        existing_columns.add(row[1])

    required_columns = {
        "phone": "TEXT",
        "email": "TEXT",
    }

    for column_name, column_type in required_columns.items():
        if column_name not in existing_columns:
            db.session.execute(
                text(f"ALTER TABLE teachers ADD COLUMN {column_name} {column_type}")
            )
    db.session.commit()


def ensure_settings_schema() -> None:
    db.session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS system_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
    )
    db.session.commit()

    defaults = {
        "admin_access_code": app.config["ADMIN_ACCESS_CODE"],
        "duty_start_time": "08:00",
        "duty_end_time": "15:00",
        "weekly_followup_weekday": "thursday",
        "weekly_followup_time": "14:00",
        "current_teaching_week": "1",
    }
    for key, value in defaults.items():
        exists = db.session.execute(
            text("SELECT 1 FROM system_settings WHERE key = :key LIMIT 1"),
            {"key": key},
        ).fetchone()
        if not exists:
            set_system_setting(key, value)
    db.session.commit()


def ensure_level_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(levels)"))
    for row in result:
        existing_columns.add(row[1])

    needs_commit = False

    if "order_index" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN order_index INTEGER"))
        needs_commit = True
    if "zoom_email" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN zoom_email TEXT"))
        needs_commit = True
    if "zoom_link" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN zoom_link TEXT"))
        needs_commit = True
    if "zoom_meeting_id" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN zoom_meeting_id TEXT"))
        needs_commit = True
    if "zoom_passcode" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN zoom_passcode TEXT"))
        needs_commit = True
    if "homework_padlet_url" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN homework_padlet_url TEXT"))
        needs_commit = True
    if "announcements_padlet_url" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN announcements_padlet_url TEXT"))
        needs_commit = True
    if "syllabus_edit_open" not in existing_columns:
        db.session.execute(text("ALTER TABLE levels ADD COLUMN syllabus_edit_open BOOLEAN NOT NULL DEFAULT 0"))
        needs_commit = True

    if needs_commit:
        db.session.commit()


def ensure_student_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(students)"))
    for row in result:
        existing_columns.add(row[1])

    needs_commit = False

    if "level_id" not in existing_columns:
        db.session.execute(text("ALTER TABLE students ADD COLUMN level_id INTEGER"))
        needs_commit = True
    if "parent_email" not in existing_columns:
        db.session.execute(text("ALTER TABLE students ADD COLUMN parent_email TEXT"))
        needs_commit = True
    if "parent_whatsapp" not in existing_columns:
        db.session.execute(text("ALTER TABLE students ADD COLUMN parent_whatsapp TEXT"))
        needs_commit = True

    if needs_commit:
        db.session.commit()


def ensure_attendance_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(attendance)"))
    for row in result:
        existing_columns.add(row[1])

    if "level_id" not in existing_columns:
        db.session.execute(text("ALTER TABLE attendance ADD COLUMN level_id INTEGER"))
        db.session.commit()


def ensure_recordings_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(class_recordings)"))
    for row in result:
        existing_columns.add(row[1])

    needs_commit = False
    if "summary" not in existing_columns:
        db.session.execute(text("ALTER TABLE class_recordings ADD COLUMN summary TEXT"))
        needs_commit = True
    if "homework" not in existing_columns:
        db.session.execute(text("ALTER TABLE class_recordings ADD COLUMN homework TEXT"))
        needs_commit = True
    if needs_commit:
        db.session.commit()


def ensure_curriculum_items_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(curriculum_items)"))
    for row in result:
        existing_columns.add(row[1])

    needs_commit = False
    if "visibility_scope" not in existing_columns:
        db.session.execute(
            text(
                "ALTER TABLE curriculum_items ADD COLUMN visibility_scope TEXT NOT NULL DEFAULT 'student_and_teacher'"
            )
        )
        needs_commit = True
    if needs_commit:
        db.session.commit()


def ensure_syllabus_plan_schema() -> None:
    existing_columns = set()
    result = db.session.execute(text("PRAGMA table_info(syllabus_plan_entries)"))
    for row in result:
        existing_columns.add(row[1])

    needs_commit = False
    if existing_columns and "session_number" not in existing_columns:
        db.session.execute(text("ALTER TABLE syllabus_plan_entries ADD COLUMN session_number INTEGER"))
        needs_commit = True
    if existing_columns and "book_name" not in existing_columns:
        db.session.execute(text("ALTER TABLE syllabus_plan_entries ADD COLUMN book_name TEXT"))
        needs_commit = True
    if existing_columns and "unit_name" not in existing_columns:
        db.session.execute(text("ALTER TABLE syllabus_plan_entries ADD COLUMN unit_name TEXT"))
        needs_commit = True
    if existing_columns and "completed_on" not in existing_columns:
        db.session.execute(text("ALTER TABLE syllabus_plan_entries ADD COLUMN completed_on DATE"))
        needs_commit = True
    if needs_commit:
        db.session.commit()


def seed_default_academic_calendar() -> None:
    if AcademicCalendarEvent.query.count():
        return

    for row in DEFAULT_ACADEMIC_CALENDAR_2025_26:
        db.session.add(AcademicCalendarEvent(**row))
    db.session.commit()


PRIMARY_WEEKLY_PLAN_LEVEL_NAMES = {
    "Qaeda",
    "Primary Beginner",
    "Primary Intermediate",
    "Primary Advance",
}


def get_level_syllabus_lessons_per_week(level: Level | None) -> int:
    if not level:
        return 4
    return 1 if (level.name or "").strip() in PRIMARY_WEEKLY_PLAN_LEVEL_NAMES else 4


def build_syllabus_plan_summary(levels: list[Level] | None = None) -> list[dict]:
    target_levels = levels or Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    level_ids = [level.id for level in target_levels]
    if not level_ids:
        return []
    current_week = get_current_teaching_week()

    rows = (
        SyllabusPlanEntry.query.filter(SyllabusPlanEntry.level_id.in_(level_ids))
        .order_by(SyllabusPlanEntry.level_id.asc(), SyllabusPlanEntry.week_number.asc(), SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
        .all()
    )
    rows_by_level: dict[int, list[SyllabusPlanEntry]] = {}
    for row in rows:
        rows_by_level.setdefault(row.level_id, []).append(row)

    summary_rows = []
    for level in target_levels:
        lessons_per_week = get_level_syllabus_lessons_per_week(level)
        current_slot_count = get_completed_instructional_slot_count(lessons_per_week=lessons_per_week)
        level_rows = rows_by_level.get(level.id, [])
        scheduled_rows = [row for row in level_rows if row.week_number and row.session_number]
        unscheduled_rows = [row for row in level_rows if not row.week_number or not row.session_number]
        total_rows = len(level_rows)
        planned_count = sum(1 for row in level_rows if row.status == "planned")
        in_progress_count = sum(1 for row in level_rows if row.status == "in_progress")
        completed_count = sum(1 for row in level_rows if row.status == "completed")
        postponed_count = sum(1 for row in level_rows if row.status == "postponed")
        ordered_rows = sorted(
            scheduled_rows,
            key=lambda item: ((item.order_index or 0), (item.week_number or 0), (item.session_number or 0), item.id),
        )
        expected_count = min(len(ordered_rows), current_slot_count)
        completed_expected_count = sum(
            1 for row in ordered_rows[:expected_count]
            if row.status == "completed"
        )
        delayed_count = max(expected_count - completed_expected_count, 0)
        remaining_count = max(total_rows - completed_count, 0)
        progress_pct = round((completed_count / total_rows) * 100, 1) if total_rows else 0.0
        last_completed_row = next(
            (
                row
                for row in sorted(level_rows, key=lambda item: ((item.week_number or 0), (item.session_number or 0)), reverse=True)
                if row.status == "completed"
            ),
            None,
        )
        summary_rows.append(
            {
                "level": level,
                "teacher": level.teacher,
                "rows": level_rows,
                "scheduled_rows": scheduled_rows,
                "unscheduled_rows": unscheduled_rows,
                "total_rows": total_rows,
                "planned_count": planned_count,
                "in_progress_count": in_progress_count,
                "completed_count": completed_count,
                "postponed_count": postponed_count,
                "expected_count": expected_count,
                "delayed_count": delayed_count,
                "remaining_count": remaining_count,
                "unscheduled_count": len(unscheduled_rows),
                "status_label_key": "on_track_status" if delayed_count == 0 else "delayed_status",
                "progress_pct": progress_pct,
                "last_completed_label": (
                    f"W{last_completed_row.week_number} / S{last_completed_row.session_number} - {last_completed_row.lesson_title}"
                    if last_completed_row
                    else ""
                ),
                "current_week": current_week,
                "lessons_per_week": lessons_per_week,
            }
        )

    return summary_rows


def build_teaching_week_schedule(total_weeks: int = 40, lessons_per_week: int = 4) -> dict[int, dict]:
    start_date = get_academic_year_start_date()
    if not start_date:
        return {}

    calendar_events = get_academic_calendar_events()
    calendar_end = max((event.end_date for event in calendar_events), default=start_date + timedelta(days=365))
    teaching_days = build_teaching_days(start_date, calendar_end)
    teaching_day_set = set(teaching_days)
    schedule = {}
    week_start = start_date
    academic_week_number = 0
    teaching_week_number = 0

    while week_start <= calendar_end:
        academic_week_number += 1
        display_end = week_start + timedelta(days=6)
        week_teaching_days = [
            day for day in teaching_days
            if week_start <= day <= display_end
        ]
        is_instructional_week = bool(week_teaching_days)
        if is_instructional_week:
            teaching_week_number += 1
        overlapping_events = []
        dominant_event_type = ""
        for event in calendar_events:
            if event.start_date <= display_end and event.end_date >= week_start:
                overlapping_events.append(
                    {
                        "title": event.title,
                        "event_type": event.event_type,
                        "type_label": get_academic_event_type_label(event.event_type),
                    }
                )
        for event_type in ("exam", "holiday", "school_resume", "parents_meeting", "teacher_training", "term_start", "event_day"):
            if any(event["event_type"] == event_type for event in overlapping_events):
                dominant_event_type = event_type
                break
        if not dominant_event_type and not is_instructional_week:
            dominant_event_type = "non_teaching_period"
        schedule[academic_week_number] = {
            "academic_week_number": academic_week_number,
            "teaching_week_number": teaching_week_number if is_instructional_week else None,
            "week_start": week_start,
            "display_end": display_end,
            "date_label": (
                f"{week_start.strftime('%Y-%m-%d')} → {display_end.strftime('%Y-%m-%d')}"
            ),
            "events": overlapping_events,
            "is_instructional_week": is_instructional_week,
            "teaching_days_count": len(week_teaching_days),
            "slot_count": min(len(week_teaching_days), lessons_per_week),
            "is_current_week": week_start <= date.today() <= display_end,
            "dominant_event_type": dominant_event_type,
        }
        week_start = display_end + timedelta(days=1)
        if teaching_week_number >= total_weeks and week_start > calendar_end:
            break

    return schedule


def build_instructional_slot_sequence(total_weeks: int = 40, lessons_per_week: int = 4) -> list[dict]:
    schedule = build_teaching_week_schedule(total_weeks=total_weeks, lessons_per_week=lessons_per_week)
    slots: list[dict] = []
    slot_order = 0
    for academic_week_number, week_meta in schedule.items():
        if not week_meta.get("is_instructional_week"):
            continue
        slot_count = week_meta.get("slot_count", lessons_per_week) or 0
        week_start = week_meta.get("week_start")
        for session_number in range(1, slot_count + 1):
            slot_date = week_start + timedelta(days=session_number - 1) if week_start else None
            slot_order += 1
            slots.append(
                {
                    "slot_order": slot_order,
                    "academic_week_number": academic_week_number,
                    "session_number": session_number,
                    "date": slot_date,
                }
            )
    return slots


def get_completed_instructional_slot_count(target_date: date | None = None, total_weeks: int = 40, lessons_per_week: int = 4) -> int:
    resolved_date = target_date or date.today()
    slots = build_instructional_slot_sequence(total_weeks=total_weeks, lessons_per_week=lessons_per_week)
    return sum(1 for slot in slots if slot["date"] and slot["date"] <= resolved_date)


def redistribute_syllabus_entries_for_level(level_id: int, total_weeks: int = 40, lessons_per_week: int = 4) -> int:
    rows = (
        SyllabusPlanEntry.query.filter_by(level_id=level_id)
        .order_by(SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
        .all()
    )
    if not rows:
        return 0

    slots = build_instructional_slot_sequence(total_weeks=total_weeks, lessons_per_week=lessons_per_week)
    for index, row in enumerate(rows):
        if index >= len(slots):
            row.week_number = None
            row.session_number = None
            row.order_index = index + 1
            continue
        slot = slots[index]
        row.week_number = slot["academic_week_number"]
        row.session_number = slot["session_number"]
        row.order_index = slot["slot_order"]
    return min(len(rows), len(slots))


def build_syllabus_plan_grid(level: Level, total_weeks: int = 40, lessons_per_week: int | None = None) -> list[dict]:
    resolved_lessons_per_week = lessons_per_week or get_level_syllabus_lessons_per_week(level)
    rows = (
        SyllabusPlanEntry.query.filter_by(level_id=level.id)
        .order_by(
            SyllabusPlanEntry.week_number.asc(),
            SyllabusPlanEntry.session_number.asc(),
            SyllabusPlanEntry.order_index.asc(),
            SyllabusPlanEntry.id.asc(),
        )
        .all()
    )
    entry_map = {}
    for row in rows:
        if row.week_number and row.session_number:
            entry_map[(row.week_number, row.session_number)] = row

    incomplete_before = False
    current_slot_count = get_completed_instructional_slot_count(total_weeks=total_weeks, lessons_per_week=resolved_lessons_per_week)
    week_schedule = build_teaching_week_schedule(total_weeks=total_weeks, lessons_per_week=resolved_lessons_per_week)
    grid_rows = []
    for academic_week_number, week_meta in week_schedule.items():
        plan_week_number = week_meta.get("teaching_week_number")
        sessions = []
        for session_number in range(1, resolved_lessons_per_week + 1):
            is_available_slot = bool(plan_week_number and session_number <= (week_meta.get("slot_count") or 0))
            entry = entry_map.get((plan_week_number, session_number)) if is_available_slot else None
            has_incomplete_previous = bool(entry and incomplete_before)
            sessions.append(
                {
                    "session_number": session_number,
                    "entry": entry,
                    "is_available_slot": is_available_slot,
                    "is_expected": bool(entry and (entry.order_index or 999999) <= current_slot_count),
                    "is_review": bool(entry and "مراجعة" in (entry.lesson_title or "")),
                    "has_incomplete_previous": has_incomplete_previous,
                }
            )
            if entry and entry.status != "completed":
                incomplete_before = True
        grid_rows.append(
            {
                "week_number": plan_week_number,
                "academic_week_number": academic_week_number,
                "is_instructional_week": bool(plan_week_number),
                "week_meta": week_meta,
                "sessions": sessions,
                "lessons_per_week": resolved_lessons_per_week,
            }
        )
    return grid_rows


SYLLABUS_TEMPLATE_COLUMNS = [
    "academic_week_number",
    "teaching_week_number",
    "session_number",
    "date_range",
    "calendar_events",
    "book_name",
    "unit_name",
    "lesson_title",
    "source_reference",
    "learning_objective",
    "planned_homework",
    "note_text",
]

SYLLABUS_IMPORT_PREVIEW_DIR = Path("/tmp/hikmah_syllabus_import_previews")


def _syllabus_import_preview_store(preview_payload: dict) -> str:
    SYLLABUS_IMPORT_PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid4().hex
    preview_path = SYLLABUS_IMPORT_PREVIEW_DIR / f"{token}.json"
    preview_path.write_text(json.dumps(preview_payload, ensure_ascii=False), encoding="utf-8")
    return token


def _syllabus_import_preview_load(token: str) -> dict | None:
    if not token:
        return None
    preview_path = SYLLABUS_IMPORT_PREVIEW_DIR / f"{token}.json"
    if not preview_path.exists():
        return None
    try:
        return json.loads(preview_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _syllabus_import_preview_delete(token: str) -> None:
    if not token:
        return
    preview_path = SYLLABUS_IMPORT_PREVIEW_DIR / f"{token}.json"
    try:
        preview_path.unlink(missing_ok=True)
    except Exception:
        pass


def _get_admin_syllabus_preview_token_map() -> dict:
    value = session.get("admin_syllabus_import_previews")
    return value if isinstance(value, dict) else {}


def _set_admin_syllabus_preview_token(level_id: int, token: str | None) -> None:
    token_map = _get_admin_syllabus_preview_token_map()
    if token:
        token_map[str(level_id)] = token
    else:
        token_map.pop(str(level_id), None)
    session["admin_syllabus_import_previews"] = token_map


def build_syllabus_import_comparison(level: Level, preview_payload: dict | None) -> dict | None:
    if not preview_payload:
        return None

    current_rows = (
        SyllabusPlanEntry.query.filter_by(level_id=level.id)
        .order_by(SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
        .all()
    )
    current_map = {
        (row.week_number, row.session_number): row
        for row in current_rows
        if row.week_number is not None and row.session_number is not None
    }
    preview_rows = list(preview_payload.get("scheduled_rows") or [])
    preview_map = {
        (row.get("week_number"), row.get("session_number")): row
        for row in preview_rows
        if row.get("week_number") is not None and row.get("session_number") is not None
    }

    changed_rows = []
    for slot_key in sorted(set(current_map.keys()) | set(preview_map.keys()), key=lambda item: ((item[0] or 999), (item[1] or 999))):
        current_row = current_map.get(slot_key)
        preview_row = preview_map.get(slot_key)
        current_lesson = (current_row.lesson_title or "").strip() if current_row else ""
        preview_lesson = (str(preview_row.get("lesson_title") or "").strip() if preview_row else "")
        current_unit = (current_row.unit_name or "").strip() if current_row else ""
        preview_unit = (str(preview_row.get("unit_name") or "").strip() if preview_row else "")
        current_book = (current_row.book_name or "").strip() if current_row else ""
        preview_book = (str(preview_row.get("book_name") or "").strip() if preview_row else "")
        if (current_lesson, current_unit, current_book) == (preview_lesson, preview_unit, preview_book):
            continue
        changed_rows.append(
            {
                "week_number": slot_key[0],
                "session_number": slot_key[1],
                "current_book": current_book or "-",
                "current_unit": current_unit or "-",
                "current_lesson": current_lesson or "-",
                "new_book": preview_book or "-",
                "new_unit": preview_unit or "-",
                "new_lesson": preview_lesson or "-",
            }
        )

    return {
        "changed_count": len(changed_rows),
        "changed_rows": changed_rows[:10],
    }


def build_syllabus_template_workbook(level: Level) -> BytesIO:
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws.append(SYLLABUS_TEMPLATE_COLUMNS)

    schedule = build_teaching_week_schedule(total_weeks=40, lessons_per_week=lessons_per_week)
    entry_map = {
        (row.week_number, row.session_number): row
        for row in SyllabusPlanEntry.query.filter_by(level_id=level.id).all()
        if row.week_number and row.session_number
    }

    for academic_week_number, week_meta in schedule.items():
        if not week_meta.get("is_instructional_week"):
            continue
        teaching_week_number = week_meta.get("teaching_week_number")
        for session_number in range(1, lessons_per_week + 1):
            if session_number > (week_meta.get("slot_count") or 0):
                continue
            entry = entry_map.get((teaching_week_number, session_number))
            ws.append(
                [
                    academic_week_number,
                    teaching_week_number,
                    session_number,
                    week_meta.get("date_label") or "",
                    " | ".join(event.get("title", "") for event in week_meta.get("events", [])),
                    entry.book_name if entry else "",
                    entry.unit_name if entry else "",
                    entry.lesson_title if entry else "",
                    entry.source_reference if entry else "",
                    entry.learning_objective if entry else "",
                    entry.planned_homework if entry else "",
                    entry.note_text if entry else "",
                ]
            )

    reserve_ws = wb.create_sheet("Reserve")
    reserve_ws.append(["book_name", "unit_name", "lesson_title", "source_reference", "learning_objective", "planned_homework", "note_text"])
    unscheduled_rows = (
        SyllabusPlanEntry.query.filter_by(level_id=level.id)
        .filter((SyllabusPlanEntry.week_number.is_(None)) | (SyllabusPlanEntry.session_number.is_(None)))
        .order_by(SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
        .all()
    )
    for row in unscheduled_rows:
        reserve_ws.append([
            row.book_name or "",
            row.unit_name or "",
            row.lesson_title or "",
            row.source_reference or "",
            row.learning_objective or "",
            row.planned_homework or "",
            row.note_text or "",
        ])

    guide_ws = wb.create_sheet("Guide")
    guide_ws.append(["field", "notes"])
    guide_rows = [
        ("academic_week_number", "Read-only reference from the academic calendar."),
        ("teaching_week_number", "Read-only teaching week index used by the site."),
        ("session_number", "Lesson slot within the teaching week. For primary plans this is usually 1."),
        ("book_name", "Editable"),
        ("unit_name", "Editable"),
        ("lesson_title", "Editable and required"),
        ("source_reference", "Editable"),
        ("learning_objective", "Editable"),
        ("planned_homework", "Editable"),
        ("note_text", "Editable"),
        ("Reserve sheet", "Optional extra rows that do not fit the current teaching weeks."),
    ]
    for row in guide_rows:
        guide_ws.append(list(row))

    for sheet in (ws, reserve_ws, guide_ws):
        sheet.freeze_panes = "A2"
        for column_letter, width in {
            "A": 18, "B": 18, "C": 16, "D": 26, "E": 28,
            "F": 28, "G": 26, "H": 42, "I": 26, "J": 42,
            "K": 34, "L": 34,
        }.items():
            if column_letter in sheet.column_dimensions:
                sheet.column_dimensions[column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_syllabus_template_workbook(level: Level, workbook_file) -> dict:
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    wb = load_workbook(workbook_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    if headers[: len(SYLLABUS_TEMPLATE_COLUMNS)] != SYLLABUS_TEMPLATE_COLUMNS:
        raise ValueError("Invalid syllabus template columns.")

    slot_sequence = build_instructional_slot_sequence(lessons_per_week=lessons_per_week)
    valid_slots = {(slot["academic_week_number"], slot["session_number"]): slot for slot in slot_sequence}
    imported_rows: list[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        values = {headers[idx]: ws.cell(row_idx, idx + 1).value for idx in range(len(headers))}
        has_content = any(str(values.get(key) or "").strip() for key in ["book_name", "unit_name", "lesson_title", "source_reference", "learning_objective", "planned_homework", "note_text"])
        if not has_content:
            continue
        academic_week_number = parse_whole_number(values.get("academic_week_number"))
        session_number = parse_whole_number(values.get("session_number")) or 1
        slot = valid_slots.get((academic_week_number, session_number))
        if not slot:
            continue
        imported_rows.append(
            {
                "slot_order": slot["slot_order"],
                "week_number": slot["academic_week_number"],
                "session_number": slot["session_number"],
                "book_name": (str(values.get("book_name") or "").strip() or None),
                "unit_name": (str(values.get("unit_name") or "").strip() or None),
                "lesson_title": (str(values.get("lesson_title") or "").strip()),
                "source_reference": (str(values.get("source_reference") or "").strip() or None),
                "learning_objective": (str(values.get("learning_objective") or "").strip() or None),
                "planned_homework": (str(values.get("planned_homework") or "").strip() or None),
                "note_text": (str(values.get("note_text") or "").strip() or None),
            }
        )

    reserve_rows: list[dict] = []
    if "Reserve" in wb.sheetnames:
        reserve_ws = wb["Reserve"]
        reserve_headers = [str(cell.value or "").strip() for cell in reserve_ws[1]]
        for row_idx in range(2, reserve_ws.max_row + 1):
            values = {reserve_headers[idx]: reserve_ws.cell(row_idx, idx + 1).value for idx in range(len(reserve_headers))}
            lesson_title = (str(values.get("lesson_title") or "").strip())
            if not lesson_title:
                continue
            reserve_rows.append(
                {
                    "book_name": (str(values.get("book_name") or "").strip() or None),
                    "unit_name": (str(values.get("unit_name") or "").strip() or None),
                    "lesson_title": lesson_title,
                    "source_reference": (str(values.get("source_reference") or "").strip() or None),
                    "learning_objective": (str(values.get("learning_objective") or "").strip() or None),
                    "planned_homework": (str(values.get("planned_homework") or "").strip() or None),
                    "note_text": (str(values.get("note_text") or "").strip() or None),
                }
            )

    return {
        "scheduled_rows": sorted(imported_rows, key=lambda item: item["slot_order"]),
        "reserve_rows": reserve_rows,
        "scheduled_count": len(imported_rows),
        "reserve_count": len(reserve_rows),
        "created_count": len(imported_rows) + len(reserve_rows),
    }


def apply_syllabus_template_import(level: Level, parsed_import: dict) -> dict:
    imported_rows = list(parsed_import.get("scheduled_rows") or [])
    reserve_rows = list(parsed_import.get("reserve_rows") or [])
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    slot_sequence = build_instructional_slot_sequence(lessons_per_week=lessons_per_week)

    for existing in SyllabusPlanEntry.query.filter_by(level_id=level.id).all():
        db.session.delete(existing)

    created_count = 0
    for row in imported_rows:
        db.session.add(
            SyllabusPlanEntry(
                level_id=level.id,
                week_number=row["week_number"],
                session_number=row["session_number"],
                book_name=row["book_name"],
                unit_name=row["unit_name"],
                lesson_title=row["lesson_title"],
                source_reference=row["source_reference"],
                learning_objective=row["learning_objective"],
                planned_homework=row["planned_homework"],
                note_text=row["note_text"],
                status="planned",
                order_index=row["slot_order"],
            )
        )
        created_count += 1

    next_order = len(slot_sequence) + 1
    for reserve in reserve_rows:
        db.session.add(
            SyllabusPlanEntry(
                level_id=level.id,
                week_number=None,
                session_number=None,
                book_name=reserve["book_name"],
                unit_name=reserve["unit_name"],
                lesson_title=reserve["lesson_title"],
                source_reference=reserve["source_reference"],
                learning_objective=reserve["learning_objective"],
                planned_homework=reserve["planned_homework"],
                note_text=reserve["note_text"],
                status="planned",
                order_index=next_order,
            )
        )
        created_count += 1
        next_order += 1

    return {
        "scheduled_count": len(imported_rows),
        "reserve_count": len(reserve_rows),
        "created_count": created_count,
    }


def import_syllabus_template_workbook(level: Level, workbook_file) -> dict:
    parsed_import = parse_syllabus_template_workbook(level, workbook_file)
    return apply_syllabus_template_import(level, parsed_import)


def build_syllabus_plan_excel_workbook(level: Level) -> BytesIO:
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    summary = build_syllabus_plan_summary([level])[0]
    academic_calendar_status = build_academic_calendar_status()
    grid_rows = build_syllabus_plan_grid(level, lessons_per_week=lessons_per_week)

    wb = Workbook()
    overview_ws = wb.active
    overview_ws.title = "Overview"
    plan_ws = wb.create_sheet("Plan")
    reserve_ws = wb.create_sheet("Reserve")

    header_fill = PatternFill("solid", fgColor="4A99AD")
    soft_fill = PatternFill("solid", fgColor="EEF8FF")
    review_fill = PatternFill("solid", fgColor="FFF5E8")
    holiday_fill = PatternFill("solid", fgColor="F5F7FA")
    current_fill = PatternFill("solid", fgColor="E9F8F3")
    border = Border(
        left=Side(style="thin", color="D7E4EF"),
        right=Side(style="thin", color="D7E4EF"),
        top=Side(style="thin", color="D7E4EF"),
        bottom=Side(style="thin", color="D7E4EF"),
    )
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrapped = Alignment(vertical="top", wrap_text=True)

    overview_ws.merge_cells("B2:E2")
    overview_ws["B2"] = "Hikmah Academy | أكاديمية الحكمة"
    overview_ws["B2"].font = Font(size=20, bold=True, color="1F2937")
    overview_ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    overview_ws.merge_cells("B3:E3")
    overview_ws["B3"] = "Arabic Department Syllabus Plan | خطة قسم اللغة العربية"
    overview_ws["B3"].font = Font(size=13, bold=True, color="4A99AD")
    overview_ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    logo_path = os.path.join(app.static_folder or "static", "academy-logo.jpg")
    if os.path.exists(logo_path):
        try:
            logo = XLImage(logo_path)
            logo.width = 80
            logo.height = 80
            overview_ws.add_image(logo, "A1")
        except Exception:
            pass

    overview_cards = [
        ("Level | الفصل", level.name, "B5", "C5"),
        ("Teacher | المعلم", level.teacher.full_name if level.teacher else "-", "D5", "E5"),
        ("Lessons / Week | الحصص أسبوعياً", str(lessons_per_week), "B7", "C7"),
        ("Progress % | نسبة الإنجاز", f"{summary['progress_pct']}%", "D7", "E7"),
        ("Expected By Now | المفترض حتى الآن", str(summary["expected_count"]), "B9", "C9"),
        ("Delayed Lessons | الحصص المتأخرة", str(summary["delayed_count"]), "D9", "E9"),
        ("Completed Lessons | الحصص المكتملة", str(summary["completed_count"]), "B11", "C11"),
        ("Unscheduled Rows | غير المجدول", str(summary["unscheduled_count"]), "D11", "E11"),
        ("Current Academic Week | الأسبوع الأكاديمي الحالي", str(academic_calendar_status["current_week"]), "B13", "C13"),
        ("Current Academic Event | الحدث الأكاديمي الحالي", academic_calendar_status["current_event"].title if academic_calendar_status["current_event"] else "Teaching In Session | الدراسة جارية", "D13", "E13"),
    ]
    for label, value, label_cell, value_cell in overview_cards:
        overview_ws[label_cell] = label
        overview_ws[value_cell] = value
        overview_ws[label_cell].fill = header_fill
        overview_ws[label_cell].font = header_font
        overview_ws[label_cell].alignment = centered
        overview_ws[value_cell].fill = soft_fill
        overview_ws[value_cell].font = Font(bold=True, color="1F2937")
        overview_ws[value_cell].alignment = wrapped
        overview_ws[label_cell].border = border
        overview_ws[value_cell].border = border

    overview_ws.column_dimensions["A"].width = 14
    overview_ws.column_dimensions["B"].width = 24
    overview_ws.column_dimensions["C"].width = 26
    overview_ws.column_dimensions["D"].width = 24
    overview_ws.column_dimensions["E"].width = 30
    overview_ws.row_dimensions[2].height = 28
    overview_ws.row_dimensions[3].height = 22

    plan_ws.merge_cells("A1:M1")
    plan_ws["A1"] = f"{level.name} | {level.teacher.full_name if level.teacher else '-'} | Syllabus Plan | خطة المنهج"
    plan_ws["A1"].fill = header_fill
    plan_ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    plan_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    plan_ws.row_dimensions[1].height = 26

    plan_ws.merge_cells("A2:M2")
    plan_ws["A2"] = (
        f"Academic Week {academic_calendar_status['current_week']} | "
        f"Expected {summary['expected_count']} | "
        f"Completed {summary['completed_count']} | "
        f"Delayed {summary['delayed_count']} | "
        f"Unscheduled {summary['unscheduled_count']}"
    )
    plan_ws["A2"].fill = soft_fill
    plan_ws["A2"].font = Font(bold=True, color="1F2937")
    plan_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    plan_ws.row_dimensions[2].height = 22

    plan_headers = [
        "Academic Week | الأسبوع الأكاديمي",
        "Teaching Week | الأسبوع التدريسي",
        "Date Range | المدة الزمنية",
        "Calendar Events | أحداث التقويم",
        "Slot | الحصة",
        "Book | الكتاب",
        "Unit | الوحدة",
        "Lesson | الدرس",
        "Reference | المرجع",
        "Learning Objective | هدف التعلم",
        "Planned Homework | الواجب المخطط",
        "Status | الحالة",
        "Completed On | تاريخ الإكمال",
    ]
    plan_ws.append(plan_headers)
    for cell in plan_ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = border

    row_cursor = 4
    for row in grid_rows:
        event_titles = " | ".join(event.get("title", "") for event in row["week_meta"].get("events", []))
        if not row["is_instructional_week"]:
            plan_ws.append(
                [
                    row["academic_week_number"],
                    "",
                    row["week_meta"].get("date_label") or "",
                    event_titles or "Non-Teaching Week | أسبوع غير تدريسي",
                    "",
                    "",
                    "",
                    "Non-Teaching Week | أسبوع غير تدريسي",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            excel_row = plan_ws.max_row
            for cell in plan_ws[excel_row]:
                cell.fill = holiday_fill
                cell.border = border
                cell.alignment = wrapped
            row_cursor += 1
            continue

        for session in row["sessions"]:
            entry = session["entry"]
            if not session["is_available_slot"]:
                plan_ws.append(
                    [
                        row["academic_week_number"],
                        row["week_number"],
                        row["week_meta"].get("date_label") or "",
                        event_titles,
                        session["session_number"],
                        "",
                        "",
                    "No Lesson | لا توجد حصة",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                excel_row = plan_ws.max_row
                for cell in plan_ws[excel_row]:
                    cell.fill = holiday_fill
                    cell.border = border
                    cell.alignment = wrapped
                row_cursor += 1
                continue

            plan_ws.append(
                [
                    row["academic_week_number"],
                    row["week_number"],
                    row["week_meta"].get("date_label") or "",
                    event_titles,
                    session["session_number"],
                    entry.book_name if entry else "",
                    entry.unit_name if entry else "",
                    entry.lesson_title if entry else "",
                    entry.source_reference if entry else "",
                    entry.learning_objective if entry else "",
                    entry.planned_homework if entry else "",
                    entry.status if entry else "",
                    entry.completed_on.strftime("%Y-%m-%d") if entry and entry.completed_on else "",
                ]
            )
            excel_row = plan_ws.max_row
            for cell in plan_ws[excel_row]:
                cell.border = border
                cell.alignment = wrapped
            if excel_row % 2 == 0:
                for cell in plan_ws[excel_row]:
                    if not cell.fill.fill_type:
                        cell.fill = PatternFill("solid", fgColor="FBFDFF")
            if row["week_meta"].get("is_current_week"):
                for cell in plan_ws[excel_row]:
                    cell.fill = current_fill
            if session.get("is_review"):
                for cell in plan_ws[excel_row]:
                    cell.fill = review_fill
            row_cursor += 1

    reserve_ws.append([
        "Book | الكتاب",
        "Unit | الوحدة",
        "Lesson | الدرس",
        "Reference | المرجع",
        "Learning Objective | هدف التعلم",
        "Planned Homework | الواجب المخطط",
    ])
    for cell in reserve_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = border
    for entry in summary["unscheduled_rows"]:
        reserve_ws.append(
            [
                entry.book_name or "",
                entry.unit_name or "",
                entry.lesson_title or "",
                entry.source_reference or "",
                entry.learning_objective or "",
                entry.planned_homework or "",
            ]
        )
        for cell in reserve_ws[reserve_ws.max_row]:
            cell.border = border
            cell.alignment = wrapped
            cell.fill = soft_fill

    width_map = {
        "A": 16, "B": 16, "C": 22, "D": 28, "E": 10, "F": 24, "G": 24, "H": 34, "I": 22, "J": 40, "K": 30, "L": 16, "M": 16
    }
    for column_letter, width in width_map.items():
        plan_ws.column_dimensions[column_letter].width = width
    for column_letter, width in {"A": 24, "B": 24, "C": 34, "D": 24, "E": 40, "F": 32}.items():
        reserve_ws.column_dimensions[column_letter].width = width

    reserve_ws.merge_cells("A1:F1")
    reserve_ws["A1"] = "Reserve / Unscheduled Rows | حصص احتياط / غير موزعة"
    reserve_ws["A1"].fill = header_fill
    reserve_ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    reserve_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    reserve_ws.row_dimensions[1].height = 24
    reserve_ws.insert_rows(2)
    reserve_ws["A2"] = "These rows did not fit into the current academic calendar slots. | هذه الصفوف لم تجد مكاناً داخل خانات التقويم الأكاديمي الحالية."
    reserve_ws.merge_cells("A2:F2")
    reserve_ws["A2"].fill = soft_fill
    reserve_ws["A2"].font = Font(bold=True, color="1F2937")
    reserve_ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in reserve_ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = border
    for row_idx in range(4, reserve_ws.max_row + 1):
        for cell in reserve_ws[row_idx]:
            cell.border = border
            cell.alignment = wrapped
            if row_idx % 2 == 0:
                cell.fill = soft_fill

    plan_ws.freeze_panes = "A4"
    reserve_ws.freeze_panes = "A4"
    overview_ws.freeze_panes = "A2"
    for sheet in (plan_ws, reserve_ws):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue
                cell.alignment = wrapped
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_syllabus_export_filename(level: Level, extension: str, variant: str | None = None) -> str:
    safe_level_name = "".join(char if char.isalnum() else "_" for char in (level.name or "").lower()).strip("_") or "level"
    teacher_name = level.teacher.full_name if level.teacher else "unassigned_teacher"
    safe_teacher_name = "".join(char if char.isalnum() else "_" for char in teacher_name.lower()).strip("_") or "teacher"
    stamp = date.today().strftime("%Y%m%d")
    variant_suffix = f"_{variant}" if variant else ""
    return f"hikmah_syllabus{variant_suffix}_{safe_level_name}_{safe_teacher_name}_{stamp}.{extension}"


def build_syllabus_plan_pdf(level: Level) -> bytes:
    if not REPORTLAB_AVAILABLE:
        rows = (
            SyllabusPlanEntry.query.filter_by(level_id=level.id)
            .order_by(SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
            .all()
        )
        lines = [
            f"Hikmah Academy - Syllabus Plan",
            f"Level: {level.name}",
            f"Teacher: {level.teacher.full_name if level.teacher else '-'}",
            "",
        ]
        for row in rows:
            location = (
                f"Academic Week {row.week_number} / Slot {row.session_number}"
                if row.week_number and row.session_number
                else "Unscheduled"
            )
            lines.append(f"{location} | {row.unit_name or '-'} | {row.lesson_title}")
            if row.source_reference:
                lines.append(f"Ref: {row.source_reference}")
            if row.learning_objective:
                lines.append(f"Objective: {row.learning_objective}")
            if row.planned_homework:
                lines.append(f"Homework: {row.planned_homework}")
            lines.append("")
        return build_simple_pdf(lines)

    def contains_arabic(text: str) -> bool:
        return bool(re.search(r"[\u0600-\u06FF]", text or ""))

    def pdf_text(text: str, lang: str) -> str:
        raw = str(text or "")
        if lang == "ar" or contains_arabic(raw):
            try:
                return get_display(arabic_reshaper.reshape(raw))
            except Exception:
                return raw
        return raw

    def pdf_para(text: str, style: ParagraphStyle, lang: str) -> Paragraph:
        safe = pdf_text(text, lang).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
        return Paragraph(safe, style)

    def register_pdf_fonts() -> tuple[str, str, str, str]:
        regular_candidates = [
            "/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        bold_candidates = [
            "/usr/share/fonts/truetype/noto/NotoKufiArabic-Bold.ttf",
            "/usr/share/fonts/truetype/noto/NotoKufiArabic-SemiBold.ttf",
            "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf",
            "/usr/share/fonts/truetype/noto/NotoSansArabic-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ]
        regular_font = "Helvetica"
        bold_font = "Helvetica-Bold"

        for path in regular_candidates:
            if os.path.exists(path):
                regular_font = "HikmahPdfRegular"
                if regular_font not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(regular_font, path))
                break

        for path in bold_candidates:
            if os.path.exists(path):
                bold_font = "HikmahPdfBold"
                if bold_font not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(bold_font, path))
                break

        latin_regular = "Helvetica"
        latin_bold = "Helvetica-Bold"
        if os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):
            latin_regular = "HikmahPdfLatinRegular"
            if latin_regular not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(latin_regular, "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
        if os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"):
            latin_bold = "HikmahPdfLatinBold"
            if latin_bold not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(latin_bold, "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"))

        return regular_font, bold_font, latin_regular, latin_bold

    current_lang = get_current_ui_language()
    current_dir = "rtl" if current_lang == "ar" else "ltr"
    regular_font, bold_font, latin_regular_font, latin_bold_font = register_pdf_fonts()
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        "HikmahPdfBody",
        parent=styles["BodyText"],
        fontName=regular_font,
        fontSize=10,
        leading=14,
        alignment=TA_RIGHT if current_dir == "rtl" else TA_LEFT,
        textColor=colors.HexColor("#173042"),
    )
    small_style = ParagraphStyle(
        "HikmahPdfSmall",
        parent=body_style,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#5B7083"),
    )
    title_style = ParagraphStyle(
        "HikmahPdfTitle",
        parent=body_style,
        fontName=bold_font,
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#0F766E"),
    )
    subtitle_style = ParagraphStyle(
        "HikmahPdfSubtitle",
        parent=body_style,
        fontName=bold_font,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#102A43"),
    )
    header_style = ParagraphStyle(
        "HikmahPdfHeader",
        parent=body_style,
        fontName=bold_font,
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=TA_CENTER if current_dir == "ltr" else TA_RIGHT,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Hikmah Academy - {level.name} Syllabus Plan",
    )

    story = []
    logo_path = os.path.join(app.static_folder or "static", "academy-logo.jpg")
    logo = RLImage(logo_path, width=26 * mm, height=26 * mm) if os.path.exists(logo_path) else ""

    teacher_label = "Teacher" if current_lang == "en" else "المعلم"
    level_label = "Level" if current_lang == "en" else "الفصل"
    week_label = "Current Academic Week" if current_lang == "en" else "الأسبوع الأكاديمي الحالي"
    progress_label = "Progress" if current_lang == "en" else "نسبة الإنجاز"
    completed_label = "Completed" if current_lang == "en" else "المكتمل"
    delayed_label = "Delayed" if current_lang == "en" else "المتأخر"

    title_lines = [
        pdf_para("Hikmah Academy", title_style, "en"),
        pdf_para("Arabic Department Syllabus Plan", subtitle_style, "en"),
        pdf_para("خطة المنهج لقسم اللغة العربية", subtitle_style, "ar"),
        Spacer(1, 2),
    ]
    hero = Table([[logo, title_lines]], colWidths=[30 * mm, 235 * mm])
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4FBFB")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#B8E0DE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend([hero, Spacer(1, 8)])

    plan_summary_rows = build_syllabus_plan_summary([level])
    plan_summary = plan_summary_rows[0] if plan_summary_rows else {
        "progress_pct": 0.0,
        "completed_count": 0,
        "total_rows": 0,
        "delayed_count": 0,
    }
    academic_status = build_academic_calendar_status()
    teacher_name_style = ParagraphStyle(
        "HikmahPdfTeacherName",
        parent=body_style,
        fontName=bold_font,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0B4F79"),
    )
    latin_value_style = ParagraphStyle(
        "HikmahPdfLatinValue",
        parent=body_style,
        fontName=latin_regular_font,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#173042"),
    )
    latin_bold_value_style = ParagraphStyle(
        "HikmahPdfLatinBoldValue",
        parent=body_style,
        fontName=latin_bold_font,
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#0B4F79"),
    )

    def value_style_for(text: str, emphasize: bool = False) -> ParagraphStyle:
        if contains_arabic(text):
            return teacher_name_style if emphasize else body_style
        return latin_bold_value_style if emphasize else latin_value_style

    detail_rows = [
        [pdf_para(level_label, subtitle_style, current_lang), pdf_para(level.name, value_style_for(level.name), current_lang)],
        [pdf_para(teacher_label, subtitle_style, current_lang), pdf_para(level.teacher.full_name if level.teacher else "-", value_style_for(level.teacher.full_name if level.teacher else "-", emphasize=True), current_lang)],
        [pdf_para(week_label, subtitle_style, current_lang), pdf_para(str(academic_status.get("current_week") or "-"), body_style, current_lang)],
        [pdf_para(progress_label, subtitle_style, current_lang), pdf_para(f"{plan_summary['progress_pct']}%", body_style, current_lang)],
        [pdf_para(completed_label, subtitle_style, current_lang), pdf_para(f"{plan_summary['completed_count']} / {plan_summary['total_rows']}", body_style, current_lang)],
        [pdf_para(delayed_label, subtitle_style, current_lang), pdf_para(str(plan_summary["delayed_count"]), body_style, current_lang)],
        [pdf_para("Unscheduled" if current_lang == "en" else "غير مجدول", subtitle_style, current_lang), pdf_para(str(plan_summary.get("unscheduled_count", 0)), body_style, current_lang)],
    ]
    details_table = Table(detail_rows, colWidths=[72 * mm, 78 * mm])
    details_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F9FCFD"), colors.HexColor("#EEF7F8")]),
                ("BOX", (0, 0), (-1, -1), 0.9, colors.HexColor("#C9E3E4")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#D8EDEE")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([details_table, Spacer(1, 10)])

    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    grid_rows = build_syllabus_plan_grid(level, total_weeks=40, lessons_per_week=lessons_per_week)
    table_data: list[list] = [[pdf_para("Academic Week", header_style, "en")]]
    for session_number in range(1, lessons_per_week + 1):
        table_data[0].append(pdf_para(f"Lesson {session_number}", header_style, "en"))
    table_style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#BFD9DD")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8E7EA")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
    ]

    for idx, row in enumerate(grid_rows, start=1):
        week_parts = [f"Academic Week {row['academic_week_number']}"]
        if row.get("week_number"):
            week_parts.append(f"Teaching Week {row['week_number']}")
        if row["week_meta"].get("date_label"):
            week_parts.append(row["week_meta"]["date_label"])
        if row["week_meta"].get("events"):
            week_parts.extend(event.get("title", "-") for event in row["week_meta"]["events"])
        week_cell = pdf_para("<br/>".join(week_parts), small_style, current_lang)

        if not row["is_instructional_week"]:
            non_teaching_row = [week_cell, pdf_para("Non-Teaching Week / أسبوع غير تدريسي", body_style, current_lang)]
            non_teaching_row.extend("" for _ in range(max(lessons_per_week - 1, 0)))
            table_data.append(non_teaching_row)
            table_style_commands.extend(
                [
                    ("SPAN", (1, idx), (lessons_per_week, idx)),
                    ("BACKGROUND", (0, idx), (lessons_per_week, idx), colors.HexColor("#F4F6F8")),
                ]
            )
            continue

        event_type = row["week_meta"].get("dominant_event_type") or ""
        if event_type == "exam":
            row_bg = colors.HexColor("#FFF4E8")
        elif event_type == "holiday":
            row_bg = colors.HexColor("#EEF2FF")
        elif event_type in {"school_resume", "term_start"}:
            row_bg = colors.HexColor("#EAFBF3")
        else:
            row_bg = colors.white

        row_cells = [week_cell]
        for session in row["sessions"]:
            entry = session["entry"]
            if not session["is_available_slot"]:
                cell = pdf_para("No Lesson", small_style, "en")
            elif not entry:
                cell = pdf_para("Empty Slot", small_style, "en")
            else:
                pieces = [
                    f"<b>{pdf_text(entry.unit_name or '-', current_lang)}</b>",
                    pdf_text(entry.lesson_title or "-", current_lang),
                ]
                if entry.source_reference:
                    pieces.append(pdf_text(entry.source_reference, current_lang))
                if entry.planned_homework:
                    pieces.append(pdf_text(entry.planned_homework, current_lang))
                cell = Paragraph("<br/>".join(pieces), body_style)
            row_cells.append(cell)
        table_data.append(row_cells)
        table_style_commands.append(("BACKGROUND", (0, idx), (lessons_per_week, idx), row_bg))
        if row["week_meta"].get("is_current_week"):
            table_style_commands.append(("LINEBELOW", (0, idx), (lessons_per_week, idx), 1.2, colors.HexColor("#0F766E")))

    lesson_col_width = (232 / max(lessons_per_week, 1)) * mm
    plan_table = Table(table_data, colWidths=[34 * mm] + [lesson_col_width for _ in range(lessons_per_week)], repeatRows=1)
    plan_table.setStyle(TableStyle(table_style_commands))
    story.append(plan_table)

    unscheduled_rows = plan_summary.get("unscheduled_rows") or []
    if unscheduled_rows:
        story.extend([Spacer(1, 10), pdf_para("Reserve / Unscheduled Rows" if current_lang == "en" else "حصص احتياط / غير موزعة بعد", subtitle_style, current_lang)])
        reserve_data = [[
            pdf_para("#", header_style, "en"),
            pdf_para("Book", header_style, "en"),
            pdf_para("Unit", header_style, "en"),
            pdf_para("Lesson", header_style, "en"),
        ]]
        for index, entry in enumerate(unscheduled_rows, start=1):
            reserve_data.append([
                pdf_para(str(index), small_style, "en"),
                pdf_para(entry.book_name or "-", body_style, current_lang),
                pdf_para(entry.unit_name or "-", body_style, current_lang),
                pdf_para(entry.lesson_title or "-", body_style, current_lang),
            ])
        reserve_table = Table(reserve_data, colWidths=[12 * mm, 45 * mm, 52 * mm, 150 * mm], repeatRows=1)
        reserve_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#BFD9DD")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8E7EA")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.extend([Spacer(1, 6), reserve_table])
    doc.build(story)
    return buffer.getvalue()


def build_curriculum_progress_summary(levels: list[Level] | None = None) -> list[dict]:
    target_levels = levels or Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    level_ids = [level.id for level in target_levels]
    if not level_ids:
        return []

    subjects = Subject.query.filter(Subject.level_id.in_(level_ids), Subject.is_active.is_(True)).all()
    progress_rows = CurriculumProgress.query.filter(CurriculumProgress.level_id.in_(level_ids)).all()
    progress_by_key = {
        (row.level_id, row.curriculum_item_id): row
        for row in progress_rows
    }

    summaries = []
    for level in target_levels:
        level_subjects = [
            subject for subject in subjects
            if subject.level_id == level.id and not is_teacher_resource_subject(subject.name)
        ]
        items = []
        for subject in level_subjects:
            items.extend(sorted(subject.curriculum_items, key=lambda item: ((item.order_index or 0), item.id)))

        completed_count = 0
        in_progress_count = 0
        pending_count = 0
        last_completed_title = ""
        last_completed_on = None

        for item in items:
            progress = progress_by_key.get((level.id, item.id))
            status = (progress.status if progress else "pending").strip().lower()
            if status == "completed":
                completed_count += 1
                if progress and progress.completed_on and (last_completed_on is None or progress.completed_on >= last_completed_on):
                    last_completed_on = progress.completed_on
                    last_completed_title = item.title
            elif status == "in_progress":
                in_progress_count += 1
            else:
                pending_count += 1

        total_items = len(items)
        if total_items == 0:
            continue
        progress_pct = round((completed_count / total_items) * 100, 1) if total_items else 0.0
        summaries.append(
            {
                "level": level,
                "teacher": level.teacher,
                "total_items": total_items,
                "completed_count": completed_count,
                "in_progress_count": in_progress_count,
                "pending_count": pending_count,
                "progress_pct": progress_pct,
                "last_completed_title": last_completed_title,
                "last_completed_on": last_completed_on,
            }
        )

    summaries.sort(key=lambda row: (row["progress_pct"], row["level"].order_index or 0, row["level"].name))
    return summaries


def sync_students_to_levels() -> None:
    levels_by_name = {normalize_level_display_name(level.name): level for level in Level.query.all()}
    students = Student.query.all()

    for student in students:
        student_display_name = normalize_level_display_name(student.level_name)
        level = levels_by_name.get(student_display_name)
        if level:
            student.level_id = level.id

    db.session.commit()


def seed_teachers_and_levels() -> None:
    teacher_level_pairs = [
        ("Mr.Alnoabani", "Qaeda", 1, "Year5b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/3024053210", "302 405 3210", "Year5@alhikmah"),
        ("Mr.Ammar", "Primary Beginner", 2, "Year6b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/9457642689", "945 764 2689", "Year6@alhikmah"),
        ("Mahfouz Ahmed", "Primary Intermediate", 3, "Year7b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/6784224657", "678 422 4657", "Year7@alhikmah"),
        ("Mr.Yahya Lediju", "Primary Advance", 4, "Year9b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/5535535595", "553 553 5595", "Year9@alhikmah"),
        ("Mr. Arafat", "Secondary Beginner 1", 5, "Year10b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/2228442286", "222 844 2286", "Year10@alhikmah"),
        ("Ahmed Khalid", "Secondary Beginner 2", 6, "Year8b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/j/8913798586", "891 379 8586", "Year8@alhikmah"),
        ("Hassan", "Secondary Intermediate", 7, "secondary.intermediate.b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/my/sec.ar.intermediate.b", "348 151 0931", "Teacher@123"),
        ("Mr.Yahya Mohsen", "Secondary Advance", 8, "sec.advance.b@alhikmahkidsacademy.com", "https://alhikmahkidsacademy.zoom.us/my/sec.ar.advance.b", "333 143 2774", "Teacher@123"),
    ]

    for teacher_name, level_name, order_index, zoom_email, zoom_link, zoom_meeting_id, zoom_passcode in teacher_level_pairs:
        teacher = Teacher.query.filter_by(full_name=teacher_name).first()
        if not teacher:
            teacher = Teacher(full_name=teacher_name, status="Active")
            db.session.add(teacher)
            db.session.flush()

        teacher.level_name = level_name

        level = Level.query.filter_by(name=level_name).first()
        if not level:
            for legacy_name, normalized_name in LEVEL_LEGACY_NAME_TO_DISPLAY.items():
                if normalized_name == level_name:
                    level = Level.query.filter_by(name=legacy_name).first()
                    if level:
                        level.name = level_name
                        break
        if not level:
            level = Level(
                name=level_name,
                teacher_id=teacher.id,
                order_index=order_index,
                zoom_email=zoom_email,
                zoom_link=zoom_link,
                zoom_meeting_id=zoom_meeting_id,
                zoom_passcode=zoom_passcode,
            )
            db.session.add(level)
        else:
            level.teacher_id = teacher.id
            level.order_index = order_index
            level.zoom_email = zoom_email
            level.zoom_link = zoom_link
            level.zoom_meeting_id = zoom_meeting_id
            level.zoom_passcode = zoom_passcode

    # Future-ready level with no mandatory teacher assignment.
    igcse_level = Level.query.filter_by(name="iGCSE").first()
    if not igcse_level:
        db.session.add(Level(name="iGCSE", order_index=9))
    else:
        if igcse_level.order_index is None:
            igcse_level.order_index = 9

    db.session.commit()


with app.app_context():
    db.create_all()
    ensure_settings_schema()
    ensure_teacher_schema()
    ensure_level_schema()
    ensure_student_schema()
    ensure_attendance_schema()
    ensure_recordings_schema()
    ensure_curriculum_items_schema()
    ensure_syllabus_plan_schema()
    seed_default_academic_calendar()
    seed_teachers_and_levels()
    sync_students_to_levels()


@app.route("/")
def dashboard():
    total_students = Student.query.count()
    total_levels = db.session.query(Student.level_name).distinct().count()
    total_teachers = Teacher.query.count()

    return render_template(
        "dashboard.html",
        total_students=total_students,
        total_levels=total_levels,
        total_teachers=total_teachers,
    )


@app.route("/healthz")
def healthz():
    return Response("ok", mimetype="text/plain")


@app.route("/student/dashboard")
def student_dashboard():
    student_code = (request.args.get("student_code") or "").strip()
    selected_student_exam_title = (request.args.get("exam_title") or "").strip()
    student = None
    level = None
    recordings = []
    exam_results = []
    exam_summary_groups = []
    student_exam_titles = []
    selected_student_exam_summary = None
    upcoming_exams = []
    assignment_rows = []
    subject_rows = []
    book_subject_rows = []
    extra_resource_subject_rows = []
    student_dashboard_summary = None
    student_announcements = []
    if student_code:
        student = Student.query.filter_by(student_code=student_code).first()
        if student:
            level = student.level
            if not level and student.level_name:
                level = Level.query.filter_by(name=normalize_level_display_name(student.level_name)).first()
            if level:
                recordings = (
                    ClassRecording.query.filter_by(class_id=level.id)
                    .order_by(ClassRecording.lesson_date.desc(), ClassRecording.id.desc())
                    .all()
                )
                upcoming_exams = (
                    UpcomingExam.query.filter_by(level_id=level.id)
                    .order_by(UpcomingExam.exam_date.asc(), UpcomingExam.id.asc())
                    .all()
                )
                assignments = (
                    Assignment.query.filter_by(level_id=level.id, is_active=True)
                    .order_by(Assignment.due_date.asc(), Assignment.id.desc())
                    .all()
                )
                subjects = (
                    Subject.query.filter_by(level_id=level.id, is_active=True)
                    .order_by(Subject.order_index.asc(), Subject.name.asc())
                    .all()
                )
                subject_rows = [
                    {
                        "subject": subject,
                        "curriculum_items": sorted(
                            [
                                item
                                for item in subject.curriculum_items
                                if (item.visibility_scope or "student_and_teacher") == "student_and_teacher"
                            ],
                            key=lambda item: ((item.order_index or 0), item.id),
                        ),
                    }
                    for subject in subjects
                ]
                book_subject_rows = [row for row in subject_rows if not is_teacher_resource_subject(row["subject"].name)]
                extra_resource_subject_rows = [row for row in subject_rows if is_teacher_resource_subject(row["subject"].name)]
                submissions = AssignmentSubmission.query.filter_by(student_id=student.id).all()
                submissions_by_assignment = {submission.assignment_id: submission for submission in submissions}
                assignment_rows = [
                    {
                        "assignment": assignment,
                        "submission": submissions_by_assignment.get(assignment.id),
                    }
                    for assignment in assignments
                ]
            exam_results = (
                ExamResult.query.filter_by(student_id=student.id)
                .order_by(ExamResult.exam_date.desc(), ExamResult.id.desc())
                .all()
            )
            exam_results = filter_exam_results_for_student(student, exam_results)
            exam_summary_groups = build_exam_summary_groups(exam_results)
            student_exam_titles = [exam_summary["exam_title"] for exam_summary in exam_summary_groups]
            if exam_summary_groups:
                if not selected_student_exam_title:
                    selected_student_exam_title = exam_summary_groups[0]["exam_title"]
                selected_student_exam_summary = next(
                    (
                        exam_summary
                        for exam_summary in exam_summary_groups
                        if exam_summary["exam_title"] == selected_student_exam_title
                    ),
                    None,
                )
            student_dashboard_summary = build_student_dashboard_summary(
                student=student,
                level=level,
                assignment_rows=assignment_rows,
                exam_summary_groups=exam_summary_groups,
                upcoming_exams=upcoming_exams,
                recordings=recordings,
            )
            student_announcements = build_announcement_view_rows(
                get_active_announcements("students", level.id if level else student.level_id)
            )
    return render_template(
        "student_dashboard.html",
        student=student,
        student_code=student_code,
        level=level,
        recordings=recordings,
        subject_rows=subject_rows,
        book_subject_rows=book_subject_rows,
        extra_resource_subject_rows=extra_resource_subject_rows,
        assignment_rows=assignment_rows,
        exam_results=exam_results,
        exam_summary_groups=exam_summary_groups,
        student_exam_titles=student_exam_titles,
        selected_student_exam_title=selected_student_exam_title,
        selected_student_exam_summary=selected_student_exam_summary,
        upcoming_exams=upcoming_exams,
        student_dashboard_summary=student_dashboard_summary,
        student_announcements=student_announcements,
        academic_calendar_status=(student_dashboard_summary["academic_calendar_status"] if student_dashboard_summary else build_academic_calendar_status()),
    )


@app.route("/students/<int:student_id>/exam-report.pdf")
def student_exam_pdf(student_id: int):
    student = Student.query.get_or_404(student_id)
    student_code = (request.args.get("student_code") or "").strip()
    if student_code != student.student_code:
        abort(403)

    exam_title = (request.args.get("exam_title") or "").strip()
    if not exam_title or not is_exam_visible_to_student(student, exam_title):
        abort(403)

    exam_results = (
        ExamResult.query.filter_by(student_id=student.id, exam_title=exam_title)
        .order_by(ExamResult.id.asc())
        .all()
    )
    exam_summary_groups = build_exam_summary_groups(exam_results)
    if not exam_summary_groups:
        abort(404)

    exam_summary = exam_summary_groups[0]
    level_name = student.level.name if student.level else student.level_name
    pdf_bytes = build_exam_report_pdf("Hikmah Academy Exam Report", student, level_name, exam_summary)
    safe_exam_title = re.sub(r"[^A-Za-z0-9_-]+", "_", exam_title).strip("_") or "exam_report"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{student.student_code}_{safe_exam_title}.pdf"'
        },
    )


@app.route("/teacher/students/<int:student_id>/exam-report")
@teacher_required
def teacher_student_exam_report(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not student.level_id:
        abort(403)
    level = get_teacher_level_or_403(student.level_id)

    exam_title = (request.args.get("exam_title") or "").strip()
    if not exam_title:
        abort(400)

    exam_results = (
        ExamResult.query.filter_by(student_id=student.id, exam_title=exam_title)
        .order_by(ExamResult.id.asc())
        .all()
    )
    exam_summary = get_exam_summary_for_title(exam_results, exam_title)
    if not exam_summary:
        abort(404)

    return render_template(
        "teacher_exam_report.html",
        student=student,
        level=level,
        exam_summary=exam_summary,
    )


@app.route("/teacher/students/<int:student_id>/exam-report.pdf")
@teacher_required
def teacher_student_exam_pdf(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not student.level_id:
        abort(403)
    get_teacher_level_or_403(student.level_id)

    exam_title = (request.args.get("exam_title") or "").strip()
    if not exam_title:
        abort(400)

    exam_results = (
        ExamResult.query.filter_by(student_id=student.id, exam_title=exam_title)
        .order_by(ExamResult.id.asc())
        .all()
    )
    exam_summary = get_exam_summary_for_title(exam_results, exam_title)
    if not exam_summary:
        abort(404)

    level_name = student.level.name if student.level else student.level_name
    pdf_bytes = build_exam_report_pdf("Hikmah Academy Teacher Exam Report", student, level_name, exam_summary)
    safe_exam_title = re.sub(r"[^A-Za-z0-9_-]+", "_", exam_title).strip("_") or "exam_report"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="teacher_{student.student_code}_{safe_exam_title}.pdf"'
        },
    )


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    error = ""
    next_url = request.args.get("next") or request.form.get("next") or "/teacher/dashboard"
    teachers = Teacher.query.order_by(Teacher.full_name.asc()).all()
    teacher_code_required = teacher_access_code_enabled()
    if request.method == "POST":
        teacher_id_str = (request.form.get("teacher_id") or "").strip()
        teacher_access_code = (request.form.get("teacher_access_code") or "").strip()
        if teacher_code_required and teacher_access_code != get_teacher_access_code():
            error = "Invalid teacher access code."
            return render_template(
                "teacher_login.html",
                error=error,
                next_url=next_url,
                teachers=teachers,
                teacher_code_required=teacher_code_required,
            )
        if teacher_id_str.isdigit():
            teacher = Teacher.query.get(int(teacher_id_str))
            if teacher:
                session["teacher_id"] = teacher.id
                session["teacher_name"] = teacher.full_name
                if not is_safe_next(next_url):
                    next_url = "/teacher/dashboard"
                return redirect(next_url)
        error = "Invalid teacher selection."
    return render_template(
        "teacher_login.html",
        error=error,
        next_url=next_url,
        teachers=teachers,
        teacher_code_required=teacher_code_required,
    )


@app.route("/teacher/logout")
def teacher_logout():
    session.pop("teacher_id", None)
    session.pop("teacher_name", None)
    return redirect(url_for("dashboard"))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = ""
    next_url = request.args.get("next") or request.form.get("next") or "/admin/dashboard"
    if request.method == "POST":
        code = (request.form.get("access_code") or "").strip()
        if code == get_admin_access_code():
            session["is_admin"] = True
            if not is_safe_next(next_url):
                next_url = "/admin/dashboard"
            return redirect(next_url)
        error = "Invalid access code."
    return render_template("admin_login.html", error=error, next_url=next_url)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("dashboard"))


@app.route("/head/dashboard")
@admin_required
def head_dashboard():
    requested_lang = (request.args.get("lang") or "").strip().lower()
    if requested_lang in {"ar", "en"}:
        session["head_dashboard_lang"] = requested_lang
    current_lang = session.get("head_dashboard_lang", "en")
    ui = get_head_dashboard_copy(current_lang)
    total_students = Student.query.count()
    total_levels = Level.query.count()
    total_teachers = Teacher.query.count()

    level_rows = (
        db.session.query(
            Level.id,
            Level.name,
            Level.zoom_link,
            Teacher.full_name.label("teacher_name"),
            db.func.count(Student.id).label("students_count"),
        )
        .outerjoin(Teacher, Level.teacher_id == Teacher.id)
        .outerjoin(Student, Student.level_id == Level.id)
        .group_by(Level.id, Level.name, Level.zoom_link, Teacher.full_name)
        .order_by(Level.order_index.asc(), Level.name.asc())
        .all()
    )

    exam_templates = ExamTemplate.query.order_by(ExamTemplate.title.asc()).all()
    exam_template_rows = [
        {
            "template": template,
            "branch_text": get_template_branch_text(template),
        }
        for template in exam_templates
    ]
    exam_titles = get_all_exam_titles()
    exam_publications = {
        publication.exam_title: publication
        for publication in ExamPublication.query.filter(ExamPublication.exam_title.in_(exam_titles)).all()
    }
    students = Student.query.order_by(Student.full_name.asc()).all()
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    subjects = Subject.query.order_by(Subject.level_id.asc(), Subject.order_index.asc(), Subject.name.asc()).all()
    curriculum_items = CurriculumItem.query.order_by(
        CurriculumItem.subject_id.asc(),
        CurriculumItem.order_index.asc(),
        CurriculumItem.id.asc(),
    ).all()
    curriculum_items_by_subject = {}
    for item in curriculum_items:
        curriculum_items_by_subject.setdefault(item.subject_id, []).append(item)
    subject_rows = [
        {
            "subject": subject,
            "curriculum_items": curriculum_items_by_subject.get(subject.id, []),
        }
        for subject in subjects
    ]
    book_subject_rows = [row for row in subject_rows if not is_teacher_resource_subject(row["subject"].name)]
    teacher_resource_subject_rows = [row for row in subject_rows if is_teacher_resource_subject(row["subject"].name)]
    cleanup_data = build_cleanup_center_data(levels, Teacher.query.order_by(Teacher.full_name.asc()).all(), students)
    head_snapshot = {
        "active_templates": sum(1 for template in exam_templates if template.is_active),
        "published_exams": sum(1 for publication in exam_publications.values() if publication.is_published),
        "books_count": sum(len(row["curriculum_items"]) for row in book_subject_rows),
        "teacher_resources_count": sum(len(row["curriculum_items"]) for row in teacher_resource_subject_rows),
        "levels_missing_zoom": sum(1 for row in level_rows if not row.zoom_link),
        "cleanup_critical": cleanup_data["cleanup_summary"]["critical_items"],
    }

    return render_template(
        "head_dashboard.html",
        ui=ui,
        current_lang=current_lang,
        total_students=total_students,
        total_levels=total_levels,
        total_teachers=total_teachers,
        level_rows=level_rows,
        exam_templates=exam_templates,
        exam_template_rows=exam_template_rows,
        exam_publications=exam_publications,
        students=students,
        levels=levels,
        subject_rows=subject_rows,
        book_subject_rows=book_subject_rows,
        teacher_resource_subject_rows=teacher_resource_subject_rows,
        head_snapshot=head_snapshot,
        exam_titles=exam_titles,
        exam_import_status=(request.args.get("exam_import_status") or "").strip(),
        exam_import_message=(request.args.get("exam_import_message") or "").strip(),
    )


@app.route("/head/exam-results/import", methods=["POST"])
@admin_required
def import_exam_results():
    uploaded_file = request.files.get("exam_sheet")
    if not uploaded_file or not uploaded_file.filename:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Please choose an Excel file first.",
            )
        )

    if not uploaded_file.filename.lower().endswith((".xlsx", ".xls")):
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Only Excel files (.xlsx or .xls) are supported.",
            )
        )

    try:
        saved_file_path = save_uploaded_exam_file(uploaded_file)
        with open(saved_file_path, "rb") as saved_file:
            parsed_rows, row_errors = read_excel_exam_rows(saved_file)
    except Exception as exc:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message=f"Could not read the Excel file: {exc}",
            )
        )

    imported_count, skipped_count, import_errors = import_exam_rows(
        parsed_rows,
        os.path.basename(saved_file_path),
    )
    row_errors.extend(import_errors)

    db.session.commit()

    summary_parts = [f"Imported or updated {imported_count} exam result(s)."]
    if skipped_count:
        summary_parts.append(f"Skipped {skipped_count} row(s).")
        summary_parts.append("Open Name Matching to resolve unmatched students.")
    if row_errors:
        summary_parts.append("Issues: " + " | ".join(row_errors[:5]))

    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success" if imported_count else "error",
            exam_import_message=" ".join(summary_parts),
        )
    )


@app.route("/head/exam-templates/add", methods=["POST"])
@admin_required
def add_exam_template():
    title = (request.form.get("title") or "").strip()
    exam_date_str = (request.form.get("exam_date") or "").strip()
    branches_text = (request.form.get("branches_text") or "").strip()
    if not title:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Exam template title is required.",
            )
        )

    template = ExamTemplate.query.filter_by(title=title).first()
    if not template:
        template = ExamTemplate(title=title, is_active=True)
        db.session.add(template)
        db.session.flush()

    if exam_date_str:
        try:
            template.exam_date = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect(
                url_for(
                    "head_dashboard",
                    exam_import_status="error",
                    exam_import_message="Invalid exam template date format.",
                )
            )
    else:
        template.exam_date = None

    branch_rows = parse_template_branches_input(branches_text)
    sync_exam_template_branches(template, branch_rows)
    ensure_exam_publication_row(title, default_published=False)
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Exam template '{title}' saved successfully.",
        )
    )


@app.route("/head/exam-templates/<int:template_id>/update", methods=["POST"])
@admin_required
def update_exam_template(template_id: int):
    template = ExamTemplate.query.get_or_404(template_id)
    title = (request.form.get("title") or "").strip()
    exam_date_str = (request.form.get("exam_date") or "").strip()
    branches_text = (request.form.get("branches_text") or "").strip()
    is_active = (request.form.get("is_active") or "").strip().lower() in {"1", "true", "on", "yes"}
    if not title:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Exam template title is required.",
            )
        )

    old_title = template.title
    template.title = title
    template.is_active = is_active
    if exam_date_str:
        try:
            template.exam_date = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect(
                url_for(
                    "head_dashboard",
                    exam_import_status="error",
                    exam_import_message="Invalid exam template date format.",
                )
            )
    else:
        template.exam_date = None

    sync_exam_template_branches(template, parse_template_branches_input(branches_text))

    if old_title != title:
        publication = ExamPublication.query.filter_by(exam_title=old_title).first()
        if publication:
            publication.exam_title = title
        StudentExamVisibility.query.filter_by(exam_title=old_title).update({"exam_title": title})
        ExamResult.query.filter_by(exam_title=old_title).update({"exam_title": title})

    ensure_exam_publication_row(title, default_published=False)
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Exam template '{title}' updated successfully.",
        )
    )


@app.route("/head/exam-publications/update", methods=["POST"])
@admin_required
def update_exam_publication():
    exam_title = (request.form.get("exam_title") or "").strip()
    if not exam_title:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Exam title is required for publication control.",
            )
        )

    publication = ExamPublication.query.filter_by(exam_title=exam_title).first()
    if not publication:
        publication = ExamPublication(exam_title=exam_title)
        db.session.add(publication)

    publication.is_published = (request.form.get("is_published") or "").strip().lower() in {"1", "true", "on", "yes"}
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Updated publication status for '{exam_title}'.",
        )
    )


@app.route("/head/student-exam-visibility/update", methods=["POST"])
@admin_required
def update_student_exam_visibility():
    student_id = (request.form.get("student_id") or "").strip()
    exam_title = (request.form.get("exam_title") or "").strip()
    if not student_id.isdigit() or not exam_title:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Student and exam title are required for individual visibility.",
            )
        )

    student = Student.query.get_or_404(int(student_id))
    hide_result = (request.form.get("visibility_mode") or "").strip() == "hide"
    set_student_exam_visibility_override(student.id, exam_title, hide_result)
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Updated individual visibility for '{student.full_name}' on '{exam_title}'.",
        )
    )


@app.route("/head/subjects/add", methods=["POST"])
@admin_required
def add_subject():
    level_id_str = (request.form.get("level_id") or "").strip()
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()
    order_index_str = (request.form.get("order_index") or "").strip()

    if not level_id_str.isdigit() or not name:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Level and subject name are required.",
            )
        )

    order_index = int(order_index_str) if order_index_str.isdigit() else None
    db.session.add(
        Subject(
            level_id=int(level_id_str),
            name=name,
            description=description or None,
            order_index=order_index,
            is_active=True,
        )
    )
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Subject '{name}' added successfully.",
        )
    )


@app.route("/head/curriculum-items/add", methods=["POST"])
@admin_required
def add_curriculum_item():
    level_id_str = (request.form.get("level_id") or "").strip()
    item_group = (request.form.get("item_group") or "").strip().lower()
    title = (request.form.get("title") or "").strip()
    description = (request.form.get("description") or "").strip()
    resource_link = (request.form.get("resource_link") or "").strip()
    visibility_scope = (request.form.get("visibility_scope") or "student_and_teacher").strip().lower()
    order_index_str = (request.form.get("order_index") or "").strip()
    apply_to_all_levels = (request.form.get("apply_to_all_levels") or "").strip().lower() in {"1", "true", "on", "yes"}

    if visibility_scope not in {"student_and_teacher", "teacher_only"}:
        visibility_scope = "student_and_teacher"

    if item_group == "teacher_resources":
        subject_name = "Teacher Resources"
    else:
        subject_name = "Books"

    if not level_id_str.isdigit() or not title or not resource_link:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Level, link title, and drive link are required.",
            )
        )

    target_level_ids = []
    if apply_to_all_levels:
        target_level_ids = [level.id for level in Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()]
    else:
        target_level_ids = [int(level_id_str)]

    if not target_level_ids:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="No levels found.",
            )
        )

    order_index = int(order_index_str) if order_index_str.isdigit() else None
    for level_id in target_level_ids:
        subject = (
            Subject.query.filter_by(level_id=level_id, name=subject_name)
            .order_by(Subject.id.asc())
            .first()
        )
        if not subject:
            subject = Subject(
                level_id=level_id,
                name=subject_name,
                is_active=True,
            )
            db.session.add(subject)
            db.session.flush()

        db.session.add(
            CurriculumItem(
                subject_id=subject.id,
                title=title,
                description=description or None,
                resource_link=resource_link,
                visibility_scope=visibility_scope,
                order_index=order_index,
            )
        )
    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=(
                f"Link '{title}' added to all levels."
                if apply_to_all_levels
                else f"Link '{title}' added successfully."
            ),
        )
    )


@app.route("/head/curriculum-items/<int:item_id>/delete", methods=["POST"])
@admin_required
def delete_curriculum_item(item_id: int):
    curriculum_item = CurriculumItem.query.get_or_404(item_id)
    subject = curriculum_item.subject
    item_title = curriculum_item.title

    db.session.delete(curriculum_item)
    db.session.flush()

    if subject and not subject.curriculum_items:
        db.session.delete(subject)

    db.session.commit()
    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Link '{item_title}' deleted.",
        )
    )


@app.route("/head/curriculum-items/<int:item_id>/visibility", methods=["POST"])
@admin_required
def update_curriculum_item_visibility(item_id: int):
    curriculum_item = CurriculumItem.query.get_or_404(item_id)
    visibility_scope = (request.form.get("visibility_scope") or "").strip().lower()

    if visibility_scope not in {"student_and_teacher", "teacher_only"}:
        return redirect(
            url_for(
                "head_dashboard",
                exam_import_status="error",
                exam_import_message="Invalid visibility scope selected.",
            )
        )

    curriculum_item.visibility_scope = visibility_scope
    db.session.commit()

    return redirect(
        url_for(
            "head_dashboard",
            exam_import_status="success",
            exam_import_message=f"Visibility updated for '{curriculum_item.title}'.",
        )
    )


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    total_students = Student.query.count()
    total_levels = Level.query.count()
    total_teachers = Teacher.query.count()
    dashboard_data = build_supervisor_dashboard_data()
    arabic_attendance_import_summary = build_imported_arabic_attendance_summary()
    arabic_attendance_action_rows, arabic_attendance_archive_rows = build_arabic_attendance_upload_history()
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    announcement_data = build_supervisor_announcement_summary(levels)
    calendar_data = build_calendar_admin_summary()
    return render_template(
        "admin_dashboard.html",
        total_students=total_students,
        total_levels=total_levels,
        total_teachers=total_teachers,
        unresolved_name_matches=ExamImportIssue.query.count(),
        access_code_status=(request.args.get("access_code_status") or "").strip(),
        access_code_message=(request.args.get("access_code_message") or "").strip(),
        readiness_issues=get_runtime_readiness_issues(),
        kpis=dashboard_data["kpis"],
        dashboard_alerts=dashboard_data["dashboard_alerts"],
        level_rows=dashboard_data["level_rows"],
        teacher_rows=dashboard_data["teacher_rows"],
        student_attention_rows=dashboard_data["student_attention_rows"],
        spotlight=dashboard_data["spotlight"],
        teacher_weekly_status_rows=dashboard_data["teacher_weekly_status_rows"],
        todays_gap_rows=dashboard_data["todays_gap_rows"],
        level_followup_snapshot_rows=dashboard_data["level_followup_snapshot_rows"],
        recent_action_rows=dashboard_data["recent_action_rows"],
        curriculum_plan_rows=dashboard_data["curriculum_plan_rows"],
        curriculum_plan_summary=dashboard_data["curriculum_plan_summary"],
        cleanup_summary=dashboard_data["cleanup_summary"],
        cleanup_snapshot_students=dashboard_data["cleanup_snapshot_students"],
        cleanup_snapshot_teachers=dashboard_data["cleanup_snapshot_teachers"],
        cleanup_snapshot_levels=dashboard_data["cleanup_snapshot_levels"],
        monthly_review_summary=dashboard_data["monthly_review_summary"],
        monthly_review_rows=dashboard_data["monthly_review_rows"],
        teacher_monthly_report_summary=dashboard_data["teacher_monthly_report_summary"],
        teacher_monthly_report_rows=dashboard_data["teacher_monthly_report_rows"],
        ideal_teacher_row=dashboard_data["ideal_teacher_row"],
        academic_calendar_status=dashboard_data["academic_calendar_status"],
        announcement_rows=announcement_data["announcement_rows"],
        levels=announcement_data["levels"],
        announcement_status=(request.args.get("announcement_status") or "").strip(),
        announcement_message=(request.args.get("announcement_message") or "").strip(),
        teacher_thanks_status=(request.args.get("teacher_thanks_status") or "").strip(),
        teacher_thanks_message=(request.args.get("teacher_thanks_message") or "").strip(),
        calendar_settings=calendar_data["calendar_settings"],
        holiday_rows=calendar_data["holiday_rows"],
        followup_ready=calendar_data["followup_ready"],
        followup_status_message=calendar_data["followup_status_message"],
        detected_followup_reviews=calendar_data["detected_followup_reviews"],
        last_weekly_followup_run=calendar_data["last_weekly_followup_run"],
        teaching_week_start=calendar_data["teaching_week_start"],
        teaching_week_end=calendar_data["teaching_week_end"],
        archived_weekly_reports=calendar_data["archived_weekly_reports"],
        calendar_status=(request.args.get("calendar_status") or "").strip(),
        calendar_message=(request.args.get("calendar_message") or "").strip(),
        arabic_attendance_import_status=(request.args.get("arabic_attendance_import_status") or "").strip(),
        arabic_attendance_import_message=(request.args.get("arabic_attendance_import_message") or "").strip(),
        arabic_attendance_import_summary=arabic_attendance_import_summary,
        arabic_attendance_action_rows=arabic_attendance_action_rows,
        arabic_attendance_archive_rows=arabic_attendance_archive_rows,
    )


@app.route("/admin/action-log")
@admin_required
def admin_action_log():
    role_filter = (request.args.get("actor_role") or "").strip().lower()
    action_filter = (request.args.get("action_type") or "").strip()
    level_filter = (request.args.get("level_id") or "").strip()
    search_query = (request.args.get("q") or "").strip()

    query = ActionLog.query
    if role_filter in {"admin", "teacher", "system"}:
        query = query.filter(ActionLog.actor_role == role_filter)
    if action_filter:
        query = query.filter(ActionLog.action_type == action_filter)
    if level_filter.isdigit():
        query = query.filter(ActionLog.level_id == int(level_filter))
    if search_query:
        like_value = f"%{search_query}%"
        query = query.filter(
            db.or_(
                ActionLog.actor_name.ilike(like_value),
                ActionLog.entity_label.ilike(like_value),
                ActionLog.details.ilike(like_value),
            )
        )

    rows = query.order_by(ActionLog.created_at.desc(), ActionLog.id.desc()).limit(200).all()
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    action_types = [
        row[0]
        for row in db.session.query(ActionLog.action_type)
        .distinct()
        .order_by(ActionLog.action_type.asc())
        .all()
        if row[0]
    ]
    role_labels = {
        "admin": "supervisor_role",
        "teacher": "teacher_role",
        "system": "system_role",
    }
    summary = {
        "total": len(rows),
        "admin": sum(1 for item in rows if item.actor_role == "admin"),
        "teacher": sum(1 for item in rows if item.actor_role == "teacher"),
        "system": sum(1 for item in rows if item.actor_role == "system"),
    }

    action_rows = [
        {
            "log": item,
            "role_label_key": role_labels.get(item.actor_role, "system_role"),
            "target_url": get_action_log_target(item)[0],
            "target_label_key": get_action_log_target(item)[1],
        }
        for item in rows
    ]

    return render_template(
        "action_log.html",
        action_rows=action_rows,
        summary=summary,
        action_types=action_types,
        levels=levels,
        role_filter=role_filter,
        action_filter=action_filter,
        level_filter=level_filter,
        search_query=search_query,
    )


@app.route("/admin/academic-calendar")
@admin_required
def admin_academic_calendar():
    calendar_status = build_academic_calendar_status()
    return render_template(
        "admin_academic_calendar.html",
        academic_calendar_status=calendar_status,
        calendar_events=get_academic_calendar_events(),
        event_type_options=ACADEMIC_EVENT_TYPE_OPTIONS,
        op_status=(request.args.get("op_status") or "").strip(),
        op_message=(request.args.get("op_message") or "").strip(),
    )


@app.route("/admin/arabic-attendance/refresh", methods=["POST"])
@admin_required
def admin_refresh_arabic_attendance():
    _load_imported_arabic_attendance_snapshot_cached.cache_clear()
    sync_result = sync_imported_arabic_attendance_aliases()
    imported_summary = build_imported_arabic_attendance_summary()
    refresh_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    set_system_setting("last_arabic_attendance_refresh", refresh_timestamp)
    db.session.commit()
    log_action(
        actor_role="admin",
        actor_name="Supervisor",
        action_type="arabic_attendance_refreshed",
        entity_type="attendance_import",
        entity_label="Arabic attendance workbook",
        details=(
            f"Refreshed Arabic attendance workbook. "
            f"Imported rows: {imported_summary['imported_rows_count']}. "
            f"Aliases created: {sync_result['created_count']}. "
            f"Aliases updated: {sync_result['updated_count']}."
        ),
    )
    return redirect(
        url_for(
            "admin_dashboard",
            arabic_attendance_import_status="success",
            arabic_attendance_import_message=(
                f"Arabic attendance refreshed. "
                f"Rows: {imported_summary['imported_rows_count']}. "
                f"Aliases created: {sync_result['created_count']}. "
                f"Aliases updated: {sync_result['updated_count']}."
            ),
        )
    )


@app.route("/admin/arabic-attendance/upload", methods=["POST"])
@admin_required
def admin_upload_arabic_attendance():
    uploaded_file = request.files.get("attendance_sheet")
    if not uploaded_file or not uploaded_file.filename:
        return redirect(
            url_for(
                "admin_dashboard",
                arabic_attendance_import_status="error",
                arabic_attendance_import_message="Arabic attendance file is required.",
            )
        )

    file_name = (uploaded_file.filename or "").strip()
    if not file_name.lower().endswith((".xlsx", ".xls")):
        return redirect(
            url_for(
                "admin_dashboard",
                arabic_attendance_import_status="error",
                arabic_attendance_import_message="Only Excel attendance files (.xlsx or .xls) are supported.",
            )
        )

    saved_paths = save_uploaded_arabic_attendance_file(uploaded_file)
    _load_imported_arabic_attendance_snapshot_cached.cache_clear()
    sync_result = sync_imported_arabic_attendance_aliases()
    imported_summary = build_imported_arabic_attendance_summary()
    refresh_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    set_system_setting("last_arabic_attendance_refresh", refresh_timestamp)
    db.session.commit()
    log_action(
        actor_role="admin",
        actor_name="Supervisor",
        action_type="arabic_attendance_uploaded",
        entity_type="attendance_import",
        entity_label=file_name,
        details=(
            f"Uploaded Arabic attendance workbook '{file_name}'. "
            f"Backup: {saved_paths['backup_path'] or '-'}; "
            f"Archived upload: {saved_paths['archived_upload_path']}. "
            f"Imported rows: {imported_summary['imported_rows_count']}. "
            f"Aliases created: {sync_result['created_count']}. "
            f"Aliases updated: {sync_result['updated_count']}."
        ),
    )
    return redirect(
        url_for(
            "admin_dashboard",
            arabic_attendance_import_status="success",
            arabic_attendance_import_message=(
                f"Arabic attendance file uploaded and refreshed. "
                f"Rows: {imported_summary['imported_rows_count']}. "
                f"Aliases created: {sync_result['created_count']}. "
                f"Aliases updated: {sync_result['updated_count']}."
            ),
        )
    )


@app.route("/admin/arabic-attendance/backup/latest")
@admin_required
def admin_download_latest_arabic_attendance_backup():
    backup_path = get_latest_arabic_attendance_backup()
    if not backup_path or not backup_path.exists():
        return redirect(
            url_for(
                "admin_dashboard",
                arabic_attendance_import_status="error",
                arabic_attendance_import_message="No Arabic attendance backup is available yet.",
            )
        )
    return send_file(
        backup_path,
        as_attachment=True,
        download_name=backup_path.name,
    )


@app.route("/admin/arabic-attendance/current")
@admin_required
def admin_download_current_arabic_attendance():
    workbook_path = get_arabic_attendance_workbook_path()
    if not workbook_path.exists():
        return redirect(
            url_for(
                "admin_dashboard",
                arabic_attendance_import_status="error",
                arabic_attendance_import_message="Current Arabic attendance workbook was not found.",
            )
        )
    return send_file(
        workbook_path,
        as_attachment=True,
        download_name=workbook_path.name,
    )


@app.route("/admin/reports")
@admin_required
def admin_reports_center():
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    students = Student.query.order_by(Student.full_name.asc()).all()
    selected_level_id = (request.args.get("level_id") or "").strip()
    selected_student_id = (request.args.get("student_id") or "").strip()
    student_period = (request.args.get("period") or "month").strip().lower()
    if student_period not in {"week", "month"}:
        student_period = "month"
    return render_template(
        "admin_reports_center.html",
        levels=levels,
        students=students,
        selected_level_id=selected_level_id,
        selected_student_id=selected_student_id,
        student_period=student_period,
    )


@app.route("/admin/teachers/monthly-thanks/send", methods=["POST"])
@admin_required
def send_teacher_monthly_thanks():
    teacher_id_str = (request.form.get("teacher_id") or "").strip()
    if not teacher_id_str.isdigit():
        return redirect(
            url_for(
                "admin_dashboard",
                teacher_thanks_status="error",
                teacher_thanks_message="Teacher selection is invalid.",
            )
        )

    teacher = Teacher.query.get_or_404(int(teacher_id_str))
    teacher_email = (teacher.email or "").strip()
    if not validate_email(teacher_email):
        return redirect(
            url_for(
                "admin_dashboard",
                teacher_thanks_status="error",
                teacher_thanks_message=f"Teacher email is missing or invalid for {teacher.full_name}.",
            )
        )

    monthly_report_data = build_teacher_monthly_report_rows()
    selected_row = next((row for row in monthly_report_data["rows"] if row["teacher"].id == teacher.id), None)
    if not selected_row:
        return redirect(
            url_for(
                "admin_dashboard",
                teacher_thanks_status="error",
                teacher_thanks_message="Monthly teacher report is not available for this teacher.",
            )
        )

    subject = f"Hikmah Academy - Monthly Appreciation for {teacher.full_name}"
    body = (
        f"Dear {teacher.full_name},\n\n"
        f"Thank you for your work during {selected_row['month_start'].strftime('%Y-%m')}.\n"
        f"Your monthly supervision summary is as follows:\n"
        f"- Attendance completion: {selected_row['attendance_pct']}%\n"
        f"- Zoom recordings completion: {selected_row['recordings_pct']}%\n"
        f"- Weekly homework completion: {selected_row['homework_pct']}%\n"
        f"- Pending assignment reviews: {selected_row['monthly_pending_reviews']}\n"
        f"- Overall score: {selected_row['overall_score']} ({selected_row['performance_label']})\n\n"
        f"We appreciate your consistency and effort with your students.\n\n"
        f"Hikmah Academy Supervisor"
    )

    try:
        send_email_via_smtp(teacher_email, subject, body)
    except Exception as exc:
        return redirect(
            url_for(
                "admin_dashboard",
                teacher_thanks_status="error",
                teacher_thanks_message=f"Failed to send appreciation email: {exc}",
            )
        )

    log_action(
        action_type="teacher_thanks_sent",
        entity_type="teacher",
        entity_id=teacher.id,
        entity_label=teacher.full_name,
        details=f"Sent monthly appreciation email to {teacher_email}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            teacher_thanks_status="success",
            teacher_thanks_message=f"Monthly appreciation email sent to {teacher.full_name}.",
        )
    )


@app.route("/admin/exam-name-matches")
@admin_required
def admin_exam_name_matches():
    issues = ExamImportIssue.query.order_by(
        ExamImportIssue.source_file_name.asc(),
        ExamImportIssue.level_name.asc(),
        ExamImportIssue.alias_name.asc(),
    ).all()

    levels = {
        level.name: level.id
        for level in Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    }
    students = Student.query.order_by(Student.full_name.asc()).all()

    issue_rows = []
    for issue in issues:
        suggested_students = students
        if issue.level_name:
            level_id = levels.get(issue.level_name)
            if level_id:
                suggested_students = [student for student in students if student.level_id == level_id]

        issue_rows.append(
            {
                "issue": issue,
                "suggested_students": suggested_students[:80],
            }
        )

    return render_template(
        "exam_name_matches.html",
        issue_rows=issue_rows,
        total_issues=len(issue_rows),
        op_status=(request.args.get("op_status") or "").strip(),
        op_message=(request.args.get("op_message") or "").strip(),
    )


@app.route("/admin/exam-name-matches/link", methods=["POST"])
@admin_required
def admin_link_exam_name_match():
    issue_id = (request.form.get("issue_id") or "").strip()
    student_id = (request.form.get("student_id") or "").strip()
    if not issue_id.isdigit() or not student_id.isdigit():
        return redirect(
            url_for(
                "admin_exam_name_matches",
                op_status="error",
                op_message="Issue and student are required.",
            )
        )

    issue = ExamImportIssue.query.get_or_404(int(issue_id))
    student = Student.query.get_or_404(int(student_id))
    normalized_level_name = normalize_level_display_name(issue.level_name) if issue.level_name else None

    existing_alias = StudentNameAlias.query.filter_by(
        alias_name=issue.alias_name,
        level_name=normalized_level_name,
    ).first()
    if not existing_alias:
        db.session.add(
            StudentNameAlias(
                alias_name=issue.alias_name,
                level_name=normalized_level_name,
                student_id=student.id,
            )
        )
    else:
        existing_alias.student_id = student.id

    clear_exam_import_issue(issue.source_file_name, issue.alias_name, issue.level_name)
    db.session.commit()
    return redirect(
        url_for(
            "admin_exam_name_matches",
            op_status="success",
            op_message=f"Linked '{issue.alias_name}' to '{student.full_name}'.",
        )
    )


@app.route("/admin/exam-name-matches/create-student", methods=["POST"])
@admin_required
def admin_create_student_from_exam_issue():
    issue_id = (request.form.get("issue_id") or "").strip()
    if not issue_id.isdigit():
        return redirect(
            url_for(
                "admin_exam_name_matches",
                op_status="error",
                op_message="Issue is required.",
            )
        )

    issue = ExamImportIssue.query.get_or_404(int(issue_id))
    normalized_level_name = normalize_level_display_name(issue.level_name) if issue.level_name else ""
    level = Level.query.filter_by(name=normalized_level_name).first() if normalized_level_name else None

    existing_student = Student.query.filter(
        db.or_(
            Student.full_name == issue.alias_name,
            db.func.lower(Student.full_name) == issue.alias_name.lower(),
        ),
        db.or_(
            Student.level_id == (level.id if level else None),
            Student.level_name == normalized_level_name,
        ),
    ).first()
    if existing_student:
        return redirect(
            url_for(
                "admin_exam_name_matches",
                op_status="error",
                op_message=f"Student '{issue.alias_name}' already exists in the selected level.",
            )
        )

    student = Student(
        student_code=generate_next_student_code(),
        full_name=issue.alias_name,
        level_name=normalized_level_name or (issue.level_name or ""),
        level_id=level.id if level else None,
        status="Active",
    )
    db.session.add(student)
    db.session.flush()

    alias = StudentNameAlias.query.filter_by(
        alias_name=issue.alias_name,
        level_name=normalized_level_name or None,
    ).first()
    if not alias:
        db.session.add(
            StudentNameAlias(
                alias_name=issue.alias_name,
                level_name=normalized_level_name or None,
                student_id=student.id,
            )
        )
    else:
        alias.student_id = student.id

    clear_exam_import_issue(issue.source_file_name, issue.alias_name, issue.level_name)
    db.session.commit()
    return redirect(
        url_for(
            "admin_exam_name_matches",
            op_status="success",
            op_message=f"Created new student '{student.full_name}' with code {student.student_code}.",
        )
    )


@app.route("/admin/exam-name-matches/reimport", methods=["POST"])
@admin_required
def admin_reimport_exam_file():
    source_file_name = (request.form.get("source_file_name") or "").strip()
    if not source_file_name:
        return redirect(
            url_for(
                "admin_exam_name_matches",
                op_status="error",
                op_message="Missing source file name.",
            )
        )

    file_path = os.path.join(app.config["EXAM_UPLOAD_DIR"], os.path.basename(source_file_name))
    if not os.path.exists(file_path):
        return redirect(
            url_for(
                "admin_exam_name_matches",
                op_status="error",
                op_message=f"Saved file not found: {source_file_name}",
            )
        )

    with open(file_path, "rb") as saved_file:
        parsed_rows, row_errors = read_excel_exam_rows(saved_file)
    imported_count, skipped_count, import_errors = import_exam_rows(parsed_rows, os.path.basename(file_path))
    row_errors.extend(import_errors)
    db.session.commit()

    message = f"Reimported {imported_count} result(s)."
    if skipped_count:
        message += f" Still unmatched: {skipped_count}."
    if row_errors:
        message += " " + " | ".join(row_errors[:3])

    return redirect(
        url_for(
            "admin_exam_name_matches",
            op_status="success" if imported_count else "error",
            op_message=message,
        )
    )


@app.route("/admin/access-code/update", methods=["POST"])
@admin_required
def update_admin_access_code():
    new_code = (request.form.get("new_access_code") or "").strip()
    if len(new_code) < 4:
        return redirect(
            url_for(
                "admin_dashboard",
                access_code_status="error",
                access_code_message="Access code must be at least 4 characters.",
            )
        )

    set_admin_access_code(new_code)
    log_action(
        action_type="admin_access_code_updated",
        entity_type="system_setting",
        entity_label="admin_access_code",
        details="Updated the supervisor access code.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            access_code_status="success",
            access_code_message="Admin access code updated successfully.",
        )
    )


@app.route("/admin/announcements/add", methods=["POST"])
@admin_required
def add_announcement():
    title = (request.form.get("title") or "").strip()
    body = (request.form.get("body") or "").strip()
    audience = (request.form.get("audience") or "all").strip().lower()
    category = (request.form.get("category") or "general").strip().lower()
    level_id_str = (request.form.get("level_id") or "").strip()
    starts_on = parse_optional_date_input(request.form.get("starts_on") or "")
    expires_on = parse_optional_date_input(request.form.get("expires_on") or "")
    is_pinned = (request.form.get("is_pinned") or "").strip().lower() in {"1", "true", "on", "yes"}

    if not title or not body:
        return redirect(
            url_for(
                "admin_dashboard",
                announcement_status="error",
                announcement_message="Announcement title and message are required.",
            )
        )

    if audience not in {"all", "teachers", "students"}:
        audience = "all"
    if category not in {"general", "reminder", "alert", "follow_up"}:
        category = "general"

    level_id = int(level_id_str) if level_id_str.isdigit() else None
    if level_id and not Level.query.get(level_id):
        return redirect(
            url_for(
                "admin_dashboard",
                announcement_status="error",
                announcement_message="Selected level was not found.",
            )
        )

    if starts_on and expires_on and expires_on < starts_on:
        return redirect(
            url_for(
                "admin_dashboard",
                announcement_status="error",
                announcement_message="Announcement expiry date cannot be before start date.",
            )
        )

    announcement = Announcement(
        title=title,
        body=body,
        audience=audience,
        category=category,
        level_id=level_id,
        starts_on=starts_on,
        expires_on=expires_on,
        is_pinned=is_pinned,
        is_active=True,
    )
    db.session.add(announcement)
    log_action(
        action_type="announcement_published",
        entity_type="announcement",
        entity_label=title,
        level=announcement.level,
        details=f"Published announcement for audience '{audience}'.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            announcement_status="success",
            announcement_message="Announcement published successfully.",
        )
    )


@app.route("/admin/announcements/<int:announcement_id>/update", methods=["POST"])
@admin_required
def update_announcement(announcement_id: int):
    announcement = Announcement.query.get_or_404(announcement_id)
    is_active = (request.form.get("is_active") or "").strip().lower() in {"1", "true", "on", "yes"}
    is_pinned = (request.form.get("is_pinned") or "").strip().lower() in {"1", "true", "on", "yes"}
    expires_on = parse_optional_date_input(request.form.get("expires_on") or "")

    if announcement.starts_on and expires_on and expires_on < announcement.starts_on:
        return redirect(
            url_for(
                "admin_dashboard",
                announcement_status="error",
                announcement_message="Announcement expiry date cannot be before start date.",
            )
        )

    announcement.is_active = is_active
    announcement.is_pinned = is_pinned
    announcement.expires_on = expires_on
    log_action(
        action_type="announcement_updated",
        entity_type="announcement",
        entity_id=announcement.id,
        entity_label=announcement.title,
        level=announcement.level,
        details="Updated announcement visibility or pin settings.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            announcement_status="success",
            announcement_message="Announcement updated successfully.",
        )
    )


@app.route("/admin/calendar-settings/update", methods=["POST"])
@admin_required
def update_calendar_settings():
    duty_start_time = (request.form.get("duty_start_time") or "").strip() or "08:00"
    duty_end_time = (request.form.get("duty_end_time") or "").strip() or "15:00"
    weekly_followup_weekday = normalize_weekday_name(request.form.get("weekly_followup_weekday") or "thursday")
    weekly_followup_time = (request.form.get("weekly_followup_time") or "").strip() or "14:00"

    set_system_setting("duty_start_time", duty_start_time)
    set_system_setting("duty_end_time", duty_end_time)
    set_system_setting("weekly_followup_weekday", weekly_followup_weekday)
    set_system_setting("weekly_followup_time", weekly_followup_time)
    log_action(
        action_type="calendar_settings_updated",
        entity_type="calendar_settings",
        entity_label="Operating Calendar",
        details=f"Updated duty window to {duty_start_time}-{duty_end_time} and weekly follow-up to {weekly_followup_weekday} {weekly_followup_time}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            calendar_status="success",
            calendar_message="Calendar settings updated successfully.",
        )
    )


@app.route("/admin/holidays/add", methods=["POST"])
@admin_required
def add_holiday_period():
    title = (request.form.get("title") or "").strip()
    start_date = parse_optional_date_input(request.form.get("start_date") or "")
    end_date = parse_optional_date_input(request.form.get("end_date") or "")

    if not title or not start_date or not end_date:
        return redirect(
            url_for(
                "admin_dashboard",
                calendar_status="error",
                calendar_message="Holiday title, start date, and end date are required.",
            )
        )

    if end_date < start_date:
        return redirect(
            url_for(
                "admin_dashboard",
                calendar_status="error",
                calendar_message="Holiday end date cannot be before start date.",
            )
        )

    holiday = HolidayPeriod(title=title, start_date=start_date, end_date=end_date, is_active=True)
    db.session.add(holiday)
    log_action(
        action_type="holiday_added",
        entity_type="holiday",
        entity_label=title,
        details=f"Added holiday from {start_date.isoformat()} to {end_date.isoformat()}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            calendar_status="success",
            calendar_message="Holiday period added successfully.",
        )
    )


@app.route("/admin/holidays/<int:holiday_id>/update", methods=["POST"])
@admin_required
def update_holiday_period(holiday_id: int):
    holiday = HolidayPeriod.query.get_or_404(holiday_id)
    holiday.is_active = (request.form.get("is_active") or "").strip().lower() in {"1", "true", "on", "yes"}
    log_action(
        action_type="holiday_updated",
        entity_type="holiday",
        entity_id=holiday.id,
        entity_label=holiday.title,
        details=f"Set active status to {holiday.is_active}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            calendar_status="success",
            calendar_message="Holiday period updated successfully.",
        )
    )


@app.route("/admin/academic-calendar/add", methods=["POST"])
@admin_required
def add_academic_calendar_event():
    title = (request.form.get("title") or "").strip()
    event_type = (request.form.get("event_type") or "").strip()
    start_date = parse_optional_date_input(request.form.get("start_date") or "")
    end_date = parse_optional_date_input(request.form.get("end_date") or "")
    note_text = (request.form.get("note_text") or "").strip()
    is_instructional = (request.form.get("is_instructional") or "").strip().lower() in {"1", "true", "on", "yes"}

    if not title or event_type not in ACADEMIC_EVENT_TYPE_OPTIONS or not start_date or not end_date:
        return redirect(
            url_for(
                "admin_academic_calendar",
                op_status="error",
                op_message=ui_text("event_dates_required"),
            )
        )

    if end_date < start_date:
        return redirect(
            url_for(
                "admin_academic_calendar",
                op_status="error",
                op_message=ui_text("event_dates_invalid"),
            )
        )

    sort_order = (db.session.query(db.func.max(AcademicCalendarEvent.sort_order)).scalar() or 0) + 10
    event = AcademicCalendarEvent(
        title=title,
        event_type=event_type,
        start_date=start_date,
        end_date=end_date,
        is_instructional=is_instructional,
        sort_order=sort_order,
        note_text=note_text or None,
    )
    db.session.add(event)
    log_action(
        action_type="academic_calendar_event_added",
        entity_type="academic_calendar_event",
        entity_label=title,
        details=f"Added {event_type} event from {start_date.isoformat()} to {end_date.isoformat()}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_academic_calendar",
            op_status="success",
            op_message=ui_text("event_saved_success"),
        )
    )


@app.route("/admin/academic-calendar/<int:event_id>/update", methods=["POST"])
@admin_required
def update_academic_calendar_event(event_id: int):
    event = AcademicCalendarEvent.query.get_or_404(event_id)
    title = (request.form.get("title") or "").strip()
    event_type = (request.form.get("event_type") or "").strip()
    start_date = parse_optional_date_input(request.form.get("start_date") or "")
    end_date = parse_optional_date_input(request.form.get("end_date") or "")
    note_text = (request.form.get("note_text") or "").strip()
    is_instructional = (request.form.get("is_instructional") or "").strip().lower() in {"1", "true", "on", "yes"}

    if not title or event_type not in ACADEMIC_EVENT_TYPE_OPTIONS or not start_date or not end_date:
        return redirect(
            url_for(
                "admin_academic_calendar",
                op_status="error",
                op_message=ui_text("event_dates_required"),
            )
        )

    if end_date < start_date:
        return redirect(
            url_for(
                "admin_academic_calendar",
                op_status="error",
                op_message=ui_text("event_dates_invalid"),
            )
        )

    event.title = title
    event.event_type = event_type
    event.start_date = start_date
    event.end_date = end_date
    event.note_text = note_text or None
    event.is_instructional = is_instructional
    log_action(
        action_type="academic_calendar_event_updated",
        entity_type="academic_calendar_event",
        entity_id=event.id,
        entity_label=event.title,
        details=f"Updated {event.event_type} event to {start_date.isoformat()} - {end_date.isoformat()}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_academic_calendar",
            op_status="success",
            op_message=ui_text("event_updated_success"),
        )
    )


@app.route("/admin/weekly-followup/generate", methods=["POST"])
@admin_required
def generate_weekly_followup():
    target_date = parse_optional_date_input(request.form.get("target_date") or "") or date.today()
    can_run, status_message = should_run_weekly_followup(target_date)
    if not can_run:
        return redirect(
            url_for(
                "admin_dashboard",
                calendar_status="error",
                calendar_message=status_message,
            )
        )

    result = run_scheduled_weekly_followup(
        now=datetime.combine(target_date, datetime.min.time()),
        force=True,
    )
    log_action(
        action_type="weekly_followup_generated",
        entity_type="weekly_followup",
        entity_label=target_date.isoformat(),
        details=result["message"],
    )
    db.session.commit()
    return redirect(
        url_for(
            "admin_dashboard",
            calendar_status="success",
            calendar_message=result["message"],
        )
    )


@app.route("/teacher/dashboard")
@teacher_required
def teacher_dashboard():
    teacher = get_current_teacher()
    if not teacher:
        return redirect(url_for("teacher_login"))
    dashboard_data = build_teacher_dashboard_data(teacher)
    teacher_level_ids = [row["level"].id for row in dashboard_data["level_rows"]]
    teacher_announcements = build_announcement_view_rows(get_active_announcements("teachers", teacher_level_ids))
    return render_template(
        "teacher_dashboard.html",
        teacher=teacher,
        kpis=dashboard_data["kpis"],
        quick_tasks=dashboard_data["quick_tasks"],
        level_rows=dashboard_data["level_rows"],
        student_focus_rows=dashboard_data["student_focus_rows"],
        teacher_announcements=teacher_announcements,
        weekly_review_rows=dashboard_data["weekly_review_rows"],
        weekly_review_summary=dashboard_data["weekly_review_summary"],
        syllabus_plan_rows=dashboard_data["syllabus_plan_rows"],
        syllabus_plan_summary=dashboard_data["syllabus_plan_summary"],
        academic_calendar_status=dashboard_data["academic_calendar_status"],
    )


@app.route("/teacher/levels/<int:level_id>")
@teacher_required
def teacher_level_workspace(level_id: int):
    level = get_teacher_level_or_403(level_id)
    students = Student.query.filter_by(level_id=level.id).order_by(Student.full_name).all()
    today = date.today()
    followup_status_filter = (request.args.get("followup_status") or "").strip()
    followup_q = (request.args.get("followup_q") or "").strip()
    followup_sort = (request.args.get("followup_sort") or "").strip()
    exam_templates = ExamTemplate.query.filter_by(is_active=True).order_by(ExamTemplate.title.asc()).all()
    subjects = (
        Subject.query.filter_by(level_id=level.id, is_active=True)
        .order_by(Subject.order_index.asc(), Subject.name.asc())
        .all()
    )
    subject_item_ids = [item.id for subject in subjects for item in subject.curriculum_items]
    progress_by_item_id = {
        row.curriculum_item_id: row
        for row in CurriculumProgress.query.filter(
            CurriculumProgress.level_id == level.id,
            CurriculumProgress.curriculum_item_id.in_(subject_item_ids) if subject_item_ids else db.text("1=0"),
        ).all()
    } if subject_item_ids else {}
    subject_rows = [
        {
            "subject": subject,
            "curriculum_items": [
                {
                    "item": item,
                    "progress": progress_by_item_id.get(item.id),
                    "status": ((progress_by_item_id.get(item.id).status if progress_by_item_id.get(item.id) else "pending") or "pending").strip().lower(),
                    "note_text": (progress_by_item_id.get(item.id).note_text if progress_by_item_id.get(item.id) else "") or "",
                }
                for item in sorted(
                    subject.curriculum_items,
                    key=lambda item: ((item.order_index or 0), item.id),
                )
            ],
        }
        for subject in subjects
    ]
    book_subject_rows = [row for row in subject_rows if not is_teacher_resource_subject(row["subject"].name)]
    teacher_resource_subject_rows = [row for row in subject_rows if is_teacher_resource_subject(row["subject"].name)]
    curriculum_plan_rows = build_curriculum_progress_summary([level])
    curriculum_plan_summary = curriculum_plan_rows[0] if curriculum_plan_rows else {
        "total_items": 0,
        "completed_count": 0,
        "in_progress_count": 0,
        "pending_count": 0,
        "progress_pct": 0.0,
        "last_completed_title": "",
        "last_completed_on": None,
    }
    syllabus_plan_rows = build_syllabus_plan_summary([level])
    syllabus_plan_summary = syllabus_plan_rows[0] if syllabus_plan_rows else {
        "total_rows": 0,
        "planned_count": 0,
        "in_progress_count": 0,
        "completed_count": 0,
        "postponed_count": 0,
        "expected_count": 0,
        "delayed_count": 0,
        "remaining_count": 0,
        "status_label_key": "on_track_status",
        "progress_pct": 0.0,
        "last_completed_label": "",
        "current_week": get_current_teaching_week(),
    }
    assignments = (
        Assignment.query.filter_by(level_id=level.id)
        .order_by(Assignment.due_date.asc(), Assignment.id.desc())
        .all()
    )
    recordings = (
        ClassRecording.query.filter_by(class_id=level.id)
        .order_by(ClassRecording.lesson_date.desc(), ClassRecording.id.desc())
        .all()
    )
    upcoming_exams = (
        UpcomingExam.query.filter_by(level_id=level.id)
        .order_by(UpcomingExam.exam_date.asc(), UpcomingExam.id.asc())
        .all()
    )
    exam_results = (
        ExamResult.query.filter_by(level_id=level.id)
        .join(Student, ExamResult.student_id == Student.id)
        .order_by(Student.full_name.asc(), ExamResult.exam_title.asc(), ExamResult.id.asc())
        .all()
    )
    active_assignments = [assignment for assignment in assignments if assignment.is_active]
    student_exam_rows = []
    student_followup_rows = filter_level_followup_rows(
        build_level_followup_register(level, today),
        status_filter=followup_status_filter,
        search_query=followup_q,
        sort_by=followup_sort,
    )
    student_followup_summary = summarize_level_followup_rows(student_followup_rows)
    for student in students:
        student_results = [result for result in exam_results if result.student_id == student.id]
        exam_groups = build_exam_summary_groups(student_results)
        student_exam_rows.append(
            {
                "student": student,
                "exam_groups": exam_groups,
            }
        )
    selected_template_id = (request.args.get("exam_template_id") or "").strip()
    selected_overview_exam_title = (request.args.get("overview_exam_title") or "").strip()
    selected_template = ExamTemplate.query.get(int(selected_template_id)) if selected_template_id.isdigit() else None
    selected_template_student_rows = []
    selected_template_branches = []
    if selected_template:
        selected_template_branches = sorted(
            selected_template.branches,
            key=lambda item: ((item.order_index or 0), item.id),
        )
        existing_template_results = ExamResult.query.filter_by(
            level_id=level.id,
            exam_title=selected_template.title,
        ).all()
        template_results_by_student = {}
        shared_notes_by_student = {}
        for result in existing_template_results:
            template_results_by_student.setdefault(result.student_id, {})[result.subject_name] = result
            if result.notes and result.student_id not in shared_notes_by_student:
                shared_notes_by_student[result.student_id] = result.notes

        for student in students:
            selected_template_student_rows.append(
                {
                    "student": student,
                    "branch_results": template_results_by_student.get(student.id, {}),
                    "shared_note": shared_notes_by_student.get(student.id, ""),
                }
            )
    level_exam_titles = sorted(
        {
            result.exam_title
            for result in exam_results
            if result.exam_title
        }
    )
    overview_exam_title = selected_overview_exam_title or (level_exam_titles[0] if level_exam_titles else "")
    selected_results_exam_title = (request.args.get("results_exam_title") or "").strip()
    results_exam_title = selected_results_exam_title or overview_exam_title or (level_exam_titles[0] if level_exam_titles else "")
    overview_rows = []
    if overview_exam_title:
        for row in student_exam_rows:
            summary = next(
                (exam_group for exam_group in row["exam_groups"] if exam_group["exam_title"] == overview_exam_title),
                None,
            )
            overview_rows.append(
                {
                    "student": row["student"],
                    "summary": summary,
                }
            )
    results_exam_rows = []
    results_exam_branches = []
    if results_exam_title:
        branch_names_in_order = []
        seen_branch_names = set()
        for result in exam_results:
            if result.exam_title != results_exam_title or not should_display_exam_component(result.subject_name):
                continue
            if result.subject_name not in seen_branch_names:
                seen_branch_names.add(result.subject_name)
                branch_names_in_order.append(result.subject_name)

        results_exam_branches = branch_names_in_order
        for row in student_exam_rows:
            summary = next(
                (exam_group for exam_group in row["exam_groups"] if exam_group["exam_title"] == results_exam_title),
                None,
            )
            components_by_subject = {}
            shared_note = ""
            if summary:
                components_by_subject = {
                    component.subject_name: component
                    for component in summary["components"]
                }
                shared_note = next(
                    (component.notes for component in summary["components"] if component.notes),
                    "",
                )

            results_exam_rows.append(
                {
                    "student": row["student"],
                    "summary": summary,
                    "components_by_subject": components_by_subject,
                    "shared_note": shared_note,
                }
            )
    selected_assignment_id = (request.args.get("assignment_id") or "").strip()
    selected_assignment = Assignment.query.get(int(selected_assignment_id)) if selected_assignment_id.isdigit() else None
    if selected_assignment and selected_assignment.level_id != level.id:
        selected_assignment = None
    selected_assignment_rows = []
    selected_assignment_has_submissions = False
    if selected_assignment:
        submissions = AssignmentSubmission.query.filter_by(assignment_id=selected_assignment.id).all()
        submissions_by_student = {submission.student_id: submission for submission in submissions}
        selected_assignment_has_submissions = bool(submissions)
        for student in students:
            selected_assignment_rows.append(
                {
                    "student": student,
                    "submission": submissions_by_student.get(student.id),
                }
            )
    requested_workspace_panel = (request.args.get("workspace") or "").strip().lower()
    if requested_workspace_panel not in {"class-tools", "assessment", "scheduling", "recordings", "assignments", "students"}:
        requested_workspace_panel = ""
    default_workspace_panel = requested_workspace_panel or ("assignments" if selected_assignment else "class-tools")
    op_status = (request.args.get("op_status") or "").strip()
    op_message = (request.args.get("op_message") or "").strip()
    return render_template(
        "teacher_level_workspace.html",
        level=level,
        students=students,
        subject_rows=subject_rows,
        book_subject_rows=book_subject_rows,
        teacher_resource_subject_rows=teacher_resource_subject_rows,
        curriculum_plan_summary=curriculum_plan_summary,
        syllabus_plan_summary=syllabus_plan_summary,
        assignments=assignments,
        selected_assignment=selected_assignment,
        selected_assignment_rows=selected_assignment_rows,
        selected_assignment_has_submissions=selected_assignment_has_submissions,
        default_workspace_panel=default_workspace_panel,
        student_exam_rows=student_exam_rows,
        student_followup_rows=student_followup_rows,
        exam_templates=exam_templates,
        selected_template=selected_template,
        selected_template_branches=selected_template_branches,
        selected_template_student_rows=selected_template_student_rows,
        level_exam_titles=level_exam_titles,
        overview_exam_title=overview_exam_title,
        overview_rows=overview_rows,
        results_exam_title=results_exam_title,
        results_exam_rows=results_exam_rows,
        results_exam_branches=results_exam_branches,
        recordings=recordings,
        upcoming_exams=upcoming_exams,
        student_followup_summary=student_followup_summary,
        followup_status_filter=followup_status_filter,
        followup_q=followup_q,
        followup_sort=followup_sort,
        op_status=op_status,
        op_message=op_message,
    )


@app.route("/admin/levels/<int:level_id>/follow-up")
@admin_required
def admin_level_followup(level_id: int):
    level = Level.query.get_or_404(level_id)
    followup_status_filter = (request.args.get("followup_status") or "").strip()
    followup_q = (request.args.get("followup_q") or "").strip()
    followup_sort = (request.args.get("followup_sort") or "").strip()
    followup_rows = filter_level_followup_rows(
        build_level_followup_register(level),
        status_filter=followup_status_filter,
        search_query=followup_q,
        sort_by=followup_sort,
    )
    followup_summary = summarize_level_followup_rows(followup_rows)
    return render_template(
        "admin_level_followup.html",
        level=level,
        followup_rows=followup_rows,
        followup_summary=followup_summary,
        followup_status_filter=followup_status_filter,
        followup_q=followup_q,
        followup_sort=followup_sort,
    )


@app.route("/teacher/curriculum-items/<int:item_id>/progress", methods=["POST"])
@teacher_required
def teacher_update_curriculum_progress(item_id: int):
    curriculum_item = CurriculumItem.query.get_or_404(item_id)
    subject = curriculum_item.subject
    level = get_teacher_level_or_403(subject.level_id)
    status = (request.form.get("status") or "pending").strip().lower()
    note_text = (request.form.get("note_text") or "").strip()

    if status not in {"pending", "in_progress", "completed"}:
        status = "pending"

    progress = CurriculumProgress.query.filter_by(
        level_id=level.id,
        curriculum_item_id=curriculum_item.id,
    ).first()
    if not progress:
        progress = CurriculumProgress(level_id=level.id, curriculum_item_id=curriculum_item.id)
        db.session.add(progress)

    progress.status = status
    progress.note_text = note_text or None
    progress.completed_on = date.today() if status == "completed" else None
    log_action(
        action_type="curriculum_progress_updated",
        entity_type="curriculum_item",
        entity_id=curriculum_item.id,
        entity_label=curriculum_item.title,
        level=level,
        details=f"Set curriculum progress to {status}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Curriculum progress updated.",
        )
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan")
@teacher_required
def teacher_syllabus_plan(level_id: int):
    level = get_teacher_level_or_403(level_id)
    plan_rows = (
        SyllabusPlanEntry.query.filter_by(level_id=level.id)
        .order_by(SyllabusPlanEntry.week_number.asc(), SyllabusPlanEntry.session_number.asc(), SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
        .all()
    )
    plan_summary_rows = build_syllabus_plan_summary([level])
    plan_summary = plan_summary_rows[0] if plan_summary_rows else None
    academic_calendar_status = build_academic_calendar_status()
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    return render_template(
        "teacher_syllabus_plan.html",
        level=level,
        plan_rows=plan_rows,
        grid_rows=build_syllabus_plan_grid(level, lessons_per_week=lessons_per_week),
        session_numbers=list(range(1, lessons_per_week + 1)),
        plan_summary=plan_summary,
        current_teaching_week=academic_calendar_status["current_week"],
        academic_calendar_status=academic_calendar_status,
        plan_edit_open=bool(level.syllabus_edit_open),
        op_status=(request.args.get("op_status") or "").strip(),
        op_message=(request.args.get("op_message") or "").strip(),
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan/template.xlsx")
@teacher_required
def teacher_syllabus_plan_template(level_id: int):
    level = get_teacher_level_or_403(level_id)
    output = build_syllabus_template_workbook(level)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=build_syllabus_export_filename(level, "xlsx", variant="template"),
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan.xlsx")
@teacher_required
def teacher_syllabus_plan_excel(level_id: int):
    level = get_teacher_level_or_403(level_id)
    output = build_syllabus_plan_excel_workbook(level)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=build_syllabus_export_filename(level, "xlsx"),
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan/import", methods=["POST"])
@teacher_required
def teacher_import_syllabus_plan_template(level_id: int):
    level = get_teacher_level_or_403(level_id)
    if not level.syllabus_edit_open:
        return redirect(
            url_for(
                "teacher_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="Plan editing is locked.",
            )
        )
    uploaded_file = request.files.get("plan_template_file")
    if not uploaded_file or not (uploaded_file.filename or "").strip():
        return redirect(
            url_for(
                "teacher_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="Excel plan file is required.",
            )
        )
    try:
        result = import_syllabus_template_workbook(level, uploaded_file)
    except ValueError as exc:
        return redirect(
            url_for(
                "teacher_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message=str(exc),
            )
        )

    log_action(
        action_type="plan_template_imported",
        entity_type="syllabus_plan",
        entity_label=level.name,
        level=level,
        details=f"Imported {result['scheduled_count']} scheduled row(s) and {result['reserve_count']} reserve row(s) from Excel template.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message=f"Excel plan imported: {result['scheduled_count']} scheduled row(s), {result['reserve_count']} reserve row(s).",
        )
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan/add", methods=["POST"])
@teacher_required
def teacher_add_syllabus_plan_entry(level_id: int):
    level = get_teacher_level_or_403(level_id)
    lesson_title = (request.form.get("lesson_title") or "").strip()
    book_name = (request.form.get("book_name") or "").strip()
    unit_name = (request.form.get("unit_name") or "").strip()
    week_number_str = (request.form.get("week_number") or "").strip()
    source_reference = (request.form.get("source_reference") or "").strip()
    learning_objective = (request.form.get("learning_objective") or "").strip()
    planned_homework = (request.form.get("planned_homework") or "").strip()
    note_text = (request.form.get("note_text") or "").strip()

    if not lesson_title:
        return redirect(
            url_for(
                "teacher_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="Lesson title is required.",
            )
        )

    week_number = int(week_number_str) if week_number_str.isdigit() else None
    order_index = (db.session.query(db.func.max(SyllabusPlanEntry.order_index)).filter_by(level_id=level.id).scalar() or 0) + 1
    entry = SyllabusPlanEntry(
        level_id=level.id,
        week_number=week_number,
        session_number=1,
        book_name=book_name or None,
        unit_name=unit_name or None,
        lesson_title=lesson_title,
        source_reference=source_reference or None,
        learning_objective=learning_objective or None,
        planned_homework=planned_homework or None,
        note_text=note_text or None,
        status="planned",
        order_index=order_index,
    )
    db.session.add(entry)
    log_action(
        action_type="plan_entry_added",
        entity_type="syllabus_plan",
        entity_label=lesson_title,
        level=level,
        details=f"Added syllabus plan row for week {week_number or '-'}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message="Plan row added.",
        )
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan/save", methods=["POST"])
@teacher_required
def teacher_save_syllabus_plan(level_id: int):
    level = get_teacher_level_or_403(level_id)
    if not level.syllabus_edit_open:
        return redirect(
            url_for(
                "teacher_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="Plan editing is locked.",
            )
        )
    existing_entries = {
        (entry.week_number, entry.session_number): entry
        for entry in SyllabusPlanEntry.query.filter_by(level_id=level.id).all()
    }
    lessons_per_week = get_level_syllabus_lessons_per_week(level)
    slot_sequence = build_instructional_slot_sequence(lessons_per_week=lessons_per_week)
    valid_slots = {(slot["academic_week_number"], slot["session_number"]): slot["slot_order"] for slot in slot_sequence}
    updated_count = 0
    for week_number in range(1, max((slot["academic_week_number"] for slot in slot_sequence), default=0) + 1):
        for session_number in range(1, lessons_per_week + 1):
            slot_order = valid_slots.get((week_number, session_number))
            if not slot_order:
                continue
            book_name = (request.form.get(f"book_name_{week_number}_{session_number}") or "").strip()
            unit_name = (request.form.get(f"unit_name_{week_number}_{session_number}") or "").strip()
            lesson_title = (request.form.get(f"lesson_title_{week_number}_{session_number}") or "").strip()
            source_reference = (request.form.get(f"source_reference_{week_number}_{session_number}") or "").strip()
            learning_objective = (request.form.get(f"learning_objective_{week_number}_{session_number}") or "").strip()
            planned_homework = (request.form.get(f"planned_homework_{week_number}_{session_number}") or "").strip()
            note_text = (request.form.get(f"note_text_{week_number}_{session_number}") or "").strip()

            existing = existing_entries.get((week_number, session_number))
            has_content = any([book_name, unit_name, lesson_title, source_reference, learning_objective, planned_homework, note_text])

            if not has_content and not existing:
                continue

            if not has_content and existing:
                db.session.delete(existing)
                updated_count += 1
                continue

            if not existing:
                existing = SyllabusPlanEntry(
                    level_id=level.id,
                    week_number=week_number,
                    session_number=session_number,
                    book_name=book_name or None,
                    unit_name=unit_name or None,
                    lesson_title=lesson_title,
                    status="planned",
                    order_index=slot_order,
                )
                db.session.add(existing)

            existing.week_number = week_number
            existing.session_number = session_number
            existing.book_name = book_name or None
            existing.unit_name = unit_name or None
            existing.lesson_title = lesson_title
            existing.source_reference = source_reference or None
            existing.learning_objective = learning_objective or None
            existing.planned_homework = planned_homework or None
            existing.note_text = note_text or None
            if existing.status == "completed" and not existing.completed_on:
                existing.completed_on = date.today()
            existing.order_index = slot_order
            updated_count += 1

    log_action(
        action_type="plan_sheet_updated",
        entity_type="syllabus_plan",
        entity_label=level.name,
        level=level,
        details=f"Updated {updated_count} syllabus plan cell(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message="Plan sheet updated.",
        )
    )


@app.route("/teacher/levels/<int:level_id>/syllabus-plan.pdf")
@teacher_required
def teacher_syllabus_plan_pdf(level_id: int):
    level = get_teacher_level_or_403(level_id)
    pdf_bytes = build_syllabus_plan_pdf(level)
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{build_syllabus_export_filename(level, "pdf")}"'},
    )


@app.route("/teacher/syllabus-plan/<int:entry_id>/complete", methods=["POST"])
@teacher_required
def teacher_complete_syllabus_entry(entry_id: int):
    entry = SyllabusPlanEntry.query.get_or_404(entry_id)
    level = get_teacher_level_or_403(entry.level_id)
    complete_previous = (request.form.get("complete_previous") or "").strip() == "1"
    completed_rows = []
    if complete_previous:
        completed_rows = (
            SyllabusPlanEntry.query.filter(
                SyllabusPlanEntry.level_id == level.id,
                SyllabusPlanEntry.order_index <= entry.order_index,
            )
            .order_by(SyllabusPlanEntry.order_index.asc(), SyllabusPlanEntry.id.asc())
            .all()
        )
    else:
        completed_rows = [entry]

    for row in completed_rows:
        row.status = "completed"
        row.completed_on = date.today()

    log_action(
        action_type="plan_entry_completed",
        entity_type="syllabus_plan",
        entity_id=entry.id,
        entity_label=entry.lesson_title,
        level=level,
        details=(
            f"Completed week {entry.week_number} lesson {entry.session_number}"
            + (f" and {max(len(completed_rows) - 1, 0)} previous lesson(s)." if complete_previous else ".")
        ),
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message=(
                f"Lesson marked as completed with {len(completed_rows)} lesson(s) updated."
                if complete_previous
                else "Lesson marked as completed."
            ),
        )
    )


@app.route("/admin/levels/<int:level_id>/syllabus-plan-editing", methods=["POST"])
@admin_required
def admin_toggle_syllabus_plan_editing(level_id: int):
    level = Level.query.get_or_404(level_id)
    level.syllabus_edit_open = (request.form.get("syllabus_edit_open") or "").strip() == "1"
    log_action(
        action_type="syllabus_editing_toggled",
        entity_type="level",
        entity_id=level.id,
        entity_label=level.name,
        level=level,
        details=f"Syllabus editing {'opened' if level.syllabus_edit_open else 'locked'} for teachers.",
    )
    db.session.commit()
    return redirect(url_for("admin_syllabus_plan", level_id=level.id))


@app.route("/admin/levels/<int:level_id>/syllabus-plan/template.xlsx")
@admin_required
def admin_syllabus_plan_template(level_id: int):
    level = Level.query.get_or_404(level_id)
    output = build_syllabus_template_workbook(level)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=build_syllabus_export_filename(level, "xlsx", variant="template"),
    )


@app.route("/admin/levels/<int:level_id>/syllabus-plan.xlsx")
@admin_required
def admin_syllabus_plan_excel(level_id: int):
    level = Level.query.get_or_404(level_id)
    output = build_syllabus_plan_excel_workbook(level)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=build_syllabus_export_filename(level, "xlsx"),
    )


@app.route("/admin/levels/<int:level_id>/syllabus-plan/import", methods=["POST"])
@admin_required
def admin_import_syllabus_plan_template(level_id: int):
    level = Level.query.get_or_404(level_id)
    uploaded_file = request.files.get("plan_template_file")
    if not uploaded_file or not (uploaded_file.filename or "").strip():
        return redirect(
            url_for(
                "admin_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="Excel plan file is required.",
            )
        )
    try:
        parsed_import = parse_syllabus_template_workbook(level, uploaded_file)
    except ValueError as exc:
        return redirect(
            url_for(
                "admin_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message=str(exc),
            )
        )

    existing_token = _get_admin_syllabus_preview_token_map().get(str(level.id))
    if existing_token:
        _syllabus_import_preview_delete(existing_token)
    preview_token = _syllabus_import_preview_store(
        {
            "level_id": level.id,
            "scheduled_rows": parsed_import["scheduled_rows"],
            "reserve_rows": parsed_import["reserve_rows"],
            "scheduled_count": parsed_import["scheduled_count"],
            "reserve_count": parsed_import["reserve_count"],
            "created_count": parsed_import["created_count"],
        }
    )
    _set_admin_syllabus_preview_token(level.id, preview_token)
    return redirect(
        url_for(
            "admin_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message="Import preview is ready. Review it, then apply the import.",
        )
    )


@app.route("/admin/levels/<int:level_id>/syllabus-plan/import/apply", methods=["POST"])
@admin_required
def admin_apply_syllabus_plan_template_import(level_id: int):
    level = Level.query.get_or_404(level_id)
    preview_token = _get_admin_syllabus_preview_token_map().get(str(level.id))
    preview_payload = _syllabus_import_preview_load(preview_token or "")
    if not preview_payload:
        return redirect(
            url_for(
                "admin_syllabus_plan",
                level_id=level.id,
                op_status="error",
                op_message="No import preview is available for this class.",
            )
        )

    result = apply_syllabus_template_import(level, preview_payload)
    log_action(
        action_type="plan_template_imported_by_admin",
        entity_type="syllabus_plan",
        entity_label=level.name,
        level=level,
        details=f"Admin imported {result['scheduled_count']} scheduled row(s) and {result['reserve_count']} reserve row(s) from Excel template.",
    )
    _syllabus_import_preview_delete(preview_token or "")
    _set_admin_syllabus_preview_token(level.id, None)
    db.session.commit()
    return redirect(
        url_for(
            "admin_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message=f"Excel plan imported: {result['scheduled_count']} scheduled row(s), {result['reserve_count']} reserve row(s).",
        )
    )


@app.route("/admin/levels/<int:level_id>/syllabus-plan/import/discard", methods=["POST"])
@admin_required
def admin_discard_syllabus_plan_template_import(level_id: int):
    level = Level.query.get_or_404(level_id)
    preview_token = _get_admin_syllabus_preview_token_map().get(str(level.id))
    _syllabus_import_preview_delete(preview_token or "")
    _set_admin_syllabus_preview_token(level.id, None)
    return redirect(
        url_for(
            "admin_syllabus_plan",
            level_id=level.id,
            op_status="success",
            op_message="Import preview discarded.",
        )
    )


@app.route("/admin/syllabus-plan")
@admin_required
def admin_syllabus_plan():
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    selected_level_id = (request.args.get("level_id") or "").strip()
    summary_rows = build_syllabus_plan_summary(levels)
    selected_summary = None
    academic_calendar_status = build_academic_calendar_status()
    selected_import_preview = None
    selected_import_comparison = None
    if selected_level_id.isdigit():
        selected_summary = next((row for row in summary_rows if row["level"].id == int(selected_level_id)), None)
        preview_token = _get_admin_syllabus_preview_token_map().get(selected_level_id)
        selected_import_preview = _syllabus_import_preview_load(preview_token or "")
        if selected_import_preview and int(selected_import_preview.get("level_id") or 0) != int(selected_level_id):
            selected_import_preview = None
        if selected_summary and selected_import_preview:
            selected_import_comparison = build_syllabus_import_comparison(selected_summary["level"], selected_import_preview)
    selected_lessons_per_week = get_level_syllabus_lessons_per_week(selected_summary["level"]) if selected_summary else 4
    return render_template(
        "admin_syllabus_plan.html",
        levels=levels,
        summary_rows=summary_rows,
        selected_summary=selected_summary,
        selected_grid_rows=build_syllabus_plan_grid(selected_summary["level"], lessons_per_week=selected_lessons_per_week) if selected_summary else [],
        selected_session_numbers=list(range(1, selected_lessons_per_week + 1)),
        selected_level_id=selected_level_id,
        selected_import_preview=selected_import_preview,
        selected_import_comparison=selected_import_comparison,
        current_teaching_week=academic_calendar_status["current_week"],
        academic_calendar_status=academic_calendar_status,
        op_status=(request.args.get("op_status") or "").strip(),
        op_message=(request.args.get("op_message") or "").strip(),
    )


@app.route("/admin/syllabus-plan/current-week", methods=["POST"])
@admin_required
def admin_update_current_teaching_week():
    week_value = (request.form.get("current_teaching_week") or "").strip()
    week_number = max(1, min(40, int(week_value))) if week_value.isdigit() else 1
    set_system_setting("current_teaching_week", str(week_number))
    log_action(
        action_type="calendar_settings_updated",
        entity_type="teaching_week",
        entity_label="current_teaching_week",
        details=f"Updated current teaching week to {week_number}.",
    )
    db.session.commit()
    return redirect(url_for("admin_syllabus_plan", level_id=(request.form.get("level_id") or "").strip() or None))


@app.route("/teacher/levels/<int:level_id>/padlet", methods=["POST"])
@teacher_required
def teacher_update_level_padlet(level_id: int):
    level = get_teacher_level_or_403(level_id)
    homework_padlet_url = (request.form.get("homework_padlet_url") or "").strip()
    announcements_padlet_url = (request.form.get("announcements_padlet_url") or "").strip()

    level.homework_padlet_url = homework_padlet_url or None
    level.announcements_padlet_url = announcements_padlet_url or None
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Padlet links updated.",
        )
    )


@app.route("/teacher/levels/<int:level_id>/recordings/add", methods=["POST"])
@teacher_required
def teacher_add_recording(level_id: int):
    level = get_teacher_level_or_403(level_id)
    title = (request.form.get("title") or "").strip()
    recording_url = (request.form.get("recording_url") or "").strip()
    lesson_date_str = (request.form.get("lesson_date") or "").strip()
    summary = (request.form.get("summary") or "").strip()
    homework = (request.form.get("homework") or "").strip()

    if not title or not recording_url or not lesson_date_str:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Title, URL, and lesson date are required.",
            )
        )

    try:
        lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
    except ValueError:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Invalid lesson date format.",
            )
        )

    db.session.add(
        ClassRecording(
            class_id=level.id,
            title=title,
            recording_url=recording_url,
            lesson_date=lesson_date,
            summary=summary or None,
            homework=homework or None,
        )
    )
    log_action(
        action_type="recording_added",
        entity_type="recording",
        entity_label=title,
        level=level,
        details=f"Added lesson recording for {lesson_date.isoformat()}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Recording added.",
        )
    )


@app.route("/teacher/levels/<int:level_id>/assignments/add", methods=["POST"])
@teacher_required
def teacher_add_assignment(level_id: int):
    level = get_teacher_level_or_403(level_id)
    title = (request.form.get("title") or "").strip()
    due_date_str = (request.form.get("due_date") or "").strip()
    instructions = (request.form.get("instructions") or "").strip()
    resource_link = (request.form.get("resource_link") or "").strip()
    resource_file = request.files.get("resource_file")

    if not title:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Assignment title is required.",
            )
        )

    due_date = None
    if due_date_str:
        try:
            due_date = datetime.strptime(due_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect(
                url_for(
                    "teacher_level_workspace",
                    level_id=level.id,
                    op_status="error",
                    op_message="Invalid assignment due date format.",
                )
            )

    resource_file_name = None
    resource_file_path = None
    if resource_file and (resource_file.filename or "").strip():
        resource_file_path, resource_file_name = save_assignment_uploaded_file(resource_file, "resources")

    assignment = Assignment(
        level_id=level.id,
        title=title,
        instructions=instructions or None,
        resource_link=resource_link or None,
        resource_file_name=resource_file_name,
        resource_file_path=resource_file_path,
        due_date=due_date,
        is_active=True,
    )
    db.session.add(assignment)
    log_action(
        action_type="assignment_added",
        entity_type="assignment",
        entity_label=title,
        level=level,
        details=f"Added assignment{f' due {due_date.isoformat()}' if due_date else ''}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            assignment_id=assignment.id,
            op_status="success",
            op_message="Assignment added.",
        )
    )


@app.route("/students/<int:student_id>/assignments/<int:assignment_id>/submit", methods=["POST"])
def student_submit_assignment(student_id: int, assignment_id: int):
    student = Student.query.get_or_404(student_id)
    student_code = (request.form.get("student_code") or "").strip()
    if student_code != student.student_code:
        abort(403)

    assignment = Assignment.query.get_or_404(assignment_id)
    if not student.level_id or assignment.level_id != student.level_id:
        abort(403)

    submission_text = (request.form.get("submission_text") or "").strip()
    submission_link = (request.form.get("submission_link") or "").strip()
    submission_file = request.files.get("submission_file")

    existing_submission = AssignmentSubmission.query.filter_by(
        assignment_id=assignment.id,
        student_id=student.id,
    ).first()
    if not existing_submission:
        existing_submission = AssignmentSubmission(
            assignment_id=assignment.id,
            student_id=student.id,
        )
        db.session.add(existing_submission)

    if submission_file and (submission_file.filename or "").strip():
        submission_file_path, submission_file_name = save_assignment_uploaded_file(submission_file, "submissions")
        existing_submission.submission_file_path = submission_file_path
        existing_submission.submission_file_name = submission_file_name

    existing_submission.submission_text = submission_text or None
    existing_submission.submission_link = submission_link or None
    existing_submission.status = "Submitted"
    existing_submission.submitted_at = datetime.utcnow()

    db.session.commit()
    return redirect(
        url_for(
            "student_dashboard",
            student_code=student.student_code,
        )
    )


@app.route("/teacher/assignments/submissions/bulk-update", methods=["POST"])
@teacher_required
def teacher_bulk_update_assignment_submissions():
    submission_ids = [int(value) for value in request.form.getlist("submission_ids") if value.isdigit()]
    assignment_id_str = (request.form.get("assignment_id") or "").strip()
    if not assignment_id_str.isdigit():
        return redirect(url_for("teacher_dashboard"))

    assignment = Assignment.query.get_or_404(int(assignment_id_str))
    level = get_teacher_level_or_403(assignment.level_id)
    if not submission_ids:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                assignment_id=assignment.id,
                op_status="error",
                op_message="No assignment submissions were available to update.",
            )
        )

    submissions = AssignmentSubmission.query.filter(AssignmentSubmission.id.in_(submission_ids)).all()
    updated_submissions = 0
    for submission in submissions:
        if submission.assignment_id != assignment.id:
            abort(403)
        submission.score_value = (request.form.get(f"score_value_{submission.id}") or "").strip() or None
        submission.teacher_notes = (request.form.get(f"teacher_notes_{submission.id}") or "").strip() or None
        status = (request.form.get(f"status_{submission.id}") or "").strip()
        submission.status = status or submission.status or "Submitted"
        if submission.score_value or submission.teacher_notes or submission.status == "Reviewed":
            submission.reviewed_at = datetime.utcnow()
        updated_submissions += 1

    log_action(
        action_type="assignment_reviews_updated",
        entity_type="assignment_review",
        entity_id=assignment.id,
        entity_label=assignment.title,
        level=level,
        details=f"Updated {updated_submissions} submission review record(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            assignment_id=assignment.id,
            op_status="success",
            op_message="Assignment submissions updated.",
        )
    )


@app.route("/teacher/assignments/<int:assignment_id>/resource")
@teacher_required
def teacher_assignment_resource_file(assignment_id: int):
    assignment = Assignment.query.get_or_404(assignment_id)
    get_teacher_level_or_403(assignment.level_id)
    absolute_path = assignment_file_absolute_path(assignment.resource_file_path)
    if not absolute_path or not os.path.exists(absolute_path):
        abort(404)
    return send_file(absolute_path, as_attachment=True, download_name=assignment.resource_file_name or "assignment_resource")


@app.route("/students/<int:student_id>/assignments/<int:assignment_id>/resource")
def student_assignment_resource_file(student_id: int, assignment_id: int):
    student = Student.query.get_or_404(student_id)
    student_code = (request.args.get("student_code") or "").strip()
    if student_code != student.student_code:
        abort(403)
    assignment = Assignment.query.get_or_404(assignment_id)
    if not student.level_id or assignment.level_id != student.level_id:
        abort(403)
    absolute_path = assignment_file_absolute_path(assignment.resource_file_path)
    if not absolute_path or not os.path.exists(absolute_path):
        abort(404)
    return send_file(absolute_path, as_attachment=True, download_name=assignment.resource_file_name or "assignment_resource")


@app.route("/teacher/assignment-submissions/<int:submission_id>/file")
@teacher_required
def teacher_assignment_submission_file(submission_id: int):
    submission = AssignmentSubmission.query.get_or_404(submission_id)
    assignment = Assignment.query.get_or_404(submission.assignment_id)
    get_teacher_level_or_403(assignment.level_id)
    absolute_path = assignment_file_absolute_path(submission.submission_file_path)
    if not absolute_path or not os.path.exists(absolute_path):
        abort(404)
    return send_file(absolute_path, as_attachment=True, download_name=submission.submission_file_name or "student_submission")


@app.route("/students/<int:student_id>/assignment-submissions/<int:submission_id>/file")
def student_assignment_submission_file(student_id: int, submission_id: int):
    student = Student.query.get_or_404(student_id)
    student_code = (request.args.get("student_code") or "").strip()
    if student_code != student.student_code:
        abort(403)
    submission = AssignmentSubmission.query.get_or_404(submission_id)
    if submission.student_id != student.id:
        abort(403)
    absolute_path = assignment_file_absolute_path(submission.submission_file_path)
    if not absolute_path or not os.path.exists(absolute_path):
        abort(404)
    return send_file(absolute_path, as_attachment=True, download_name=submission.submission_file_name or "my_submission")


@app.route("/teacher/levels/<int:level_id>/template-results/save", methods=["POST"])
@teacher_required
def teacher_save_template_results(level_id: int):
    level = get_teacher_level_or_403(level_id)
    template_id = (request.form.get("exam_template_id") or "").strip()
    if not template_id.isdigit():
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Exam template is required.",
            )
        )

    template = ExamTemplate.query.get_or_404(int(template_id))
    ensure_exam_publication_row(template.title, default_published=False)
    level_students = Student.query.filter_by(level_id=level.id).all()
    students_by_id = {student.id: student for student in level_students}
    student_ids = [
        int(student_id)
        for student_id in request.form.getlist("student_ids")
        if student_id.isdigit() and int(student_id) in students_by_id
    ]
    if not student_ids:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                exam_template_id=template.id,
                op_status="error",
                op_message="No students were submitted for exam entry.",
            )
        )

    template_branches = sorted(template.branches, key=lambda item: ((item.order_index or 0), item.id))
    existing_results = ExamResult.query.filter(
        ExamResult.student_id.in_(student_ids),
        ExamResult.exam_title == template.title,
    ).all()
    existing_results_by_key = {
        (result.student_id, result.subject_name): result
        for result in existing_results
    }

    saved_students = 0
    for student_id in student_ids:
        student = students_by_id[student_id]
        shared_note = (request.form.get(f"student_{student.id}_note") or "").strip()
        student_row_saved = False

        for branch in template_branches:
            score_value = (request.form.get(f"student_{student.id}_branch_{branch.id}_score") or "").strip()
            existing_result = existing_results_by_key.get((student.id, branch.branch_name))

            if not score_value and not existing_result:
                continue

            if not existing_result:
                existing_result = ExamResult(
                    student_id=student.id,
                    level_id=level.id,
                    exam_title=template.title,
                    subject_name=branch.branch_name,
                )
                db.session.add(existing_result)
                existing_results_by_key[(student.id, branch.branch_name)] = existing_result

            existing_result.level_id = level.id
            existing_result.exam_date = template.exam_date
            existing_result.score_value = score_value or "0"
            existing_result.max_score = branch.max_score or None
            existing_result.notes = shared_note or None
            student_row_saved = True

        if student_row_saved:
            saved_students += 1

    log_action(
        action_type="template_results_saved",
        entity_type="exam_template",
        entity_id=template.id,
        entity_label=template.title,
        level=level,
        details=f"Saved exam template results for {saved_students} student(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            exam_template_id=template.id,
            op_status="success",
            op_message=f"Saved {template.title} results for {saved_students} student(s).",
        )
    )


@app.route("/teacher/levels/<int:level_id>/upcoming-exams/add", methods=["POST"])
@teacher_required
def teacher_add_upcoming_exam(level_id: int):
    level = get_teacher_level_or_403(level_id)
    title = (request.form.get("title") or "").strip()
    subject_name = (request.form.get("subject_name") or "").strip()
    exam_date_str = (request.form.get("exam_date") or "").strip()
    exam_time = (request.form.get("exam_time") or "").strip()
    notes = (request.form.get("notes") or "").strip()

    if not title or not exam_date_str:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Upcoming exam title and date are required.",
            )
        )

    try:
        exam_date = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
    except ValueError:
        return redirect(
            url_for(
                "teacher_level_workspace",
                level_id=level.id,
                op_status="error",
                op_message="Invalid upcoming exam date format.",
            )
        )

    db.session.add(
        UpcomingExam(
            level_id=level.id,
            title=title,
            subject_name=subject_name or None,
            exam_date=exam_date,
            exam_time=exam_time or None,
            notes=notes or None,
        )
    )
    log_action(
        action_type="upcoming_exam_added",
        entity_type="upcoming_exam",
        entity_label=title,
        level=level,
        details=f"Added upcoming exam for {exam_date.isoformat()}.",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Upcoming exam added.",
        )
    )


@app.route("/teacher/upcoming-exams/bulk-update", methods=["POST"])
@teacher_required
def teacher_bulk_update_upcoming_exams():
    upcoming_exam_ids = [int(value) for value in request.form.getlist("upcoming_exam_ids") if value.isdigit()]
    if not upcoming_exam_ids:
        return redirect(url_for("teacher_dashboard"))

    upcoming_exams = UpcomingExam.query.filter(UpcomingExam.id.in_(upcoming_exam_ids)).all()
    if not upcoming_exams:
        return redirect(url_for("teacher_dashboard"))

    level = get_teacher_level_or_403(upcoming_exams[0].level_id)
    for upcoming_exam in upcoming_exams:
        if upcoming_exam.level_id != level.id:
            abort(403)
        title = (request.form.get(f"title_{upcoming_exam.id}") or "").strip()
        subject_name = (request.form.get(f"subject_name_{upcoming_exam.id}") or "").strip()
        exam_date_str = (request.form.get(f"exam_date_{upcoming_exam.id}") or "").strip()
        exam_time = (request.form.get(f"exam_time_{upcoming_exam.id}") or "").strip()
        notes = (request.form.get(f"notes_{upcoming_exam.id}") or "").strip()

        if not title or not exam_date_str:
            return redirect(
                url_for(
                    "teacher_level_workspace",
                    level_id=level.id,
                    op_status="error",
                    op_message="Each upcoming exam needs a title and date.",
                )
            )

        try:
            upcoming_exam.exam_date = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect(
                url_for(
                    "teacher_level_workspace",
                    level_id=level.id,
                    op_status="error",
                    op_message="Invalid upcoming exam date format.",
                )
            )

        upcoming_exam.title = title
        upcoming_exam.subject_name = subject_name or None
        upcoming_exam.exam_time = exam_time or None
        upcoming_exam.notes = notes or None

    log_action(
        action_type="upcoming_exams_updated",
        entity_type="upcoming_exam",
        entity_label=level.name,
        level=level,
        details=f"Updated {len(upcoming_exams)} upcoming exam record(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Upcoming exams updated.",
        )
    )


@app.route("/teacher/upcoming-exams/<int:upcoming_exam_id>/delete", methods=["POST"])
@teacher_required
def teacher_delete_upcoming_exam(upcoming_exam_id: int):
    upcoming_exam = UpcomingExam.query.get_or_404(upcoming_exam_id)
    level = get_teacher_level_or_403(upcoming_exam.level_id)
    db.session.delete(upcoming_exam)
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Upcoming exam deleted.",
        )
    )


@app.route("/teacher/exam-results/bulk-update", methods=["POST"])
@teacher_required
def teacher_bulk_update_exam_results():
    exam_result_ids = [
        int(value)
        for value in request.form.getlist("exam_result_ids")
        if value.isdigit()
    ]
    if not exam_result_ids:
        return redirect(url_for("teacher_dashboard"))

    exam_results = ExamResult.query.filter(ExamResult.id.in_(exam_result_ids)).all()
    if not exam_results:
        return redirect(url_for("teacher_dashboard"))

    first_result = exam_results[0]
    if first_result.level_id:
        level = get_teacher_level_or_403(first_result.level_id)
    else:
        student = Student.query.get_or_404(first_result.student_id)
        if not student.level_id:
            abort(403)
        level = get_teacher_level_or_403(student.level_id)

    for exam_result in exam_results:
        if exam_result.level_id and exam_result.level_id != level.id:
            abort(403)
        if is_exam_total_subject(exam_result.subject_name) or is_exam_percentage_subject(exam_result.subject_name):
            continue
        exam_result.score_value = (request.form.get(f"score_value_{exam_result.id}") or "").strip()
        exam_result.max_score = (request.form.get(f"max_score_{exam_result.id}") or "").strip() or None
        exam_result.notes = (request.form.get(f"notes_{exam_result.id}") or "").strip() or None

    results_exam_title = (request.form.get("results_exam_title") or "").strip()
    log_action(
        action_type="exam_results_updated",
        entity_type="exam_result",
        entity_label=results_exam_title or first_result.exam_title,
        level=level,
        details=f"Updated {len(exam_results)} exam result component(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            results_exam_title=results_exam_title or None,
            op_status="success",
            op_message="Exam results updated.",
        )
    )


@app.route("/teacher/recordings/bulk-update", methods=["POST"])
@teacher_required
def teacher_bulk_update_recordings():
    recording_ids = [int(value) for value in request.form.getlist("recording_ids") if value.isdigit()]
    if not recording_ids:
        return redirect(url_for("teacher_dashboard"))

    recordings = ClassRecording.query.filter(ClassRecording.id.in_(recording_ids)).all()
    if not recordings:
        return redirect(url_for("teacher_dashboard"))

    level = get_teacher_level_or_403(recordings[0].class_id)
    for recording in recordings:
        if recording.class_id != level.id:
            abort(403)
        title = (request.form.get(f"title_{recording.id}") or "").strip()
        recording_url = (request.form.get(f"recording_url_{recording.id}") or "").strip()
        lesson_date_str = (request.form.get(f"lesson_date_{recording.id}") or "").strip()
        summary = (request.form.get(f"summary_{recording.id}") or "").strip()
        homework = (request.form.get(f"homework_{recording.id}") or "").strip()

        if not title or not recording_url or not lesson_date_str:
            return redirect(
                url_for(
                    "teacher_level_workspace",
                    level_id=level.id,
                    op_status="error",
                    op_message="Each recording needs title, URL, and lesson date.",
                )
            )

        try:
            lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect(
                url_for(
                    "teacher_level_workspace",
                    level_id=level.id,
                    op_status="error",
                    op_message="Invalid lesson date format.",
                )
            )

        recording.title = title
        recording.recording_url = recording_url
        recording.lesson_date = lesson_date
        recording.summary = summary or None
        recording.homework = homework or None

    log_action(
        action_type="recordings_updated",
        entity_type="recording",
        entity_label=level.name,
        level=level,
        details=f"Updated {len(recordings)} lesson recording row(s).",
    )
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Recordings updated.",
        )
    )


@app.route("/teacher/recordings/<int:recording_id>/delete", methods=["POST"])
@teacher_required
def teacher_delete_recording(recording_id: int):
    recording = ClassRecording.query.get_or_404(recording_id)
    level = get_teacher_level_or_403(recording.class_id)
    db.session.delete(recording)
    db.session.commit()
    return redirect(
        url_for(
            "teacher_level_workspace",
            level_id=level.id,
            op_status="success",
            op_message="Recording deleted.",
        )
    )


@app.route("/head/data-root")
@admin_required
def head_data_root():
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    teachers = Teacher.query.order_by(Teacher.full_name).all()
    students_query = Student.query.outerjoin(Level, Student.level_id == Level.id)

    search = (request.args.get("search") or "").strip()
    level_id_filter = (request.args.get("level_id") or "").strip()
    email_filter = (request.args.get("email_filter") or "all").strip().lower()
    sort_by = (request.args.get("sort_by") or "name_asc").strip().lower()
    student_status_filter = (request.args.get("student_status_filter") or "all").strip().lower()
    student_year_filter = (request.args.get("student_year_filter") or "all").strip().lower()
    teacher_search = (request.args.get("teacher_search") or "").strip()
    teacher_status_filter = (request.args.get("teacher_status_filter") or "all").strip().lower()
    level_search = (request.args.get("level_search") or "").strip()
    level_health_filter = (request.args.get("level_health_filter") or "all").strip().lower()
    cleanup_scope = (request.args.get("cleanup_scope") or "all").strip().lower()
    cleanup_priority_filter = (request.args.get("cleanup_priority_filter") or "all").strip().lower()

    if search:
        pattern = f"%{search}%"
        students_query = students_query.filter(
            db.or_(Student.full_name.ilike(pattern), Student.student_code.ilike(pattern))
        )

    if level_id_filter.isdigit():
        students_query = students_query.filter(Student.level_id == int(level_id_filter))

    if student_status_filter != "all":
        students_query = students_query.filter(db.func.lower(db.func.coalesce(Student.status, "")) == student_status_filter)

    if student_year_filter != "all":
        students_query = students_query.filter(db.func.lower(db.func.coalesce(Student.student_year, "")) == student_year_filter)

    if email_filter == "missing":
        students_query = students_query.filter(
            db.or_(Student.parent_email.is_(None), Student.parent_email == "")
        )
    elif email_filter == "has":
        students_query = students_query.filter(
            Student.parent_email.isnot(None), Student.parent_email != ""
        )

    if sort_by == "name_desc":
        students_query = students_query.order_by(Student.full_name.desc())
    elif sort_by == "code_asc":
        students_query = students_query.order_by(Student.student_code.asc())
    elif sort_by == "code_desc":
        students_query = students_query.order_by(Student.student_code.desc())
    elif sort_by == "level_asc":
        students_query = students_query.order_by(Level.order_index.asc(), Student.full_name.asc())
    elif sort_by == "level_desc":
        students_query = students_query.order_by(Level.order_index.desc(), Student.full_name.asc())
    else:
        students_query = students_query.order_by(Student.full_name.asc())

    students = students_query.all()
    student_rows = []
    for student in students:
        student_flags = []
        if not student.parent_email:
            student_flags.append("missing_parent_email")
        if not student.level_id:
            student_flags.append("missing_level")
        if not student.status:
            student_flags.append("missing_status")
        if not student.student_year:
            student_flags.append("missing_year")
        student_rows.append(
            {
                "student": student,
                "flags": student_flags,
            }
        )

    all_students = Student.query.all()
    student_summary = {
        "total": len(all_students),
        "missing_parent_email": sum(1 for student in all_students if not student.parent_email),
        "missing_parent_whatsapp": sum(1 for student in all_students if not student.parent_whatsapp),
        "missing_status": sum(1 for student in all_students if not student.status),
        "missing_year": sum(1 for student in all_students if not student.student_year),
        "filtered_total": len(student_rows),
    }
    level_students_count = {
        row.level_id: row.students_count
        for row in (
            db.session.query(
                Student.level_id,
                db.func.count(Student.id).label("students_count"),
            )
            .group_by(Student.level_id)
            .all()
        )
        if row.level_id
    }
    filtered_teachers = teachers
    if teacher_search:
        teacher_search_lower = teacher_search.lower()
        filtered_teachers = [
            teacher
            for teacher in filtered_teachers
            if teacher_search_lower in (teacher.full_name or "").lower()
            or teacher_search_lower in (teacher.email or "").lower()
            or teacher_search_lower in (teacher.phone or "").lower()
            or teacher_search_lower in (teacher.subject_name or "").lower()
        ]
    if teacher_status_filter != "all":
        filtered_teachers = [
            teacher
            for teacher in filtered_teachers
            if (teacher.status or "").strip().lower() == teacher_status_filter
        ]

    teacher_rows = []
    for teacher in filtered_teachers:
        assigned_levels = [level for level in levels if level.teacher_id == teacher.id]
        teacher_rows.append(
            {
                "teacher": teacher,
                "assigned_levels": assigned_levels,
                "levels_count": len(assigned_levels),
                "students_count": sum(level_students_count.get(level.id, 0) for level in assigned_levels),
            }
        )

    teacher_summary = {
        "total": len(teachers),
        "active": sum(1 for teacher in teachers if (teacher.status or "").strip().lower() == "active"),
        "without_levels": sum(1 for teacher in teachers if not any(level.teacher_id == teacher.id for level in levels)),
        "filtered_total": len(teacher_rows),
    }

    level_rows = []
    filtered_levels = levels
    if level_search:
        level_search_lower = level_search.lower()
        filtered_levels = [
            level
            for level in filtered_levels
            if level_search_lower in (level.name or "").lower()
            or level_search_lower in (level.teacher.full_name if level.teacher else "").lower()
        ]

    for level in filtered_levels:
        health_flags = []
        if not level.teacher_id:
            health_flags.append("missing_teacher")
        if not level.zoom_link:
            health_flags.append("missing_zoom")
        if not level.homework_padlet_url:
            health_flags.append("missing_homework_board")
        if not level.announcements_padlet_url:
            health_flags.append("missing_class_board")

        if level_health_filter != "all" and level_health_filter not in health_flags:
            continue

        level_rows.append(
            {
                "level": level,
                "students_count": level_students_count.get(level.id, 0),
                "health_flags": health_flags,
                "is_complete": not health_flags,
            }
        )

    level_summary = {
        "total": len(levels),
        "complete": sum(
            1
            for level in levels
            if level.teacher_id and level.zoom_link and level.homework_padlet_url and level.announcements_padlet_url
        ),
        "missing_teacher": sum(1 for level in levels if not level.teacher_id),
        "missing_zoom": sum(1 for level in levels if not level.zoom_link),
        "filtered_total": len(level_rows),
    }

    cleanup_data = build_cleanup_center_data(levels, teachers, students)
    cleanup_students = cleanup_data["cleanup_students"]
    cleanup_teachers = cleanup_data["cleanup_teachers"]
    cleanup_levels = cleanup_data["cleanup_levels"]

    if cleanup_priority_filter != "all":
        cleanup_students = [row for row in cleanup_students if row["priority"]["tone"] == cleanup_priority_filter]
        cleanup_teachers = [row for row in cleanup_teachers if row["priority"]["tone"] == cleanup_priority_filter]
        cleanup_levels = [row for row in cleanup_levels if row["priority"]["tone"] == cleanup_priority_filter]

    cleanup_sections = {
        "students": cleanup_scope in {"all", "students"},
        "teachers": cleanup_scope in {"all", "teachers"},
        "levels": cleanup_scope in {"all", "levels"},
    }

    cleanup_summary = {
        "student_records": len(cleanup_students),
        "teacher_records": len(cleanup_teachers),
        "level_records": len(cleanup_levels),
        "critical_items": sum(
            1
            for row in cleanup_students + cleanup_teachers + cleanup_levels
            if row["priority"]["tone"] == "critical"
        ),
        "total_items": len(cleanup_students) + len(cleanup_teachers) + len(cleanup_levels),
    }
    recordings = (
        db.session.query(ClassRecording, Level)
        .join(Level, ClassRecording.class_id == Level.id)
        .order_by(ClassRecording.lesson_date.desc(), Level.order_index.asc(), ClassRecording.id.desc())
        .all()
    )
    op_status = (request.args.get("op_status") or "").strip()
    op_message = (request.args.get("op_message") or "").strip()
    section = normalize_data_root_section(request.args.get("section") or "students")
    return render_template(
        "head_data_root.html",
        students=students,
        levels=levels,
        teachers=teachers,
        recordings=recordings,
        op_status=op_status,
        op_message=op_message,
        section=section,
        search=search,
        level_id_filter=level_id_filter,
        email_filter=email_filter,
        sort_by=sort_by,
        student_status_filter=student_status_filter,
        student_year_filter=student_year_filter,
        student_rows=student_rows,
        student_summary=student_summary,
        teacher_search=teacher_search,
        teacher_status_filter=teacher_status_filter,
        teacher_rows=teacher_rows,
        teacher_summary=teacher_summary,
        level_search=level_search,
        level_health_filter=level_health_filter,
        level_rows=level_rows,
        level_summary=level_summary,
        cleanup_students=cleanup_students,
        cleanup_teachers=cleanup_teachers,
        cleanup_levels=cleanup_levels,
        cleanup_summary=cleanup_summary,
        cleanup_scope=cleanup_scope,
        cleanup_priority_filter=cleanup_priority_filter,
        cleanup_sections=cleanup_sections,
    )


@app.route("/students")
@admin_required
def students_list():
    students = Student.query.order_by(Student.level_id.asc(), Student.full_name).all()
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    bulk_status = (request.args.get("bulk_status") or "").strip()
    bulk_message = (request.args.get("bulk_message") or "").strip()
    arabic_attendance_import_summary = build_imported_arabic_attendance_summary()
    return render_template(
        "students.html",
        students=students,
        levels=levels,
        bulk_status=bulk_status,
        bulk_message=bulk_message,
        arabic_attendance_import_summary=arabic_attendance_import_summary,
    )


@app.route("/students/add", methods=["POST"])
@admin_required
def add_student():
    full_name = (request.form.get("full_name") or "").strip()
    level_id_str = (request.form.get("level_id") or "").strip()
    status = (request.form.get("status") or "").strip()
    student_year = (request.form.get("student_year") or "").strip()
    parent_email = (request.form.get("parent_email") or "").strip()
    parent_whatsapp = (request.form.get("parent_whatsapp") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "students")

    if not full_name:
        return redirect_head_data_root("error", "Student name is required.", section)
    if parent_email and not validate_email(parent_email):
        return redirect_head_data_root("error", "Invalid parent email address.", section)
    if parent_whatsapp and not validate_whatsapp(parent_whatsapp):
        return redirect_head_data_root("error", "Invalid parent WhatsApp number.", section)

    level = None
    if level_id_str.isdigit():
        level = Level.query.get(int(level_id_str))
    if not level:
        return redirect_head_data_root("error", "Please select a valid level.", section)

    student = Student(
        student_code=generate_next_student_code(),
        full_name=full_name,
        status=status or None,
        student_year=student_year or None,
        parent_email=parent_email or None,
        parent_whatsapp=parent_whatsapp or None,
        level_id=level.id,
        level_name=level.name,
    )
    db.session.add(student)
    log_action(
        action_type="student_added",
        entity_type="student",
        entity_label=full_name,
        level=level,
        details=f"Added student with code {student.student_code}.",
    )
    db.session.commit()

    return redirect_head_data_root("success", f"Student added: {student.full_name}.", section)


@app.route("/students/<int:student_id>/update", methods=["POST"])
@admin_required
def update_student(student_id: int):
    student = Student.query.get_or_404(student_id)
    full_name = (request.form.get("full_name") or "").strip()
    level_id_str = (request.form.get("level_id") or "").strip()
    status = request.form.get("status") if "status" in request.form else None
    student_year = request.form.get("student_year") if "student_year" in request.form else None
    parent_email = request.form.get("parent_email") if "parent_email" in request.form else None
    parent_whatsapp = request.form.get("parent_whatsapp") if "parent_whatsapp" in request.form else None
    section = normalize_data_root_section(request.form.get("section") or "students")

    if not full_name:
        return redirect_head_data_root("error", "Student name cannot be empty.", section)
    if not level_id_str.isdigit():
        return redirect_head_data_root("error", "Please select a valid level.", section)
    if parent_email is not None and parent_email.strip() and not validate_email(parent_email.strip()):
        return redirect_head_data_root("error", "Invalid parent email address.", section)
    if parent_whatsapp is not None and parent_whatsapp.strip() and not validate_whatsapp(parent_whatsapp.strip()):
        return redirect_head_data_root("error", "Invalid parent WhatsApp number.", section)

    level = Level.query.get(int(level_id_str))
    if not level:
        return redirect_head_data_root("error", "Selected level was not found.", section)

    student.full_name = full_name
    student.level_id = level.id
    student.level_name = level.name
    if status is not None:
        student.status = status.strip() or None
    if student_year is not None:
        student.student_year = student_year.strip() or None
    if parent_email is not None:
        student.parent_email = parent_email.strip() or None
    if parent_whatsapp is not None:
        student.parent_whatsapp = parent_whatsapp.strip() or None
    log_action(
        action_type="student_updated",
        entity_type="student",
        entity_id=student.id,
        entity_label=student.full_name,
        level=level,
        details=f"Updated student profile and level assignment to {level.name}.",
    )
    db.session.commit()

    return redirect_head_data_root("success", f"Student updated: {student.full_name}.", section)


@app.route("/students/<int:student_id>/delete", methods=["POST"])
@admin_required
def delete_student(student_id: int):
    student = Student.query.get_or_404(student_id)
    student_name = student.full_name
    section = normalize_data_root_section(request.form.get("section") or "students")

    Attendance.query.filter_by(student_id=student.id).delete()
    StudentMonthlyNote.query.filter_by(student_id=student.id).delete()
    db.session.delete(student)
    db.session.commit()

    return redirect_head_data_root("success", f"Student deleted: {student_name}.", section)


@app.route("/students/<int:student_id>/parent-email", methods=["POST"])
@admin_required
def save_parent_email(student_id: int):
    student = Student.query.get_or_404(student_id)
    parent_email = (request.form.get("parent_email") or "").strip()
    parent_whatsapp = (request.form.get("parent_whatsapp") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "students")

    if parent_email and not validate_email(parent_email):
        return redirect_head_data_root("error", f"Invalid email for {student.full_name}.", section)
    if parent_whatsapp and not validate_whatsapp(parent_whatsapp):
        return redirect_head_data_root("error", f"Invalid WhatsApp for {student.full_name}.", section)

    student.parent_email = parent_email or None
    student.parent_whatsapp = parent_whatsapp or None
    log_action(
        action_type="parent_contact_saved",
        entity_type="student_contact",
        entity_id=student.id,
        entity_label=student.full_name,
        level=student.level,
        details="Saved parent email and/or WhatsApp contact.",
    )
    db.session.commit()

    return redirect_head_data_root("success", f"Parent contact saved for {student.full_name}.", section)


@app.route("/cleanup/students/<int:student_id>/quick-fix", methods=["POST"])
@admin_required
def cleanup_quick_fix_student(student_id: int):
    student = Student.query.get_or_404(student_id)
    parent_email = (request.form.get("parent_email") or "").strip()
    parent_whatsapp = (request.form.get("parent_whatsapp") or "").strip()
    status = (request.form.get("status") or "").strip()
    student_year = (request.form.get("student_year") or "").strip()

    if parent_email and not validate_email(parent_email):
        return redirect_head_data_root("error", f"Invalid parent email for {student.full_name}.", "cleanup")
    if parent_whatsapp and not validate_whatsapp(parent_whatsapp):
        return redirect_head_data_root("error", f"Invalid parent WhatsApp for {student.full_name}.", "cleanup")

    if "parent_email" in request.form:
        student.parent_email = parent_email or None
    if "parent_whatsapp" in request.form:
        student.parent_whatsapp = parent_whatsapp or None
    if "status" in request.form:
        student.status = status or None
    if "student_year" in request.form:
        student.student_year = student_year or None

    db.session.commit()
    return redirect_head_data_root("success", f"Student cleanup updated: {student.full_name}.", "cleanup")


@app.route("/teachers/add", methods=["POST"])
@admin_required
def add_teacher():
    full_name = (request.form.get("full_name") or "").strip()
    subject_name = (request.form.get("subject_name") or "").strip()
    phone = (request.form.get("phone") or "").strip()
    email = (request.form.get("email") or "").strip()
    status = (request.form.get("status") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "teachers")

    if not full_name:
        return redirect_head_data_root("error", "Teacher name is required.", section)
    if email and not validate_email(email):
        return redirect_head_data_root("error", "Invalid teacher email address.", section)

    teacher = Teacher(
        full_name=full_name,
        subject_name=subject_name or None,
        phone=phone or None,
        email=email or None,
        status=status or None,
    )
    db.session.add(teacher)
    db.session.commit()
    return redirect_head_data_root("success", f"Teacher added: {teacher.full_name}.", section)


@app.route("/teachers/<int:teacher_id>/update", methods=["POST"])
@admin_required
def update_teacher(teacher_id: int):
    teacher = Teacher.query.get_or_404(teacher_id)
    full_name = (request.form.get("full_name") or "").strip()
    subject_name = (request.form.get("subject_name") or "").strip()
    phone = (request.form.get("phone") or "").strip()
    email = (request.form.get("email") or "").strip()
    status = (request.form.get("status") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "teachers")

    if not full_name:
        return redirect_head_data_root("error", "Teacher name cannot be empty.", section)
    if email and not validate_email(email):
        return redirect_head_data_root("error", "Invalid teacher email address.", section)

    teacher.full_name = full_name
    teacher.subject_name = subject_name or None
    teacher.phone = phone or None
    teacher.email = email or None
    teacher.status = status or None
    db.session.commit()
    return redirect_head_data_root("success", f"Teacher updated: {teacher.full_name}.", section)


@app.route("/cleanup/teachers/<int:teacher_id>/quick-fix", methods=["POST"])
@admin_required
def cleanup_quick_fix_teacher(teacher_id: int):
    teacher = Teacher.query.get_or_404(teacher_id)
    phone = (request.form.get("phone") or "").strip()
    email = (request.form.get("email") or "").strip()
    status = (request.form.get("status") or "").strip()

    if email and not validate_email(email):
        return redirect_head_data_root("error", f"Invalid teacher email for {teacher.full_name}.", "cleanup")

    if "phone" in request.form:
        teacher.phone = phone or None
    if "email" in request.form:
        teacher.email = email or None
    if "status" in request.form:
        teacher.status = status or None

    db.session.commit()
    return redirect_head_data_root("success", f"Teacher cleanup updated: {teacher.full_name}.", "cleanup")


@app.route("/levels/<int:level_id>/update", methods=["POST"])
@admin_required
def update_level(level_id: int):
    level = Level.query.get_or_404(level_id)
    old_name = level.name
    name = (request.form.get("name") or "").strip()
    teacher_id_str = (request.form.get("teacher_id") or "").strip()
    order_index_str = (request.form.get("order_index") or "").strip()
    zoom_email = (request.form.get("zoom_email") or "").strip()
    zoom_link = (request.form.get("zoom_link") or "").strip()
    zoom_meeting_id = (request.form.get("zoom_meeting_id") or "").strip()
    zoom_passcode = (request.form.get("zoom_passcode") or "").strip()
    homework_padlet_url = (request.form.get("homework_padlet_url") or "").strip()
    announcements_padlet_url = (request.form.get("announcements_padlet_url") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "levels")

    if not name:
        return redirect_head_data_root("error", "Level name cannot be empty.", section)

    teacher_id = int(teacher_id_str) if teacher_id_str.isdigit() else None
    if teacher_id and not Teacher.query.get(teacher_id):
        return redirect_head_data_root("error", "Selected teacher was not found.", section)

    order_index = int(order_index_str) if order_index_str.isdigit() else None

    level.name = name
    level.teacher_id = teacher_id
    level.order_index = order_index
    level.zoom_email = zoom_email or None
    level.zoom_link = zoom_link or None
    level.zoom_meeting_id = zoom_meeting_id or None
    level.zoom_passcode = zoom_passcode or None
    level.homework_padlet_url = homework_padlet_url or None
    level.announcements_padlet_url = announcements_padlet_url or None

    # Keep legacy text field aligned for compatibility across existing pages/reports.
    if old_name != name:
        Student.query.filter_by(level_id=level.id).update({"level_name": name})

    db.session.commit()
    return redirect_head_data_root("success", f"Level updated: {name}.", section)


@app.route("/cleanup/levels/<int:level_id>/quick-fix", methods=["POST"])
@admin_required
def cleanup_quick_fix_level(level_id: int):
    level = Level.query.get_or_404(level_id)
    teacher_id_str = (request.form.get("teacher_id") or "").strip()
    zoom_link = (request.form.get("zoom_link") or "").strip()

    if "teacher_id" in request.form:
        teacher_id = int(teacher_id_str) if teacher_id_str.isdigit() else None
        if teacher_id and not Teacher.query.get(teacher_id):
            return redirect_head_data_root("error", f"Selected teacher was not found for {level.name}.", "cleanup")
        level.teacher_id = teacher_id
    if "zoom_link" in request.form:
        level.zoom_link = zoom_link or None

    db.session.commit()
    return redirect_head_data_root("success", f"Level cleanup updated: {level.name}.", "cleanup")


@app.route("/admin/recordings/add", methods=["POST"])
@admin_required
def admin_add_recording():
    class_id_str = (request.form.get("class_id") or "").strip()
    title = (request.form.get("title") or "").strip()
    recording_url = (request.form.get("recording_url") or "").strip()
    lesson_date_str = (request.form.get("lesson_date") or "").strip()
    summary = (request.form.get("summary") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "recordings")

    if not class_id_str.isdigit():
        return redirect_head_data_root("error", "Select a valid class for recording.", section)
    class_id = int(class_id_str)
    level = Level.query.get(class_id)
    if not level:
        return redirect_head_data_root("error", "Selected class was not found.", section)
    if not title or not recording_url or not lesson_date_str:
        return redirect_head_data_root("error", "Recording title, URL, and lesson date are required.", section)

    try:
        lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
    except ValueError:
        return redirect_head_data_root("error", "Invalid lesson date format.", section)

    db.session.add(
        ClassRecording(
            class_id=class_id,
            title=title,
            recording_url=recording_url,
            lesson_date=lesson_date,
            summary=summary or None,
        )
    )
    db.session.commit()
    return redirect_head_data_root("success", "Recording added.", section)


@app.route("/admin/recordings/<int:recording_id>/update", methods=["POST"])
@admin_required
def admin_update_recording(recording_id: int):
    recording = ClassRecording.query.get_or_404(recording_id)
    class_id_str = (request.form.get("class_id") or "").strip()
    title = (request.form.get("title") or "").strip()
    recording_url = (request.form.get("recording_url") or "").strip()
    lesson_date_str = (request.form.get("lesson_date") or "").strip()
    summary = (request.form.get("summary") or "").strip()
    section = normalize_data_root_section(request.form.get("section") or "recordings")

    if not class_id_str.isdigit():
        return redirect_head_data_root("error", "Select a valid class for recording.", section)
    class_id = int(class_id_str)
    if not Level.query.get(class_id):
        return redirect_head_data_root("error", "Selected class was not found.", section)
    if not title or not recording_url or not lesson_date_str:
        return redirect_head_data_root("error", "Recording title, URL, and lesson date are required.", section)

    try:
        lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
    except ValueError:
        return redirect_head_data_root("error", "Invalid lesson date format.", section)

    recording.class_id = class_id
    recording.title = title
    recording.recording_url = recording_url
    recording.lesson_date = lesson_date
    recording.summary = summary or None
    db.session.commit()
    return redirect_head_data_root("success", "Recording updated.", section)


@app.route("/admin/recordings/<int:recording_id>/delete", methods=["POST"])
@admin_required
def admin_delete_recording(recording_id: int):
    recording = ClassRecording.query.get_or_404(recording_id)
    section = normalize_data_root_section(request.form.get("section") or "recordings")
    db.session.delete(recording)
    db.session.commit()
    return redirect_head_data_root("success", "Recording deleted.", section)


@app.route("/students/reports/email-bulk", methods=["POST"])
@admin_required
def send_bulk_student_reports_email():
    period = (request.form.get("period") or "week").strip().lower()
    if period not in {"week", "month"}:
        period = "week"

    students = Student.query.order_by(Student.full_name).all()
    sent_count = 0
    skipped_count = 0
    failed_count = 0

    for student in students:
        parent_email = (student.parent_email or "").strip()
        if not validate_email(parent_email):
            skipped_count += 1
            continue

        report_data = build_student_report_data(student, period)
        subject = f"Hikmah Academy - Attendance Report for {student.full_name}"
        body = (
            f"Student Name: {student.full_name}\n"
            f"Level: {report_data['level_name']}\n"
            f"Teacher: {report_data['teacher_name']}\n"
            f"Period: {period}\n"
            f"Date Range: {report_data['start_date']} to {report_data['end_date']}\n\n"
            f"Present Count: {report_data['present_count']}\n"
            f"Absent Count: {report_data['absent_count']}\n"
            f"Late Count: {report_data['late_count']}\n"
            f"Excused Count: {report_data['excused_count']}\n"
            f"Attendance Percentage: {report_data['attendance_percentage']}%\n"
        )
        if period == "month":
            body += (
                f"Homework Given: {report_data['homework_total']}\n"
                f"Homework Submitted: {report_data['homework_submitted']}\n"
                f"Homework Reviewed: {report_data['homework_reviewed']}\n"
                f"Homework Waiting Review: {report_data['homework_pending_review']}\n"
                f"Homework Missing: {report_data['homework_missing']}\n"
                f"Homework Completion: {report_data['homework_completion_rate']}%\n"
                f"Monthly Teacher Note: {report_data['monthly_note_text'] or '-'}\n"
            )

        try:
            send_email_via_smtp(parent_email, subject, body)
            sent_count += 1
        except Exception:
            failed_count += 1

    status = "success" if failed_count == 0 else "error"
    message = (
        f"Bulk report completed. Sent: {sent_count}, "
        f"Skipped (missing/invalid email): {skipped_count}, Failed: {failed_count}."
    )
    log_action(
        action_type="bulk_reports_sent",
        entity_type="student_report",
        entity_label=period,
        details=message,
    )
    db.session.commit()
    return redirect(url_for("students_list", bulk_status=status, bulk_message=message))


@app.route("/students/reports/monthly-review")
@admin_required
def monthly_reports_review():
    review_data = build_monthly_report_review_rows()
    review_filter = (request.args.get("review_filter") or "all").strip().lower()
    review_rows = review_data["rows"]
    if review_filter == "ready":
        review_rows = [row for row in review_rows if row["send_ready"]]
    elif review_filter == "not_ready":
        review_rows = [row for row in review_rows if not row["send_ready"]]
    bulk_status = (request.args.get("bulk_status") or "").strip()
    bulk_message = (request.args.get("bulk_message") or "").strip()
    return render_template(
        "monthly_reports_review.html",
        review_rows=review_rows,
        summary=review_data["summary"],
        review_filter=review_filter,
        bulk_status=bulk_status,
        bulk_message=bulk_message,
    )


@app.route("/students/reports/monthly-send", methods=["POST"])
@admin_required
def send_monthly_reports_reviewed():
    selected_student_ids = {
        int(value)
        for value in request.form.getlist("selected_student_ids")
        if str(value).isdigit()
    }

    if not selected_student_ids:
        return redirect(
            url_for(
                "monthly_reports_review",
                bulk_status="error",
                bulk_message="No students were selected for sending.",
            )
        )

    students = (
        Student.query.filter(Student.id.in_(selected_student_ids))
        .order_by(Student.full_name.asc())
        .all()
    )
    sent_count = 0
    skipped_count = 0
    failed_count = 0

    for student in students:
        report_data = build_student_report_data(student, "month")
        parent_email = (student.parent_email or "").strip()
        if not validate_email(parent_email) or not (report_data["monthly_note_text"] or "").strip():
            skipped_count += 1
            continue

        subject = f"Hikmah Academy - Monthly Report for {student.full_name}"
        body = (
            f"Student Name: {student.full_name}\n"
            f"Level: {report_data['level_name']}\n"
            f"Teacher: {report_data['teacher_name']}\n"
            f"Period: month\n"
            f"Date Range: {report_data['start_date']} to {report_data['end_date']}\n\n"
            f"Present Count: {report_data['present_count']}\n"
            f"Absent Count: {report_data['absent_count']}\n"
            f"Late Count: {report_data['late_count']}\n"
            f"Excused Count: {report_data['excused_count']}\n"
            f"Attendance Percentage: {report_data['attendance_percentage']}%\n"
            f"Homework Given: {report_data['homework_total']}\n"
            f"Homework Submitted: {report_data['homework_submitted']}\n"
            f"Homework Reviewed: {report_data['homework_reviewed']}\n"
            f"Homework Waiting Review: {report_data['homework_pending_review']}\n"
            f"Homework Missing: {report_data['homework_missing']}\n"
            f"Homework Completion: {report_data['homework_completion_rate']}%\n"
            f"Monthly Teacher Note: {report_data['monthly_note_text']}\n"
        )

        try:
            send_email_via_smtp(parent_email, subject, body)
            sent_count += 1
        except Exception:
            failed_count += 1

    status = "success" if failed_count == 0 else "error"
    message = (
        f"Monthly sending completed. Sent: {sent_count}, "
        f"Skipped: {skipped_count}, Failed: {failed_count}."
    )
    log_action(
        action_type="monthly_reports_sent",
        entity_type="student_report",
        entity_label="month",
        details=message,
    )
    db.session.commit()
    return redirect(url_for("monthly_reports_review", bulk_status=status, bulk_message=message))


@app.route("/students/<int:student_id>/report")
def student_report(student_id: int):
    student = Student.query.get_or_404(student_id)
    period = (request.args.get("period") or "week").strip().lower()
    if period not in {"week", "month"}:
        period = "week"

    report_data = build_student_report_data(student, period)
    email_status = (request.args.get("email_status") or "").strip()
    email_message = (request.args.get("email_message") or "").strip()
    parent_email = (request.args.get("parent_email") or student.parent_email or "").strip()
    parent_whatsapp = (request.args.get("parent_whatsapp") or student.parent_whatsapp or "").strip()
    note_status = (request.args.get("note_status") or "").strip()
    note_message = (request.args.get("note_message") or "").strip()
    can_manage_actions = can_manage_student_report(student)

    return render_template(
        "student_report.html",
        student=student,
        level_name=report_data["level_name"],
        teacher_name=report_data["teacher_name"],
        period=period,
        start_date=report_data["start_date"],
        end_date=report_data["end_date"],
        present_count=report_data["present_count"],
        absent_count=report_data["absent_count"],
        late_count=report_data["late_count"],
        excused_count=report_data["excused_count"],
        attendance_percentage=report_data["attendance_percentage"],
        imported_arabic_attendance=report_data["imported_arabic_attendance"],
        month_key=report_data["month_key"],
        monthly_note_text=report_data["monthly_note_text"],
        email_status=email_status,
        email_message=email_message,
        parent_email=parent_email,
        parent_whatsapp=parent_whatsapp,
        note_status=note_status,
        note_message=note_message,
        can_manage_actions=can_manage_actions,
    )


@app.route("/students/<int:student_id>/report.xlsx")
def student_report_excel(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not can_manage_student_report(student):
        abort(403)
    period = (request.args.get("period") or "month").strip().lower()
    if period not in {"week", "month"}:
        period = "month"
    excel_file, file_name = build_student_full_excel_file(student, period)
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/students/<int:student_id>/report.pdf")
def student_report_pdf(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not can_manage_student_report(student):
        abort(403)
    period = (request.args.get("period") or "month").strip().lower()
    if period not in {"week", "month"}:
        period = "month"
    pdf_bytes = build_student_full_pdf(student, period)
    safe_student = re.sub(r"[^A-Za-z0-9_-]+", "_", student.full_name or "student").strip("_") or "student"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="hikmah_student_{safe_student}_{period}.pdf"'},
    )


@app.route("/students/<int:student_id>/monthly-note", methods=["POST"])
def save_student_monthly_note(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not can_manage_student_report(student):
        abort(403)
    month_key = (request.form.get("month_key") or "").strip()
    note_text = (request.form.get("note_text") or "").strip()

    if not month_key or len(month_key) != 7 or month_key[4] != "-":
        return redirect(
            url_for(
                "student_report",
                student_id=student.id,
                period="month",
                note_status="error",
                note_message="Invalid month value.",
            )
        )

    note = StudentMonthlyNote.query.filter_by(student_id=student.id, month_key=month_key).first()
    if not note:
        note = StudentMonthlyNote(student_id=student.id, month_key=month_key)
        db.session.add(note)

    note.note_text = note_text
    log_action(
        action_type="monthly_note_saved",
        entity_type="monthly_note",
        entity_id=student.id,
        entity_label=student.full_name,
        level=student.level,
        details=f"Saved monthly note for {month_key}.",
    )
    db.session.commit()

    return redirect(
        url_for(
            "student_report",
            student_id=student.id,
            period="month",
            note_status="success",
            note_message="Monthly teacher note saved.",
        )
    )


@app.route("/students/<int:student_id>/report/email", methods=["POST"])
def send_student_report_email(student_id: int):
    student = Student.query.get_or_404(student_id)
    if not can_manage_student_report(student):
        abort(403)
    period = (request.form.get("period") or "week").strip().lower()
    if period not in {"week", "month"}:
        period = "week"

    parent_email = (request.form.get("parent_email") or "").strip()
    if not validate_email(parent_email):
        return redirect(
            url_for(
                "student_report",
                student_id=student.id,
                period=period,
                email_status="error",
                email_message="Invalid parent email address.",
                parent_email=parent_email,
            )
        )

    student.parent_email = parent_email
    db.session.commit()

    report_data = build_student_report_data(student, period)
    subject = f"Hikmah Academy - Attendance Report for {student.full_name}"
    body = (
        f"Student Name: {student.full_name}\n"
        f"Level: {report_data['level_name']}\n"
        f"Teacher: {report_data['teacher_name']}\n"
        f"Period: {period}\n"
        f"Date Range: {report_data['start_date']} to {report_data['end_date']}\n\n"
        f"Present Count: {report_data['present_count']}\n"
        f"Absent Count: {report_data['absent_count']}\n"
        f"Late Count: {report_data['late_count']}\n"
        f"Excused Count: {report_data['excused_count']}\n"
        f"Attendance Percentage: {report_data['attendance_percentage']}%\n"
    )
    if period == "month":
        body += (
            f"Homework Given: {report_data['homework_total']}\n"
            f"Homework Submitted: {report_data['homework_submitted']}\n"
            f"Homework Reviewed: {report_data['homework_reviewed']}\n"
            f"Homework Waiting Review: {report_data['homework_pending_review']}\n"
            f"Homework Missing: {report_data['homework_missing']}\n"
            f"Homework Completion: {report_data['homework_completion_rate']}%\n"
            f"Monthly Teacher Note: {report_data['monthly_note_text'] or '-'}\n"
        )

    try:
        send_email_via_smtp(parent_email, subject, body)
        log_action(
            action_type="student_report_sent",
            entity_type="student_report",
            entity_id=student.id,
            entity_label=student.full_name,
            level=student.level,
            details=f"Sent {period} report to {parent_email}.",
        )
        db.session.commit()
        return redirect(
            url_for(
                "student_report",
                student_id=student.id,
                period=period,
                email_status="sent",
                email_message="Report sent successfully.",
                parent_email=parent_email,
            )
        )
    except Exception as exc:
        return redirect(
            url_for(
                "student_report",
                student_id=student.id,
                period=period,
                email_status="error",
                email_message=f"Email sending failed: {exc}",
                parent_email=parent_email,
            )
        )


@app.route("/teachers")
def teachers_list():
    teachers = Teacher.query.order_by(Teacher.full_name).all()
    return render_template("teachers.html", teachers=teachers)


@app.route("/levels")
def levels_list():
    levels = Level.query.order_by(Level.order_index.asc(), Level.name.asc()).all()
    levels_view = []
    for level in levels:
        levels_view.append(
            {
                "id": level.id,
                "name": normalize_level_display_name(level.name),
                "teacher": level.teacher,
                "zoom_link": level.zoom_link,
            }
        )
    return render_template("levels.html", levels=levels_view)


@app.route("/levels/<int:level_id>")
def level_detail(level_id: int):
    level = Level.query.get_or_404(level_id)
    display_level_name = normalize_level_display_name(level.name)
    students = Student.query.filter_by(level_id=level.id).order_by(Student.full_name).all()
    students_count = len(students)
    return render_template(
        "level_detail.html",
        level=level,
        display_level_name=display_level_name,
        students=students,
        students_count=students_count,
    )


@app.route("/levels/<int:level_id>/attendance-report")
def level_attendance_report(level_id: int):
    level = Level.query.get_or_404(level_id)
    period = (request.args.get("period") or "today").strip().lower()
    if period not in {"today", "week", "month"}:
        period = "today"

    end_date = date.today()
    if period == "today":
        start_date = end_date
    elif period == "week":
        start_date = end_date - timedelta(days=end_date.weekday())
    else:
        start_date = end_date.replace(day=1)

    students = Student.query.filter_by(level_id=level.id).order_by(Student.full_name).all()
    student_ids = [student.id for student in students]

    counts_map = {}
    if student_ids:
        rows = (
            db.session.query(
                Attendance.student_id,
                db.func.sum(db.case((Attendance.status == "Present", 1), else_=0)).label("present_count"),
                db.func.sum(db.case((Attendance.status == "Absent", 1), else_=0)).label("absent_count"),
                db.func.sum(db.case((Attendance.status == "Late", 1), else_=0)).label("late_count"),
                db.func.sum(db.case((Attendance.status == "Excused", 1), else_=0)).label("excused_count"),
                db.func.count(Attendance.id).label("total_count"),
            )
            .filter(
                Attendance.level_id == level.id,
                Attendance.student_id.in_(student_ids),
                Attendance.attendance_date >= start_date,
                Attendance.attendance_date <= end_date,
            )
            .group_by(Attendance.student_id)
            .all()
        )
        counts_map = {row.student_id: row for row in rows}

    report_rows = []
    totals = {
        "present": 0,
        "absent": 0,
        "late": 0,
        "excused": 0,
        "records": 0,
    }

    for student in students:
        row = counts_map.get(student.id)
        present_count = int(row.present_count or 0) if row else 0
        absent_count = int(row.absent_count or 0) if row else 0
        late_count = int(row.late_count or 0) if row else 0
        excused_count = int(row.excused_count or 0) if row else 0
        total_count = int(row.total_count or 0) if row else 0
        attendance_percentage = round((present_count / total_count) * 100, 1) if total_count else 0.0

        report_rows.append(
            {
                "student_name": student.full_name,
                "present_count": present_count,
                "absent_count": absent_count,
                "late_count": late_count,
                "excused_count": excused_count,
                "attendance_percentage": attendance_percentage,
            }
        )

        totals["present"] += present_count
        totals["absent"] += absent_count
        totals["late"] += late_count
        totals["excused"] += excused_count
        totals["records"] += total_count

    return render_template(
        "level_attendance_report.html",
        level=level,
        teacher_name=level.teacher.full_name if level.teacher else "-",
        period=period,
        start_date=start_date,
        end_date=end_date,
        students_count=len(students),
        report_rows=report_rows,
        totals=totals,
    )


@app.route("/monthly-report")
def monthly_report():
    total_students = Student.query.count()
    total_levels = db.session.query(Student.level_name).distinct().count()
    students_per_level = (
        db.session.query(
            Student.level_name,
            db.func.count(Student.id).label("students_count"),
        )
        .group_by(Student.level_name)
        .order_by(Student.level_name)
        .all()
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return render_template(
        "monthly_report.html",
        total_students=total_students,
        total_levels=total_levels,
        students_per_level=students_per_level,
        generated_at=generated_at,
    )


@app.route("/admin/weekly-report")
@admin_required
def supervisor_weekly_report():
    report_data = build_supervisor_weekly_report_data()
    return render_template(
        "supervisor_weekly_report.html",
        summary=report_data["summary"],
        teacher_status_rows=report_data["teacher_status_rows"],
        follow_up_teachers=report_data["follow_up_teachers"],
        on_track_teachers=report_data["on_track_teachers"],
        level_attention_rows=report_data["level_attention_rows"],
        student_attention_rows=report_data["student_attention_rows"],
    )


@app.route("/admin/weekly-report.pdf")
@admin_required
def supervisor_weekly_report_pdf():
    report_data = build_supervisor_weekly_report_data()
    pdf_bytes = build_supervisor_weekly_report_pdf(report_data)
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="hikmah_weekly_supervisor_report.pdf"'},
    )


@app.route("/admin/reports/teachers.xlsx")
@admin_required
def admin_teachers_master_excel():
    excel_file, file_name = build_teachers_master_excel_file()
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/reports/teachers.pdf")
@admin_required
def admin_teachers_master_pdf():
    pdf_bytes = build_teachers_master_pdf()
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="hikmah_teachers_master.pdf"'},
    )


@app.route("/admin/levels/<int:level_id>/follow-up.xlsx")
@admin_required
def admin_level_followup_excel(level_id: int):
    level = Level.query.get_or_404(level_id)
    excel_file, file_name = build_level_full_excel_file(level)
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/levels/<int:level_id>/follow-up.pdf")
@admin_required
def admin_level_followup_pdf(level_id: int):
    level = Level.query.get_or_404(level_id)
    pdf_bytes = build_level_full_pdf(level)
    safe_level_name = re.sub(r"[^A-Za-z0-9_-]+", "_", level.name or "level").strip("_") or "level"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="hikmah_level_{safe_level_name}.pdf"'},
    )


@app.route("/admin/weekly-report/archive/<path:file_name>")
@admin_required
def supervisor_weekly_report_archive_file(file_name: str):
    safe_name = os.path.basename(file_name)
    file_path = weekly_report_archive_path(safe_name)
    if not os.path.exists(file_path):
        abort(404)
    return send_file(file_path, as_attachment=True, download_name=safe_name)


@app.route("/attendance", methods=["GET", "POST"])
def attendance_page():
    is_admin = bool(session.get("is_admin"))
    teacher = get_current_teacher()
    if not is_admin and not teacher:
        return redirect(url_for("teacher_login", next=request.path))

    levels_query = Level.query.order_by(Level.order_index.asc(), Level.name.asc())
    if not is_admin and teacher:
        levels_query = levels_query.filter(Level.teacher_id == teacher.id)
    levels = levels_query.all()
    allowed_level_ids = {level.id for level in levels}

    today_str = date.today().isoformat()

    if request.method == "POST":
        selected_level_id_str = (request.form.get("level_id") or "").strip()
        attendance_date_str = request.form.get("attendance_date") or today_str
    else:
        selected_level_id_str = (request.args.get("level_id") or "").strip()
        attendance_date_str = request.args.get("attendance_date") or today_str

    selected_level_id = int(selected_level_id_str) if selected_level_id_str.isdigit() else None
    saved = request.args.get("saved") == "1"
    error_message = None
    students = []
    existing_statuses = {}
    attendance_summary = {
        "present": 0,
        "absent": 0,
        "late": 0,
        "excused": 0,
        "present_rate": 0.0,
    }
    selected_level = Level.query.get(selected_level_id) if selected_level_id else None
    if selected_level and selected_level.id not in allowed_level_ids:
        selected_level = None
        error_message = "You are not allowed to access this level."

    attendance_date = None
    try:
        attendance_date = datetime.strptime(attendance_date_str, "%Y-%m-%d").date()
    except ValueError:
        error_message = "Invalid date format. Please use YYYY-MM-DD."

    if selected_level:
        students = Student.query.filter_by(level_id=selected_level.id).order_by(Student.full_name).all()

    if attendance_date and students:
        existing_records = (
            Attendance.query.filter(
                Attendance.attendance_date == attendance_date,
                Attendance.student_id.in_([student.id for student in students]),
            ).all()
        )
        existing_statuses = {record.student_id: record.status for record in existing_records}
        for status in existing_statuses.values():
            normalized = (status or "").strip().lower()
            if normalized == "present":
                attendance_summary["present"] += 1
            elif normalized == "absent":
                attendance_summary["absent"] += 1
            elif normalized == "late":
                attendance_summary["late"] += 1
            elif normalized == "excused":
                attendance_summary["excused"] += 1
        attendance_summary["present_rate"] = (
            round((attendance_summary["present"] / len(students)) * 100, 1)
            if students
            else 0.0
        )

    if request.method == "POST" and not error_message:
        if not selected_level:
            error_message = "Please choose a level first."
        elif not students:
            error_message = "No students found for the selected level."
        else:
            status_counts = {status: 0 for status in ATTENDANCE_STATUSES}
            for student in students:
                status = request.form.get(f"status_{student.id}")
                if status not in ATTENDANCE_STATUSES:
                    status = "Present"
                status_counts[status] += 1

                record = Attendance.query.filter_by(
                    student_id=student.id,
                    attendance_date=attendance_date,
                ).first()

                if record:
                    record.status = status
                    record.level_id = selected_level.id
                else:
                    db.session.add(
                        Attendance(
                            student_id=student.id,
                            level_id=selected_level.id,
                            attendance_date=attendance_date,
                            status=status,
                        )
                    )

            log_action(
                action_type="attendance_saved",
                entity_type="attendance",
                entity_label=selected_level.name,
                level=selected_level,
                details=(
                    f"Saved attendance for {attendance_date.isoformat()} "
                    f"(Present: {status_counts['Present']}, Absent: {status_counts['Absent']}, "
                    f"Late: {status_counts['Late']}, Excused: {status_counts['Excused']})."
                ),
            )
            db.session.commit()
            return redirect(
                url_for(
                    "attendance_page",
                    level_id=selected_level.id,
                    attendance_date=attendance_date_str,
                    saved=1,
                )
            )

    return render_template(
        "attendance.html",
        levels=levels,
        selected_level=selected_level,
        attendance_date=attendance_date_str,
        students=students,
        statuses=ATTENDANCE_STATUSES,
        existing_statuses=existing_statuses,
        attendance_summary=attendance_summary,
        saved=saved,
        error_message=error_message,
    )


@app.route("/attendance/list")
def attendance_list():
    is_admin = bool(session.get("is_admin"))
    teacher = get_current_teacher()
    if not is_admin and not teacher:
        return redirect(url_for("teacher_login", next=request.path))

    records_query = (
        db.session.query(Attendance, Student, Level)
        .join(Student, Attendance.student_id == Student.id)
        .outerjoin(Level, Attendance.level_id == Level.id)
    )
    if not is_admin and teacher:
        records_query = records_query.filter(Level.teacher_id == teacher.id)

    records = records_query.order_by(Attendance.attendance_date.desc(), Student.full_name.asc()).all()
    return render_template("attendance_list.html", records=records)


if __name__ == "__main__":
    debug_mode = os.getenv("FLASK_DEBUG", "").strip().lower() in {"1", "true", "yes", "on"}
    if not debug_mode:
        with app.app_context():
            readiness_issues = get_runtime_readiness_issues()
        if readiness_issues:
            raise RuntimeError(
                "Production readiness check failed:\n- " + "\n- ".join(readiness_issues)
            )

    app.run(
        host=os.getenv("FLASK_HOST", "127.0.0.1"),
        port=int(os.getenv("FLASK_PORT", "5000")),
        debug=debug_mode,
    )
